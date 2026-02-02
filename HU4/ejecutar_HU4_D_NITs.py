#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
================================================================================
MODULO: ejecutar_HU4_D_NITs.py
================================================================================

Descripcion General:
--------------------
    Este modulo ejecuta el Punto D del proceso HU4 - Validacion de NITs contra
    un listado maestro de proveedores. Lee NITs desde un archivo Excel y los
    valida contra el Stored Procedure [CxP].[HU4_D_NITs].

Autor: Diego Ivan Lopez Ochoa
Version: 1.0.0
Fecha de Creacion: 2025

Stored Procedure:
-----------------
    [CxP].[HU4_D_NITs]
    
    Parametros:
        @DiasMaximos INT - Dias maximos para filtrar registros
        @BatchSize INT - Tamano del lote de procesamiento
        @ListadoNITS NVARCHAR(MAX) - Lista de NITs separados por coma

Dependencias:
-------------
    - asyncio: Ejecucion asincrona
    - os: Manejo de rutas de archivo
    - pyodbc: Conexion a SQL Server
    - openpyxl: Lectura de archivos Excel
    - json, ast: Parseo de configuracion
    - traceback: Captura de errores
    - struct: Manejo de DATETIMEOFFSET
    - datetime: Conversion de fechas

================================================================================
DIAGRAMA DE FLUJO
================================================================================

    +-------------------------------------------------------------+
    |                        INICIO                               |
    |               ejecutar_HU4_D_NITs()                         |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  reset_vars() - Inicializar variables                       |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  Parsear configuracion:                                     |
    |  - ServidorBaseDatos, NombreBaseDatos                       |
    |  - RutaInsumoNitMaestros + NombreArchivoNitsMaestros        |
    |  - DiasMaximosPuntoD, BatchSizePuntoD                       |
    +-------------------------+-----------------------------------+
                              |
              +---------------+---------------+
              |    ¿Archivo Excel existe?     |
              +---------------+---------------+
                     |                |
                     | NO             | SI
                     v                v
    +------------------------+   +--------------------------------+
    |  FileNotFoundError     |   |  Cargar workbook con openpyxl  |
    |  return False, None    |   |  load_workbook(data_only=True) |
    +------------------------+   +-----------------+--------------+
                                                   |
                                                   v
    +-------------------------------------------------------------+
    |  Buscar hoja "SIN MANDATORIOS"                              |
    |  ¿Existe la hoja?                                           |
    +-------------------------+-----------------------------------+
                     |                |
                     | NO             | SI
                     v                v
    +------------------------+   +--------------------------------+
    |  ValueError            |   |  Leer NITs de columna A        |
    |  "Hoja no existe"      |   |  (fila 2 en adelante)          |
    +------------------------+   +-----------------+--------------+
                                                   |
                                                   v
    +-------------------------------------------------------------+
    |  Procesar NITs:                                             |
    |  - Limpiar espacios (strip)                                 |
    |  - Eliminar duplicados (set)                                |
    |  - Ordenar (sorted)                                         |
    |  - Unir con comas (join)                                    |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  Conectar a SQL Server:                                     |
    |  - ODBC Driver 17                                           |
    |  - Trusted_Connection=yes                                   |
    |  - Registrar converter para DATETIMEOFFSET (-155)           |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  EXEC [CxP].[HU4_D_NITs]                                    |
    |  @DiasMaximos, @BatchSize, @ListadoNITS                     |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  ResultSet 1: RESUMEN                                       |
    |  - TotalCandidatos, TotalIDsProcesados                      |
    |  - TotalIDsConNITEnListado                                  |
    |  - TotalUpdatesDocumentsProcessing                          |
    |  - TotalUpdatesComparativa_* |
    |  - TotalGruposNITFactura, TotalIDsSinFactura                |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  ResultSet 2: DETALLE (opcional)                            |
    |  - Contar filas totales                                     |
    |  - Contar filas con matched=1                               |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  build_sp_summary_from_row()                                |
    |  Generar resumen con todas las metricas                     |
    +-------------------------+-----------------------------------+
                              |
                              v
    +-------------------------------------------------------------+
    |  SetVar("vLocStrResultadoSP", ok)                           |
    |  SetVar("vLocStrResumenSP", resumen_txt)                    |
    |  return ok, payload                                         |
    |                          FIN                                |
    +-------------------------------------------------------------+

================================================================================
VARIABLES DE ENTRADA/SALIDA
================================================================================

Variables de Entrada (GetVar):
------------------------------
    vLocDicConfig : dict o str
        {
            "ServidorBaseDatos": "SERVIDOR\\INSTANCIA",
            "NombreBaseDatos": "CxP_Database",
            "RutaInsumoNitMaestros": "C:\\Insumos\\NITs",
            "NombreArchivoNitsMaestros": "NITs_Maestros.xlsx",
            "DiasMaximosPuntoD": 120,
            "BatchSizePuntoD": 500
        }

Variables de Salida (SetVar):
-----------------------------
    vLocStrResultadoSP : bool
        True si exito, False si error.
        
    vLocStrResumenSP : str
        Ejemplo:
        "OK | CantidadNITsLeidos=150 | TotalCandidatos=5000 | 
         TotalProcesados=4800 | TotalCoincidenciasDeNIT=3500 | ..."
        
    vGblStrMensajeError : str
        Mensaje de error (vacio si exito).
        
    vGblStrSystemError : str
        Stack trace (vacio si exito).

================================================================================
ESTRUCTURA DEL ARCHIVO EXCEL
================================================================================

Archivo: NITs_Maestros.xlsx
Hoja: "SIN MANDATORIOS"

    +-------------+
    |      A      |
    +-------------+
    |    NIT      |  <- Fila 1 (encabezado, se ignora)
    +-------------+
    |  900123456  |  <- Fila 2 (primer NIT)
    |  800987654  |  <- Fila 3
    |  900111222  |  <- Fila 4
    |     ...     |
    +-------------+

================================================================================
"""



async def ejecutar_HU4_D_NITs():
    """
    Ejecuta Punto D - Validacion NITs contra el SP [CxP].[HU4_D_NITs].
    
    Lee NITs desde un archivo Excel (hoja "SIN MANDATORIOS"), los procesa
    para eliminar duplicados, y ejecuta el SP de validacion contra la
    base de datos de CxP.
    
    Returns:
        tuple: (bool, dict|None)
            - bool: True si ejecucion exitosa, False si error
            - dict: Payload con resumen y conteos, o None si error
                {
                    "resumen": {...},
                    "detalle_rows": int,
                    "detalle_matched_rows": int
                }
    
    Side Effects:
        Lee:
            - vLocDicConfig
            - Archivo Excel de NITs maestros
        Escribe:
            - vLocStrResultadoSP
            - vLocStrResumenSP
            - vGblStrMensajeError
            - vGblStrSystemError
    
    Example:
        Configuracion::
        
            SetVar("vLocDicConfig", {
                "ServidorBaseDatos": "SQLPROD\\CXP",
                "NombreBaseDatos": "CuentasPorPagar",
                "RutaInsumoNitMaestros": "C:\\\\Insumos",
                "NombreArchivoNitsMaestros": "NITs_Proveedores.xlsx",
                "DiasMaximosPuntoD": 90,
                "BatchSizePuntoD": 1000
            })
        
        Ejecucion::
        
            ok, payload = await ejecutar_HU4_D_NITs()
            
            if ok:
                print(f"NITs procesados: {payload['resumen']['TotalIDsProcesados']}")
                print(f"Coincidencias: {payload['resumen']['TotalIDsConNITEnListado']}")
    
    Note:
        - El archivo Excel debe tener una hoja llamada "SIN MANDATORIOS"
        - Los NITs se leen de la columna A, comenzando en fila 2
        - Se eliminan duplicados automaticamente
        - Se maneja DATETIMEOFFSET con converter personalizado
    """
    # ==========================================================================
    # IMPORTS
    # ==========================================================================
    import asyncio
    import os
    import pyodbc
    import openpyxl
    import json
    import ast
    import traceback
    import struct
    from datetime import datetime, timedelta, timezone

    # ==========================================================================
    # FUNCIONES AUXILIARES
    # ==========================================================================
    
    def safe_str(v):
        """Convierte valor a string de forma segura."""
        try:
            return "" if v is None else str(v)
        except Exception:
            return ""

    def safe_int(v, default=0):
        """Convierte valor a entero con default."""
        try:
            if v is None:
                return default
            return int(v)
        except Exception:
            return default

    def reset_vars():
        """Inicializa variables de salida."""
        try:
            SetVar("vGblStrMensajeError", "")
            SetVar("vGblStrSystemError", "")
            SetVar("vLocStrResultadoSP", "")
            SetVar("vLocStrResumenSP", "")
        except Exception:
            pass

    def set_error_vars(user_msg, exc=None):
        """Establece variables de error."""
        try:
            SetVar("vGblStrMensajeError", safe_str(user_msg))
            SetVar("vGblStrSystemError", "" if exc is None else traceback.format_exc())
            SetVar("vLocStrResultadoSP", False)
            SetVar("vLocStrResumenSP", "")
        except Exception:
            pass

    def parse_config(raw):
        """Parsea configuracion desde JSON o literal."""
        if isinstance(raw, dict):
            return raw
        text = safe_str(raw).strip()
        if not text:
            raise ValueError("vLocDicConfig vacio")
        try:
            return json.loads(text)
        except Exception:
            return ast.literal_eval(text)

    def fetch_resultset_as_dict(cursor):
        """
        Obtiene una fila del cursor como diccionario.
        
        Args:
            cursor: Cursor de pyodbc.
        
        Returns:
            dict: Fila como diccionario o {} si no hay datos.
        """
        cols = [c[0] for c in cursor.description] if cursor.description else []
        r = cursor.fetchone()
        if not r:
            return {}
        return {cols[i]: r[i] for i in range(len(cols))}

    def build_sp_summary_from_row(row_dict, total_nits, detalle_rows=0, detalle_matched=0):
        """
        Construye resumen de ejecucion del SP.
        
        Args:
            row_dict: Diccionario con datos del ResultSet 1.
            total_nits: Cantidad de NITs unicos leidos del Excel.
            detalle_rows: Total de filas en ResultSet 2.
            detalle_matched: Filas con matched=1 en ResultSet 2.
        
        Returns:
            str: Resumen formateado.
        
        Example:
            >>> build_sp_summary_from_row(
            ...     {"TotalCandidatos": 100, "TotalIDsProcesados": 95, ...},
            ...     total_nits=50,
            ...     detalle_rows=95,
            ...     detalle_matched=80
            ... )
            "OK | CantidadNITsLeidos=50 | TotalCandidatos=100 | ..."
        """
        candidatos = safe_int(row_dict.get("TotalCandidatos"))
        procesados = safe_int(row_dict.get("TotalIDsProcesados"))
        coincidencias = safe_int(row_dict.get("TotalIDsConNITEnListado"))
        actualizaciones_dp = safe_int(row_dict.get("TotalUpdatesDocumentsProcessing"))
        actualizaciones_comparativa_observaciones = safe_int(
            row_dict.get("TotalUpdatesComparativa_Observaciones")
        )
        actualizaciones_comparativa_estado = safe_int(
            row_dict.get("TotalUpdatesComparativa_EstadoValidacion")
        )
        grupos_nit_factura = safe_int(row_dict.get("TotalGruposNITFactura"))
        sin_factura = safe_int(row_dict.get("TotalIDsSinFactura"))

        return (
            f"OK | CantidadNITsLeidos={int(total_nits)} | "
            f"TotalCandidatos={candidatos} | TotalProcesados={procesados} | "
            f"TotalCoincidenciasDeNIT={coincidencias} | "
            f"TotalActualizacionesEnDocumentsProcessing={actualizaciones_dp} | "
            f"TotalActualizacionesEnComparativaObservaciones={actualizaciones_comparativa_observaciones} | "
            f"TotalActualizacionesEnComparativaEstadoValidacion={actualizaciones_comparativa_estado} | "
            f"TotalGruposPorNITYFactura={grupos_nit_factura} | "
            f"TotalRegistrosSinFactura={sin_factura} | "
            f"TotalFilasEnDetalle={int(detalle_rows)} | "
            f"TotalFilasCoincidentesEnDetalle={int(detalle_matched)}"
        )

    # ==========================================================================
    # INICIO
    # ==========================================================================
    reset_vars()

    # ==========================================================================
    # CONFIGURACION
    # ==========================================================================
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        servidor = safe_str(cfg["ServidorBaseDatos"]).replace("\\\\", "\\")
        db = safe_str(cfg["NombreBaseDatos"])
        
        # Construir ruta completa del archivo Excel
        ruta = os.path.join(
            safe_str(cfg["RutaInsumoNitMaestros"]),
            safe_str(cfg["NombreArchivoNitsMaestros"])
        )
        
        dias_maximos = int(cfg.get("DiasMaximosPuntoD", 120))
        batch_size = int(cfg.get("BatchSizePuntoD", 500))
        
    except Exception as e:
        set_error_vars("Error configuracion Punto D NITs", e)
        try:
            SetVar("vLocStrResultadoSP", False)
            SetVar("vLocStrResumenSP", "ERROR | configuracion | Punto D NITs | Ver vGblStrSystemError")
        except Exception:
            pass
        return False, None

    # ==========================================================================
    # EJECUCION SP (SINCRONA EN THREAD)
    # ==========================================================================
    def run_sp_sync():
        """
        Ejecuta el SP de forma sincrona.
        
        1. Valida existencia del archivo Excel
        2. Lee NITs de la hoja "SIN MANDATORIOS"
        3. Elimina duplicados y construye lista CSV
        4. Conecta a SQL Server con converter para DATETIMEOFFSET
        5. Ejecuta SP y procesa ResultSets
        
        Returns:
            tuple: (bool, dict, int)
                - bool: Exito
                - dict: Payload con resumen
                - int: Total de NITs unicos leidos
        """
        # Validar existencia del archivo
        if not os.path.exists(ruta):
            raise FileNotFoundError(f"Archivo NITs no existe: {ruta}")

        # Cargar workbook
        wb = openpyxl.load_workbook(ruta, data_only=True)
        
        # Validar existencia de la hoja
        if "SIN MANDATORIOS" not in wb.sheetnames:
            raise ValueError("Hoja 'SIN MANDATORIOS' no existe en el Excel")

        # Leer NITs de columna A, fila 2 en adelante
        ws = wb["SIN MANDATORIOS"]
        nits = []
        for r in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if r and r[0]:
                nit_txt = str(r[0]).strip()
                if nit_txt:
                    nits.append(nit_txt)

        # Eliminar duplicados y ordenar
        nits_unicos = sorted(set(nits))
        
        # Crear lista CSV para el SP
        lista = ",".join(nits_unicos)

        # Connection string
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={servidor};"
            f"DATABASE={db};"
            "Trusted_Connection=yes;"
        )

        def handle_datetimeoffset(dto_value):
            """
            Converter para DATETIMEOFFSET (ODBC type -155).
            
            Convierte bytes de DATETIMEOFFSET a objeto datetime de Python.
            
            Estructura del valor:
                - 6 shorts: year, month, day, hour, minute, second
                - 1 int: nanoseconds
                - 2 shorts: offset_hours, offset_minutes
            
            Args:
                dto_value: Bytes del valor DATETIMEOFFSET.
            
            Returns:
                datetime: Objeto datetime con timezone.
            """
            tup = struct.unpack("<6hI2h", dto_value)
            return datetime(
                tup[0], tup[1], tup[2], tup[3], tup[4], tup[5],
                tup[6] // 1000,  # Convertir nanosegundos a microsegundos
                timezone(timedelta(hours=tup[7], minutes=tup[8]))
            )

        with pyodbc.connect(conn_str) as c:
            c.autocommit = True

            # Registrar converter para DATETIMEOFFSET ANTES de ejecutar
            c.add_output_converter(-155, handle_datetimeoffset)

            cur = c.cursor()
            cur.execute(
                "EXEC [CxP].[HU4_D_NITs] @DiasMaximos=?, @BatchSize=?, @ListadoNITS=?",
                dias_maximos, batch_size, lista
            )

            # ResultSet 1: Resumen (1 fila)
            resumen_row = fetch_resultset_as_dict(cur)

            # ResultSet 2: Detalle (contar filas y matches)
            detalle_count = 0
            detalle_matched = 0

            if cur.nextset() and cur.description:
                cols = [c[0] for c in cur.description]
                
                # Buscar indice de columna "matched"
                idx_matched = None
                for i, name in enumerate(cols):
                    if str(name).lower() == "matched":
                        idx_matched = i
                        break

                # Contar filas y matches
                for row in cur:
                    detalle_count += 1
                    if idx_matched is not None:
                        try:
                            if row[idx_matched] in (1, True, "1"):
                                detalle_matched += 1
                        except Exception:
                            pass

            payload = {
                "resumen": resumen_row,
                "detalle_rows": detalle_count,
                "detalle_matched_rows": detalle_matched
            }
            return True, payload, len(nits_unicos)

    # ==========================================================================
    # WRAPPER ASYNC + SETVARS
    # ==========================================================================
    try:
        loop = asyncio.get_running_loop()
        ok, payload, total_nits = await loop.run_in_executor(None, run_sp_sync)

        resumen_row = payload.get("resumen", {}) if isinstance(payload, dict) else {}
        resumen_txt = build_sp_summary_from_row(
            resumen_row,
            total_nits,
            detalle_rows=payload.get("detalle_rows", 0),
            detalle_matched=payload.get("detalle_matched_rows", 0),
        )

        try:
            SetVar("vLocStrResultadoSP", ok)
            SetVar("vLocStrResumenSP", resumen_txt)
        except Exception:
            pass

        return ok, payload

    except Exception as e:
        set_error_vars("Error ejecucion Punto D NITs", e)
        try:
            SetVar("vLocStrResultadoSP", False)
            SetVar("vLocStrResumenSP", "ERROR | ejecucion | Punto D NITs | Ver vGblStrSystemError")
        except Exception:
            pass
        return False, None


# ==============================================================================
# EJEMPLOS DE USO
# ==============================================================================
"""
EJEMPLO 1: Ejecucion basica
---------------------------
    # Archivo Excel: C:\\Insumos\\NITs_Proveedores.xlsx
    # Hoja: SIN MANDATORIOS
    # Columna A: 900123456, 800654321, 900111222, ...
    
    SetVar("vLocDicConfig", {
        "ServidorBaseDatos": "SQLPROD\\CXP",
        "NombreBaseDatos": "CuentasPorPagar",
        "RutaInsumoNitMaestros": "C:\\\\Insumos",
        "NombreArchivoNitsMaestros": "NITs_Proveedores.xlsx",
        "DiasMaximosPuntoD": 90,
        "BatchSizePuntoD": 1000
    })
    
    ok, payload = await ejecutar_HU4_D_NITs()
    
    if ok:
        print(f"Resumen: {GetVar('vLocStrResumenSP')}")
        # Output:
        # "OK | CantidadNITsLeidos=150 | TotalCandidatos=5000 | 
        #  TotalProcesados=4800 | TotalCoincidenciasDeNIT=3500 | ..."

EJEMPLO 2: Verificar coincidencias
----------------------------------
    ok, payload = await ejecutar_HU4_D_NITs()
    
    if ok:
        coincidencias = payload['resumen']['TotalIDsConNITEnListado']
        procesados = payload['resumen']['TotalIDsProcesados']
        
        porcentaje = (coincidencias / procesados * 100) if procesados > 0 else 0
        print(f"Porcentaje de coincidencia: {porcentaje:.2f}%")
        
        # Verificar filas del detalle
        print(f"Filas en detalle: {payload['detalle_rows']}")
        print(f"Filas con match: {payload['detalle_matched_rows']}")

EJEMPLO 3: Manejo de errores comunes
------------------------------------
    # Error: Archivo no existe
    SetVar("vLocDicConfig", {
        ...
        "NombreArchivoNitsMaestros": "archivo_inexistente.xlsx"
    })
    
    ok, _ = await ejecutar_HU4_D_NITs()
    # ok = False
    # GetVar("vGblStrMensajeError") = "Error ejecucion Punto D NITs"
    # GetVar("vGblStrSystemError") contiene "FileNotFoundError: ..."
    
    # Error: Hoja no existe
    # Si el archivo existe pero no tiene hoja "SIN MANDATORIOS"
    # GetVar("vGblStrSystemError") contiene "ValueError: Hoja 'SIN MANDATORIOS' no existe"
"""
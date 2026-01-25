# -*- coding: utf-8 -*-

import asyncio


async def generar_ruta_logs():
    import json
    import os
    from datetime import datetime
    import ast

    def safe_str(v):
        try:
            if v is None:
                return ""
            if isinstance(v, str):
                return v
            if isinstance(v, (bytes, bytearray)):
                return v.decode("utf-8", errors="replace")
            return str(v)
        except Exception:
            return ""

    def parse_config(raw):
        if isinstance(raw, dict):
            return raw
        t = safe_str(raw).strip()
        if not t:
            raise ValueError("vLocDicConfig vacio")
        try:
            return json.loads(t)
        except Exception:
            return ast.literal_eval(t)

    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        ruta_base = cfg.get("RutaLogs", "")
    except Exception as e:
        print("Error leyendo configuracion:", e)
        return

    fecha = datetime.now()
    ano = fecha.strftime("%Y")
    mes = fecha.strftime("%m")
    dia = fecha.strftime("%d")

    ruta_final = os.path.join(ruta_base, ano, mes, dia)

    try:
        os.makedirs(ruta_final, exist_ok=True)
    except Exception as e:
        print("No se pudo crear la ruta:", e)

    SetVar("vGblStrRutaLogs", ruta_final + os.sep)

def buscarCandidatos():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from itertools import combinations
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # NUEVA FUNCIÓN: Truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))

        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )
        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def tabla_existe(cx, schema, tabla):
        query = """
            SELECT COUNT(*) 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        """
        cur = cx.cursor()
        cur.execute(query, (schema, tabla))
        count = cur.fetchone()[0]
        cur.close()
        return count > 0
    
    def crear_tabla_candidatos(cx, df_muestra):
        col_defs = []
        for col in df_muestra.columns:
            col_safe = "[" + col + "]"
            col_defs.append(col_safe + " NVARCHAR(MAX)")
        
        create_sql = (
            "CREATE TABLE [CxP].[HU41_CandidatosValidacion] (\n    " +
            ",\n    ".join(col_defs) +
            "\n)"
        )
        
        print("[DEBUG] Tabla [CxP].[HU41_CandidatosValidacion] creada - TODAS las columnas NVARCHAR(MAX)")
        
        cur = cx.cursor()
        cur.execute(create_sql)
        cx.commit()
        cur.close()
    
    def insertar_candidatos(cx, df):
        if df.empty:
            return
        
        try:
            columns = df.columns.tolist()
            
            placeholders = []
            for col in columns:
                placeholders.append("CAST(? AS NVARCHAR(MAX))")
            
            placeholders_str = ','.join(placeholders)
            columns_str = ','.join(['[' + col + ']' for col in columns])
            
            insert_sql = (
                "INSERT INTO [CxP].[HU41_CandidatosValidacion] (" +
                columns_str + ") VALUES (" + placeholders_str + ")"
            )
            
            print("[DEBUG] Preparando insercion de " + str(len(df)) + " registros")
            print("[DEBUG] TODAS las columnas se insertan como NVARCHAR(MAX)")
            
            cur = cx.cursor()
            
            rows_inserted = 0
            for idx, row in df.iterrows():
                try:
                    values = []
                    for col in columns:
                        val = row[col]
                        if pd.isna(val):
                            values.append(None)
                        else:
                            values.append(safe_str(val))
                    
                    cur.execute(insert_sql, tuple(values))
                    rows_inserted += 1
                    
                    if rows_inserted % 50 == 0:
                        print("[DEBUG] Insertados " + str(rows_inserted) + " registros...")
                        
                except Exception as e_row:
                    print("[ERROR] Error en fila " + str(idx) + ": " + str(e_row))
                    print("[DEBUG] Primeros 5 valores:")
                    for i, col in enumerate(columns[:5]):
                        v = str(row[col])[:50] if not pd.isna(row[col]) else "NULL"
                        print("  " + col + ": " + v)
                    raise
            
            cx.commit()
            cur.close()
            print("[DEBUG] Commit exitoso. Filas insertadas: " + str(rows_inserted))
            
        except Exception as e:
            print("[ERROR] Error insertando candidatos: " + str(e))
            cx.rollback()
            raise
    
    def unir_valores(df, columna):
        if df.empty or columna not in df.columns:
            return ""
        try:
            vals = df[columna].values
            str_vals = [safe_str(v) for v in vals]
            return '|'.join(str_vals)
        except:
            return ""
    
    def buscar_combinacion(valores, objetivo, cantidad, tolerancia=500):
        try:
            n = len(valores)
            if n < cantidad or cantidad <= 0:
                return None
            if cantidad == 1:
                for i, v in enumerate(valores):
                    if abs(v - objetivo) <= tolerancia:
                        return [i]
                return None
            valores_sorted = sorted(valores)
            suma_min = sum(valores_sorted[:cantidad])
            suma_max = sum(valores_sorted[-cantidad:])
            if objetivo < suma_min - tolerancia or objetivo > suma_max + tolerancia:
                return None
            if abs(objetivo - suma_min) <= tolerancia:
                return [valores.index(v) for v in valores_sorted[:cantidad]]
            if abs(objetivo - suma_max) <= tolerancia:
                return [valores.index(v) for v in valores_sorted[-cantidad:]]
            promedio = objetivo / cantidad
            indexed = [(i, v, abs(v - promedio)) for i, v in enumerate(valores)]
            indexed_sorted = sorted(indexed, key=lambda x: x[2])
            indices_cercanos = [x[0] for x in indexed_sorted[:min(cantidad*2, n)]]
            if len(indices_cercanos) >= cantidad:
                for combo_idx in combinations(range(len(indices_cercanos)), cantidad):
                    indices = [indices_cercanos[i] for i in combo_idx]
                    suma = sum(valores[i] for i in indices)
                    if abs(suma - objetivo) <= tolerancia:
                        return indices
            indexed_vals = [(i, v) for i, v in enumerate(valores)]
            for combo in combinations(indexed_vals, cantidad):
                indices = [c[0] for c in combo]
                suma = sum(c[1] for c in combo)
                if abs(suma - objetivo) <= tolerancia:
                    return indices
            return None
        except:
            return None
    
    # CORRECCIÓN: batch_update ahora trunca observaciones
    def batch_update(cx, query, params_list, max_retries=3):
        if not params_list:
            return
        for attempt in range(max_retries):
            try:
                cur = cx.cursor()
                cur.fast_executemany = True
                for params in params_list:
                    # CORRECCIÓN: Truncar observaciones (segundo parámetro)
                    safe_params = list(params)
                    if len(safe_params) >= 2:
                        # Truncar el parámetro de observaciones (índice 1)
                        safe_params[1] = truncar_observacion(safe_params[1])
                    
                    # Convertir a tupla después de truncar
                    safe_params_tuple = tuple(
                        safe_str(p) if isinstance(p, str) else p 
                        for p in safe_params
                    )
                    cur.execute(query, safe_params_tuple)
                cx.commit()
                cur.close()
                return
            except pyodbc.Error:
                cx.rollback()
                if attempt < max_retries - 1:
                    time.sleep(0.5 * (attempt + 1))
                    continue
                raise
    
    def crear_candidato(row_dp, df_ddp, df_hoc):
        candidato = {}
        for col, val in row_dp.items():
            candidato[col + "_dp"] = val
        for col in df_ddp.columns:
            candidato[col + "_ddp"] = unir_valores(df_ddp, col)
        for col in df_hoc.columns:
            candidato[col + "_hoc"] = unir_valores(df_hoc, col)
        candidato["indices_ddp"] = '|'.join(map(str, df_ddp.index.tolist()))
        candidato["indices_hoc"] = '|'.join(map(str, df_hoc.index.tolist()))
        return candidato
    
    def read_sql_safe(query, cx):
        try:
            return pd.read_sql(query, cx)
        except UnicodeDecodeError:
            cur = cx.cursor()
            cur.execute(query)
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]
            data = []
            for row in rows:
                safe_row = []
                for val in row:
                    if isinstance(val, str):
                        safe_row.append(val)
                    elif isinstance(val, bytes):
                        safe_row.append(safe_str(val))
                    else:
                        safe_row.append(val)
                data.append(safe_row)
            cur.close()
            return pd.DataFrame(data, columns=columns)
    
    stats = {
        "total": 0, "candidatos": 0, "sin_oc": 0,
        "no_encontrados": 0, "sin_combinacion": 0, "errores": 0,
        "tiempo_carga": 0, "tiempo_procesamiento": 0,
        "tiempo_updates": 0, "tiempo_tabla": 0, "tiempo_total": 0
    }
    
    updates_dp = []
    updates_comp = []
    candidatos_lista = []
    t_inicio = time.time()
    
    print("="*80)
    print("[INICIO] Funcion buscarCandidatos() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("="*80)
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        print("[DEBUG] Servidor: " + cfg.get('ServidorBaseDatos', 'NO DEFINIDO'))
        print("[DEBUG] Base de datos: " + cfg.get('NombreBaseDatos', 'NO DEFINIDO'))
        tolerancia = cfg.get("Tolerancia", 500)
        max_retries = cfg.get("MaxRetries", 3)
        
        t_carga = time.time()
        
        with crear_conexion_db(cfg, max_retries) as cx:
            
            q1 = (
                "SELECT * FROM [CxP].[DocumentsProcessing] WITH (NOLOCK) "
                "WHERE documenttype = 'FV' "
                "AND (EstadoFinalFase_5 IS NULL OR LTRIM(RTRIM(EstadoFinalFase_5)) = '') "
                "AND (ResultadoFinalAntesEventos IS NULL OR ResultadoFinalAntesEventos NOT IN ("
                "'APROBADO','APROBADO CONTADO Y/O EVENTO MANUAL','APROBADO SIN CONTABILIZACION',"
                "'RECHAZADO - RETORNADO','RECLASIFICAR','EXCLUIDO IMPORTACIONES',"
                "'EXCLUIDO COSTO INDIRECTO FLETES','EXCLUIDO GRANOS','EXCLUIDO MAIZ',"
                "'NO EXITOSO','RECHAZADO','CON NOVEDAD'))"
            )
            
            q2 = "SELECT * FROM [CxP].[DocumentsDetailProcessing] WITH (NOLOCK)"
            
            q3 = (
                "SELECT * FROM [CxP].[HistoricoOrdenesCompra] WITH (NOLOCK) "
                "WHERE Marca IS NULL OR Marca != 'PROCESADO'"
            )
            
            df_dp = read_sql_safe(q1, cx)
            df_ddp = read_sql_safe(q2, cx)
            df_hoc = read_sql_safe(q3, cx)
            
            stats["total"] = len(df_dp)
            stats["tiempo_carga"] = time.time() - t_carga
            
            if not df_ddp.empty:
                df_ddp.set_index(['nit_emisor_o_nit_del_proveedor', 'numero_de_factura'],
                                inplace=True, drop=False)
                df_ddp.sort_index(inplace=True)
            if not df_hoc.empty:
                df_hoc.set_index(['NitCedula', 'DocCompra'], inplace=True, drop=False)
                df_hoc.sort_index(inplace=True)
            
            t_proc = time.time()
            
            for idx, row in df_dp.iterrows():
                try:
                    nit = safe_str(row["nit_emisor_o_nit_del_proveedor"])
                    factura = safe_str(row["numero_de_factura"])
                    oc = row.get("numero_de_liquidacion_u_orden_de_compra")
                    forma_pago = row.get("valor_a_pagar")
                    oc_str = safe_str(oc)
                    
                    if not oc_str:
                        stats["sin_oc"] += 1
                        updates_dp.append((
                            "VALIDACION DATOS DE FACTURACION: Exitoso",
                            "Registro NO cuenta con Orden de compra",
                            "CON NOVEDAD", nit, factura
                        ))
                        updates_comp.append((
                            "SIN ORDEN DE COMPRA", "CON NOVEDAD",
                            "Registro NO cuenta con Orden de compra", nit, factura
                        ))
                        continue
                    
                    try:
                        hoc = (df_hoc.loc[[(nit, oc_str)]].copy().reset_index(drop=True)
                               if (nit, oc_str) in df_hoc.index else pd.DataFrame())
                    except:
                        hoc = pd.DataFrame()
                    
                    try:
                        ddp = (df_ddp.loc[[(nit, factura)]].copy().reset_index(drop=True)
                               if (nit, factura) in df_ddp.index else pd.DataFrame())
                    except:
                        ddp = pd.DataFrame()
                    
                    if hoc.empty or ddp.empty:
                        stats["no_encontrados"] += 1
                        estado = "EN ESPERA - CONTADO" if forma_pago in (1,"1","01") else "EN ESPERA"
                        obs = "No se encuentra registro en historico"
                        updates_dp.append((
                            "VALIDACION DATOS DE FACTURACION: Exitoso", obs, estado, nit, factura
                        ))
                        updates_comp.append(("LLAVES NO ENCONTRADAS", estado, obs, nit, factura))
                        continue
                    
                    cant_hoc, cant_ddp = len(hoc), len(ddp)
                    
                    if cant_hoc <= cant_ddp:
                        candidatos_lista.append(crear_candidato(row, ddp, hoc))
                        stats["candidatos"] += 1
                    else:
                        suma_lea = ddp["Valor de la Compra LEA"].sum()
                        valores_hoc = hoc["PorCalcular"].tolist()
                        combo = buscar_combinacion(valores_hoc, suma_lea, cant_ddp, tolerancia)
                        
                        if combo is not None:
                            hoc_sel = hoc.iloc[combo].copy()
                            candidatos_lista.append(crear_candidato(row, ddp, hoc_sel))
                            stats["candidatos"] += 1
                        else:
                            stats["sin_combinacion"] += 1
                            estado = "EN ESPERA - CONTADO" if forma_pago in (1,"1","01") else "EN ESPERA"
                            obs = "No se encuentra combinacion valida"
                            updates_dp.append((
                                "VALIDACION DATOS DE FACTURACION: Exitoso", obs, estado, nit, factura
                            ))
                            updates_comp.append(("LLAVES NO ENCONTRADAS", estado, obs, nit, factura))
                
                except Exception as e:
                    stats["errores"] += 1
                    continue
            
            stats["tiempo_procesamiento"] = time.time() - t_proc
            t_updates = time.time()
            
            # CORRECCIÓN: batch_update ahora trunca automáticamente
            if updates_dp:
                batch_update(cx,
                    "UPDATE [CxP].[DocumentsProcessing] SET EstadoFinalFase_4=?, "
                    "ObservacionesFase_4=?, ResultadoFinalAntesEventos=? "
                    "WHERE nit_emisor_o_nit_del_proveedor=? AND numero_de_factura=?",
                    updates_dp, max_retries)
            
            if updates_comp:
                # Para Comparativa, el tercer parámetro es Valor_XML
                params_comp_truncado = []
                for params in updates_comp:
                    params_lista = list(params)
                    if len(params_lista) >= 3:
                        params_lista[2] = truncar_observacion(params_lista[2])
                    params_comp_truncado.append(tuple(params_lista))
                
                batch_update(cx,
                    "UPDATE [dbo].[CxP.Comparativa] SET Orden_de_Compra=?, "
                    "Estado_validacion_antes_de_eventos=?, "
                    "Valor_XML=CASE WHEN Item='Observaciones' THEN ? ELSE Valor_XML END "
                    "WHERE NIT=? AND Factura=?",
                    params_comp_truncado, max_retries)
            
            stats["tiempo_updates"] = time.time() - t_updates
            
            if candidatos_lista:
                df_candidatos = pd.DataFrame(candidatos_lista)
                df_candidatos = df_candidatos.where(pd.notna(df_candidatos), None)
            else:
                df_candidatos = pd.DataFrame()
            
            t_tabla = time.time()
            
            if not df_candidatos.empty:
                print("[DEBUG] Iniciando proceso de tabla SQL...")
                print("[DEBUG] Candidatos a insertar: " + str(len(df_candidatos)))
                
                existe = tabla_existe(cx, "CxP", "HU41_CandidatosValidacion")
                print("[DEBUG] Tabla existe: " + str(existe))
                
                if existe:
                    print("[DEBUG] Borrando tabla existente para recrear...")
                    cur = cx.cursor()
                    cur.execute("DROP TABLE [CxP].[HU41_CandidatosValidacion]")
                    cx.commit()
                    cur.close()
                    print("[DEBUG] Tabla borrada OK")
                
                print("[DEBUG] Creando tabla con estructura actual...")
                crear_tabla_candidatos(cx, df_candidatos)
                print("[DEBUG] Tabla creada OK")
                
                print("[DEBUG] Insertando " + str(len(df_candidatos)) + " registros...")
                insertar_candidatos(cx, df_candidatos)
                print("[DEBUG] Registros insertados OK")
            else:
                print("[DEBUG] DataFrame de candidatos VACIO - no se insertara nada")
            
            stats["tiempo_tabla"] = time.time() - t_tabla
            print("[DEBUG] Tiempo tabla: " + str(stats["tiempo_tabla"]) + "s")
        
        stats["tiempo_total"] = time.time() - t_inicio
        
        msg = ("Done. Total:" + str(stats['total']) + " Candidatos:" + str(stats['candidatos']) +
               " SinOC:" + str(stats['sin_oc']) + " NoEnc:" + str(stats['no_encontrados']) +
               " SinComb:" + str(stats['sin_combinacion']) + " Err:" + str(stats['errores']) +
               " Time:" + str(round(stats['tiempo_total'], 2)) + "s")
        
        print("[RESULTADO] " + msg)
        print("[ESTADISTICAS DETALLADAS]")
        print("  Total procesados: " + str(stats['total']))
        print("  Candidatos validos: " + str(stats['candidatos']))
        print("  Sin orden de compra: " + str(stats['sin_oc']))
        print("  No encontrados: " + str(stats['no_encontrados']))
        print("  Sin combinacion: " + str(stats['sin_combinacion']))
        print("  Errores: " + str(stats['errores']))
        print("  Tiempo total: " + str(stats['tiempo_total']) + "s")
        
        print("="*80)
        print("[EXITO] Funcion completada exitosamente")
        print("="*80)
        
        SetVar("vLocDfCandidatosJson", df_candidatos.to_json(orient="records"))
        SetVar("vLocDicEstadisticas", json.dumps(stats))
        
        SetVar("vGblStrDetalleError", "")
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        
        SetVar("vLocStrResultadoSP", True)
        SetVar("vLocStrResumenSP", msg)
        
        print("[DEBUG] Variables Rocketbot configuradas:")
        print("  vLocStrResultadoSP = True")
        print("  vLocStrResumenSP = " + msg)
        
        return (True, msg, df_candidatos, stats)
    
    except Exception as e:
        stats["tiempo_total"] = time.time() - t_inicio
        err = "Error: " + str(e)
        
        print("="*80)
        print("[ERROR CRITICO] La funcion fallo")
        print("="*80)
        print("[ERROR] Tipo de error: " + type(e).__name__)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("="*80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        
        SetVar("vLocStrResultadoSP", False)
        
        return (False, err, None, stats)

def ZPRE_ValidarCOP():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPRE_ValidarCOP() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    def actualizar_items_comparativa(id_reg, cx, nit, factura, nombre_item, valores_lista, actualizar_valor_xml=False, valor_xml=None,actualizar_aprobado=False, valor_aprobado=None):
        cur = cx.cursor()
        
        query_count = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(query_count, (nit, factura, nombre_item))
        count_actual = cur.fetchone()[0]
        
        count_necesario = len(valores_lista)
        
        if count_actual == 0:
            for i, valor in enumerate(valores_lista):
                insert_query = """
                INSERT INTO [dbo].[CxP.Comparativa] (
                    ID_registro, NIT, Factura, Item, Valor_Orden_de_Compra, Valor_XML, Aprobado
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                vxml = valor_xml if actualizar_valor_xml else None
                vaprob = valor_aprobado if actualizar_aprobado else None
                cur.execute(insert_query, (id_reg, nit, factura, nombre_item, valor, vxml, vaprob))
        
        elif count_actual < count_necesario:
            for i in range(count_actual):
                update_query = "UPDATE [dbo].[CxP.Comparativa] SET Valor_Orden_de_Compra = ?"
                params = [valores_lista[i]]
                
                if actualizar_valor_xml:
                    update_query += ", Valor_XML = ?"
                    params.append(valor_xml)
                
                if actualizar_aprobado:
                    update_query += ", Aprobado = ?"
                    params.append(valor_aprobado)
                
                # CORRECCIÓN: Eliminar TOP 1 (incompatible con OFFSET/FETCH)
                update_query += """
                WHERE NIT = ? AND Factura = ? AND Item = ?
                  AND ID_registro IN (
                    SELECT ID_registro FROM [dbo].[CxP.Comparativa]
                    WHERE NIT = ? AND Factura = ? AND Item = ?
                    ORDER BY ID_registro
                    OFFSET ? ROWS FETCH NEXT 1 ROWS ONLY
                  )
                """
                params.extend([nit, factura, nombre_item, nit, factura, nombre_item, i])
                cur.execute(update_query, params)
            
            for i in range(count_actual, count_necesario):
                insert_query = """
                INSERT INTO [dbo].[CxP.Comparativa] (
                    ID_registro, NIT, Factura, Item, Valor_Orden_de_Compra, Valor_XML, Aprobado
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                vxml = valor_xml if actualizar_valor_xml else None
                vaprob = valor_aprobado if actualizar_aprobado else None
                cur.execute(insert_query, (id_reg, nit, factura, nombre_item, valores_lista[i], vxml, vaprob))
        
        else:
            for i, valor in enumerate(valores_lista):
                update_query = "UPDATE [dbo].[CxP.Comparativa] SET Valor_Orden_de_Compra = ?"
                params = [valor]
                
                if actualizar_valor_xml:
                    update_query += ", Valor_XML = ?"
                    params.append(valor_xml)
                
                if actualizar_aprobado:
                    update_query += ", Aprobado = ?"
                    params.append(valor_aprobado)
                
                # CORRECCIÓN: Eliminar TOP 1 (incompatible con OFFSET/FETCH)
                update_query += """
                WHERE NIT = ? AND Factura = ? AND Item = ?
                  AND ID_registro IN (
                    SELECT ID_registro FROM [dbo].[CxP.Comparativa]
                    WHERE NIT = ? AND Factura = ? AND Item = ?
                    ORDER BY ID_registro
                    OFFSET ? ROWS FETCH NEXT 1 ROWS ONLY
                  )
                """
                params.extend([nit, factura, nombre_item, nit, factura, nombre_item, i])
                cur.execute(update_query, params)
        
        cur.close()
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def convertir_a_numero(valor_str):
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return 0.0
        try:
            v = str(valor_str).strip()
            v = v.replace(',', '')
            return float(v)
        except:
            print("[WARNING] No se pudo convertir '" + str(valor_str) + "' a numero, usando 0.0")
            return 0.0
    
    def sumar_valores(valor_str, nombre_campo="campo"):
        valores = split_valores(valor_str)
        suma = 0.0
        errores = []
        
        for i, v in enumerate(valores):
            try:
                v_limpio = v.strip()
                v_limpio = v_limpio.replace(',', '')
                valor_num = float(v_limpio)
                suma += valor_num
                
                if i < 3:
                    print("[DEBUG] " + nombre_campo + "[" + str(i) + "]: '" + v + "' -> " + str(valor_num))
                    
            except ValueError as e:
                errores.append("Valor[" + str(i) + "]='" + v + "' no convertible: " + str(e))
            except Exception as e:
                errores.append("Error[" + str(i) + "]='" + v + "': " + str(e))
        
        if errores:
            print("[WARNING] Errores en conversion de " + nombre_campo + ":")
            for err in errores[:5]:
                print("  " + err)
        
        print("[DEBUG] " + nombre_campo + " - Total valores: " + str(len(valores)) + " | Suma: " + str(suma))
        
        return suma
    
    def contiene_valor(campo, valor_buscado):
        valores = split_valores(campo)
        return valor_buscado in valores
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        print("[DEBUG] Servidor: " + cfg.get('ServidorBaseDatos', 'N/A'))
        print("[DEBUG] Base de datos: " + cfg.get('NombreBaseDatos', 'N/A'))
        
        tolerancia = float(cfg.get('Tolerancia', 500))
        print("[DEBUG] Tolerancia configurada: " + str(tolerancia))
        
        stats = {
            'total_registros': 0,
            'aprobados': 0,
            'con_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                nombre_emisor_dp,
                ClaseDePedido_hoc,
                PorCalcular_hoc,
                [Valor de la Compra LEA_ddp],
                Posicion_hoc,
                TipoNif_hoc,
                Acreedor_hoc,
                FecDoc_hoc,
                FecReg_hoc,
                FecContGasto_hoc,
                IndicadorImpuestos_hoc,
                TextoBreve_hoc,
                ClaseDeImpuesto_hoc,
                Cuenta_hoc,
                CiudadProveedor_hoc,
                DocFiEntrada_hoc,
                Cuenta26_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                Moneda_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            print("[PASO 2] Aplicando filtros...")
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPRE') or contiene_valor(x, '45') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPRE o 45: " + str(mask_clase.sum()))
            
            def es_cop_o_vacio(valor):
                if pd.isna(valor) or valor == "":
                    return True
                valores = split_valores(valor)
                return 'COP' in valores or len(valores) == 0
            
            mask_cop = df_candidatos['Moneda_hoc'].apply(es_cop_o_vacio)
            
            print("[DEBUG] Registros con Moneda COP/vacio: " + str(mask_cop.sum()))
            
            df_filtrado = df_candidatos[mask_clase & mask_cop].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPRE/45 y COP")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            print("")
            print("[PASO 3] Procesando registros...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "]")
                    
                    id_reg = safe_str(row['ID_dp'])
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    print("")
                    
                    print("[CALCULO] Sumando valores de PorCalcular_hoc...")
                    suma_por_calcular = sumar_valores(row['PorCalcular_hoc'], "PorCalcular_hoc")
                    
                    print("[CALCULO] Sumando valores de Valor de la Compra LEA_ddp...")
                    suma_valor_compra = sumar_valores(row['Valor de la Compra LEA_ddp'], "Valor_Compra_LEA")
                    
                    diferencia = abs(suma_por_calcular - suma_valor_compra)
                    print("")
                    
                    print("[DEBUG] Suma PorCalcular: " + str(suma_por_calcular))
                    print("[DEBUG] Suma Valor Compra: " + str(suma_valor_compra))
                    print("[DEBUG] Diferencia: " + str(diferencia))
                    print("[DEBUG] Tolerancia: " + str(tolerancia))
                    
                    if diferencia <= tolerancia:
                        print("[RESULTADO] APROBADO (diferencia <= " + str(tolerancia) + ")")
                        stats['aprobados'] += 1
                        
                        print("[UPDATE] Actualizando tabla CxP.Comparativa...")
                        
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        valores_posicion = split_valores(row['Posicion_hoc'])
                        valores_tiponif = split_valores(row['TipoNif_hoc'])
                        valores_acreedor = split_valores(row['Acreedor_hoc'])
                        valores_fecdoc = split_valores(row['FecDoc_hoc'])
                        valores_fecreg = split_valores(row['FecReg_hoc'])
                        valores_feccontgasto = split_valores(row['FecContGasto_hoc'])
                        valores_indicadorimpuestos = split_valores(row['IndicadorImpuestos_hoc'])
                        valores_textobreve = split_valores(row['TextoBreve_hoc'])
                        valores_clasedeimpuesto = split_valores(row['ClaseDeImpuesto_hoc'])
                        valores_cuenta = split_valores(row['Cuenta_hoc'])
                        valores_ciudadproveedor = split_valores(row['CiudadProveedor_hoc'])
                        valores_docfientrada = split_valores(row['DocFiEntrada_hoc'])
                        valores_cuenta26 = split_valores(row['Cuenta26_hoc'])
                        
                        actualizar_items_comparativa(
                            id_reg, cx, nit, factura,
                            'LineExtensionAmount',
                            valores_porcalcular,
                            actualizar_valor_xml=True,
                            valor_xml=str(suma_valor_compra),
                            actualizar_aprobado=True,
                            valor_aprobado='SI'
                        )
                        
                        if valores_posicion:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'Posicion', valores_posicion)
                        
                        if valores_tiponif:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'TipoNIF', valores_tiponif)
                        
                        if valores_acreedor:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'Acreedor', valores_acreedor)
                        
                        if valores_fecdoc:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'FecDoc', valores_fecdoc)
                        
                        if valores_fecreg:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'FecReg', valores_fecreg)
                        
                        if valores_feccontgasto:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'FecContGasto', valores_feccontgasto)
                        
                        if valores_indicadorimpuestos:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'IndicadorImpuestos', valores_indicadorimpuestos)
                        
                        if valores_textobreve:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'TextoBreve', valores_textobreve)
                        
                        if valores_clasedeimpuesto:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'ClaseDeImpuesto', valores_clasedeimpuesto)
                        
                        if valores_cuenta:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'Cuenta', valores_cuenta)
                        
                        if valores_ciudadproveedor:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'CiudadProveedor', valores_ciudadproveedor)
                        
                        if valores_docfientrada:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'DocFIEntrada', valores_docfientrada)
                        
                        if valores_cuenta26:
                            actualizar_items_comparativa(id_reg, cx, nit, factura, 'Cuenta26', valores_cuenta26)
                        
                        print("[UPDATE] Tabla CxP.Comparativa actualizada OK")
                        
                    else:
                        print("[RESULTADO] CON NOVEDAD (diferencia > " + str(tolerancia) + ")")
                        stats['con_novedad'] += 1
                        
                        if forma_pago == '1' or forma_pago == '01':
                            estado_final = 'CON NOVEDAD - CONTADO'
                        else:
                            estado_final = 'CON NOVEDAD'
                        
                        print("[DEBUG] Estado final: " + estado_final)
                        
                        cur = cx.cursor()
                        
                        print("[UPDATE] Actualizando tabla DocumentsProcessing...")
                        
                        update_fase4 = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_fase4, (nit, factura, oc))
                        
                        select_obs = """
                        SELECT ObservacionesFase_4
                        FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(select_obs, (nit, factura, oc))
                        result_obs = cur.fetchone()
                        
                        obs_actual = safe_str(result_obs[0]) if result_obs and result_obs[0] else ""
                        nueva_obs = "No se encuentra coincidencia del Valor a pagar de la factura"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        update_obs = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_obs, (obs_final, nit, factura, oc))
                        
                        update_resultado = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_resultado, (estado_final, nit, factura, oc))
                        
                        print("[DEBUG] DocumentsProcessing actualizado OK")
                        
                        print("[UPDATE] Actualizando tabla CxP.Comparativa...")
                        
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        
                        actualizar_items_comparativa(
                            id_reg, cx, nit, factura,
                            'LineExtensionAmount',
                            valores_porcalcular,
                            actualizar_valor_xml=True,
                            valor_xml=str(suma_valor_compra),
                            actualizar_aprobado=True,
                            valor_aprobado='NO'
                        )
                        
                        cur = cx.cursor()
                        select_obs_comp = """
                        SELECT Valor_XML
                        FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(select_obs_comp, (nit, factura))
                        result_obs_comp = cur.fetchone()
                        
                        obs_comp_actual = safe_str(result_obs_comp[0]) if result_obs_comp and result_obs_comp[0] else ""
                        nueva_obs_comp = "No se encuentra coincidencia del Valor a pagar de la factura"
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs_comp + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs_comp
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        update_obs_comp = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(update_obs_comp, (obs_comp_final, nit, factura))
                        
                        update_estado_todos = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ?
                          AND Factura = ?
                        """
                        cur.execute(update_estado_todos, (estado_final, nit, factura))
                        cur.close()
                        
                        print("[DEBUG] CxP.Comparativa actualizado OK")
                        
                        print("[UPDATE] Actualizando tabla HistoricoOrdenesCompra...")
                        
                        valores_doccompra = split_valores(row['DocCompra_hoc'])
                        valores_nitcedula = split_valores(row['NitCedula_hoc'])
                        valores_porcalcular_hoc = split_valores(row['PorCalcular_hoc'])
                        valores_textobreve = split_valores(row['TextoBreve_hoc'])
                        
                        cur = cx.cursor()
                        num_actualizados = 0
                        for i in range(max(len(valores_doccompra), len(valores_nitcedula), 
                                          len(valores_porcalcular_hoc), len(valores_textobreve))):
                            
                            doccompra_val = valores_doccompra[i] if i < len(valores_doccompra) else ""
                            nitcedula_val = valores_nitcedula[i] if i < len(valores_nitcedula) else ""
                            porcalcular_val = valores_porcalcular_hoc[i] if i < len(valores_porcalcular_hoc) else ""
                            textobreve_val = valores_textobreve[i] if i < len(valores_textobreve) else ""
                            
                            if doccompra_val and nitcedula_val:
                                update_marca = """
                                UPDATE [CxP].[HistoricoOrdenesCompra]
                                SET Marca = 'PROCESADO'
                                WHERE DocCompra = ?
                                  AND NitCedula = ?
                                  AND PorCalcular = ?
                                  AND TextoBreve = ?
                                """
                                cur.execute(update_marca, (doccompra_val, nitcedula_val, porcalcular_val, textobreve_val))
                                num_actualizados += 1
                        
                        print("[DEBUG] HistoricoOrdenesCompra actualizado: " + str(num_actualizados) + " registros")
                        
                        cur.close()
                        print("[UPDATE] Todas las tablas actualizadas OK")
                        
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Aprobados: " + str(stats['aprobados']))
            print("  Con novedad: " + str(stats['con_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = "Proceso OK. Total:" + str(stats['total_registros']) + " Aprobados:" + str(stats['aprobados']) + " ConNovedad:" + str(stats['con_novedad'])
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            SetVar("vGblStrDetalleError", "")
            SetVar("vGblStrSystemError", "")
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ZPRE_ValidarUSD():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPRE_ValidarUSD() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    def actualizar_items_comparativa(id_reg, cx, nit, factura, nombre_item, valores_lista,actualizar_valor_xml=False, valor_xml=None,actualizar_aprobado=False, valor_aprobado=None):
        cur = cx.cursor()
        
        query_count = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(query_count, (nit, factura, nombre_item))
        count_actual = cur.fetchone()[0]
        
        count_necesario = len(valores_lista)
        
        if count_actual == 0:
            for i, valor in enumerate(valores_lista):
                insert_query = """
                INSERT INTO [dbo].[CxP.Comparativa] (
                   ID_registro, NIT, Factura, Item, Valor_Orden_de_Compra, Valor_XML, Aprobado
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                vxml = valor_xml if actualizar_valor_xml else None
                vaprob = valor_aprobado if actualizar_aprobado else None
                cur.execute(insert_query, (id_reg, nit, factura, nombre_item, valor, vxml, vaprob))
        
        elif count_actual < count_necesario:
            for i in range(count_actual):
                update_query = "UPDATE [dbo].[CxP.Comparativa] SET Valor_Orden_de_Compra = ?"
                params = [valores_lista[i]]
                
                if actualizar_valor_xml:
                    update_query += ", Valor_XML = ?"
                    params.append(valor_xml)
                
                if actualizar_aprobado:
                    update_query += ", Aprobado = ?"
                    params.append(valor_aprobado)
                
                update_query += """
                WHERE NIT = ? AND Factura = ? AND Item = ?
                  AND ID_registro IN (
                    SELECT ID_registro FROM [dbo].[CxP.Comparativa]
                    WHERE NIT = ? AND Factura = ? AND Item = ?
                    ORDER BY ID_registro
                    OFFSET ? ROWS FETCH NEXT 1 ROWS ONLY
                  )
                """
                params.extend([nit, factura, nombre_item, nit, factura, nombre_item, i])
                cur.execute(update_query, params)
            
            for i in range(count_actual, count_necesario):
                insert_query = """
                INSERT INTO [dbo].[CxP.Comparativa] (
                    ID_registro, NIT, Factura, Item, Valor_Orden_de_Compra, Valor_XML, Aprobado
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """
                vxml = valor_xml if actualizar_valor_xml else None
                vaprob = valor_aprobado if actualizar_aprobado else None
                cur.execute(insert_query, (id_reg, nit, factura, nombre_item, valores_lista[i], vxml, vaprob))
        
        else:
            for i, valor in enumerate(valores_lista):
                update_query = "UPDATE [dbo].[CxP.Comparativa] SET Valor_Orden_de_Compra = ?"
                params = [valor]
                
                if actualizar_valor_xml:
                    update_query += ", Valor_XML = ?"
                    params.append(valor_xml)
                
                if actualizar_aprobado:
                    update_query += ", Aprobado = ?"
                    params.append(valor_aprobado)
                
                update_query += """
                WHERE NIT = ? AND Factura = ? AND Item = ?
                  AND id IN (
                    SELECT ID_registro FROM [dbo].[CxP.Comparativa]
                    WHERE NIT = ? AND Factura = ? AND Item = ?
                    ORDER BY id
                    OFFSET ? ROWS FETCH NEXT 1 ROWS ONLY
                  )
                """
                params.extend([nit, factura, nombre_item, nit, factura, nombre_item, i])
                cur.execute(update_query, params)
        
        cur.close()
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def sumar_valores(valor_str, nombre_campo="campo"):
        valores = split_valores(valor_str)
        suma = 0.0
        for i, v in enumerate(valores):
            try:
                v_limpio = v.strip().replace(',', '')
                suma += float(v_limpio)
            except:
                pass
        print("[DEBUG] " + nombre_campo + " suma: " + str(suma))
        return suma
    
    def contiene_valor(campo, valor_buscado):
        valores = split_valores(campo)
        return valor_buscado in valores
    
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        tolerancia = float(cfg.get('Tolerancia', 500))
        stats = {'total_registros': 0, 'aprobados': 0, 'con_novedad': 0, 'errores': 0}
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            query_candidatos = """
            SELECT ID_dp, nit_emisor_o_nit_del_proveedor_dp, numero_de_factura_dp,
                   numero_de_liquidacion_u_orden_de_compra_dp, forma_de_pago_dp,
                   ClaseDePedido_hoc, PorCalcular_hoc, VlrPagarCop_dp,
                   Posicion_hoc, TipoNif_hoc, Acreedor_hoc, FecDoc_hoc, FecReg_hoc,
                   FecContGasto_hoc, IndicadorImpuestos_hoc, TextoBreve_hoc,
                   ClaseDeImpuesto_hoc, Cuenta_hoc, CiudadProveedor_hoc,
                   DocFiEntrada_hoc, Cuenta26_hoc, DocCompra_hoc, NitCedula_hoc, Moneda_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            """
            df_candidatos = pd.read_sql(query_candidatos, cx)
            
            if df_candidatos.empty:
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros")
                return True, "No hay registros", None, stats
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPRE') or contiene_valor(x, '45') if pd.notna(x) else False
            )
            mask_usd = df_candidatos['Moneda_hoc'].apply(
                lambda x: contiene_valor(x, 'USD') if pd.notna(x) else False
            )
            df_filtrado = df_candidatos[mask_clase & mask_usd].copy()
            
            if df_filtrado.empty:
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros USD")
                return True, "No hay registros USD", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            for idx, row in df_filtrado.iterrows():
                try:
                    id_reg = safe_str(row['ID_dp'])
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    suma_por_calcular = sumar_valores(row['PorCalcular_hoc'], "PorCalcular")
                    vlr_pagar_cop = sumar_valores(row['VlrPagarCop_dp'], "VlrPagarCop")
                    diferencia = abs(suma_por_calcular - vlr_pagar_cop)
                    
                    if diferencia <= tolerancia:
                        stats['aprobados'] += 1
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        actualizar_items_comparativa(id_reg, cx, nit, factura, 'LineExtensionAmount',
                                                   valores_porcalcular, True, str(vlr_pagar_cop), True, 'SI')
                    else:
                        stats['con_novedad'] += 1
                        estado_final = 'CON NOVEDAD - CONTADO' if forma_pago in ('1', '01') else 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        
                        cur.execute("""
                        SELECT ObservacionesFase_4 FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        result_obs = cur.fetchone()
                        obs_actual = safe_str(result_obs[0]) if result_obs and result_obs[0] else ""
                        nueva_obs = "No se encuentra coincidencia del Valor a pagar COP de la factura"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing] SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (obs_final, nit, factura, oc))
                        
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing] SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (estado_final, nit, factura, oc))
                        
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        actualizar_items_comparativa(id_reg, cx, nit, factura, 'LineExtensionAmount',
                                                   valores_porcalcular, True, str(vlr_pagar_cop), True, 'NO')
                        
                        cur.execute("""
                        SELECT Valor_XML FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (nit, factura))
                        result_obs_comp = cur.fetchone()
                        obs_comp_actual = safe_str(result_obs_comp[0]) if result_obs_comp and result_obs_comp[0] else ""
                        nueva_obs_comp = "No se encuentra coincidencia del Valor a pagar COP de la factura"
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs_comp + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs_comp
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        cur.execute("""
                        UPDATE [dbo].[CxP.Comparativa] SET Valor_XML = ?
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (obs_comp_final, nit, factura))
                        
                        cur.execute("""
                        UPDATE [dbo].[CxP.Comparativa] SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ? AND Factura = ?
                        """, (estado_final, nit, factura))
                        cur.close()
                
                except Exception as e_row:
                    print("[ERROR] Registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
            
            stats['tiempo_total'] = time.time() - t_inicio
            msg = "OK. Total:" + str(stats['total_registros']) + " Aprobados:" + str(stats['aprobados'])
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vGblStrDetalleError", "")
            return True, msg, None, stats
    
    except Exception as e:
        print("[ERROR] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPRE_ValidarTRM():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            return json.loads(text)
        except:
            return ast.literal_eval(text)
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                break
            except:
                if attempt < max_retries - 1:
                    time.sleep(1)
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
        except:
            if cx:
                cx.rollback()
            raise
        finally:
            if cx:
                cx.close()
    
    def split_valores(valor_str):
        if not valor_str or pd.isna(valor_str):
            return []
        return [v.strip() for v in str(valor_str).split('|') if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        return valor_buscado in split_valores(campo)
    
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        tolerancia_trm = float(cfg.get('ToleranciaTRM', 10))
        stats = {'total': 0, 'aprobados': 0, 'con_novedad': 0}
        
        with crear_conexion_db(cfg) as cx:
            query = """
            SELECT nit_emisor_o_nit_del_proveedor_dp, numero_de_factura_dp,
                   numero_de_liquidacion_u_orden_de_compra_dp, forma_de_pago_dp,
                   ClaseDePedido_hoc, CalculationRate_dp, TRM_hoc, Moneda_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            """
            df = pd.read_sql(query, cx)
            
            if df.empty:
                SetVar("vLocStrResultadoSP", "True")
                return True, "No hay registros", None, stats
            
            mask_clase = df['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPRE') or contiene_valor(x, '45') if pd.notna(x) else False
            )
            mask_usd = df['Moneda_hoc'].apply(lambda x: contiene_valor(x, 'USD') if pd.notna(x) else False)
            df_filtrado = df[mask_clase & mask_usd].copy()
            
            stats['total'] = len(df_filtrado)
            
            for idx, row in df_filtrado.iterrows():
                try:
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    trm_dp = float(row['CalculationRate_dp']) if pd.notna(row['CalculationRate_dp']) else 0
                    trms_hoc = split_valores(row['TRM_hoc'])
                    
                    trm_match = False
                    for trm_str in trms_hoc:
                        try:
                            trm_hoc = float(trm_str.replace(',', ''))
                            if abs(trm_dp - trm_hoc) <= tolerancia_trm:
                                trm_match = True
                                break
                        except:
                            pass
                    
                    if trm_match:
                        stats['aprobados'] += 1
                    else:
                        stats['con_novedad'] += 1
                        estado_final = 'CON NOVEDAD - CONTADO' if forma_pago in ('1', '01') else 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        
                        cur.execute("""
                        SELECT ObservacionesFase_4 FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        result = cur.fetchone()
                        obs_actual = safe_str(result[0]) if result and result[0] else ""
                        nueva_obs = "No se encuentra coincidencia de TRM"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing] SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (obs_final, nit, factura, oc))
                        
                        cur.execute("""
                        UPDATE [CxP].[DocumentsProcessing] SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (estado_final, nit, factura, oc))
                        
                        cur.execute("""
                        SELECT Valor_XML FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (nit, factura))
                        result_comp = cur.fetchone()
                        obs_comp_actual = safe_str(result_comp[0]) if result_comp and result_comp[0] else ""
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        cur.execute("""
                        UPDATE [dbo].[CxP.Comparativa] SET Valor_XML = ?
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (obs_comp_final, nit, factura))
                        
                        cur.execute("""
                        UPDATE [dbo].[CxP.Comparativa] SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ? AND Factura = ?
                        """, (estado_final, nit, factura))
                        cur.close()
                
                except Exception as e:
                    print("[ERROR] " + str(e))
            
            msg = "OK. Total:" + str(stats['total']) + " Aprobados:" + str(stats['aprobados'])
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            return True, msg, None, stats
    
    except Exception as e:
        print(traceback.format_exc())
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPRE_ValidarCantidadPrecio():
    import json, ast, traceback, pyodbc, pandas as pd, numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time, warnings
    warnings.filterwarnings('ignore')
    
    def safe_str(v):
        if v is None: return ""
        if isinstance(v, str): return v.strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v): return ""
            return str(v)
        try: return str(v).strip()
        except: return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs: return ""
        obs_str = safe_str(obs)
        return obs_str[:3900] if len(obs_str) > 3900 else obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict): return raw
        try: return json.loads(safe_str(raw))
        except: return ast.literal_eval(safe_str(raw))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    
    def split_valores(v):
        if not v or pd.isna(v): return []
        return [x.strip() for x in str(v).split('|') if x.strip()]
    
    def contiene(campo, val):
        return val in split_valores(campo)
    
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        tol = float(cfg.get('Tolerancia', 500))
        stats = {'total': 0, 'aprobados': 0, 'con_novedad': 0}
        
        with crear_conexion_db(cfg) as cx:
            df = pd.read_sql("""
            SELECT nit_emisor_o_nit_del_proveedor_dp, numero_de_factura_dp,
                   numero_de_liquidacion_u_orden_de_compra_dp, forma_de_pago_dp,
                   ClaseDePedido_hoc, PrecioUnit_hoc, CantProd_hoc, PorCalcular_hoc,
                   [Precio Unitario del producto_ddp], [Cantidad de producto_ddp]
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            """, cx)
            
            if df.empty:
                SetVar("vLocStrResultadoSP", "True")
                return True, "No hay registros", None, stats
            
            mask = df['ClaseDePedido_hoc'].apply(
                lambda x: contiene(x, 'ZPRE') or contiene(x, '45') if pd.notna(x) else False
            )
            df = df[mask].copy()
            stats['total'] = len(df)
            
            for idx, row in df.iterrows():
                try:
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    precios_hoc = [float(x.replace(',', '')) for x in split_valores(row['PrecioUnit_hoc'])]
                    cantidades_hoc = [float(x.replace(',', '')) for x in split_valores(row['CantProd_hoc'])]
                    precios_ddp = [float(x.replace(',', '')) for x in split_valores(row['Precio Unitario del producto_ddp'])]
                    cantidades_ddp = [float(x.replace(',', '')) for x in split_valores(row['Cantidad de producto_ddp'])]
                    
                    precio_ok = all(abs(precios_hoc[i] - precios_ddp[i]) <= tol for i in range(min(len(precios_hoc), len(precios_ddp))))
                    cantidad_ok = all(abs(cantidades_hoc[i] - cantidades_ddp[i]) <= tol for i in range(min(len(cantidades_hoc), len(cantidades_ddp))))
                    
                    if precio_ok and cantidad_ok:
                        stats['aprobados'] += 1
                    else:
                        stats['con_novedad'] += 1
                        estado = 'CON NOVEDAD - CONTADO' if forma_pago in ('1', '01') else 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        
                        cur.execute("""SELECT ObservacionesFase_4 FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        r = cur.fetchone()
                        obs_act = safe_str(r[0]) if r and r[0] else ""
                        nueva = "No coincide cantidad o precio unitario"
                        obs_final = (nueva + ", " + obs_act) if obs_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (obs_final, nit, factura, oc))
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (estado, nit, factura, oc))
                        
                        cur.execute("""SELECT Valor_XML FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (nit, factura))
                        rc = cur.fetchone()
                        obs_comp_act = safe_str(rc[0]) if rc and rc[0] else ""
                        obs_comp_final = (nueva + ", " + obs_comp_act) if obs_comp_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Valor_XML = ?
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (obs_comp_final, nit, factura))
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ? AND Factura = ?
                        """, (estado, nit, factura))
                        cur.close()
                except: pass
            
            msg = "OK. Total:" + str(stats['total'])
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            return True, msg, None, stats
    except Exception as e:
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPRE_ValidarEmisor():
    import json, ast, traceback, pyodbc, pandas as pd, numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time, warnings, unicodedata
    warnings.filterwarnings('ignore')
    
    def safe_str(v):
        if v is None: return ""
        if isinstance(v, str): return v.strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v): return ""
            return str(v)
        try: return str(v).strip()
        except: return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs: return ""
        obs_str = safe_str(obs)
        return obs_str[:3900] if len(obs_str) > 3900 else obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict): return raw
        try: return json.loads(safe_str(raw))
        except: return ast.literal_eval(safe_str(raw))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    
    def split_valores(v):
        if not v or pd.isna(v): return []
        return [x.strip() for x in str(v).split('|') if x.strip()]
    
    def contiene(campo, val):
        return val in split_valores(campo)
    
    def normalizar(texto):
        if not texto: return ""
        t = ''.join(c for c in unicodedata.normalize('NFD', texto.upper()) if unicodedata.category(c) != 'Mn')
        return ''.join(c if c.isalnum() or c.isspace() else '' for c in t).strip()
    
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        stats = {'total': 0, 'aprobados': 0, 'con_novedad': 0}
        
        with crear_conexion_db(cfg) as cx:
            df = pd.read_sql("""
            SELECT nit_emisor_o_nit_del_proveedor_dp, numero_de_factura_dp,
                   numero_de_liquidacion_u_orden_de_compra_dp, forma_de_pago_dp,
                   ClaseDePedido_hoc, nombre_emisor_dp, Acreedor_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            """, cx)
            
            if df.empty:
                SetVar("vLocStrResultadoSP", "True")
                return True, "No hay registros", None, stats
            
            mask = df['ClaseDePedido_hoc'].apply(
                lambda x: contiene(x, 'ZPRE') or contiene(x, '45') if pd.notna(x) else False
            )
            df = df[mask].copy()
            stats['total'] = len(df)
            
            for idx, row in df.iterrows():
                try:
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    nombre_dp = normalizar(safe_str(row['nombre_emisor_dp']))
                    nombres_hoc = [normalizar(x) for x in split_valores(row['Acreedor_hoc'])]
                    
                    match = any(nombre_dp == n for n in nombres_hoc)
                    
                    if match:
                        stats['aprobados'] += 1
                    else:
                        stats['con_novedad'] += 1
                        estado = 'CON NOVEDAD - CONTADO' if forma_pago in ('1', '01') else 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        
                        cur.execute("""SELECT ObservacionesFase_4 FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        r = cur.fetchone()
                        obs_act = safe_str(r[0]) if r and r[0] else ""
                        nueva = "No coincide nombre del emisor"
                        obs_final = (nueva + ", " + obs_act) if obs_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (obs_final, nit, factura, oc))
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (estado, nit, factura, oc))
                        
                        cur.execute("""SELECT Valor_XML FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (nit, factura))
                        rc = cur.fetchone()
                        obs_comp_act = safe_str(rc[0]) if rc and rc[0] else ""
                        obs_comp_final = (nueva + ", " + obs_comp_act) if obs_comp_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Valor_XML = ?
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (obs_comp_final, nit, factura))
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ? AND Factura = ?
                        """, (estado, nit, factura))
                        cur.close()
                except: pass
            
            msg = "OK. Total:" + str(stats['total'])
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            return True, msg, None, stats
    except Exception as e:
        print(traceback.format_exc())
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarCOP():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarCOP() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def obtener_primer_valor(campo):
        valores = split_valores(campo)
        if valores:
            return valores[0]
        return ""
    
    def sumar_valores(campo):
        valores = split_valores(campo)
        total = 0.0
        for val in valores:
            try:
                total += float(val)
            except:
                pass
        return total
    
    # CORRECCIÓN SQL: Obtener min_id primero, sin subconsulta
    def verificar_y_crear_item(cx, nit, factura, item_name):
        cur = cx.cursor()
        
        check_query = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(check_query, (nit, factura, item_name))
        count = cur.fetchone()[0]
        
        if count > 0:
            print("[DEBUG] Item '" + item_name + "' ya existe")
            cur.close()
            return True
        
        print("[INFO] Creando Item '" + item_name + "'...")
        
        # CORRECCIÓN: Obtener min_id primero
        get_min_id = """
        SELECT MIN(ID_registro)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
        """
        cur.execute(get_min_id, (nit, factura))
        result = cur.fetchone()
        
        if not result or not result[0]:
            print("[ERROR] No se encontro registro base")
            cur.close()
            return False
        
        min_id = result[0]
        
        # CORRECCIÓN: Usar min_id directamente (sin subconsulta)
        insert_query = """
        INSERT INTO [dbo].[CxP.Comparativa] (
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            Item, Valor_XML, Valor_Orden_de_Compra,
            Valor_Orden_de_Compra_Comercializados, Aprobado,
            Estado_validacion_antes_de_eventos,
            Fecha_de_retoma_contabilizacion, Estado_contabilizacion,
            Fecha_de_retoma_compensacion, Estado_compensacion
        )
        SELECT 
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            ?, NULL, NULL,
            NULL, NULL,
            NULL,
            NULL, NULL,
            NULL, NULL
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND ID_registro = ?
        """
        cur.execute(insert_query, (item_name, nit, factura, min_id))
        print("[DEBUG] Item '" + item_name + "' creado exitosamente")
        
        cur.close()
        return True
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        stats = {
            'total_registros': 0,
            'aprobados': 0,
            'con_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                ClaseDePedido_hoc,
                Moneda_hoc,
                PorCalcular_hoc,
                [Valor de la Compra LEA_ddp],
                DocCompra_hoc,
                NitCedula_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            print("[PASO 2] Aplicando filtros...")
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            mask_moneda = df_candidatos['Moneda_hoc'].apply(
                lambda x: contiene_valor(x, 'COP') or contiene_valor(x, '') if pd.notna(x) else True
            )
            
            df_filtrado = df_candidatos[mask_clase & mask_moneda].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42 y Moneda COP/vacio")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            print("")
            print("[PASO 3] Procesando VALIDACION: Suma de valores...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    suma_porcalcular = sumar_valores(row['PorCalcular_hoc'])
                    suma_valor_compra = sumar_valores(row['Valor de la Compra LEA_ddp'])
                    
                    diferencia = abs(suma_porcalcular - suma_valor_compra)
                    tolerancia = 500.0
                    
                    print("")
                    print("[DEBUG] PorCalcular sum: " + str(suma_porcalcular))
                    print("[DEBUG] Valor_Compra_LEA sum: " + str(suma_valor_compra))
                    print("[DEBUG] Diferencia: " + str(diferencia))
                    print("[DEBUG] Tolerancia: " + str(tolerancia))
                    
                    if diferencia > tolerancia:
                        stats['con_novedad'] += 1
                        
                        if forma_pago == '1' or forma_pago == '01':
                            estado_final = 'CON NOVEDAD - CONTADO'
                        else:
                            estado_final = 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        
                        verificar_y_crear_item(cx, nit, factura, 'Valor')
                        verificar_y_crear_item(cx, nit, factura, 'Observaciones')
                        
                        update_fase4 = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_fase4, (nit, factura, oc))
                        
                        select_obs = """
                        SELECT ObservacionesFase_4
                        FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(select_obs, (nit, factura, oc))
                        result_obs = cur.fetchone()
                        
                        obs_actual = safe_str(result_obs[0]) if result_obs and result_obs[0] else ""
                        nueva_obs = "No se encuentra coincidencia en el valor total de la factura vs la informacion reportada en SAP"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        obs_final = truncar_observacion(obs_final)
                        
                        update_obs = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_obs, (obs_final, nit, factura, oc))
                        
                        update_resultado = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_resultado, (estado_final, nit, factura, oc))
                        
                        update_valor_no = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Aprobado = 'NO'
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Valor'
                        """
                        cur.execute(update_valor_no, (nit, factura))
                        
                        select_obs_comp = """
                        SELECT Valor_XML
                        FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(select_obs_comp, (nit, factura))
                        result_obs_comp = cur.fetchone()
                        
                        obs_comp_actual = safe_str(result_obs_comp[0]) if result_obs_comp and result_obs_comp[0] else ""
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs
                        
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        update_obs_comp = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(update_obs_comp, (obs_comp_final, nit, factura))
                        
                        update_estado_todos = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ?
                          AND Factura = ?
                        """
                        cur.execute(update_estado_todos, (estado_final, nit, factura))
                        
                        valores_doccompra = split_valores(row['DocCompra_hoc'])
                        valores_nitcedula = split_valores(row['NitCedula_hoc'])
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        valores_textobreve = split_valores(row['TextoBreve_hoc'])
                        
                        for i in range(max(len(valores_doccompra), len(valores_nitcedula), len(valores_porcalcular), len(valores_textobreve))):
                            doccompra_val = valores_doccompra[i] if i < len(valores_doccompra) else ""
                            nitcedula_val = valores_nitcedula[i] if i < len(valores_nitcedula) else ""
                            porcalcular_val = valores_porcalcular[i] if i < len(valores_porcalcular) else ""
                            textobreve_val = valores_textobreve[i] if i < len(valores_textobreve) else ""
                            
                            if doccompra_val and nitcedula_val:
                                update_marca = """
                                UPDATE [CxP].[HistoricoOrdenesCompra]
                                SET Marca = 'PROCESADO'
                                WHERE DocCompra = ?
                                  AND NitCedula = ?
                                  AND PorCalcular = ?
                                  AND TextoBreve = ?
                                """
                                cur.execute(update_marca, (doccompra_val, nitcedula_val, porcalcular_val, textobreve_val))
                        
                        cur.close()
                    else:
                        stats['aprobados'] += 1
                
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " Aprobados:" + str(stats['aprobados']) + 
                   " ConNovedad:" + str(stats['con_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        print(traceback.format_exc())
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarUSD():
    import json, ast, traceback, pyodbc, pandas as pd, numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time, warnings
    warnings.filterwarnings('ignore')
    
    def safe_str(v):
        if v is None: return ""
        if isinstance(v, str): return v.strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v): return ""
            return str(v)
        try: return str(v).strip()
        except: return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs: return ""
        obs_str = safe_str(obs)
        return obs_str[:3900] if len(obs_str) > 3900 else obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict): return raw
        try: return json.loads(safe_str(raw))
        except: return ast.literal_eval(safe_str(raw))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    
    def split_valores(v):
        if not v or pd.isna(v): return []
        return [x.strip() for x in str(v).split('|') if x.strip()]
    
    def contiene(campo, val):
        return val in split_valores(campo)
    
    def sumar(valor_str):
        suma = 0.0
        for v in split_valores(valor_str):
            try:
                suma += float(v.replace(',', ''))
            except:
                pass
        return suma
    
    try:
        cfg = parse_config(GetVar("vLocDicConfig"))
        tol = float(cfg.get('Tolerancia', 500))
        stats = {'total': 0, 'aprobados': 0, 'con_novedad': 0}
        
        with crear_conexion_db(cfg) as cx:
            df = pd.read_sql("""
            SELECT nit_emisor_o_nit_del_proveedor_dp, numero_de_factura_dp,
                   numero_de_liquidacion_u_orden_de_compra_dp, forma_de_pago_dp,
                   ClaseDePedido_hoc, PorCalcular_hoc, VlrPagarCop_dp, Moneda_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            """, cx)
            
            if df.empty:
                SetVar("vLocStrResultadoSP", "True")
                return True, "No hay registros", None, stats
            
            mask_clase = df['ClaseDePedido_hoc'].apply(
                lambda x: contiene(x, 'ZPCN') or contiene(x, 'ZPPA') or contiene(x, '42') if pd.notna(x) else False
            )
            mask_usd = df['Moneda_hoc'].apply(
                lambda x: contiene(x, 'USD') if pd.notna(x) else False
            )
            df = df[mask_clase & mask_usd].copy()
            stats['total'] = len(df)
            
            for idx, row in df.iterrows():
                try:
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    suma_hoc = sumar(row['PorCalcular_hoc'])
                    vlr_cop = sumar(row['VlrPagarCop_dp'])
                    diff = abs(suma_hoc - vlr_cop)
                    
                    if diff <= tol:
                        stats['aprobados'] += 1
                    else:
                        stats['con_novedad'] += 1
                        estado = 'CON NOVEDAD - CONTADO' if forma_pago in ('1', '01') else 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        
                        cur.execute("""SELECT ObservacionesFase_4 FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (nit, factura, oc))
                        r = cur.fetchone()
                        obs_act = safe_str(r[0]) if r and r[0] else ""
                        nueva = "No se encuentra coincidencia del Valor a pagar COP de la factura"
                        obs_final = (nueva + ", " + obs_act) if obs_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (obs_final, nit, factura, oc))
                        
                        cur.execute("""UPDATE [CxP].[DocumentsProcessing] SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ? AND numero_de_factura = ? AND numero_de_liquidacion_u_orden_de_compra = ?
                        """, (estado, nit, factura, oc))
                        
                        cur.execute("""SELECT Valor_XML FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (nit, factura))
                        rc = cur.fetchone()
                        obs_comp_act = safe_str(rc[0]) if rc and rc[0] else ""
                        obs_comp_final = (nueva + ", " + obs_comp_act) if obs_comp_act else nueva
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Valor_XML = ?
                        WHERE NIT = ? AND Factura = ? AND Item = 'Observaciones'
                        """, (obs_comp_final, nit, factura))
                        
                        cur.execute("""UPDATE [dbo].[CxP.Comparativa] SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ? AND Factura = ?
                        """, (estado, nit, factura))
                        cur.close()
                except: pass
            
            msg = "OK. Total:" + str(stats['total'])
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            return True, msg, None, stats
    except Exception as e:
        print(traceback.format_exc())
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vLocStrResultadoSP", "False")
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarTRM():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarTRM() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        """Dividir string por | y retornar lista de valores"""
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        """Verificar si campo (que puede tener valores separados por |) contiene valor buscado"""
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def obtener_primer_valor(campo):
        """Obtener solo el primer valor de un campo que puede tener multiples valores separados por |"""
        valores = split_valores(campo)
        if valores:
            return valores[0]
        return ""
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        print("[DEBUG] Servidor: " + cfg.get('ServidorBaseDatos', 'N/A'))
        print("[DEBUG] Base de datos: " + cfg.get('NombreBaseDatos', 'N/A'))
        
        stats = {
            'total_registros': 0,
            'aprobados': 0,
            'con_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                nombre_emisor_dp,
                ClaseDePedido_hoc,
                Moneda_hoc,
                Trm_hoc,
                CalculationRate_dp,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            print("[PASO 2] Aplicando filtros...")
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPPA, ZPCN o 42: " + str(mask_clase.sum()))
            
            mask_moneda = df_candidatos['Moneda_hoc'].apply(
                lambda x: contiene_valor(x, 'USD') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con Moneda = USD: " + str(mask_moneda.sum()))
            
            mask_final = mask_clase & mask_moneda
            df_filtrado = df_candidatos[mask_final].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42 y Moneda USD")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            print("")
            print("[PASO 3] Procesando VALIDACION: TRM vs CalculationRate...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "] - VALIDACION TRM")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    
                    trm_hoc_primer_valor = obtener_primer_valor(row['Trm_hoc'])
                    calculation_rate = safe_str(row['CalculationRate_dp'])
                    
                    print("[DEBUG] Trm_hoc (primer valor): '" + trm_hoc_primer_valor + "'")
                    print("[DEBUG] CalculationRate_dp: '" + calculation_rate + "'")
                    
                    if trm_hoc_primer_valor != calculation_rate:
                        print("[RESULTADO] CON NOVEDAD - TRM (valores diferentes)")
                        stats['con_novedad'] += 1
                        
                        if forma_pago == '1' or forma_pago == '01':
                            estado_final = 'CON NOVEDAD - CONTADO'
                        else:
                            estado_final = 'CON NOVEDAD'
                        
                        print("[DEBUG] Estado final: " + estado_final)
                        
                        cur = cx.cursor()
                        
                        print("[UPDATE] Actualizando tabla DocumentsProcessing...")
                        
                        update_fase4 = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_fase4, (nit, factura, oc))
                        
                        select_obs = """
                        SELECT ObservacionesFase_4
                        FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(select_obs, (nit, factura, oc))
                        result_obs = cur.fetchone()
                        
                        obs_actual = safe_str(result_obs[0]) if result_obs and result_obs[0] else ""
                        nueva_obs = "No se encuentra coincidencia en el campo TRM de la factura vs la informacion reportada en SAP"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        update_obs = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_obs, (obs_final, nit, factura, oc))
                        
                        update_resultado = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_resultado, (estado_final, nit, factura, oc))
                        
                        print("[DEBUG] DocumentsProcessing actualizado OK")
                        
                        print("[UPDATE] Actualizando tabla CxP.Comparativa...")
                        
                        select_obs_comp = """
                        SELECT Valor_XML
                        FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(select_obs_comp, (nit, factura))
                        result_obs_comp = cur.fetchone()
                        
                        obs_comp_actual = safe_str(result_obs_comp[0]) if result_obs_comp and result_obs_comp[0] else ""
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        update_obs_comp = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(update_obs_comp, (obs_comp_final, nit, factura))
                        
                        update_estado_todos = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ?
                          AND Factura = ?
                        """
                        cur.execute(update_estado_todos, (estado_final, nit, factura))
                        
                        print("[DEBUG] CxP.Comparativa actualizado OK")
                        
                        print("[UPDATE] Actualizando tabla HistoricoOrdenesCompra...")
                        
                        valores_doccompra = split_valores(row['DocCompra_hoc'])
                        valores_nitcedula = split_valores(row['NitCedula_hoc'])
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        valores_textobreve = split_valores(row['TextoBreve_hoc'])
                        
                        num_actualizados = 0
                        for i in range(max(len(valores_doccompra), len(valores_nitcedula), 
                                          len(valores_porcalcular), len(valores_textobreve))):
                            
                            doccompra_val = valores_doccompra[i] if i < len(valores_doccompra) else ""
                            nitcedula_val = valores_nitcedula[i] if i < len(valores_nitcedula) else ""
                            porcalcular_val = valores_porcalcular[i] if i < len(valores_porcalcular) else ""
                            textobreve_val = valores_textobreve[i] if i < len(valores_textobreve) else ""
                            
                            if doccompra_val and nitcedula_val:
                                update_marca = """
                                UPDATE [CxP].[HistoricoOrdenesCompra]
                                SET Marca = 'PROCESADO'
                                WHERE DocCompra = ?
                                  AND NitCedula = ?
                                  AND PorCalcular = ?
                                  AND TextoBreve = ?
                                """
                                cur.execute(update_marca, (doccompra_val, nitcedula_val, porcalcular_val, textobreve_val))
                                num_actualizados += 1
                        
                        print("[DEBUG] HistoricoOrdenesCompra actualizado: " + str(num_actualizados) + " registros")
                        
                        cur.close()
                        print("[UPDATE] Todas las tablas actualizadas OK (TRM)")
                        
                    else:
                        print("[RESULTADO] APROBADO - TRM (valores iguales)")
                        stats['aprobados'] += 1
                    
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + " (TRM): " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Aprobados: " + str(stats['aprobados']))
            print("  Con novedad: " + str(stats['con_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " Aprobados:" + str(stats['aprobados']) + 
                   " ConNovedad:" + str(stats['con_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarEmisor():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    import re
    import unicodedata
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarEmisor() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        """Dividir string por | y retornar lista de valores"""
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        """Verificar si campo contiene valor buscado"""
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def quitar_tildes(texto):
        """
        Quitar tildes y acentos de un texto
        
        Ejemplos:
        - INTERES -> INTERES
        - JOSE -> JOSE
        - Maria -> MARIA
        """
        if not texto:
            return ""
        
        # Normalizar a NFD (descomponer caracteres con tildes)
        nfd = unicodedata.normalize('NFD', texto)
        
        # Filtrar solo caracteres que NO sean marcas diacriticas
        sin_tildes = ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')
        
        # Re-normalizar a NFC (forma canonica compuesta)
        return unicodedata.normalize('NFC', sin_tildes)
    
    def normalizar_nombre(nombre):
        """
        Normalizar nombre de empresa segun reglas especificas
        
        MEJORAS V3:
        1. Quitar tildes/acentos (INTERES -> INTERES)
        2. Normalizar '&' a 'Y' (ANGEL & DG -> ANGEL Y DG)
        3. Eliminar 'Y' que esten solas
        
        Normalizaciones completas:
        - S.A.S. = SAS
        - S.A.S = SAS
        - S. A. S. = SAS
        - S A S = SAS
        - S. A. = SA
        - S.A = SA
        - Limitada = LTDA
        - Ltda = LTDA
        - S. EN C. = SENC
        - & = Y (se convierte y luego se elimina si esta sola)
        """
        if not nombre or nombre == "":
            return ""
        
        # Convertir a string y hacer copia para trabajar
        texto = safe_str(nombre)
        
        # PASO 1: Quitar tildes ANTES de convertir a mayusculas
        texto = quitar_tildes(texto)
        
        # PASO 2: Convertir a mayusculas
        texto = texto.upper()
        
        # PASO 2.5: Reemplazar '&' por 'Y' ANTES de otras normalizaciones
        # Esto normaliza "ANGEL & DG" -> "ANGEL Y DG"
        # Luego la 'Y' se eliminara si esta sola
        texto = texto.replace('&', 'Y')
        
        # PASO 3: Aplicar normalizaciones especificas ANTES de quitar caracteres especiales
        # IMPORTANTE: Orden de mas especifico a mas general
        normalizaciones = [
            # SAS - todas las variantes
            ('S. A. S.', 'SAS'),  # S. A. S.
            ('S. A. S', 'SAS'),   # Sin punto final
            ('S.A.S.', 'SAS'),    # S.A.S.
            ('S.A.S', 'SAS'),     # S.A.S (sin punto final)
            ('S, A. S.', 'SAS'),  # S, A. S.
            ('S. A, S.', 'SAS'),  # S. A, S.
            ('S,A.S', 'SAS'),     # S,A.S
            ('S A S', 'SAS'),     # S A S (con espacios)
            # SA - variantes (NUEVO para S.A, S.A.)
            ('S. A.', 'SA'),      # S. A.
            ('S.A.', 'SA'),       # S.A.
            ('S. A', 'SA'),       # S. A (sin punto final)
            ('S.A', 'SA'),        # S.A (sin punto final)
            # LTDA - todas las variantes
            ('LIMITADA', 'LTDA'),
            ('LTDA.', 'LTDA'),
            ('LTDA,', 'LTDA'),
            ('LTDA', 'LTDA'),  # Ya normalizado
            # SENC - todas las variantes
            ('S. EN C A', 'SENC'),
            ('S. EN C.', 'SENC'),
            ('S. EN C', 'SENC'),
            ('S EN C A', 'SENC'),
            ('S EN C', 'SENC')
        ]
        
        for patron, reemplazo in normalizaciones:
            texto = texto.replace(patron, reemplazo)
        
        # PASO 4: Quitar caracteres especiales, solo dejar letras, numeros y espacios
        texto = re.sub(r'[^A-Z0-9\s]', '', texto)
        
        # PASO 5: Normalizar espacios multiples a uno solo
        texto = re.sub(r'\s+', ' ', texto)
        
        # PASO 6: Eliminar 'Y' que esten solas (palabras completas)
        # Esto elimina " Y " pero no "YUCA" ni "YESO"
        # Usar word boundaries \b para asegurar que sea una palabra completa
        texto = re.sub(r'\bY\b', '', texto)
        
        # PASO 7: Limpiar espacios nuevamente (pueden quedar espacios dobles despues de eliminar Y)
        texto = re.sub(r'\s+', ' ', texto)
        
        return texto.strip()
    
    def comparar_nombres(nombre1, nombre2):
        """
        Comparar dos nombres normalizados segun reglas especificas
        
        Pasos:
        1. Normalizar ambos nombres
        2. Separar por espacios
        3. Verificar que tengan EXACTAMENTE las mismas palabras
        
        IMPORTANTE: NO se permiten subconjuntos
        "FERRICENTROS SAS" != "FERRICENTROS SAS COLOMBIA"
        
        Retorna: (coincide: bool, nombre1_norm: str, nombre2_norm: str, detalle: str)
        """
        # Normalizar
        norm1 = normalizar_nombre(nombre1)
        norm2 = normalizar_nombre(nombre2)
        
        print("[DEBUG] Nombre 1 original: '" + safe_str(nombre1) + "'")
        print("[DEBUG] Nombre 1 normalizado: '" + norm1 + "'")
        print("[DEBUG] Nombre 2 original: '" + safe_str(nombre2) + "'")
        print("[DEBUG] Nombre 2 normalizado: '" + norm2 + "'")
        
        # Separar por espacios
        items1 = [item for item in norm1.split(' ') if item]
        items2 = [item for item in norm2.split(' ') if item]
        
        print("[DEBUG] Items nombre 1: " + str(items1))
        print("[DEBUG] Items nombre 2: " + str(items2))
        
        # Crear sets de palabras
        items1_set = set(items1)
        items2_set = set(items2)
        
        # CORRECCION: Solo coincide si los sets son EXACTAMENTE iguales
        # NO se permiten subconjuntos
        if items1_set == items2_set:
            detalle = "Coincidencia exacta - mismas palabras"
            coincide = True
        else:
            # Calcular diferencias
            faltantes_en_2 = items1_set - items2_set
            faltantes_en_1 = items2_set - items1_set
            
            if faltantes_en_2 and not faltantes_en_1:
                detalle = "NO COINCIDE: Nombre 2 le falta: " + str(faltantes_en_2)
            elif faltantes_en_1 and not faltantes_en_2:
                detalle = "NO COINCIDE: Nombre 1 le falta: " + str(faltantes_en_1)
            else:
                detalle = "NO COINCIDE: Diferentes palabras. Falta en 2: " + str(faltantes_en_2) + ", Falta en 1: " + str(faltantes_en_1)
            
            coincide = False
        
        print("[DEBUG] Detalle comparacion: " + detalle)
        
        return coincide, norm1, norm2, detalle
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        print("[DEBUG] Servidor: " + cfg.get('ServidorBaseDatos', 'N/A'))
        print("[DEBUG] Base de datos: " + cfg.get('NombreBaseDatos', 'N/A'))
        
        stats = {
            'total_registros': 0,
            'aprobados': 0,
            'con_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 1: Consultar candidatos de HU41_CandidatosValidacion
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                nombre_emisor_dp,
                Acreedor_hoc,
                ClaseDePedido_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            # ================================================================
            # PASO 2: Aplicar filtros
            # ================================================================
            
            print("[PASO 2] Aplicando filtros...")
            
            # Filtro: ClaseDePedido_hoc contiene 'ZPPA', 'ZPCN' o '42'
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPPA, ZPCN o 42: " + str(mask_clase.sum()))
            
            # Aplicar filtro
            df_filtrado = df_candidatos[mask_clase].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            # ================================================================
            # PASO 3: Procesar cada registro - VALIDACION NOMBRE EMISOR
            # ================================================================
            
            print("")
            print("[PASO 3] Procesando validacion: Nombre Emisor...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "]")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    
                    # Obtener valores
                    nombre_emisor = safe_str(row['nombre_emisor_dp'])
                    acreedor_completo = safe_str(row['Acreedor_hoc'])
                    
                    # Obtener PRIMER valor de Acreedor_hoc
                    valores_acreedor = split_valores(acreedor_completo)
                    primer_acreedor = valores_acreedor[0] if valores_acreedor else ""
                    
                    print("")
                    print("[VALIDACION] Comparando Nombre Emisor vs Acreedor...")
                    print("[DEBUG] Nombre Emisor (dp): '" + nombre_emisor + "'")
                    print("[DEBUG] Acreedor completo (hoc): '" + acreedor_completo + "'")
                    print("[DEBUG] Primer Acreedor (a comparar): '" + primer_acreedor + "'")
                    
                    # ========================================================
                    # COMPARAR NOMBRES CON NORMALIZACION
                    # ========================================================
                    
                    coincide, norm1, norm2, detalle = comparar_nombres(nombre_emisor, primer_acreedor)
                    
                    print("[RESULTADO] " + ("COINCIDEN" if coincide else "NO COINCIDEN"))
                    print("[DETALLE] " + detalle)
                    
                    # ========================================================
                    # DECISION: COINCIDEN O NO?
                    # ========================================================
                    
                    if coincide:
                        # ====================================================
                        # CASO: NOMBRES COINCIDEN (APROBADO)
                        # ====================================================
                        
                        print("")
                        print("[RESULTADO FINAL] APROBADO (nombres coinciden)")
                        stats['aprobados'] += 1
                        
                        # ====================================================
                        # 2. ACTUALIZAR [dbo].[CxP.Comparativa] - SOLO ESTO
                        # ====================================================
                        
                        print("[UPDATE] Actualizando tabla CxP.Comparativa...")
                        
                        cur = cx.cursor()
                        
                        # 2.1.1: Actualizar Valor_XML
                        update_xml = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_xml, (nombre_emisor, nit, factura))
                        
                        # 2.1.3: Actualizar Aprobado = 'SI'
                        update_aprobado = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Aprobado = 'SI'
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_aprobado, (nit, factura))
                        
                        # 2.1.4: Actualizar Valor_Orden_de_Compra
                        update_voc = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_Orden_de_Compra = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_voc, (primer_acreedor, nit, factura))
                        
                        cur.close()
                        print("[UPDATE] Tabla CxP.Comparativa actualizada OK (APROBADO)")
                        
                    else:
                        # ====================================================
                        # CASO: NOMBRES NO COINCIDEN (CON NOVEDAD)
                        # ====================================================
                        
                        print("")
                        print("[RESULTADO FINAL] CON NOVEDAD (nombres NO coinciden)")
                        stats['con_novedad'] += 1
                        
                        # Determinar estado segun forma de pago
                        if forma_pago == '1' or forma_pago == '01':
                            estado_final = 'CON NOVEDAD - CONTADO'
                        else:
                            estado_final = 'CON NOVEDAD'
                        
                        cur = cx.cursor()
                        
                        # ====================================================
                        # 3.1 ACTUALIZAR [CxP].[DocumentsProcessing]
                        # ====================================================
                        
                        print("[UPDATE] Actualizando tabla DocumentsProcessing...")
                        
                        update_fase4 = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_fase4, (nit, factura, oc))
                        
                        select_obs = """
                        SELECT ObservacionesFase_4
                        FROM [CxP].[DocumentsProcessing]
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(select_obs, (nit, factura, oc))
                        result_obs = cur.fetchone()
                        
                        obs_actual = safe_str(result_obs[0]) if result_obs and result_obs[0] else ""
                        nueva_obs = "No se encuentra coincidencia en Nombre Emisor de la factura vs la informacion reportada en SAP"
                        
                        if obs_actual:
                            obs_final = nueva_obs + ", " + obs_actual
                        else:
                            obs_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_final = truncar_observacion(obs_final)
                        
                        update_obs = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ObservacionesFase_4 = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_obs, (obs_final, nit, factura, oc))
                        
                        update_resultado = """
                        UPDATE [CxP].[DocumentsProcessing]
                        SET ResultadoFinalAntesEventos = ?
                        WHERE nit_emisor_o_nit_del_proveedor = ?
                          AND numero_de_factura = ?
                          AND numero_de_liquidacion_u_orden_de_compra = ?
                        """
                        cur.execute(update_resultado, (estado_final, nit, factura, oc))
                        
                        print("[DEBUG] DocumentsProcessing actualizado OK")
                        
                        # ====================================================
                        # 3.2 ACTUALIZAR [dbo].[CxP.Comparativa]
                        # ====================================================
                        
                        print("[UPDATE] Actualizando tabla CxP.Comparativa...")
                        
                        # 3.2.1: Actualizar Valor_XML NombreEmisor
                        update_xml = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_xml, (nombre_emisor, nit, factura))
                        
                        # 3.2.3: Actualizar Aprobado = 'NO'
                        update_aprobado = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Aprobado = 'NO'
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_aprobado, (nit, factura))
                        
                        # 3.2.4: Actualizar Valor_Orden_de_Compra
                        update_voc = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_Orden_de_Compra = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'NombreEmisor'
                        """
                        cur.execute(update_voc, (primer_acreedor, nit, factura))
                        
                        # 3.2.2: Actualizar Observaciones
                        select_obs_comp = """
                        SELECT Valor_XML
                        FROM [dbo].[CxP.Comparativa]
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(select_obs_comp, (nit, factura))
                        result_obs_comp = cur.fetchone()
                        
                        obs_comp_actual = safe_str(result_obs_comp[0]) if result_obs_comp and result_obs_comp[0] else ""
                        
                        if obs_comp_actual:
                            obs_comp_final = nueva_obs + ", " + obs_comp_actual
                        else:
                            obs_comp_final = nueva_obs
                        
                        # CORRECCIÓN: Truncar antes de UPDATE
                        obs_comp_final = truncar_observacion(obs_comp_final)
                        
                        update_obs_comp = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Valor_XML = ?
                        WHERE NIT = ?
                          AND Factura = ?
                          AND Item = 'Observaciones'
                        """
                        cur.execute(update_obs_comp, (obs_comp_final, nit, factura))
                        
                        # 3.2.5: Actualizar Estado_validacion_antes_de_eventos (TODOS LOS ITEMS)
                        update_estado_todos = """
                        UPDATE [dbo].[CxP.Comparativa]
                        SET Estado_validacion_antes_de_eventos = ?
                        WHERE NIT = ?
                          AND Factura = ?
                        """
                        cur.execute(update_estado_todos, (estado_final, nit, factura))
                        
                        print("[DEBUG] CxP.Comparativa actualizado OK")
                        
                        # ====================================================
                        # 3.3 ACTUALIZAR [CxP].[HistoricoOrdenesCompra]
                        # ====================================================
                        
                        print("[UPDATE] Actualizando tabla HistoricoOrdenesCompra...")
                        
                        valores_doccompra = split_valores(row['DocCompra_hoc'])
                        valores_nitcedula = split_valores(row['NitCedula_hoc'])
                        valores_porcalcular = split_valores(row['PorCalcular_hoc'])
                        valores_textobreve = split_valores(row['TextoBreve_hoc'])
                        
                        num_actualizados = 0
                        for i in range(max(len(valores_doccompra), len(valores_nitcedula), 
                                          len(valores_porcalcular), len(valores_textobreve))):
                            
                            doccompra_val = valores_doccompra[i] if i < len(valores_doccompra) else ""
                            nitcedula_val = valores_nitcedula[i] if i < len(valores_nitcedula) else ""
                            porcalcular_val = valores_porcalcular[i] if i < len(valores_porcalcular) else ""
                            textobreve_val = valores_textobreve[i] if i < len(valores_textobreve) else ""
                            
                            if doccompra_val and nitcedula_val:
                                update_marca = """
                                UPDATE [CxP].[HistoricoOrdenesCompra]
                                SET Marca = 'PROCESADO'
                                WHERE DocCompra = ?
                                  AND NitCedula = ?
                                  AND PorCalcular = ?
                                  AND TextoBreve = ?
                                """
                                cur.execute(update_marca, (doccompra_val, nitcedula_val, porcalcular_val, textobreve_val))
                                num_actualizados += 1
                        
                        print("[DEBUG] HistoricoOrdenesCompra actualizado: " + str(num_actualizados) + " registros")
                        
                        cur.close()
                        print("[UPDATE] Todas las tablas actualizadas OK (CON NOVEDAD)")
                    
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Aprobados: " + str(stats['aprobados']))
            print("  Con novedad: " + str(stats['con_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " Aprobados:" + str(stats['aprobados']) + 
                   " ConNovedad:" + str(stats['con_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarOrdenRegistro():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarOrdenRegistro() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        """Dividir string por | y retornar lista de valores"""
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        """Verificar si campo (que puede tener valores separados por |) contiene valor buscado"""
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def verificar_y_crear_item(cx, nit, factura, item_name):
        cur = cx.cursor()
        
        check_query = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(check_query, (nit, factura, item_name))
        count = cur.fetchone()[0]
        
        if count > 0:
            print("[DEBUG] Item '" + item_name + "' ya existe")
            cur.close()
            return True
        
        print("[INFO] Creando Item '" + item_name + "'...")
        
        # CORRECCIÓN: Obtener min_id primero
        get_min_id = """
        SELECT MIN(ID_registro)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
        """
        cur.execute(get_min_id, (nit, factura))
        result = cur.fetchone()
        
        if not result or not result[0]:
            print("[ERROR] No se encontro registro base")
            cur.close()
            return False
        
        min_id = result[0]
        
        # CORRECCIÓN: Usar min_id directamente (sin subconsulta)
        insert_query = """
        INSERT INTO [dbo].[CxP.Comparativa] (
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            Item, Valor_XML, Valor_Orden_de_Compra,
            Valor_Orden_de_Compra_Comercializados, Aprobado,
            Estado_validacion_antes_de_eventos,
            Fecha_de_retoma_contabilizacion, Estado_contabilizacion,
            Fecha_de_retoma_compensacion, Estado_compensacion
        )
        SELECT 
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            ?, NULL, NULL,
            NULL, NULL,
            NULL,
            NULL, NULL,
            NULL, NULL
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND ID_registro = ?
        """
        cur.execute(insert_query, (item_name, nit, factura, min_id))
        print("[DEBUG] Item '" + item_name + "' creado exitosamente")
        
        cur.close()
        return True
    

    def actualizar_item_comparativa(cx, nit, factura, item_name, aprobado=None, observacion=None, estado=None):
        """Actualizar campos de un Item en Comparativa"""
        cur = cx.cursor()
        
        if aprobado is not None:
            update_aprobado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Aprobado = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = ?
            """
            cur.execute(update_aprobado, (aprobado, nit, factura, item_name))
            print("[UPDATE] Item '" + item_name + "' Aprobado = " + aprobado)
        
        if observacion is not None:
            # Obtener observaciones actuales
            select_obs = """
            SELECT Valor_XML
            FROM [dbo].[CxP.Comparativa]
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(select_obs, (nit, factura))
            result = cur.fetchone()
            
            obs_actual = safe_str(result[0]) if result and result[0] else ""
            
            if obs_actual:
                obs_final = observacion + ", " + obs_actual
            else:
                obs_final = observacion
            
            # CORRECCIÓN: Truncar antes de UPDATE
            obs_final = truncar_observacion(obs_final)
            
            update_obs = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Valor_XML = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(update_obs, (obs_final, nit, factura))
            print("[UPDATE] Observaciones actualizadas")
        
        if estado is not None:
            update_estado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Estado_validacion_antes_de_eventos = ?
            WHERE NIT = ?
              AND Factura = ?
            """
            cur.execute(update_estado, (estado, nit, factura))
            print("[UPDATE] Estado_validacion_antes_de_eventos = " + estado)
        
        cur.close()
    
    def actualizar_documents_processing(cx, nit, factura, oc, observacion, forma_pago):
        """Actualizar DocumentsProcessing con novedad"""
        cur = cx.cursor()
        
        # Determinar estado
        if forma_pago == '1' or forma_pago == '01':
            estado_final = 'CON NOVEDAD - CONTADO'
        else:
            estado_final = 'CON NOVEDAD'
        
        # Actualizar EstadoFinalFase_4
        update_fase4 = """
        UPDATE [CxP].[DocumentsProcessing]
        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_fase4, (nit, factura, oc))
        
        # Obtener observaciones actuales
        select_obs = """
        SELECT ObservacionesFase_4
        FROM [CxP].[DocumentsProcessing]
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(select_obs, (nit, factura, oc))
        result = cur.fetchone()
        
        obs_actual = safe_str(result[0]) if result and result[0] else ""
        
        if obs_actual:
            obs_final = observacion + ", " + obs_actual
        else:
            obs_final = observacion
        
        # CORRECCIÓN: Truncar antes de UPDATE
        obs_final = truncar_observacion(obs_final)
        
        # Actualizar observaciones
        update_obs = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ObservacionesFase_4 = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_obs, (obs_final, nit, factura, oc))
        
        # Actualizar ResultadoFinalAntesEventos
        update_resultado = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ResultadoFinalAntesEventos = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_resultado, (estado_final, nit, factura, oc))
        
        cur.close()
        print("[UPDATE] DocumentsProcessing actualizado (CON NOVEDAD)")
    
    def actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list):
        """Actualizar HistoricoOrdenesCompra marcando como PROCESADO"""
        cur = cx.cursor()
        
        num_actualizados = 0
        max_len = max(len(doccompra_list), len(nitcedula_list), len(porcalcular_list), len(textobreve_list))
        
        for i in range(max_len):
            doccompra = doccompra_list[i] if i < len(doccompra_list) else ""
            nitcedula = nitcedula_list[i] if i < len(nitcedula_list) else ""
            porcalcular = porcalcular_list[i] if i < len(porcalcular_list) else ""
            textobreve = textobreve_list[i] if i < len(textobreve_list) else ""
            
            if doccompra and nitcedula:
                update_marca = """
                UPDATE [CxP].[HistoricoOrdenesCompra]
                SET Marca = 'PROCESADO'
                WHERE DocCompra = ?
                  AND NitCedula = ?
                  AND PorCalcular = ?
                  AND TextoBreve = ?
                """
                cur.execute(update_marca, (doccompra, nitcedula, porcalcular, textobreve))
                num_actualizados += 1
        
        cur.close()
        print("[UPDATE] HistoricoOrdenesCompra: " + str(num_actualizados) + " registros marcados como PROCESADO")
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        stats = {
            'total_registros': 0,
            'validaciones_ok': 0,
            'validaciones_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                ClaseDePedido_hoc,
                Orden_hoc,
                IndicadorImpuestos_hoc,
                CentroDeCoste_hoc,
                Cuenta_hoc,
                ClaseDeOrden_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            print("[PASO 2] Aplicando filtros...")
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPPA, ZPCN o 42: " + str(mask_clase.sum()))
            
            df_filtrado = df_candidatos[mask_clase].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            print("")
            print("[PASO 3] Procesando validaciones por posicion...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "]")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    
                    ordenes = split_valores(row['Orden_hoc'])
                    ind_impuestos = split_valores(row['IndicadorImpuestos_hoc'])
                    centros_coste = split_valores(row['CentroDeCoste_hoc'])
                    cuentas = split_valores(row['Cuenta_hoc'])
                    clases_orden = split_valores(row['ClaseDeOrden_hoc'])
                    
                    doccompra_list = split_valores(row['DocCompra_hoc'])
                    nitcedula_list = split_valores(row['NitCedula_hoc'])
                    porcalcular_list = split_valores(row['PorCalcular_hoc'])
                    textobreve_list = split_valores(row['TextoBreve_hoc'])
                    
                    print("[DEBUG] Total posiciones a procesar: " + str(len(ordenes)))
                    
                    hay_novedad = False
                    
                    for pos in range(len(ordenes)):
                        print("")
                        print("[POSICION " + str(pos + 1) + "/" + str(len(ordenes)) + "]")
                        
                        orden = ordenes[pos] if pos < len(ordenes) else ""
                        ind_imp = ind_impuestos[pos] if pos < len(ind_impuestos) else ""
                        centro = centros_coste[pos] if pos < len(centros_coste) else ""
                        cuenta = cuentas[pos] if pos < len(cuentas) else ""
                        clase_ord = clases_orden[pos] if pos < len(clases_orden) else ""
                        
                        print("[DEBUG] Orden: '" + orden + "'")
                        print("[DEBUG] IndicadorImpuestos: '" + ind_imp + "'")
                        print("[DEBUG] CentroDeCoste: '" + centro + "'")
                        print("[DEBUG] Cuenta: '" + cuenta + "'")
                        print("[DEBUG] ClaseDeOrden: '" + clase_ord + "'")
                        
                        if not orden:
                            print("[INFO] Orden vacio, saltando posicion")
                            continue
                        
                        if pos > 0:
                            verificar_y_crear_item(cx, nit, factura, 'IndicadorImpuestos')
                            verificar_y_crear_item(cx, nit, factura, 'CentroCoste')
                            verificar_y_crear_item(cx, nit, factura, 'Cuenta')
                            verificar_y_crear_item(cx, nit, factura, 'ClaseOrden')
                            verificar_y_crear_item(cx, nit, factura, 'Observaciones')
                        
                        if len(orden) == 9 and orden.startswith('15'):
                            print("[VALIDACION] Orden 15 detectada (9 caracteres)")
                            
                            ind_validos = ['H4', 'H5', 'H6', 'H7', 'VP', 'CO', 'IC', 'CR']
                            
                            if ind_imp in ind_validos:
                                print("[OK] IndicadorImpuestos '" + ind_imp + "' es valido")
                                actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='SI')
                            
                            elif not ind_imp or ind_imp not in ind_validos:
                                print("[NOVEDAD] IndicadorImpuestos NO valido o vacio")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero campo 'Indicador impuestos' NO se encuentra diligenciado o NO corresponde alguna de las opciones H4, H5, H6, H7, VP, CO, IC, CR"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            if not centro:
                                print("[OK] CentroDeCoste vacio (correcto para Orden 15)")
                                actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                            
                            else:
                                print("[NOVEDAD] CentroDeCoste tiene valor (incorrecto para Orden 15)")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Centro de coste' se encuentra diligenciado cuando NO debe estarlo"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            if cuenta == '5199150001':
                                print("[OK] Cuenta es 5199150001 (correcto)")
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                            
                            else:
                                print("[NOVEDAD] Cuenta diferente a 5199150001")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Cuenta' es diferente a 5199150001"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            clase_correcta = False
                            obs_clase = ""
                            
                            if not clase_ord:
                                obs_clase = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Clase orden' NO se encuentra diligenciado"
                            elif ind_imp in ['H4', 'H5']:
                                clase_correcta = (clase_ord == 'ZINV')
                                if not clase_correcta:
                                    obs_clase = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Clase orden' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = ZINV', 'H6 y H7 = ZADM' o 'VP, CO, CR o IC = ZINV o ZADM'"
                            elif ind_imp in ['H6', 'H7']:
                                clase_correcta = (clase_ord == 'ZADM')
                                if not clase_correcta:
                                    obs_clase = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Clase orden' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = ZINV', 'H6 y H7 = ZADM' o 'VP, CO, CR o IC = ZINV o ZADM'"
                            elif ind_imp in ['VP', 'CO', 'CR', 'IC']:
                                clase_correcta = (clase_ord in ['ZINV', 'ZADM'])
                                if not clase_correcta:
                                    obs_clase = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 15, pero Campo 'Clase orden' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = ZINV', 'H6 y H7 = ZADM' o 'VP, CO, CR o IC = ZINV o ZADM'"
                            else:
                                clase_correcta = True
                            
                            if clase_correcta:
                                print("[OK] ClaseDeOrden correcta segun IndicadorImpuestos")
                                actualizar_item_comparativa(cx, nit, factura, 'ClaseOrden', aprobado='SI')
                            else:
                                print("[NOVEDAD] ClaseDeOrden incorrecta")
                                hay_novedad = True
                                actualizar_documents_processing(cx, nit, factura, oc, obs_clase, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'ClaseOrden', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs_clase)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        elif len(orden) == 8 and orden.isdigit():
                            print("[VALIDACION] Orden de 8 digitos detectada")
                            
                            if orden.startswith('53'):
                                print("[DEBUG] Orden 53 (ESTADISTICAS)")
                                
                                if centro:
                                    print("[OK] CentroDeCoste tiene valor (correcto para Orden 53)")
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                                
                                else:
                                    print("[NOVEDAD] CentroDeCoste vacio (incorrecto para Orden 53)")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden 53, pero Campo 'Centro de coste' se encuentra vacio para pedidos ESTADISTICAS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            else:
                                print("[DEBUG] Orden NO 53 (NO ESTADISTICAS)")
                                
                                if not centro:
                                    print("[OK] CentroDeCoste vacio (correcto para Orden NO 53)")
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                                
                                else:
                                    print("[NOVEDAD] CentroDeCoste tiene valor (incorrecto para Orden NO 53)")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden diferente a 53, pero Campo 'Centro de coste' se encuentra diligenciado para pedidos NO ESTADISTICAS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                cuenta_ok = (cuenta == '5299150099') or (len(cuenta) == 10 and cuenta.startswith('73'))
                                
                                if cuenta_ok:
                                    print("[OK] Cuenta correcta (5299150099 o 10 digitos empezando con 73)")
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                                
                                else:
                                    print("[NOVEDAD] Cuenta incorrecta")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden diferente a 53, pero Campo 'Cuenta' es diferente a 5299150099 y/o NO cumple regla 'inicia con 7 y tiene 10 digitos'"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        elif len(orden) == 10 and orden.startswith('73'):
                            print("[VALIDACION] Orden de 10 digitos empezando con 73")
                            cuenta_ok = (cuenta == '5299150099') or (len(cuenta) == 10 and cuenta.startswith('73'))
                            
                            if cuenta_ok:
                                print("[OK] Cuenta correcta")
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                            else:
                                print("[NOVEDAD] Cuenta incorrecta")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA y cuenta con Orden diferente a 53, pero Campo 'Cuenta' es diferente a 5299150099 y/o NO cumple regla 'inicia con 7 y tiene 10 digitos'"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        else:
                            print("[INFO] Orden no cumple ningun criterio de validacion")
                    
                    if hay_novedad:
                        stats['validaciones_novedad'] += 1
                    else:
                        stats['validaciones_ok'] += 1
                    
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Validaciones OK: " + str(stats['validaciones_ok']))
            print("  Validaciones con novedad: " + str(stats['validaciones_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " OK:" + str(stats['validaciones_ok']) + 
                   " Novedad:" + str(stats['validaciones_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarElementoPEP():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarElementoPEP() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    # CORRECCIÓN: Función para truncar observaciones
    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def verificar_y_crear_item(cx, nit, factura, item_name):
        cur = cx.cursor()
        
        check_query = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(check_query, (nit, factura, item_name))
        count = cur.fetchone()[0]
        
        if count > 0:
            print("[DEBUG] Item '" + item_name + "' ya existe")
            cur.close()
            return True
        
        print("[INFO] Creando Item '" + item_name + "'...")
        
        # CORRECCIÓN: Obtener min_id primero
        get_min_id = """
        SELECT MIN(ID_registro)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
        """
        cur.execute(get_min_id, (nit, factura))
        result = cur.fetchone()
        
        if not result or not result[0]:
            print("[ERROR] No se encontro registro base")
            cur.close()
            return False
        
        min_id = result[0]
        
        # CORRECCIÓN: Usar min_id directamente (sin subconsulta)
        insert_query = """
        INSERT INTO [dbo].[CxP.Comparativa] (
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            Item, Valor_XML, Valor_Orden_de_Compra,
            Valor_Orden_de_Compra_Comercializados, Aprobado,
            Estado_validacion_antes_de_eventos,
            Fecha_de_retoma_contabilizacion, Estado_contabilizacion,
            Fecha_de_retoma_compensacion, Estado_compensacion
        )
        SELECT 
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            ?, NULL, NULL,
            NULL, NULL,
            NULL,
            NULL, NULL,
            NULL, NULL
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND ID_registro = ?
        """
        cur.execute(insert_query, (item_name, nit, factura, min_id))
        print("[DEBUG] Item '" + item_name + "' creado exitosamente")
        
        cur.close()
        return True
    

    def actualizar_item_comparativa(cx, nit, factura, item_name, aprobado=None, observacion=None, estado=None):
        cur = cx.cursor()
        
        if aprobado is not None:
            update_aprobado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Aprobado = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = ?
            """
            cur.execute(update_aprobado, (aprobado, nit, factura, item_name))
            print("[UPDATE] Item '" + item_name + "' Aprobado = " + aprobado)
        
        if observacion is not None:
            select_obs = """
            SELECT Valor_XML
            FROM [dbo].[CxP.Comparativa]
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(select_obs, (nit, factura))
            result = cur.fetchone()
            
            obs_actual = safe_str(result[0]) if result and result[0] else ""
            
            if obs_actual:
                obs_final = observacion + ", " + obs_actual
            else:
                obs_final = observacion
            
            # CORRECCIÓN: Truncar antes de UPDATE
            obs_final = truncar_observacion(obs_final)
            
            update_obs = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Valor_XML = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(update_obs, (obs_final, nit, factura))
            print("[UPDATE] Observaciones actualizadas")
        
        if estado is not None:
            update_estado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Estado_validacion_antes_de_eventos = ?
            WHERE NIT = ?
              AND Factura = ?
            """
            cur.execute(update_estado, (estado, nit, factura))
            print("[UPDATE] Estado_validacion_antes_de_eventos = " + estado)
        
        cur.close()
    
    def actualizar_documents_processing(cx, nit, factura, oc, observacion, forma_pago):
        cur = cx.cursor()
        
        if forma_pago == '1' or forma_pago == '01':
            estado_final = 'CON NOVEDAD - CONTADO'
        else:
            estado_final = 'CON NOVEDAD'
        
        update_fase4 = """
        UPDATE [CxP].[DocumentsProcessing]
        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_fase4, (nit, factura, oc))
        
        select_obs = """
        SELECT ObservacionesFase_4
        FROM [CxP].[DocumentsProcessing]
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(select_obs, (nit, factura, oc))
        result = cur.fetchone()
        
        obs_actual = safe_str(result[0]) if result and result[0] else ""
        
        if obs_actual:
            obs_final = observacion + ", " + obs_actual
        else:
            obs_final = observacion
        
        # CORRECCIÓN: Truncar antes de UPDATE
        obs_final = truncar_observacion(obs_final)
        
        update_obs = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ObservacionesFase_4 = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_obs, (obs_final, nit, factura, oc))
        
        update_resultado = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ResultadoFinalAntesEventos = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_resultado, (estado_final, nit, factura, oc))
        
        cur.close()
        print("[UPDATE] DocumentsProcessing actualizado (CON NOVEDAD)")
    
    def actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list):
        cur = cx.cursor()
        
        num_actualizados = 0
        max_len = max(len(doccompra_list), len(nitcedula_list), len(porcalcular_list), len(textobreve_list))
        
        for i in range(max_len):
            doccompra = doccompra_list[i] if i < len(doccompra_list) else ""
            nitcedula = nitcedula_list[i] if i < len(nitcedula_list) else ""
            porcalcular = porcalcular_list[i] if i < len(porcalcular_list) else ""
            textobreve = textobreve_list[i] if i < len(textobreve_list) else ""
            
            if doccompra and nitcedula:
                update_marca = """
                UPDATE [CxP].[HistoricoOrdenesCompra]
                SET Marca = 'PROCESADO'
                WHERE DocCompra = ?
                  AND NitCedula = ?
                  AND PorCalcular = ?
                  AND TextoBreve = ?
                """
                cur.execute(update_marca, (doccompra, nitcedula, porcalcular, textobreve))
                num_actualizados += 1
        
        cur.close()
        print("[UPDATE] HistoricoOrdenesCompra: " + str(num_actualizados) + " registros marcados como PROCESADO")
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        stats = {
            'total_registros': 0,
            'posiciones_procesadas': 0,
            'validaciones_ok': 0,
            'validaciones_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                ClaseDePedido_hoc,
                ElementoPEP_hoc,
                IndicadorImpuestos_hoc,
                CentroDeCoste_hoc,
                Cuenta_hoc,
                Emplazamiento_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            print("[PASO 2] Aplicando filtros...")
            
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPPA, ZPCN o 42: " + str(mask_clase.sum()))
            
            df_filtrado = df_candidatos[mask_clase].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            print("")
            print("[PASO 3] Procesando validaciones por posicion...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "]")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    
                    elementos_pep = split_valores(row['ElementoPEP_hoc'])
                    ind_impuestos = split_valores(row['IndicadorImpuestos_hoc'])
                    centros_coste = split_valores(row['CentroDeCoste_hoc'])
                    cuentas = split_valores(row['Cuenta_hoc'])
                    emplazamientos = split_valores(row['Emplazamiento_hoc'])
                    
                    doccompra_list = split_valores(row['DocCompra_hoc'])
                    nitcedula_list = split_valores(row['NitCedula_hoc'])
                    porcalcular_list = split_valores(row['PorCalcular_hoc'])
                    textobreve_list = split_valores(row['TextoBreve_hoc'])
                    
                    print("[DEBUG] Total posiciones en ElementoPEP: " + str(len(elementos_pep)))
                    
                    hay_novedad = False
                    posiciones_procesadas = 0
                    
                    for pos in range(len(elementos_pep)):
                        elemento_pep = elementos_pep[pos] if pos < len(elementos_pep) else ""
                        
                        if not elemento_pep:
                            print("[INFO] Posicion " + str(pos + 1) + ": ElementoPEP vacio, saltando...")
                            continue
                        
                        print("")
                        print("[POSICION " + str(pos + 1) + "/" + str(len(elementos_pep)) + "]")
                        
                        ind_imp = ind_impuestos[pos] if pos < len(ind_impuestos) else ""
                        centro = centros_coste[pos] if pos < len(centros_coste) else ""
                        cuenta = cuentas[pos] if pos < len(cuentas) else ""
                        emplazamiento = emplazamientos[pos] if pos < len(emplazamientos) else ""
                        
                        print("[DEBUG] ElementoPEP: '" + elemento_pep + "'")
                        print("[DEBUG] IndicadorImpuestos: '" + ind_imp + "'")
                        print("[DEBUG] CentroDeCoste: '" + centro + "'")
                        print("[DEBUG] Cuenta: '" + cuenta + "'")
                        print("[DEBUG] Emplazamiento: '" + emplazamiento + "'")
                        
                        posiciones_procesadas += 1
                        
                        if pos > 0:
                            verificar_y_crear_item(cx, nit, factura, 'IndicadorImpuestos')
                            verificar_y_crear_item(cx, nit, factura, 'CentroCoste')
                            verificar_y_crear_item(cx, nit, factura, 'Cuenta')
                            verificar_y_crear_item(cx, nit, factura, 'Emplazamiento')
                            verificar_y_crear_item(cx, nit, factura, 'Observaciones')
                        
                        ind_validos = ['H4', 'H5', 'H6', 'H7', 'VP', 'CO', 'IC', 'CR']
                        
                        if ind_imp in ind_validos:
                            print("[OK] IndicadorImpuestos '" + ind_imp + "' es valido")
                            actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='SI')
                        elif not ind_imp:
                            print("[NOVEDAD] IndicadorImpuestos vacio")
                            hay_novedad = True
                            obs = "Pedido corresponde a ZPCN o ZPPA con Elemento PEP, pero campo 'Indicador impuestos' NO se encuentra diligenciado"
                            actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                            actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        else:
                            print("[NOVEDAD] IndicadorImpuestos NO valido: '" + ind_imp + "'")
                            hay_novedad = True
                            obs = "Pedido corresponde a ZPCN o ZPPA con Elemento PEP, pero campo 'Indicador impuestos' NO corresponde alguna de las opciones H4, H5, H6, H7, VP, CO, IC, CR"
                            actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                            actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        if not centro:
                            print("[OK] CentroDeCoste vacio (correcto)")
                            actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                        else:
                            print("[NOVEDAD] CentroDeCoste tiene valor (incorrecto)")
                            hay_novedad = True
                            obs = "Pedido corresponde a ZPCN o ZPPA con Elemento PEP, pero Campo 'Centro de coste' se encuentra diligenciado cuando NO debe estarlo"
                            actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                            actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        if cuenta == '5199150001':
                            print("[OK] Cuenta es 5199150001 (correcto)")
                            actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                        else:
                            print("[NOVEDAD] Cuenta diferente a 5199150001")
                            hay_novedad = True
                            obs = "Pedido corresponde a ZPCN o ZPPA con Elemento PEP, pero Campo 'Cuenta' es diferente a 5199150001"
                            actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                            actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        emplazamiento_correcto = False
                        obs_emplazamiento = ""
                        
                        if not emplazamiento:
                            obs_emplazamiento = "Pedido corresponde a ZPCN o ZPPA y cuenta con Elemento PEP, pero Campo 'Emplazamiento' NO se encuentra diligenciado"
                        elif ind_imp in ['H4', 'H5']:
                            emplazamiento_correcto = (emplazamiento == 'DCTO_01')
                            if not emplazamiento_correcto:
                                obs_emplazamiento = "Pedido corresponde a ZPCN o ZPPA y cuenta con Elemento PEP, pero Campo 'Emplazamiento' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = DCTO_01', 'H6 y H7 = GTO_02' o 'VP, CO, CR o IC = DCTO_01 o GTO_02'"
                        elif ind_imp in ['H6', 'H7']:
                            emplazamiento_correcto = (emplazamiento == 'GTO_02')
                            if not emplazamiento_correcto:
                                obs_emplazamiento = "Pedido corresponde a ZPCN o ZPPA y cuenta con Elemento PEP, pero Campo 'Emplazamiento' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = DCTO_01', 'H6 y H7 = GTO_02' o 'VP, CO, CR o IC = DCTO_01 o GTO_02'"
                        elif ind_imp in ['VP', 'CO', 'CR', 'IC']:
                            emplazamiento_correcto = (emplazamiento in ['DCTO_01', 'GTO_02'])
                            if not emplazamiento_correcto:
                                obs_emplazamiento = "Pedido corresponde a ZPCN o ZPPA y cuenta con Elemento PEP, pero Campo 'Emplazamiento' NO se encuentra aplicado correctamente segun reglas 'H4 y H5 = DCTO_01', 'H6 y H7 = GTO_02' o 'VP, CO, CR o IC = DCTO_01 o GTO_02'"
                        else:
                            emplazamiento_correcto = True
                        
                        if emplazamiento_correcto:
                            print("[OK] Emplazamiento correcto segun IndicadorImpuestos")
                            actualizar_item_comparativa(cx, nit, factura, 'Emplazamiento', aprobado='SI')
                        else:
                            print("[NOVEDAD] Emplazamiento incorrecto")
                            hay_novedad = True
                            actualizar_documents_processing(cx, nit, factura, oc, obs_emplazamiento, forma_pago)
                            actualizar_item_comparativa(cx, nit, factura, 'Emplazamiento', aprobado='NO')
                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs_emplazamiento)
                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                    
                    stats['posiciones_procesadas'] += posiciones_procesadas
                    if hay_novedad:
                        stats['validaciones_novedad'] += 1
                    else:
                        stats['validaciones_ok'] += 1
                    
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Posiciones procesadas: " + str(stats['posiciones_procesadas']))
            print("  Validaciones OK: " + str(stats['validaciones_ok']))
            print("  Validaciones con novedad: " + str(stats['validaciones_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " Posiciones:" + str(stats['posiciones_procesadas']) +
                   " OK:" + str(stats['validaciones_ok']) + 
                   " Novedad:" + str(stats['validaciones_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ZPCN_ZPPA_ValidarActivoFijo():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    import os
    import unicodedata
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ZPCN_ZPPA_ValidarActivoFijo() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""

    def truncar_observacion(obs):
        """Truncar observación a 3900 caracteres para prevenir overflow"""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > 3900:
            return obs_str[:3900]
        return obs_str
    
    def normalizar_columna(nombre):
        """Quitar tildes y normalizar nombre de columna"""
        if not nombre:
            return ""
        # Quitar tildes
        nfd = unicodedata.normalize('NFD', nombre)
        sin_tildes = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
        return unicodedata.normalize('NFC', sin_tildes)
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        """Dividir string por | y retornar lista de valores"""
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valor_buscado):
        """Verificar si campo (que puede tener valores separados por |) contiene valor buscado"""
        valores = split_valores(campo)
        return valor_buscado in valores
    
    def verificar_y_crear_item(cx, nit, factura, item_name):
        cur = cx.cursor()
        
        check_query = """
        SELECT COUNT(*)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND Item = ?
        """
        cur.execute(check_query, (nit, factura, item_name))
        count = cur.fetchone()[0]
        
        if count > 0:
            print("[DEBUG] Item '" + item_name + "' ya existe")
            cur.close()
            return True
        
        print("[INFO] Creando Item '" + item_name + "'...")
        
        # CORRECCIÓN: Obtener min_id primero
        get_min_id = """
        SELECT MIN(ID_registro)
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
        """
        cur.execute(get_min_id, (nit, factura))
        result = cur.fetchone()
        
        if not result or not result[0]:
            print("[ERROR] No se encontro registro base")
            cur.close()
            return False
        
        min_id = result[0]
        
        # CORRECCIÓN: Usar min_id directamente (sin subconsulta)
        insert_query = """
        INSERT INTO [dbo].[CxP.Comparativa] (
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            Item, Valor_XML, Valor_Orden_de_Compra,
            Valor_Orden_de_Compra_Comercializados, Aprobado,
            Estado_validacion_antes_de_eventos,
            Fecha_de_retoma_contabilizacion, Estado_contabilizacion,
            Fecha_de_retoma_compensacion, Estado_compensacion
        )
        SELECT 
            Fecha_de_ejecucion, Fecha_de_retoma_antes_de_contabilizacion,
            ID_ejecucion, Tipo_de_documento, Orden_de_Compra,
            Clase_de_pedido, NIT, Nombre_Proveedor, Factura,
            ?, NULL, NULL,
            NULL, NULL,
            NULL,
            NULL, NULL,
            NULL, NULL
        FROM [dbo].[CxP.Comparativa]
        WHERE NIT = ?
          AND Factura = ?
          AND ID_registro = ?
        """
        cur.execute(insert_query, (item_name, nit, factura, min_id))
        print("[DEBUG] Item '" + item_name + "' creado exitosamente")
        
        cur.close()
        return True
    
    def actualizar_item_comparativa(cx, nit, factura, item_name, aprobado=None, observacion=None, estado=None):
        """Actualizar campos de un Item en Comparativa"""
        cur = cx.cursor()
        
        if aprobado is not None:
            update_aprobado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Aprobado = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = ?
            """
            cur.execute(update_aprobado, (aprobado, nit, factura, item_name))
            print("[UPDATE] Item '" + item_name + "' Aprobado = " + aprobado)
        
        if observacion is not None:
            # Obtener observaciones actuales
            select_obs = """
            SELECT Valor_XML
            FROM [dbo].[CxP.Comparativa]
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(select_obs, (nit, factura))
            result = cur.fetchone()
            
            obs_actual = safe_str(result[0]) if result and result[0] else ""
            
            if obs_actual:
                obs_final = observacion + ", " + obs_actual
            else:
                obs_final = observacion

            obs_final = truncar_observacion(obs_final)
            
            update_obs = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Valor_XML = ?
            WHERE NIT = ?
              AND Factura = ?
              AND Item = 'Observaciones'
            """
            cur.execute(update_obs, (obs_final, nit, factura))
            print("[UPDATE] Observaciones actualizadas")
        
        if estado is not None:
            update_estado = """
            UPDATE [dbo].[CxP.Comparativa]
            SET Estado_validacion_antes_de_eventos = ?
            WHERE NIT = ?
              AND Factura = ?
            """
            cur.execute(update_estado, (estado, nit, factura))
            print("[UPDATE] Estado_validacion_antes_de_eventos = " + estado)
        
        cur.close()
    
    def actualizar_documents_processing(cx, nit, factura, oc, observacion, forma_pago):
        """Actualizar DocumentsProcessing con novedad"""
        cur = cx.cursor()
        
        # Determinar estado
        if forma_pago == '1' or forma_pago == '01':
            estado_final = 'CON NOVEDAD - CONTADO'
        else:
            estado_final = 'CON NOVEDAD'
        
        # Actualizar EstadoFinalFase_4
        update_fase4 = """
        UPDATE [CxP].[DocumentsProcessing]
        SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso'
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_fase4, (nit, factura, oc))
        
        # Obtener observaciones actuales
        select_obs = """
        SELECT ObservacionesFase_4
        FROM [CxP].[DocumentsProcessing]
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(select_obs, (nit, factura, oc))
        result = cur.fetchone()
        
        obs_actual = safe_str(result[0]) if result and result[0] else ""
        
        if obs_actual:
            obs_final = observacion + ", " + obs_actual
        else:
            obs_final = observacion

        obs_final = truncar_observacion(obs_final)
        
        # Actualizar observaciones
        update_obs = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ObservacionesFase_4 = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_obs, (obs_final, nit, factura, oc))
        
        # Actualizar ResultadoFinalAntesEventos
        update_resultado = """
        UPDATE [CxP].[DocumentsProcessing]
        SET ResultadoFinalAntesEventos = ?
        WHERE nit_emisor_o_nit_del_proveedor = ?
          AND numero_de_factura = ?
          AND numero_de_liquidacion_u_orden_de_compra = ?
        """
        cur.execute(update_resultado, (estado_final, nit, factura, oc))
        
        cur.close()
        print("[UPDATE] DocumentsProcessing actualizado (CON NOVEDAD)")
    
    def actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list):
        """Actualizar HistoricoOrdenesCompra marcando como PROCESADO"""
        cur = cx.cursor()
        
        num_actualizados = 0
        max_len = max(len(doccompra_list), len(nitcedula_list), len(porcalcular_list), len(textobreve_list))
        
        for i in range(max_len):
            doccompra = doccompra_list[i] if i < len(doccompra_list) else ""
            nitcedula = nitcedula_list[i] if i < len(nitcedula_list) else ""
            porcalcular = porcalcular_list[i] if i < len(porcalcular_list) else ""
            textobreve = textobreve_list[i] if i < len(textobreve_list) else ""
            
            if doccompra and nitcedula:
                update_marca = """
                UPDATE [CxP].[HistoricoOrdenesCompra]
                SET Marca = 'PROCESADO'
                WHERE DocCompra = ?
                  AND NitCedula = ?
                  AND PorCalcular = ?
                  AND TextoBreve = ?
                """
                cur.execute(update_marca, (doccompra, nitcedula, porcalcular, textobreve))
                num_actualizados += 1
        
        cur.close()
        print("[UPDATE] HistoricoOrdenesCompra: " + str(num_actualizados) + " registros marcados como PROCESADO")
    
    def cargar_excel_impuestos(ruta_archivo):
        """
        Cargar archivo Excel de impuestos especiales.
        FIXED: Normaliza nombres de columnas para evitar problemas de encoding
        Retorna DataFrame o genera error critico.
        """
        print("")
        print("[EXCEL] Cargando archivo de Impuestos Especiales...")
        print("[EXCEL] Ruta: " + ruta_archivo)
        
        # Verificar existencia del archivo
        if not os.path.exists(ruta_archivo):
            raise ValueError("ERROR CRITICO: Archivo Excel no existe en ruta: " + ruta_archivo)
        
        # Verificar que no esta vacio
        if os.path.getsize(ruta_archivo) == 0:
            raise ValueError("ERROR CRITICO: Archivo Excel esta vacio: " + ruta_archivo)
        
        try:
            # Intentar cargar el archivo
            df = pd.read_excel(ruta_archivo, sheet_name='IVA CECO')
            
            if df.empty:
                raise ValueError("ERROR CRITICO: Hoja 'IVA CECO' esta vacia")
            
            print("[EXCEL] Archivo cargado exitosamente")
            print("[EXCEL] Registros: " + str(len(df)))
            
            # FIXED: Normalizar nombres de columnas (quitar tildes)
            print("[EXCEL] Normalizando nombres de columnas...")
            df.columns = [normalizar_columna(col) for col in df.columns]
            
            print("[EXCEL] Columnas despues de normalizar:")
            for col in df.columns:
                print("  - " + col)
            
            # Verificar estructura (SIN TILDES)
            columnas_requeridas = ['CECO', 'Codigo Ind. Iva aplicable']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                print("[ERROR] Columnas disponibles: " + str(list(df.columns)))
                raise ValueError("ERROR CRITICO: Estructura incorrecta. Columnas faltantes: " + ', '.join(columnas_faltantes))
            
            print("[EXCEL] Estructura verificada OK")
            return df
            
        except ValueError as ve:
            # Re-lanzar errores de validacion
            raise
        except Exception as e:
            raise ValueError("ERROR CRITICO: Error al cargar Excel: " + str(e))
    
    def buscar_indicadores_permitidos(df_excel, centro_coste):
        """
        Buscar en Excel los indicadores permitidos para un Centro de Coste.
        FIXED: Usa columna sin tildes
        Retorna lista de indicadores permitidos o None si no se encuentra.
        """
        try:
            # Convertir centro_coste a int para buscar
            ceco_int = int(centro_coste) if centro_coste.isdigit() else None
            
            if ceco_int is None:
                print("[EXCEL] CentroDeCoste no es numerico: '" + centro_coste + "'")
                return None
            
            # Buscar en DataFrame
            resultado = df_excel[df_excel['CECO'] == ceco_int]
            
            if resultado.empty:
                print("[EXCEL] CECO " + str(ceco_int) + " NO encontrado en Excel")
                return None
            
            # Extraer codigo (SIN TILDES)
            codigo = str(resultado.iloc[0]['Codigo Ind. Iva aplicable'])
            print("[EXCEL] CECO " + str(ceco_int) + " encontrado -> Codigo: '" + codigo + "'")
            
            # Separar por guion
            if '-' in codigo:
                indicadores = codigo.split('-')
            else:
                indicadores = [codigo]
            
            print("[EXCEL] Indicadores permitidos: " + str(indicadores))
            return indicadores
            
        except Exception as e:
            print("[ERROR] Error buscando en Excel: " + str(e))
            return None
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        # Verificar ruta del Excel
        if 'DocImpuestosEspeciales' not in cfg:
            raise ValueError("ERROR CRITICO: Configuracion no contiene 'DocImpuestosEspeciales'")
        
        ruta_excel = cfg['DocImpuestosEspeciales']
        
        # Cargar Excel (puede generar error critico)
        df_impuestos = cargar_excel_impuestos(ruta_excel)
        
        stats = {
            'total_registros': 0,
            'posiciones_procesadas': 0,
            'validaciones_ok': 0,
            'validaciones_novedad': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 1: Consultar candidatos de HU41_CandidatosValidacion
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando tabla HU41_CandidatosValidacion...")
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                forma_de_pago_dp,
                ClaseDePedido_hoc,
                ActivoFijo_hoc,
                IndicadorImpuestos_hoc,
                CentroDeCoste_hoc,
                Cuenta_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Registros consultados: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            # ================================================================
            # PASO 2: Aplicar filtros
            # ================================================================
            
            print("[PASO 2] Aplicando filtros...")
            
            # Filtro: ClaseDePedido_hoc contiene 'ZPPA', 'ZPCN' o '42'
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, 'ZPPA') or contiene_valor(x, 'ZPCN') or contiene_valor(x, '42') if pd.notna(x) else False
            )
            
            print("[DEBUG] Registros con ClaseDePedido = ZPPA, ZPCN o 42: " + str(mask_clase.sum()))
            
            df_filtrado = df_candidatos[mask_clase].copy()
            
            print("[DEBUG] Registros despues de filtros: " + str(len(df_filtrado)))
            
            if df_filtrado.empty:
                print("[INFO] No hay registros que cumplan los filtros")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido ZPPA/ZPCN/42")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros que cumplan filtros", None, stats
            
            stats['total_registros'] = len(df_filtrado)
            
            # ================================================================
            # PASO 3: Procesar cada registro - VALIDACIONES POR POSICION
            # ================================================================
            
            print("")
            print("[PASO 3] Procesando validaciones por posicion...")
            
            for idx, row in df_filtrado.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_filtrado)) + "]")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    forma_pago = safe_str(row['forma_de_pago_dp'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    
                    # Obtener listas de valores
                    activos_fijos = split_valores(row['ActivoFijo_hoc'])
                    ind_impuestos = split_valores(row['IndicadorImpuestos_hoc'])
                    centros_coste = split_valores(row['CentroDeCoste_hoc'])
                    cuentas = split_valores(row['Cuenta_hoc'])
                    
                    # Para HistoricoOrdenesCompra
                    doccompra_list = split_valores(row['DocCompra_hoc'])
                    nitcedula_list = split_valores(row['NitCedula_hoc'])
                    porcalcular_list = split_valores(row['PorCalcular_hoc'])
                    textobreve_list = split_valores(row['TextoBreve_hoc'])
                    
                    max_posiciones = max(len(activos_fijos), len(ind_impuestos), len(centros_coste), len(cuentas))
                    print("[DEBUG] Total posiciones: " + str(max_posiciones))
                    
                    hay_novedad = False
                    posiciones_procesadas = 0
                    
                    # PROCESAR POR POSICION
                    for pos in range(max_posiciones):
                        print("")
                        print("[POSICION " + str(pos + 1) + "/" + str(max_posiciones) + "]")
                        
                        activo_fijo = activos_fijos[pos] if pos < len(activos_fijos) else ""
                        ind_imp = ind_impuestos[pos] if pos < len(ind_impuestos) else ""
                        centro = centros_coste[pos] if pos < len(centros_coste) else ""
                        cuenta = cuentas[pos] if pos < len(cuentas) else ""
                        
                        print("[DEBUG] ActivoFijo: '" + activo_fijo + "'")
                        print("[DEBUG] IndicadorImpuestos: '" + ind_imp + "'")
                        print("[DEBUG] CentroDeCoste: '" + centro + "'")
                        print("[DEBUG] Cuenta: '" + cuenta + "'")
                        
                        posiciones_procesadas += 1
                        
                        # Verificar/crear Items necesarios (a partir de segunda iteracion)
                        if pos > 0:
                            verificar_y_crear_item(cx, nit, factura, 'IndicadorImpuestos')
                            verificar_y_crear_item(cx, nit, factura, 'CentroCoste')
                            verificar_y_crear_item(cx, nit, factura, 'Cuenta')
                            verificar_y_crear_item(cx, nit, factura, 'Observaciones')
                        
                        # ====================================================
                        # CAMINO 1: ACTIVO FIJO CON VALOR
                        # ====================================================
                        
                        if activo_fijo:
                            # Verificar tipo de Activo Fijo
                            es_diferido = (len(activo_fijo) == 10 and activo_fijo.startswith('2000'))
                            es_bonos = (len(activo_fijo) == 9 and activo_fijo.startswith('8000'))
                            
                            if not es_diferido and not es_bonos:
                                print("[INFO] ActivoFijo '" + activo_fijo + "' no es DIFERIDO ni BONOS, ignorando posicion")
                                continue
                            
                            # ============================================
                            # CAMINO 1.1: DIFERIDO (2000* - 10 digitos)
                            # ============================================
                            
                            if es_diferido:
                                print("[VALIDACION] Activo Fijo DIFERIDO detectado")
                                
                                # Validacion 1: IndicadorImpuestos
                                ind_validos_diferido = ['C1', 'FA', 'VP', 'CO', 'CR']
                                
                                if ind_imp in ind_validos_diferido:
                                    print("[OK] IndicadorImpuestos '" + ind_imp + "' valido para DIFERIDO")
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='SI')
                                elif not ind_imp:
                                    print("[NOVEDAD] IndicadorImpuestos vacio para DIFERIDO")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero campo 'Indicador impuestos' NO se encuentra diligenciado para pedido DIFERIDO"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                else:
                                    print("[NOVEDAD] IndicadorImpuestos NO valido para DIFERIDO: '" + ind_imp + "'")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA Activo Fijo, pero campo 'Indicador impuestos' NO corresponde alguna de las opciones 'C1', 'FA', 'VP', 'CO' o 'CR' para pedido DIFERIDO"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                # Validacion 2: CentroDeCoste (debe estar vacio)
                                if not centro:
                                    print("[OK] CentroDeCoste vacio (correcto para DIFERIDO)")
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                                else:
                                    print("[NOVEDAD] CentroDeCoste tiene valor para DIFERIDO")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero Campo 'Centro de coste' se encuentra diligenciado cuando NO debe estarlo para pedido DIFERIDO"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                # Validacion 3: Cuenta (debe ser 2695950020)
                                if cuenta == '2695950020':
                                    print("[OK] Cuenta es 2695950020 (correcto para DIFERIDO)")
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                                else:
                                    print("[NOVEDAD] Cuenta diferente a 2695950020 para DIFERIDO")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero Campo 'Cuenta' no es igual 2695950020 pedido DIFERIDO"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            # ============================================
                            # CAMINO 1.2: BONOS (8000* - 9 digitos)
                            # ============================================
                            
                            elif es_bonos:
                                print("[VALIDACION] Activo Fijo BONOS detectado")
                                
                                # Validacion 1: IndicadorImpuestos (sin FA)
                                ind_validos_bonos = ['C1', 'VP', 'CO', 'CR']
                                
                                if ind_imp in ind_validos_bonos:
                                    print("[OK] IndicadorImpuestos '" + ind_imp + "' valido para BONOS")
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='SI')
                                elif not ind_imp:
                                    print("[NOVEDAD] IndicadorImpuestos vacio para BONOS")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero campo 'Indicador impuestos' NO se encuentra diligenciado para pedido EQUIVALENTE AL EFECTIVO - BONOS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                else:
                                    print("[NOVEDAD] IndicadorImpuestos NO valido para BONOS: '" + ind_imp + "'")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero campo 'Indicador impuestos' NO corresponde alguna de las opciones 'C1', 'VP', 'CO' o 'CR' para pedido EQUIVALENTE AL EFECTIVO - BONOS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                # Validacion 2: CentroDeCoste (debe estar vacio)
                                if not centro:
                                    print("[OK] CentroDeCoste vacio (correcto para BONOS)")
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                                else:
                                    print("[NOVEDAD] CentroDeCoste tiene valor para BONOS")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero Campo 'Centro de coste' se encuentra diligenciado cuando NO debe estarlo para pedido EQUIVALENTE AL EFECTIVO - BONOS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                # Validacion 3: Cuenta (debe ser 2695950020)
                                if cuenta == '2695950020':
                                    print("[OK] Cuenta es 2695950020 (correcto para BONOS)")
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                                else:
                                    print("[NOVEDAD] Cuenta diferente a 2695950020 para BONOS")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA con Activo Fijo, pero Campo 'Cuenta' NO ES IGUAL A 2695950020, para pedido EQUIVALENTE AL EFECTIVO - BONOS"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                        
                        # ====================================================
                        # CAMINO 2: ACTIVO FIJO VACIO (GENERALES)
                        # ====================================================
                        
                        else:
                            print("[VALIDACION] Sin Activo Fijo (GENERALES)")
                            
                            # Validacion 1: Cuenta debe tener valor
                            if cuenta:
                                print("[OK] Cuenta tiene valor (correcto para GENERALES)")
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='SI')
                            else:
                                print("[NOVEDAD] Cuenta vacia para GENERALES")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA sin Activo Fijo, pero Campo 'Cuenta' NO se encuentra diligenciado cuando debe estarlo para pedido GENERALES"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'Cuenta', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                continue  # No continuar con validaciones de GENERALES
                            
                            # Validacion 2: IndicadorImpuestos
                            if ind_imp:
                                print("[DEBUG] IndicadorImpuestos tiene valor: '" + ind_imp + "'")
                                
                                # Validacion 3: CentroDeCoste
                                if centro:
                                    print("[DEBUG] CentroDeCoste tiene valor: '" + centro + "'")
                                    
                                    # Marcar CentroCoste como SI inicialmente
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='SI')
                                    
                                    # Buscar en Excel
                                    indicadores_permitidos = buscar_indicadores_permitidos(df_impuestos, centro)
                                    
                                    if indicadores_permitidos is None:
                                        # CentroDeCoste NO encontrado en Excel
                                        print("[NOVEDAD] CentroDeCoste NO encontrado en Excel")
                                        hay_novedad = True
                                        obs = "Pedido corresponde a ZPCN o ZPPA sin Activo Fijo, pero Campo 'Centro de coste' NO se encuentra diligenciado en el archivo Impuestos especiales CXP, cuando debe estarlo para pedido GENERALES"
                                        actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                        actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                        actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                        actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                        estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                        actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                        actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                    else:
                                        # Validar que IndicadorImpuestos este en lista permitida
                                        if ind_imp in indicadores_permitidos:
                                            print("[OK] IndicadorImpuestos '" + ind_imp + "' es valido segun Excel")
                                            actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='SI')
                                        else:
                                            print("[NOVEDAD] IndicadorImpuestos NO valido segun Excel")
                                            hay_novedad = True
                                            obs = "Pedido corresponde a ZPCN o ZPPA sin Activo Fijo, pero campo 'Indicador impuestos' NO se encuentra diligenciado correctamente segun los indicadores (" + ', '.join(indicadores_permitidos) + ")"
                                            actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                            actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                            actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                            estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                            actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                            actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                                
                                else:
                                    # CentroDeCoste vacio
                                    print("[NOVEDAD] CentroDeCoste vacio para GENERALES")
                                    hay_novedad = True
                                    obs = "Pedido corresponde a ZPCN o ZPPA sin Activo Fijo, pero Campo 'Centro de coste' NO se encuentra diligenciado cuando debe estarlo para pedido GENERALES"
                                    actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                    actualizar_item_comparativa(cx, nit, factura, 'CentroCoste', aprobado='NO')
                                    actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                    estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                    actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                    actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                            
                            else:
                                # IndicadorImpuestos vacio
                                print("[NOVEDAD] IndicadorImpuestos vacio para GENERALES")
                                hay_novedad = True
                                obs = "Pedido corresponde a ZPCN o ZPPA sin Activo Fijo, pero campo 'Indicador impuestos' NO se encuentra diligenciado para pedido GENERALES"
                                actualizar_documents_processing(cx, nit, factura, oc, obs, forma_pago)
                                actualizar_item_comparativa(cx, nit, factura, 'IndicadorImpuestos', aprobado='NO')
                                actualizar_item_comparativa(cx, nit, factura, None, observacion=obs)
                                estado = 'CON NOVEDAD - CONTADO' if forma_pago in ['1', '01'] else 'CON NOVEDAD'
                                actualizar_item_comparativa(cx, nit, factura, None, estado=estado)
                                actualizar_historico_ordenes(cx, doccompra_list, nitcedula_list, porcalcular_list, textobreve_list)
                    
                    # Actualizar estadisticas
                    stats['posiciones_procesadas'] += posiciones_procesadas
                    if hay_novedad:
                        stats['validaciones_novedad'] += 1
                    else:
                        stats['validaciones_ok'] += 1
                    
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Proceso completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  Posiciones procesadas: " + str(stats['posiciones_procesadas']))
            print("  Validaciones OK: " + str(stats['validaciones_ok']))
            print("  Validaciones con novedad: " + str(stats['validaciones_novedad']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Proceso OK. Total:" + str(stats['total_registros']) + 
                   " Posiciones:" + str(stats['posiciones_procesadas']) +
                   " OK:" + str(stats['validaciones_ok']) + 
                   " Novedad:" + str(stats['validaciones_novedad']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def PostProcesamiento_EstadosFinales():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion PostProcesamiento_EstadosFinales() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def split_valores(valor_str):
        """Dividir string por | y retornar lista de valores"""
        if not valor_str or valor_str == "" or pd.isna(valor_str):
            return []
        valores = str(valor_str).split('|')
        return [v.strip() for v in valores if v.strip()]
    
    def contiene_valor(campo, valores_buscados):
        """Verificar si campo contiene alguno de los valores buscados"""
        valores = split_valores(campo)
        for v in valores_buscados:
            if v in valores:
                return True
        return False
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        stats = {
            'total_registros': 0,
            'con_novedad_clase31': 0,
            'aprobado_sin_contab': 0,
            'aprobado_contado': 0,
            'aprobado': 0,
            'errores': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        fecha_ejecucion = datetime.now()
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 1: Consultar registros de HU41_CandidatosValidacion
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando HU41_CandidatosValidacion...")
            
            # FILTRO MODIFICABLE: Filtrar por clases de pedido
            clases_pedido_filtro = ['ZPRE', 'ZPPA', 'ZPCN', '45', '42']
            
            query_candidatos = """
            SELECT 
                ID_dp,
                nit_emisor_o_nit_del_proveedor_dp,
                numero_de_factura_dp,
                numero_de_liquidacion_u_orden_de_compra_dp,
                ClaseDePedido_hoc,
                ClaseDeImpuesto_hoc,
                DocCompra_hoc,
                NitCedula_hoc,
                PorCalcular_hoc,
                TextoBreve_hoc
            FROM [CxP].[HU41_CandidatosValidacion] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_candidatos = pd.read_sql(query_candidatos, cx)
            print("[DEBUG] Total registros en HU41_CandidatosValidacion: " + str(len(df_candidatos)))
            
            if df_candidatos.empty:
                print("[INFO] No hay registros en HU41_CandidatosValidacion")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros para procesar")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros para procesar", None, stats
            
            # Aplicar filtro de ClaseDePedido
            print("[DEBUG] Aplicando filtro de ClaseDePedido: " + str(clases_pedido_filtro))
            mask_clase = df_candidatos['ClaseDePedido_hoc'].apply(
                lambda x: contiene_valor(x, clases_pedido_filtro) if pd.notna(x) else False
            )
            
            df_candidatos_filtrado = df_candidatos[mask_clase].copy()
            print("[DEBUG] Registros despues de filtro: " + str(len(df_candidatos_filtrado)))
            
            if df_candidatos_filtrado.empty:
                print("[INFO] No hay registros que cumplan el filtro de ClaseDePedido")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay registros con ClaseDePedido valido")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay registros con ClaseDePedido valido", None, stats
            
            # ================================================================
            # PASO 2: Consultar DocumentsProcessing
            # ================================================================
            
            print("")
            print("[PASO 2] Consultando DocumentsProcessing...")
            
            query_dp = """
            SELECT 
                nit_emisor_o_nit_del_proveedor,
                numero_de_factura,
                numero_de_liquidacion_u_orden_de_compra,
                forma_de_pago,
                ResultadoFinalAntesEventos,
                ObservacionesFase_4
            FROM [CxP].[DocumentsProcessing] WITH (NOLOCK)
            WHERE 1=1
            """
            
            df_dp = pd.read_sql(query_dp, cx)
            print("[DEBUG] Registros en DocumentsProcessing: " + str(len(df_dp)))
            
            # ================================================================
            # PASO 3: Hacer INNER JOIN
            # ================================================================
            
            print("[PASO 3] Haciendo INNER JOIN entre tablas...")
            
            df_merged = pd.merge(
                df_candidatos_filtrado,
                df_dp,
                left_on=['nit_emisor_o_nit_del_proveedor_dp', 'numero_de_factura_dp', 'numero_de_liquidacion_u_orden_de_compra_dp'],
                right_on=['nit_emisor_o_nit_del_proveedor', 'numero_de_factura', 'numero_de_liquidacion_u_orden_de_compra'],
                how='inner'
            )
            
            print("[DEBUG] Registros despues de INNER JOIN: " + str(len(df_merged)))
            
            if df_merged.empty:
                print("[INFO] No hay registros que coincidan entre HU41 y DocumentsProcessing")
                SetVar("vLocStrResultadoSP", "True")
                SetVar("vLocStrResumenSP", "No hay coincidencias entre tablas")
                SetVar("vLocDicEstadisticas", str(stats))
                return True, "No hay coincidencias", None, stats
            
            stats['total_registros'] = len(df_merged)
            
            # ================================================================
            # PASO 4: Procesar cada registro
            # ================================================================
            
            print("")
            print("[PASO 4] Procesando post-validaciones...")
            
            for idx, row in df_merged.iterrows():
                try:
                    print("")
                    print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_merged)) + "]")
                    
                    nit = safe_str(row['nit_emisor_o_nit_del_proveedor_dp'])
                    factura = safe_str(row['numero_de_factura_dp'])
                    oc = safe_str(row['numero_de_liquidacion_u_orden_de_compra_dp'])
                    
                    resultado_actual = safe_str(row['ResultadoFinalAntesEventos'])
                    forma_pago = safe_str(row['forma_de_pago'])
                    clase_impuesto_completo = safe_str(row['ClaseDeImpuesto_hoc'])
                    obs_actual = safe_str(row['ObservacionesFase_4'])
                    
                    print("[DEBUG] NIT: " + nit)
                    print("[DEBUG] Factura: " + factura)
                    print("[DEBUG] OC: " + oc)
                    print("[DEBUG] ResultadoFinalAntesEventos: '" + resultado_actual + "'")
                    print("[DEBUG] forma_de_pago: '" + forma_pago + "'")
                    print("[DEBUG] ClaseDeImpuesto_hoc: '" + clase_impuesto_completo + "'")
                    
                    # CORRECCION: Verificar si ClaseDeImpuesto_hoc contiene '31'
                    tiene_clase31 = contiene_valor(clase_impuesto_completo, ['31'])
                    print("[DEBUG] Contiene Clase 31: " + str(tiene_clase31))
                    
                    tiene_con_novedad = "CON NOVEDAD" in resultado_actual.upper()
                    
                    # ====================================================
                    # CASO 1: TIENE CON NOVEDAD
                    # ====================================================
                    
                    if tiene_con_novedad:
                        print("[VALIDACION] Registro con CON NOVEDAD")
                        
                        if tiene_clase31:
                            print("[ACCION] ClaseDeImpuesto contiene 31 (ZOMAC-ZESE)")
                            stats['con_novedad_clase31'] += 1
                            
                            # Actualizar ResultadoFinalAntesEventos
                            # Agregar "EXCLUIDOS CONTABILIZACION" al final
                            nuevo_resultado = resultado_actual
                            if "EXCLUIDOS CONTABILIZACION" not in nuevo_resultado:
                                nuevo_resultado = resultado_actual + " EXCLUIDOS CONTABILIZACION"
                            
                            # Actualizar observaciones (PREPEND)
                            nueva_obs = "Factura corresponde a Clase de impuesto 31 (ZOMAC-ZESE)"
                            if obs_actual:
                                obs_final = nueva_obs + ", " + obs_actual
                            else:
                                obs_final = nueva_obs
                            
                            # Actualizar DocumentsProcessing
                            cur = cx.cursor()
                            
                            update_dp = """
                            UPDATE [CxP].[DocumentsProcessing]
                            SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso',
                                ObservacionesFase_4 = ?,
                                ResultadoFinalAntesEventos = ?
                            WHERE nit_emisor_o_nit_del_proveedor = ?
                              AND numero_de_factura = ?
                              AND numero_de_liquidacion_u_orden_de_compra = ?
                            """
                            cur.execute(update_dp, (obs_final, nuevo_resultado, nit, factura, oc))
                            print("[UPDATE] DocumentsProcessing actualizado")
                            
                            # Actualizar Comparativa - ESTADO en todas las filas
                            update_comparativa = """
                            UPDATE [dbo].[CxP.Comparativa]
                            SET Estado_validacion_antes_de_eventos = ?
                            WHERE NIT = ?
                              AND Factura = ?
                            """
                            cur.execute(update_comparativa, (nuevo_resultado, nit, factura))
                            print("[UPDATE] Comparativa actualizada (ESTADO)")
                            
                            cur.close()
                        
                        else:
                            print("[INFO] ClaseDeImpuesto NO contiene 31, no se procesa")
                    
                    # ====================================================
                    # CASO 2: NO TIENE CON NOVEDAD (APROBADO o vacio)
                    # ====================================================
                    
                    else:
                        print("[VALIDACION] Registro SIN CON NOVEDAD")
                        
                        # SUBCASO 2.1: ClaseDeImpuesto contiene 31
                        if tiene_clase31:
                            print("[ACCION] ClaseDeImpuesto contiene 31 (ZOMAC-ZESE)")
                            stats['aprobado_sin_contab'] += 1
                            
                            nuevo_resultado = "APROBADO SIN CONTABILIZACION"
                            
                            # Actualizar observaciones (PREPEND)
                            nueva_obs = "Factura corresponde a Clase de impuesto 31 (ZOMAC-ZESE)"
                            if obs_actual:
                                obs_final = nueva_obs + ", " + obs_actual
                            else:
                                obs_final = nueva_obs
                            
                            # Actualizar DocumentsProcessing
                            cur = cx.cursor()
                            
                            update_dp = """
                            UPDATE [CxP].[DocumentsProcessing]
                            SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso',
                                ObservacionesFase_4 = ?,
                                ResultadoFinalAntesEventos = ?
                            WHERE nit_emisor_o_nit_del_proveedor = ?
                              AND numero_de_factura = ?
                              AND numero_de_liquidacion_u_orden_de_compra = ?
                            """
                            cur.execute(update_dp, (obs_final, nuevo_resultado, nit, factura, oc))
                            print("[UPDATE] DocumentsProcessing actualizado")
                            
                            # Actualizar Comparativa
                            update_comparativa = """
                            UPDATE [dbo].[CxP.Comparativa]
                            SET Estado_validacion_antes_de_eventos = ?
                            WHERE NIT = ?
                              AND Factura = ?
                            """
                            cur.execute(update_comparativa, (nuevo_resultado, nit, factura))
                            print("[UPDATE] Comparativa actualizada (ESTADO)")
                            
                            cur.close()
                        
                        # SUBCASO 2.2: ClaseDeImpuesto NO contiene 31
                        else:
                            print("[VALIDACION] ClaseDeImpuesto NO contiene 31")
                            
                            # SUBCASO 2.2.1: forma_de_pago = "1" o "01"
                            if forma_pago in ['1', '01']:
                                print("[ACCION] forma_de_pago = " + forma_pago + " (CONTADO)")
                                stats['aprobado_contado'] += 1
                                
                                nuevo_resultado = "APROBADO CONTADO"
                                
                                # Actualizar observaciones (PREPEND)
                                nueva_obs = "Factura cuenta con forma de pago de contado"
                                if obs_actual:
                                    obs_final = nueva_obs + ", " + obs_actual
                                else:
                                    obs_final = nueva_obs
                                
                                # Actualizar DocumentsProcessing
                                cur = cx.cursor()
                                
                                update_dp = """
                                UPDATE [CxP].[DocumentsProcessing]
                                SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso',
                                    ObservacionesFase_4 = ?,
                                    ResultadoFinalAntesEventos = ?
                                WHERE nit_emisor_o_nit_del_proveedor = ?
                                  AND numero_de_factura = ?
                                  AND numero_de_liquidacion_u_orden_de_compra = ?
                                """
                                cur.execute(update_dp, (obs_final, nuevo_resultado, nit, factura, oc))
                                print("[UPDATE] DocumentsProcessing actualizado")
                                
                                # Actualizar Comparativa
                                update_comparativa = """
                                UPDATE [dbo].[CxP.Comparativa]
                                SET Estado_validacion_antes_de_eventos = ?
                                WHERE NIT = ?
                                  AND Factura = ?
                                """
                                cur.execute(update_comparativa, (nuevo_resultado, nit, factura))
                                print("[UPDATE] Comparativa actualizada (ESTADO)")
                                
                                # Actualizar HistoricoOrdenesCompra (solo columna Marca)
                                doccompra_list = split_valores(row['DocCompra_hoc'])
                                nitcedula_list = split_valores(row['NitCedula_hoc'])
                                porcalcular_list = split_valores(row['PorCalcular_hoc'])
                                textobreve_list = split_valores(row['TextoBreve_hoc'])
                                
                                num_actualizados = 0
                                max_len = max(len(doccompra_list), len(nitcedula_list), len(porcalcular_list), len(textobreve_list))
                                
                                for i in range(max_len):
                                    doccompra = doccompra_list[i] if i < len(doccompra_list) else ""
                                    nitcedula = nitcedula_list[i] if i < len(nitcedula_list) else ""
                                    porcalcular = porcalcular_list[i] if i < len(porcalcular_list) else ""
                                    textobreve = textobreve_list[i] if i < len(textobreve_list) else ""
                                    
                                    if doccompra and nitcedula:
                                        update_hoc = """
                                        UPDATE [CxP].[HistoricoOrdenesCompra]
                                        SET Marca = 'PROCESADO'
                                        WHERE DocCompra = ?
                                          AND NitCedula = ?
                                          AND PorCalcular = ?
                                          AND TextoBreve = ?
                                        """
                                        cur.execute(update_hoc, (doccompra, nitcedula, porcalcular, textobreve))
                                        num_actualizados += 1
                                
                                print("[UPDATE] HistoricoOrdenesCompra: " + str(num_actualizados) + " registros actualizados")
                                cur.close()
                            
                            # SUBCASO 2.2.2: forma_de_pago != "1" ni "01"
                            else:
                                print("[ACCION] forma_de_pago != 1/01 (CREDITO)")
                                stats['aprobado'] += 1
                                
                                nuevo_resultado = "APROBADO"
                                
                                # Actualizar DocumentsProcessing
                                cur = cx.cursor()
                                
                                update_dp = """
                                UPDATE [CxP].[DocumentsProcessing]
                                SET EstadoFinalFase_4 = 'VALIDACION DATOS DE FACTURACION: Exitoso',
                                    ResultadoFinalAntesEventos = ?
                                WHERE nit_emisor_o_nit_del_proveedor = ?
                                  AND numero_de_factura = ?
                                  AND numero_de_liquidacion_u_orden_de_compra = ?
                                """
                                cur.execute(update_dp, (nuevo_resultado, nit, factura, oc))
                                print("[UPDATE] DocumentsProcessing actualizado")
                                
                                # Actualizar Comparativa
                                update_comparativa = """
                                UPDATE [dbo].[CxP.Comparativa]
                                SET Estado_validacion_antes_de_eventos = ?
                                WHERE NIT = ?
                                  AND Factura = ?
                                """
                                cur.execute(update_comparativa, (nuevo_resultado, nit, factura))
                                print("[UPDATE] Comparativa actualizada (ESTADO)")
                                
                                # Actualizar HistoricoOrdenesCompra (solo columna Marca)
                                doccompra_list = split_valores(row['DocCompra_hoc'])
                                nitcedula_list = split_valores(row['NitCedula_hoc'])
                                porcalcular_list = split_valores(row['PorCalcular_hoc'])
                                textobreve_list = split_valores(row['TextoBreve_hoc'])
                                
                                num_actualizados = 0
                                max_len = max(len(doccompra_list), len(nitcedula_list), len(porcalcular_list), len(textobreve_list))
                                
                                for i in range(max_len):
                                    doccompra = doccompra_list[i] if i < len(doccompra_list) else ""
                                    nitcedula = nitcedula_list[i] if i < len(nitcedula_list) else ""
                                    porcalcular = porcalcular_list[i] if i < len(porcalcular_list) else ""
                                    textobreve = textobreve_list[i] if i < len(textobreve_list) else ""
                                    
                                    if doccompra and nitcedula:
                                        update_hoc = """
                                        UPDATE [CxP].[HistoricoOrdenesCompra]
                                        SET Marca = 'PROCESADO'
                                        WHERE DocCompra = ?
                                          AND NitCedula = ?
                                          AND PorCalcular = ?
                                          AND TextoBreve = ?
                                        """
                                        cur.execute(update_hoc, (doccompra, nitcedula, porcalcular, textobreve))
                                        num_actualizados += 1
                                
                                print("[UPDATE] HistoricoOrdenesCompra: " + str(num_actualizados) + " registros actualizados")
                                cur.close()
                
                except Exception as e_row:
                    print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                    stats['errores'] += 1
                    continue
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Post-procesamiento completado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros: " + str(stats['total_registros']))
            print("  CON NOVEDAD + Clase 31: " + str(stats['con_novedad_clase31']))
            print("  APROBADO SIN CONTABILIZACION: " + str(stats['aprobado_sin_contab']))
            print("  APROBADO CONTADO: " + str(stats['aprobado_contado']))
            print("  APROBADO: " + str(stats['aprobado']))
            print("  Errores: " + str(stats['errores']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Post-procesamiento OK. Total:" + str(stats['total_registros']) + 
                   " Clase31:" + str(stats['con_novedad_clase31'] + stats['aprobado_sin_contab']) +
                   " Contado:" + str(stats['aprobado_contado']) + 
                   " Aprobado:" + str(stats['aprobado']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}       

def GenerarReporte_Retorno():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion GenerarReporte_Retorno() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def crear_excel_reporte(df, ruta_completa):
        """
        Crear archivo Excel con formato profesional
        """
        print("[EXCEL] Creando archivo Excel...")
        print("[EXCEL] Ruta: " + ruta_completa)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FV"
        
        # Definir estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Escribir encabezados
        headers = ['ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Escribir datos
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border_thin
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 12,  # ID
            'B': 15,  # Fecha_Carga
            'C': 15,  # Nit
            'D': 35,  # Nombre Proveedor
            'E': 18,  # Orden_de_compra
            'F': 18,  # Numero_factura
            'G': 50,  # Estado_CXP_Bot
            'H': 60   # Observaciones
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Guardar archivo
        try:
            wb.save(ruta_completa)
            print("[EXCEL] Archivo guardado exitosamente")
            return True, ruta_completa
        except Exception as e:
            print("[ERROR] Error guardando archivo: " + str(e))
            
            # Intentar con "1" al final
            base_name = os.path.splitext(ruta_completa)[0]
            ext = os.path.splitext(ruta_completa)[1]
            ruta_alternativa = base_name + "1" + ext
            
            try:
                wb.save(ruta_alternativa)
                print("[EXCEL] Archivo guardado con nombre alternativo")
                return True, ruta_alternativa
            except Exception as e2:
                print("[ERROR] Error guardando con nombre alternativo: " + str(e2))
                return False, None
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        # Verificar parámetros de configuración
        if 'RutaReporteNovedades' not in cfg:
            raise ValueError("Configuracion no contiene 'RutaReporteNovedades'")
        
        if 'NombreReporteNovedades' not in cfg:
            raise ValueError("Configuracion no contiene 'NombreReporteNovedades'")
        
        ruta_base = cfg['RutaReporteNovedades']
        nombre_base = cfg['NombreReporteNovedades']
        
        # Verificar que la ruta base exista
        if not os.path.exists(ruta_base):
            print("[WARNING] Ruta base no existe, creando directorio: " + ruta_base)
            os.makedirs(ruta_base, exist_ok=True)
        
        stats = {
            'total_registros_consultados': 0,
            'registros_insertados': 0,
            'total_registros_tabla': 0,
            'registros_exportados': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        fecha_actual = datetime.now()
        fecha_carga = fecha_actual.strftime("%Y-%m-%d")
        timestamp = fecha_actual.strftime("%Y%m%d%H%M")
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 1: Consultar DocumentsProcessing (CON NOVEDAD)
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando DocumentsProcessing (CON NOVEDAD)...")
            
            query_dp = """
            SELECT 
                ID,
                nit_emisor_o_nit_del_proveedor AS Nit,
                nombre_emisor AS Nombre_Proveedor,
                numero_de_liquidacion_u_orden_de_compra AS Orden_de_compra,
                numero_de_factura AS Numero_factura,
                ResultadoFinalAntesEventos AS Estado_CXP_Bot,
                ObservacionesFase_4 AS Observaciones,
                documenttype
            FROM [CxP].[DocumentsProcessing] WITH (NOLOCK)
            WHERE documenttype = 'FV'
              AND ResultadoFinalAntesEventos LIKE '%CON NOVEDAD%'
            """
            
            df_dp = pd.read_sql(query_dp, cx)
            stats['total_registros_consultados'] = len(df_dp)
            
            print("[DEBUG] Registros consultados: " + str(len(df_dp)))
            
            # ================================================================
            # PASO 2: Insertar registros en [CxP].[ReporteNovedades]
            # ================================================================
            
            if not df_dp.empty:
                print("")
                print("[PASO 2] Insertando registros en [CxP].[ReporteNovedades]...")
                
                cur = cx.cursor()
                
                # Agregar columna Fecha_Carga
                df_dp['Fecha_Carga'] = fecha_carga
                
                insert_query = """
                INSERT INTO [CxP].[ReporteNovedades] (
                    ID,
                    Fecha_Carga,
                    Nit,
                    Nombre_Proveedor,
                    Orden_de_compra,
                    Numero_factura,
                    Estado_CXP_Bot,
                    Observaciones,
                    SP_Origen,
                    Fecha_Insercion
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
                """
                
                registros_insertados = 0
                sp_origen = 'GenerarReporte_Retorno'
                
                for idx, row in df_dp.iterrows():
                    try:
                        cur.execute(insert_query, (
                            safe_str(row['ID']),
                            fecha_carga,
                            safe_str(row['Nit']),
                            safe_str(row['Nombre_Proveedor']),
                            safe_str(row['Orden_de_compra']),
                            safe_str(row['Numero_factura']),
                            safe_str(row['Estado_CXP_Bot']),
                            safe_str(row['Observaciones']),
                            sp_origen
                        ))
                        registros_insertados += 1
                        
                        if registros_insertados % 50 == 0:
                            print("[DEBUG] Insertados " + str(registros_insertados) + " registros...")
                            
                    except Exception as e_row:
                        print("[WARNING] Error insertando registro ID=" + safe_str(row['ID']) + ": " + str(e_row))
                        continue
                
                cx.commit()
                cur.close()
                
                stats['registros_insertados'] = registros_insertados
                
                print("[DEBUG] Total registros insertados: " + str(registros_insertados))
            else:
                print("[INFO] No hay registros CON NOVEDAD para insertar")
                stats['registros_insertados'] = 0
            
            # ================================================================
            # PASO 3: Consultar TODOS los registros de [CxP].[ReporteNovedades]
            # ================================================================
            
            print("")
            print("[PASO 3] Consultando todos los registros de [CxP].[ReporteNovedades]...")
            
            query_reporte = """
            SELECT 
                ID,
                Fecha_Carga,
                Nit,
                Nombre_Proveedor,
                Orden_de_compra,
                Numero_factura,
                Estado_CXP_Bot,
                Observaciones
            FROM [CxP].[ReporteNovedades] WITH (NOLOCK)
            ORDER BY Fecha_Insercion DESC, RowID DESC
            """
            
            df_reporte = pd.read_sql(query_reporte, cx)
            stats['total_registros_tabla'] = len(df_reporte)
            
            print("[DEBUG] Total registros en tabla ReporteNovedades: " + str(len(df_reporte)))
            
            if df_reporte.empty:
                print("[INFO] No hay registros en ReporteNovedades para exportar")
                
                # Crear Excel vacío con solo encabezados
                df_vacio = pd.DataFrame(columns=[
                    'ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 
                    'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones'
                ])
                
                # Generar nombre de archivo
                nombre_archivo = os.path.splitext(nombre_base)[0] + "_" + timestamp + ".xlsx"
                ruta_completa = os.path.join(ruta_base, nombre_archivo)
                
                # Crear Excel
                exito, ruta_final = crear_excel_reporte(df_vacio, ruta_completa)
                
                if exito:
                    stats['tiempo_total'] = time.time() - t_inicio
                    msg = "Reporte generado (sin registros): " + ruta_final
                    
                    SetVar("vLocStrResultadoSP", "True")
                    SetVar("vLocStrResumenSP", msg)
                    SetVar("vLocDicEstadisticas", str(stats))
                    SetVar("vLocStrRutaReporte", ruta_final)
                    # ✅ CAMBIO APLICADO: Caso de éxito
                    SetVar("vGblStrDetalleError", "")
                    SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                    
                    return True, msg, ruta_final, stats
                else:
                    raise ValueError("Error al crear archivo Excel vacio")
            
            # ================================================================
            # PASO 4: Preparar datos para Excel
            # ================================================================
            
            print("")
            print("[PASO 4] Preparando datos para Excel...")
            
            # Renombrar columna para coincidir con especificación
            df_reporte.columns = [
                'ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 
                'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones'
            ]
            
            # Convertir valores a string para evitar problemas
            for col in df_reporte.columns:
                df_reporte[col] = df_reporte[col].apply(safe_str)
            
            print("[DEBUG] Registros a exportar: " + str(len(df_reporte)))
            
            stats['registros_exportados'] = len(df_reporte)
            
            # ================================================================
            # PASO 5: Generar archivo Excel
            # ================================================================
            
            print("")
            print("[PASO 5] Generando archivo Excel...")
            
            # Generar nombre de archivo con timestamp
            nombre_sin_ext = os.path.splitext(nombre_base)[0]
            extension = os.path.splitext(nombre_base)[1]
            
            if not extension:
                extension = ".xlsx"
            
            nombre_archivo = nombre_sin_ext + "_" + timestamp + extension
            ruta_completa = os.path.join(ruta_base, nombre_archivo)
            
            print("[DEBUG] Nombre archivo: " + nombre_archivo)
            
            # Crear Excel
            exito, ruta_final = crear_excel_reporte(df_reporte, ruta_completa)
            
            if not exito:
                raise ValueError("Error al crear archivo Excel")
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Reporte generado exitosamente")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Registros consultados (CON NOVEDAD): " + str(stats['total_registros_consultados']))
            print("  Registros insertados en tabla: " + str(stats['registros_insertados']))
            print("  Total registros en tabla: " + str(stats['total_registros_tabla']))
            print("  Registros exportados a Excel: " + str(stats['registros_exportados']))
            print("  Archivo generado: " + ruta_final)
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Reporte OK. Insertados:" + str(stats['registros_insertados']) + 
                   " TotalTabla:" + str(stats['total_registros_tabla']) +
                   " Exportados:" + str(stats['registros_exportados']) + 
                   " Archivo:" + os.path.basename(ruta_final))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            SetVar("vLocStrRutaReporte", ruta_final)
            # ✅ CAMBIO APLICADO: Caso de éxito
            SetVar("vGblStrDetalleError", "")
            SetVar("vGblStrSystemError", "ErrorHU4_4.1")
            
            return True, msg, ruta_final, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        # ✅ CAMBIO APLICADO: Caso de error
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

def ActualizarHistoricoNovedades():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion ActualizarHistoricoNovedades() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def verificar_tabla_existe(cx, tabla):
        """Verificar si tabla existe"""
        query = """
        SELECT COUNT(*) 
        FROM INFORMATION_SCHEMA.TABLES 
        WHERE TABLE_SCHEMA = 'CxP' 
          AND TABLE_NAME = 'HistoricoNovedades'
        """
        cur = cx.cursor()
        cur.execute(query)
        count = cur.fetchone()[0]
        cur.close()
        return count > 0
    
    def crear_tabla_historico(cx):
        """Crear tabla HistoricoNovedades si no existe"""
        print("[TABLA] Creando tabla [CxP].[HistoricoNovedades]...")
        
        create_table = """
        CREATE TABLE [CxP].[HistoricoNovedades] (
            [Fecha_ejecucion] DATETIME NULL,
            [Fecha_de_retoma] DATETIME NULL,
            [ID_ejecucion] NVARCHAR(MAX) NULL,
            [ID_registro] NVARCHAR(MAX) NULL,
            [Nit] NVARCHAR(MAX) NULL,
            [Nombre_Proveedor] NVARCHAR(MAX) NULL,
            [Orden_de_compra] NVARCHAR(MAX) NULL,
            [Factura] NVARCHAR(MAX) NULL,
            [Fec_Doc] NVARCHAR(MAX) NULL,
            [Fec_Reg] NVARCHAR(MAX) NULL,
            [Observaciones] NVARCHAR(MAX) NULL
        )
        """
        
        cur = cx.cursor()
        cur.execute(create_table)
        cx.commit()
        cur.close()
        
        print("[TABLA] Tabla creada exitosamente")
    
    def buscar_fechas_comparativa(cx, nit, factura):
        """Buscar Fec.Doc y Fec.Reg en tabla Comparativa"""
        query = """
        SELECT TOP 1
            FecDoc,
            FecReg
        FROM [CxP].[HistoricoOrdenesCompra] WITH (NOLOCK)
        WHERE NitCedula = ?
          AND DocCompra = ?
          AND (FecDoc IS NOT NULL OR FecReg IS NOT NULL)
        """
        
        cur = cx.cursor()
        cur.execute(query, (nit, factura))
        row = cur.fetchone()
        cur.close()
        
        if row:
            fec_doc = safe_str(row[0]) if row[0] else "NO ENCONTRADO"
            fec_reg = safe_str(row[1]) if row[1] else "NO ENCONTRADO"
            return fec_doc, fec_reg
        else:
            return "NO ENCONTRADO", "NO ENCONTRADO"
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        stats = {
            'total_registros': 0,
            'nuevos_insertados': 0,
            'actualizados': 0,
            'actualizados_no_encontrado': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        fecha_ejecucion = datetime.now()
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 0: Verificar/Crear tabla HistoricoNovedades
            # ================================================================
            
            print("")
            print("[PASO 0] Verificando tabla HistoricoNovedades...")
            
            if not verificar_tabla_existe(cx, '[CxP].[HistoricoNovedades]'):
                print("[INFO] Tabla no existe, creando...")
                crear_tabla_historico(cx)
            else:
                print("[INFO] Tabla existe OK")
            
            # ================================================================
            # PASO 1: Consultar registros con CON NOVEDAD
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando registros con CON NOVEDAD...")
            
            query_novedades = """
            SELECT 
                ID,
                executionNum,
                Fecha_primer_proceso,
                nit_emisor_o_nit_del_proveedor,
                nombre_emisor,
                numero_de_liquidacion_u_orden_de_compra,
                numero_de_factura,
                ObservacionesFase_4,
                documenttype
            FROM [CxP].[DocumentsProcessing] WITH (NOLOCK)
            WHERE documenttype = 'FV'
              AND ResultadoFinalAntesEventos LIKE '%CON NOVEDAD%'
            """
            
            df_novedades = pd.read_sql(query_novedades, cx)
            print("[DEBUG] Registros con CON NOVEDAD: " + str(len(df_novedades)))
            
            if df_novedades.empty:
                print("[INFO] No hay registros con CON NOVEDAD")
            else:
                stats['total_registros'] = len(df_novedades)
                
                # ============================================================
                # PASO 2: Consultar histórico existente
                # ============================================================
                
                print("")
                print("[PASO 2] Consultando HistoricoNovedades existente...")
                
                query_historico = """
                SELECT 
                    Nit,
                    Factura,
                    Fec_Doc,
                    Fec_Reg
                FROM [CxP].[HistoricoNovedades] WITH (NOLOCK)
                """
                
                df_historico = pd.read_sql(query_historico, cx)
                print("[DEBUG] Registros en histórico: " + str(len(df_historico)))
                
                # ============================================================
                # PASO 3: Procesar cada registro con CON NOVEDAD
                # ============================================================
                
                print("")
                print("[PASO 3] Procesando registros con CON NOVEDAD...")
                
                for idx, row in df_novedades.iterrows():
                    try:
                        print("")
                        print("[REGISTRO " + str(idx + 1) + "/" + str(len(df_novedades)) + "]")
                        
                        id_registro = safe_str(row['ID'])
                        id_ejecucion = safe_str(row['executionNum'])
                        fecha_retoma = row['Fecha_primer_proceso'] if pd.notna(row['Fecha_primer_proceso']) else fecha_ejecucion
                        nit = safe_str(row['nit_emisor_o_nit_del_proveedor'])
                        nombre = safe_str(row['nombre_emisor'])
                        orden = safe_str(row['numero_de_liquidacion_u_orden_de_compra'])
                        factura = safe_str(row['numero_de_factura'])
                        observaciones = safe_str(row['ObservacionesFase_4'])
                        
                        print("[DEBUG] NIT: " + nit + " | Factura: " + factura)
                        
                        # Buscar fechas en Comparativa
                        fec_doc, fec_reg = buscar_fechas_comparativa(cx, nit, orden)
                        print("[DEBUG] Fec.Doc: " + fec_doc + " | Fec.Reg: " + fec_reg)
                        
                        # Verificar si existe en histórico
                        existe = False
                        tiene_fechas = False
                        
                        if not df_historico.empty:
                            mask = (df_historico['Nit'] == nit) & (df_historico['Factura'] == factura)
                            registros_existentes = df_historico[mask]
                            
                            if not registros_existentes.empty:
                                existe = True
                                # Verificar si tiene Fec.Doc o Fec.Reg
                                fec_doc_hist = safe_str(registros_existentes.iloc[0]['Fec_Doc'])
                                fec_reg_hist = safe_str(registros_existentes.iloc[0]['Fec_Reg'])
                                
                                if (fec_doc_hist and fec_doc_hist != 'NO ENCONTRADO') or \
                                   (fec_reg_hist and fec_reg_hist != 'NO ENCONTRADO'):
                                    tiene_fechas = True
                        
                        # ====================================================
                        # CASO 1: NO existe en histórico - INSERT
                        # ====================================================
                        
                        if not existe:
                            print("[ACCION] Registro nuevo - INSERT")
                            
                            cur = cx.cursor()
                            
                            insert_query = """
                            INSERT INTO [CxP].[HistoricoNovedades] (
                                Fecha_ejecucion,
                                Fecha_de_retoma,
                                ID_ejecucion,
                                ID_registro,
                                Nit,
                                Nombre_Proveedor,
                                Orden_de_compra,
                                Factura,
                                Fec_Doc,
                                Fec_Reg,
                                Observaciones
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """
                            
                            cur.execute(insert_query, (
                                fecha_ejecucion,
                                fecha_retoma,
                                id_ejecucion,
                                id_registro,
                                nit,
                                nombre,
                                orden,
                                factura,
                                fec_doc,
                                fec_reg,
                                observaciones
                            ))
                            
                            cur.close()
                            stats['nuevos_insertados'] += 1
                            print("[INSERT] Nuevo registro insertado")
                        
                        # ====================================================
                        # CASO 2: Existe y tiene fechas - UPDATE
                        # ====================================================
                        
                        elif existe and tiene_fechas:
                            print("[ACCION] Registro existe con fechas - UPDATE")
                            
                            cur = cx.cursor()
                            
                            update_query = """
                            UPDATE [CxP].[HistoricoNovedades]
                            SET Fecha_ejecucion = ?,
                                Fec_Doc = ?,
                                Fec_Reg = ?,
                                Observaciones = ?
                            WHERE Nit = ?
                              AND Factura = ?
                            """
                            
                            cur.execute(update_query, (
                                fecha_ejecucion,
                                fec_doc,
                                fec_reg,
                                observaciones,
                                nit,
                                factura
                            ))
                            
                            cur.close()
                            stats['actualizados'] += 1
                            print("[UPDATE] Registro actualizado")
                        
                        # ====================================================
                        # CASO 3: Existe sin fechas - SKIP
                        # ====================================================
                        
                        else:
                            print("[INFO] Registro existe sin fechas - SKIP")
                    
                    except Exception as e_row:
                        print("[ERROR] Error procesando registro " + str(idx) + ": " + str(e_row))
                        continue
            
            # ================================================================
            # PASO 4: Actualizar registros con NO ENCONTRADO
            # ================================================================
            
            print("")
            print("[PASO 4] Actualizando registros con NO ENCONTRADO...")
            
            # Buscar registros en histórico con NO ENCONTRADO
            query_no_encontrado = """
            SELECT 
                Nit,
                Factura,
                Fec_Doc,
                Fec_Reg
            FROM [CxP].[HistoricoNovedades] WITH (NOLOCK)
            WHERE Fec_Doc = 'NO ENCONTRADO'
               OR Fec_Reg = 'NO ENCONTRADO'
               OR Fec_Doc IS NULL
               OR Fec_Reg IS NULL
            """
            
            df_no_encontrado = pd.read_sql(query_no_encontrado, cx)
            print("[DEBUG] Registros con NO ENCONTRADO: " + str(len(df_no_encontrado)))
            
            if not df_no_encontrado.empty:
                # Consultar registros APROBADOS
                query_aprobados = """
                SELECT 
                    nit_emisor_o_nit_del_proveedor,
                    numero_de_factura,
                    ObservacionesFase_4
                FROM [CxP].[DocumentsProcessing] WITH (NOLOCK)
                WHERE documenttype = 'FV'
                  AND (ResultadoFinalAntesEventos = 'APROBADO'
                    OR ResultadoFinalAntesEventos = 'APROBADO SIN CONTABILIZACION'
                    OR ResultadoFinalAntesEventos = 'APROBADO CONTADO')
                """
                
                df_aprobados = pd.read_sql(query_aprobados, cx)
                print("[DEBUG] Registros APROBADOS: " + str(len(df_aprobados)))
                
                # Cruzar registros
                for idx, row_hist in df_no_encontrado.iterrows():
                    nit_hist = safe_str(row_hist['Nit'])
                    factura_hist = safe_str(row_hist['Factura'])
                    
                    # Buscar en aprobados
                    mask = (df_aprobados['nit_emisor_o_nit_del_proveedor'] == nit_hist) & \
                           (df_aprobados['numero_de_factura'] == factura_hist)
                    
                    registros_aprobados = df_aprobados[mask]
                    
                    if not registros_aprobados.empty:
                        print("[MATCH] Encontrado en APROBADOS: " + nit_hist + " - " + factura_hist)
                        
                        # Buscar fechas en Comparativa
                        fec_doc, fec_reg = buscar_fechas_comparativa(cx, nit_hist, factura_hist)
                        
                        # Solo actualizar si encontró fechas
                        if fec_doc != "NO ENCONTRADO" or fec_reg != "NO ENCONTRADO":
                            observaciones = safe_str(registros_aprobados.iloc[0]['ObservacionesFase_4'])
                            
                            cur = cx.cursor()
                            
                            update_query = """
                            UPDATE [CxP].[HistoricoNovedades]
                            SET Fecha_ejecucion = ?,
                                Fec_Doc = ?,
                                Fec_Reg = ?,
                                Observaciones = ?
                            WHERE Nit = ?
                              AND Factura = ?
                            """
                            
                            cur.execute(update_query, (
                                fecha_ejecucion,
                                fec_doc,
                                fec_reg,
                                observaciones,
                                nit_hist,
                                factura_hist
                            ))
                            
                            cur.close()
                            stats['actualizados_no_encontrado'] += 1
                            print("[UPDATE] Fechas actualizadas")
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Historico de novedades actualizado")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Total registros procesados: " + str(stats['total_registros']))
            print("  Nuevos insertados: " + str(stats['nuevos_insertados']))
            print("  Actualizados: " + str(stats['actualizados']))
            print("  Actualizados (NO ENCONTRADO): " + str(stats['actualizados_no_encontrado']))
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Historico actualizado. Nuevos:" + str(stats['nuevos_insertados']) + 
                   " Actualizados:" + str(stats['actualizados'] + stats['actualizados_no_encontrado']))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            
            return True, msg, None, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}




# ============================================================
# DISPATCHER FINAL
# ============================================================

funcion_ejecutar = GetVar("vLocStrFuncionEjecutar")

if funcion_ejecutar == "generar_ruta_logs":
    import traceback
    try:
        generar_ruta_logs()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
    
elif funcion_ejecutar == "buscarCandidatos":
    import traceback
    try:
        buscarCandidatos()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPRE_ValidarCOP":
    import traceback
    try:
        ZPRE_ValidarCOP()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPRE_ValidarUSD":
    import traceback
    try:
        ZPRE_ValidarUSD()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
    ZPRE_ValidarUSD()

elif funcion_ejecutar == "ZPRE_ValidarTRM":
    import traceback
    try:
        ZPRE_ValidarTRM()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPRE_ValidarCantidadPrecio":
    import traceback
    try:
        ZPRE_ValidarCantidadPrecio()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPRE_ValidarEmisor":
    import traceback
    try:
        ZPRE_ValidarEmisor()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarCOP":
    import traceback
    try:
        ZPCN_ZPPA_ValidarCOP()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarUSD":
    import traceback
    try:
        ZPCN_ZPPA_ValidarUSD()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarTRM":
    import traceback
    try:
        ZPCN_ZPPA_ValidarTRM()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarEmisor":
    import traceback
    try:
        ZPCN_ZPPA_ValidarEmisor()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarOrdenRegistro":
    import traceback
    try:
        ZPCN_ZPPA_ValidarOrdenRegistro()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarElementoPEP":
    import traceback
    try:
        ZPCN_ZPPA_ValidarElementoPEP()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ZPCN_ZPPA_ValidarActivoFijo":
    import traceback
    try:
        ZPCN_ZPPA_ValidarActivoFijo()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "buscarCandidatos":
    import traceback
    try:
        buscarCandidatos()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
    

elif funcion_ejecutar == "PostProcesamiento_EstadosFinales":
    import traceback
    try:
        PostProcesamiento_EstadosFinales()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "GenerarReporte_Retorno":
    import traceback
    try:
        GenerarReporte_Retorno()
        if ruta_reporte:
            print("[INFO] Ruta reporte: " + ruta_reporte)
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

elif funcion_ejecutar == "ActualizarHistoricoNovedades":
    import traceback
    try:
        ActualizarHistoricoNovedades()
    except Exception as e:
        print("[ERROR FATAL] " + str(e))
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")

else:
    try:
        msg = "" if funcion_ejecutar is None else str(funcion_ejecutar)
    except Exception:
        msg = ""
    try:
        SetVar("vGblStrDetalleError", "Funcion HU4 no reconocida: " + msg)
        SetVar("vGblStrSystemError", "")
        SetVar("vLocStrResultadoSP", False)
        SetVar("vLocStrResumenSP", "")
    except Exception:
        pass
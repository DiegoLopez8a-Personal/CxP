"""
================================================================================
SCRIPT: GenerarReporte_Retorno.py
================================================================================

Descripcion General:
--------------------
    Genera reporte Excel con todas las novedades detectadas en el proceso HU4.
    Inserta registros CON NOVEDAD en la tabla [CxP].[ReporteNovedades] y
    exporta el contenido completo a un archivo Excel formateado.

Autor: Diego Ivan Lopez Ochoa
Version: 1.0.0
Plataforma: RocketBot RPA

================================================================================
DIAGRAMA DE FLUJO
================================================================================

    +-------------------------------------------------------------+
    |                        INICIO                               |
    |             GenerarReporte_Retorno()                        |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  PASO 1: Consultar registros CON NOVEDAD                    |
    |  de [CxP].[DocumentsProcessing]                             |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  PASO 2: Insertar nuevos registros en                       |
    |  [CxP].[ReporteNovedades]                                   |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  PASO 3: Consultar TODOS los registros de                   |
    |  [CxP].[ReporteNovedades]                                   |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  PASO 4: Preparar datos para Excel                          |
    |  - Renombrar columnas                                       |
    |  - Convertir valores a string                               |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  PASO 5: Crear archivo Excel con openpyxl                   |
    |  - Aplicar formato profesional                              |
    |  - Encabezados con estilo                                   |
    |  - Ajustar anchos de columna                                |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  Retornar estadisticas y ruta del archivo                   |
    +-------------------------------------------------------------+

================================================================================
ESTRUCTURA ARCHIVO EXCEL
================================================================================

    Columnas del reporte:
        - ID: Identificador del registro
        - Fecha_Carga: Fecha de carga del documento
        - Nit: NIT del proveedor
        - Nombre Proveedor: Nombre del proveedor
        - Orden_de_compra: Numero de orden de compra
        - Numero_factura: Numero de factura
        - Estado_CXP_Bot: Estado asignado por el bot
        - Observaciones: Detalle de la novedad

    Formato:
        - Encabezados: Fondo azul (#366092), texto blanco, negrita
        - Celdas: Arial 10, alineacion izquierda
        - Bordes: Linea fina negra

================================================================================
ESTRUCTURA TABLA [CxP].[ReporteNovedades]
================================================================================

    Columna              Tipo              Descripcion
    -------              ----              -----------
    RowID                INT IDENTITY      ID autoincremental
    ID                   NVARCHAR(MAX)     ID del documento
    Fecha_Carga          NVARCHAR(MAX)     Fecha de carga
    Nit                  NVARCHAR(MAX)     NIT proveedor
    Nombre_Proveedor     NVARCHAR(MAX)     Nombre proveedor
    Orden_de_compra      NVARCHAR(MAX)     Numero OC
    Numero_factura       NVARCHAR(MAX)     Numero factura
    Estado_CXP_Bot       NVARCHAR(MAX)     Estado final
    Observaciones        NVARCHAR(MAX)     Observaciones
    SP_Origen            NVARCHAR(100)     Origen del registro
    Fecha_Insercion      DATETIME          Fecha de insercion

================================================================================
VARIABLES DE ENTRADA (RocketBot)
================================================================================

    vLocDicConfig : str | dict
        - ServidorBaseDatos: Servidor SQL Server
        - NombreBaseDatos: Base de datos
        - RutaBaseReporte: Directorio para guardar el archivo
        - NombreBaseReporte: Nombre base del archivo (sin timestamp)

================================================================================
VARIABLES DE SALIDA (RocketBot)
================================================================================

    vLocStrResultadoSP : str
        "True" si exitoso, "False" si error

    vLocStrResumenSP : str
        "Reporte OK. Insertados:X TotalTabla:Y Exportados:Z Archivo:nombre.xlsx"

    vLocStrRutaReporte : str
        Ruta completa del archivo generado

    vLocDicEstadisticas : str
        Diccionario con estadisticas

================================================================================
EJEMPLOS DE USO
================================================================================

    # Configurar variables en RocketBot
    SetVar("vLocDicConfig", json.dumps({
        "ServidorBaseDatos": "servidor.ejemplo.com",
        "NombreBaseDatos": "NotificationsPaddy",
        "RutaBaseReporte": "C:/Reportes/CxP",
        "NombreBaseReporte": "Reporte_Novedades"
    }))
    
    # Ejecutar funcion
    GenerarReporte_Retorno()
    
    # Obtener ruta del archivo
    ruta = GetVar("vLocStrRutaReporte")
    # "C:/Reportes/CxP/Reporte_Novedades_20240115_143022.xlsx"

================================================================================
NOTAS TECNICAS
================================================================================

    - Timestamp se agrega automaticamente al nombre del archivo
    - Si no puede guardar, intenta con sufijo "1"
    - Usa openpyxl para formato profesional
    - Ordena por Fecha_Insercion DESC, RowID DESC
    - SP_Origen = 'GenerarReporte_Retorno'

================================================================================
"""

def GenerarReporte_Retorno():
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from contextlib import contextmanager
    import time
    import warnings
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    print("=" * 80)
    print("[INICIO] Funcion GenerarReporte_Retorno() iniciada")
    print("[INICIO] Timestamp: " + str(datetime.now()))
    print("=" * 80)
    
    # ========================================================================
    # FUNCIONES AUXILIARES
    # ========================================================================
    
    def safe_str(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and np.isnan(v):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def parse_config(raw):
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config empty")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig empty")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config empty JSON")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config empty literal")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError("Invalid config: " + str(e))
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError("Missing params: " + ', '.join(missing))
        
        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=" + cfg['ServidorBaseDatos'] + ";"
            "DATABASE=" + cfg['NombreBaseDatos'] + ";"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )

        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print("[DEBUG] Conexion SQL abierta (intento " + str(attempt + 1) + ")")
                break
            except pyodbc.Error:
                if attempt < max_retries - 1:
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final de conexion exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print("[ERROR] Rollback por error: " + str(e))
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    def crear_excel_reporte(df, ruta_completa):
        """
        Crear archivo Excel con formato profesional
        """
        print("[EXCEL] Creando archivo Excel...")
        print("[EXCEL] Ruta: " + ruta_completa)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "FV"
        
        # Definir estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        
        border_thin = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Escribir encabezados
        headers = ['ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Escribir datos
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = border_thin
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 12,  # ID
            'B': 15,  # Fecha_Carga
            'C': 15,  # Nit
            'D': 35,  # Nombre Proveedor
            'E': 18,  # Orden_de_compra
            'F': 18,  # Numero_factura
            'G': 50,  # Estado_CXP_Bot
            'H': 60   # Observaciones
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Guardar archivo
        try:
            wb.save(ruta_completa)
            print("[EXCEL] Archivo guardado exitosamente")
            return True, ruta_completa
        except Exception as e:
            print("[ERROR] Error guardando archivo: " + str(e))
            
            # Intentar con "1" al final
            base_name = os.path.splitext(ruta_completa)[0]
            ext = os.path.splitext(ruta_completa)[1]
            ruta_alternativa = base_name + "1" + ext
            
            try:
                wb.save(ruta_alternativa)
                print("[EXCEL] Archivo guardado con nombre alternativo")
                return True, ruta_alternativa
            except Exception as e2:
                print("[ERROR] Error guardando con nombre alternativo: " + str(e2))
                return False, None
    
    # ========================================================================
    # INICIO DE PROCESO
    # ========================================================================
    
    try:
        print("[DEBUG] Obteniendo configuracion...")
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[DEBUG] Configuracion obtenida OK")
        
        # Verificar parámetros de configuración
        if 'RutaReporteNovedades' not in cfg:
            raise ValueError("Configuracion no contiene 'RutaReporteNovedades'")
        
        if 'NombreReporteNovedades' not in cfg:
            raise ValueError("Configuracion no contiene 'NombreReporteNovedades'")
        
        ruta_base = cfg['RutaReporteNovedades']
        nombre_base = cfg['NombreReporteNovedades']
        
        # Verificar que la ruta base exista
        if not os.path.exists(ruta_base):
            print("[WARNING] Ruta base no existe, creando directorio: " + ruta_base)
            os.makedirs(ruta_base, exist_ok=True)
        
        stats = {
            'total_registros_consultados': 0,
            'registros_insertados': 0,
            'total_registros_tabla': 0,
            'registros_exportados': 0,
            'tiempo_total': 0
        }
        
        t_inicio = time.time()
        fecha_actual = datetime.now()
        fecha_carga = fecha_actual.strftime("%Y-%m-%d")
        timestamp = fecha_actual.strftime("%Y%m%d%H%M")
        
        with crear_conexion_db(cfg) as cx:
            
            # ================================================================
            # PASO 1: Consultar DocumentsProcessing (CON NOVEDAD)
            # ================================================================
            
            print("")
            print("[PASO 1] Consultando DocumentsProcessing (CON NOVEDAD)...")
            
            query_dp = """
            SELECT 
                ID,
                nit_emisor_o_nit_del_proveedor AS Nit,
                nombre_emisor AS Nombre_Proveedor,
                numero_de_liquidacion_u_orden_de_compra AS Orden_de_compra,
                numero_de_factura AS Numero_factura,
                ResultadoFinalAntesEventos AS Estado_CXP_Bot,
                ObservacionesFase_4 AS Observaciones,
                documenttype
            FROM [CxP].[DocumentsProcessing] WITH (NOLOCK)
            WHERE documenttype = 'FV'
              AND ResultadoFinalAntesEventos LIKE '%CON NOVEDAD%'
            """
            
            df_dp = pd.read_sql(query_dp, cx)
            stats['total_registros_consultados'] = len(df_dp)
            
            print("[DEBUG] Registros consultados: " + str(len(df_dp)))
            
            # ================================================================
            # PASO 2: Insertar registros en [CxP].[ReporteNovedades]
            # ================================================================
            
            if not df_dp.empty:
                print("")
                print("[PASO 2] Insertando registros en [CxP].[ReporteNovedades]...")
                
                cur = cx.cursor()
                
                # Agregar columna Fecha_Carga
                df_dp['Fecha_Carga'] = fecha_carga
                
                insert_query = """
                INSERT INTO [CxP].[ReporteNovedades] (
                    ID,
                    Fecha_Carga,
                    Nit,
                    Nombre_Proveedor,
                    Orden_de_compra,
                    Numero_factura,
                    Estado_CXP_Bot,
                    Observaciones,
                    SP_Origen,
                    Fecha_Insercion
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE())
                """
                
                registros_insertados = 0
                sp_origen = 'GenerarReporte_Retorno'
                
                for idx, row in df_dp.iterrows():
                    try:
                        cur.execute(insert_query, (
                            safe_str(row['ID']),
                            fecha_carga,
                            safe_str(row['Nit']),
                            safe_str(row['Nombre_Proveedor']),
                            safe_str(row['Orden_de_compra']),
                            safe_str(row['Numero_factura']),
                            safe_str(row['Estado_CXP_Bot']),
                            safe_str(row['Observaciones']),
                            sp_origen
                        ))
                        registros_insertados += 1
                        
                        if registros_insertados % 50 == 0:
                            print("[DEBUG] Insertados " + str(registros_insertados) + " registros...")
                            
                    except Exception as e_row:
                        print("[WARNING] Error insertando registro ID=" + safe_str(row['ID']) + ": " + str(e_row))
                        continue
                
                cx.commit()
                cur.close()
                
                stats['registros_insertados'] = registros_insertados
                
                print("[DEBUG] Total registros insertados: " + str(registros_insertados))
            else:
                print("[INFO] No hay registros CON NOVEDAD para insertar")
                stats['registros_insertados'] = 0
            
            # ================================================================
            # PASO 3: Consultar TODOS los registros de [CxP].[ReporteNovedades]
            # ================================================================
            
            print("")
            print("[PASO 3] Consultando todos los registros de [CxP].[ReporteNovedades]...")
            
            query_reporte = """
            SELECT 
                ID,
                Fecha_Carga,
                Nit,
                Nombre_Proveedor,
                Orden_de_compra,
                Numero_factura,
                Estado_CXP_Bot,
                Observaciones
            FROM [CxP].[ReporteNovedades] WITH (NOLOCK)
            ORDER BY Fecha_Insercion DESC, RowID DESC
            """
            
            df_reporte = pd.read_sql(query_reporte, cx)
            stats['total_registros_tabla'] = len(df_reporte)
            
            print("[DEBUG] Total registros en tabla ReporteNovedades: " + str(len(df_reporte)))
            
            if df_reporte.empty:
                print("[INFO] No hay registros en ReporteNovedades para exportar")
                
                # Crear Excel vacío con solo encabezados
                df_vacio = pd.DataFrame(columns=[
                    'ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 
                    'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones'
                ])
                
                # Generar nombre de archivo
                nombre_archivo = os.path.splitext(nombre_base)[0] + "_" + timestamp + ".xlsx"
                ruta_completa = os.path.join(ruta_base, nombre_archivo)
                
                # Crear Excel
                exito, ruta_final = crear_excel_reporte(df_vacio, ruta_completa)
                
                if exito:
                    stats['tiempo_total'] = time.time() - t_inicio
                    msg = "Reporte generado (sin registros): " + ruta_final
                    
                    SetVar("vLocStrResultadoSP", "True")
                    SetVar("vLocStrResumenSP", msg)
                    SetVar("vLocDicEstadisticas", str(stats))
                    SetVar("vLocStrRutaReporte", ruta_final)
                    # ✅ CAMBIO APLICADO: Caso de éxito
                    SetVar("vGblStrDetalleError", "")
                    SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                    
                    return True, msg, ruta_final, stats
                else:
                    raise ValueError("Error al crear archivo Excel vacio")
            
            # ================================================================
            # PASO 4: Preparar datos para Excel
            # ================================================================
            
            print("")
            print("[PASO 4] Preparando datos para Excel...")
            
            # Renombrar columna para coincidir con especificación
            df_reporte.columns = [
                'ID', 'Fecha_Carga', 'Nit', 'Nombre Proveedor', 
                'Orden_de_compra', 'Numero_factura', 'Estado_CXP_Bot', 'Observaciones'
            ]
            
            # Convertir valores a string para evitar problemas
            for col in df_reporte.columns:
                df_reporte[col] = df_reporte[col].apply(safe_str)
            
            print("[DEBUG] Registros a exportar: " + str(len(df_reporte)))
            
            stats['registros_exportados'] = len(df_reporte)
            
            # ================================================================
            # PASO 5: Generar archivo Excel
            # ================================================================
            
            print("")
            print("[PASO 5] Generando archivo Excel...")
            
            # Generar nombre de archivo con timestamp
            nombre_sin_ext = os.path.splitext(nombre_base)[0]
            extension = os.path.splitext(nombre_base)[1]
            
            if not extension:
                extension = ".xlsx"
            
            nombre_archivo = nombre_sin_ext + "_" + timestamp + extension
            ruta_completa = os.path.join(ruta_base, nombre_archivo)
            
            print("[DEBUG] Nombre archivo: " + nombre_archivo)
            
            # Crear Excel
            exito, ruta_final = crear_excel_reporte(df_reporte, ruta_completa)
            
            if not exito:
                raise ValueError("Error al crear archivo Excel")
            
            # ================================================================
            # FIN DE PROCESO
            # ================================================================
            
            stats['tiempo_total'] = time.time() - t_inicio
            
            print("")
            print("=" * 80)
            print("[FIN] Reporte generado exitosamente")
            print("=" * 80)
            print("[ESTADISTICAS]")
            print("  Registros consultados (CON NOVEDAD): " + str(stats['total_registros_consultados']))
            print("  Registros insertados en tabla: " + str(stats['registros_insertados']))
            print("  Total registros en tabla: " + str(stats['total_registros_tabla']))
            print("  Registros exportados a Excel: " + str(stats['registros_exportados']))
            print("  Archivo generado: " + ruta_final)
            print("  Tiempo total: " + str(round(stats['tiempo_total'], 2)) + "s")
            print("=" * 80)
            
            msg = ("Reporte OK. Insertados:" + str(stats['registros_insertados']) + 
                   " TotalTabla:" + str(stats['total_registros_tabla']) +
                   " Exportados:" + str(stats['registros_exportados']) + 
                   " Archivo:" + os.path.basename(ruta_final))
            
            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", msg)
            SetVar("vLocDicEstadisticas", str(stats))
            SetVar("vLocStrRutaReporte", ruta_final)
            # ✅ CAMBIO APLICADO: Caso de éxito
            SetVar("vGblStrDetalleError", "")
            SetVar("vGblStrSystemError", "ErrorHU4_4.1")
            
            return True, msg, ruta_final, stats
    
    except Exception as e:
        exc_type = type(e).__name__
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion fallo")
        print("=" * 80)
        print("[ERROR] Tipo de error: " + exc_type)
        print("[ERROR] Mensaje: " + str(e))
        print("[ERROR] Traceback completo:")
        print(traceback.format_exc())
        print("=" * 80)
        
        # ✅ CAMBIO APLICADO: Caso de error
        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        
        return False, str(e), None, {}

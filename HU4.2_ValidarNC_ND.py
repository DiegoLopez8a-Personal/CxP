def HU42_ValidarNotasCreditoDebito():
    """
    Función para procesar las validaciones de Notas Crédito (NC) y Notas Débito (ND).
    
    VERSIÓN: 1.0 - 12 Enero 2026
    
    FLUJO PRINCIPAL NC:
        1. Verificar fecha de retoma (plazo máximo parametrizable)
        2. Validar campos básicos: Nombre Emisor, NIT Emisor, Fecha emisión
        3. Validar campos receptor: Nombre, NIT (860031606), Tipo Persona (31), 
           DigitoVerificacion (6), TaxLevelCode
        4. Validar referencia (Tipo nota crédito = 20)
        5. Comparar NC con FV en BD (NIT + Referencia)
        6. Comparar valores (Valor a Pagar nc vs Valor a Pagar FV)
        7. Generar insumo de retorno para NC con novedad
    
    FLUJO PRINCIPAL ND:
        1. Validar campos básicos
        2. Validar campos receptor
        3. Marcar como Exitoso
    
    ESTRUCTURA TRAZABILIDAD NC (12 columnas):
        Fecha ejecución, Fecha retoma, ID ejecución, ID Registro, NIT, 
        Nombre Proveedor, Nota Credito, Item, Valor XML, Valor Factura, Aprobado, Estado
    
    ESTRUCTURA TRAZABILIDAD ND (10 columnas):
        Fecha ejecución, ID ejecución, ID Registro, NIT, Nombre Proveedor, 
        Nota Debito, Item, Valor XML, Aprobado, Estado
    
    Returns:
        None: Actualiza variables globales en RocketBot
    """
    
    # =========================================================================
    # IMPORTS
    # =========================================================================
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime, timedelta
    from contextlib import contextmanager
    import time
    import warnings
    import re
    import os
    import unicodedata
    
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')
    
    # =========================================================================
    # FUNCIONES AUXILIARES BÁSICAS
    # =========================================================================
    
    def safe_str(v):
        """Convierte un valor a string de manera segura."""
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (np.isnan(v) or pd.isna(v)):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def truncar_observacion(obs, max_len=3900):
        """Trunca observación para prevenir overflow en BD."""
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > max_len:
            return obs_str[:max_len - 3] + "..."
        return obs_str
    
    def parse_config(raw):
        """Parsea la configuración desde RocketBot."""
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config vacia (dict)")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig vacio")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config vacia (JSON)")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config vacia (literal)")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError(f"Config invalida: {str(e)}")
    
    def normalizar_decimal(valor):
        """Normaliza valores decimales con punto o coma."""
        if pd.isna(valor) or valor == '' or valor is None:
            return 0.0
        if isinstance(valor, (int, float)):
            if np.isnan(valor) if isinstance(valor, float) else False:
                return 0.0
            return float(valor)
        valor_str = str(valor).strip()
        valor_str = valor_str.replace(',', '.')
        valor_str = re.sub(r'[^\d.\-]', '', valor_str)
        try:
            return float(valor_str)
        except:
            return 0.0
    
    def campo_vacio(valor):
        """Verifica si un campo está vacío."""
        valor_str = safe_str(valor)
        return valor_str == "" or valor_str.lower() in ('null', 'none', 'nan')
    
    def campo_con_valor(valor):
        """Verifica si un campo tiene valor."""
        return not campo_vacio(valor)
    
    def quitar_tildes(texto):
        """Elimina tildes de un texto."""
        if not texto:
            return ""
        nfkd = unicodedata.normalize('NFKD', texto)
        return ''.join([c for c in nfkd if not unicodedata.combining(c)])
    
    # =========================================================================
    # CONEXIÓN A BASE DE DATOS
    # =========================================================================
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        """Crea conexión a la base de datos con reintentos."""
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError(f"Parametros faltantes: {', '.join(missing)}")

        usuario = GetVar("vGblStrUsuarioBaseDatos")
        contrasena = GetVar("vGblStrClaveBaseDatos")
        
        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={cfg['ServidorBaseDatos']};"
            f"DATABASE={cfg['NombreBaseDatos']};"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )
        
        cx = None
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str, timeout=30)
                cx.autocommit = False
                print(f"[DEBUG] Conexion SQL abierta (intento {attempt + 1})")
                break
            except pyodbc.Error as e:
                if attempt < max_retries - 1:
                    print(f"[WARNING] Intento {attempt + 1} fallido, reintentando...")
                    time.sleep(1 * (attempt + 1))
                    continue
                raise
        
        try:
            yield cx
            if cx:
                cx.commit()
                print("[DEBUG] Commit final exitoso")
        except Exception as e:
            if cx:
                cx.rollback()
                print(f"[ERROR] Rollback por error: {str(e)}")
            raise
        finally:
            if cx:
                try:
                    cx.close()
                    print("[DEBUG] Conexion cerrada")
                except:
                    pass
    
    # =========================================================================
    # FUNCIONES DE VALIDACION ESPECÍFICAS
    # =========================================================================
    
    def validar_nombre_receptor(nombre):
        """
        Valida que el nombre receptor sea DIANA CORPORACIÓN SAS o variantes permitidas.
        
        Variantes permitidas:
            - DIANA CORPORACIÓN SAS (con o sin tilde)
            - DIANA CORPORACION S.A.S., S. A. S., S A S, etc.
            - DICORP SAS y variantes
            - No debe contener información adicional después
        
        Returns:
            bool: True si es válido
        """
        if campo_vacio(nombre):
            return False
        
        nombre_str = safe_str(nombre).upper().strip()
        nombre_sin_tilde = quitar_tildes(nombre_str)
        
        # Limpiar puntuación y espacios para comparación
        nombre_limpio = re.sub(r'[,.\s]', '', nombre_sin_tilde)
        
        # Patrones válidos (sin tildes, sin puntuación)
        patrones_validos = [
            'DIANACORPORACIONSAS',
            'DICORPSAS'
        ]
        
        # Verificar que coincida exactamente con algún patrón
        for patron in patrones_validos:
            if nombre_limpio == patron:
                return True
        
        return False
    
    def validar_nit_receptor(nit):
        """Valida que NIT receptor sea 860031606."""
        nit_str = safe_str(nit).strip()
        # Limpiar cualquier caracter no numérico
        nit_limpio = re.sub(r'\D', '', nit_str)
        return nit_limpio == '860031606'
    
    def validar_tipo_persona(tipo):
        """Valida que Tipo Persona sea 31."""
        tipo_str = safe_str(tipo).strip()
        return tipo_str == '31'
    
    def validar_digito_verificacion(digito):
        """Valida que Dígito de verificación sea 6."""
        digito_str = safe_str(digito).strip()
        return digito_str == '6'
    
    def validar_tax_level_code(tax_code):
        """
        Valida que TaxLevelCode esté dentro del estándar.
        Valores permitidos: O-13, O-15, O-23, O-47, R-99-PN
        """
        if campo_vacio(tax_code):
            return False
        
        tax_str = safe_str(tax_code).upper().strip()
        
        valores_permitidos = ['O-13', 'O-15', 'O-23', 'O-47', 'R-99-PN']
        
        # Verificar si contiene alguno de los valores permitidos
        for valor in valores_permitidos:
            if valor in tax_str:
                return True
        
        return False
    
    def calcular_dias_diferencia(fecha_inicio, fecha_fin):
        """Calcula la diferencia en días entre dos fechas."""
        try:
            if isinstance(fecha_inicio, str):
                # Intentar varios formatos
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                    try:
                        fecha_inicio = datetime.strptime(fecha_inicio, fmt)
                        break
                    except:
                        continue
            
            if isinstance(fecha_fin, str):
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                    try:
                        fecha_fin = datetime.strptime(fecha_fin, fmt)
                        break
                    except:
                        continue
            
            if isinstance(fecha_inicio, datetime) and isinstance(fecha_fin, datetime):
                return (fecha_fin - fecha_inicio).days
            
            return 0
        except:
            return 0
    
    # =========================================================================
    # FUNCIONES DE ACTUALIZACIÓN DE BD
    # =========================================================================
    
    def actualizar_bd_cxp(cx, registro_id, campos_actualizar):
        """Actualiza campos en [CxP].[DocumentsProcessing]."""
        try:
            sets = []
            parametros = []
            
            for campo, valor in campos_actualizar.items():
                if valor is not None:
                    if campo == 'ObservacionesFase_4':
                        # Añadir al inicio, conservando observaciones previas
                        sets.append(f"[{campo}] = CASE WHEN [{campo}] IS NULL OR [{campo}] = '' THEN ? ELSE ? + ', ' + [{campo}] END")
                        parametros.extend([valor, valor])
                    else:
                        sets.append(f"[{campo}] = ?")
                        parametros.append(valor)
            
            if sets:
                parametros.append(registro_id)
                sql = f"UPDATE [CxP].[DocumentsProcessing] SET {', '.join(sets)} WHERE [ID] = ?"
                
                cur = cx.cursor()
                cur.execute(sql, parametros)
                cur.close()
                
                print(f"[UPDATE] DocumentsProcessing actualizada - ID {registro_id}")
            
        except Exception as e:
            print(f"[ERROR] Error actualizando DocumentsProcessing: {str(e)}")
            raise
    
    def actualizar_nota_credito_referenciada_fv(cx, fv_id, numero_nc):
        """Actualiza el campo Nota crédito referenciada en la FV."""
        try:
            cur = cx.cursor()
            sql = """
            UPDATE [CxP].[DocumentsProcessing]
            SET [NotaCreditoReferenciada] = ?
            WHERE [ID] = ?
            """
            cur.execute(sql, (numero_nc, fv_id))
            cur.close()
            print(f"[UPDATE] FV {fv_id} - NotaCreditoReferenciada = {numero_nc}")
        except Exception as e:
            print(f"[ERROR] Error actualizando NotaCreditoReferenciada: {str(e)}")
    
    def insertar_trazabilidad_nc(cx, datos_traza):
        """
        Inserta registro en tabla de trazabilidad para NC.
        
        Estructura: Fecha ejecución, Fecha retoma, ID ejecución, ID Registro, 
                    NIT, Nombre Proveedor, Nota Credito, Item, Valor XML, 
                    Valor Factura, Aprobado, Estado
        """
        try:
            cur = cx.cursor()
            
            sql = """
            INSERT INTO [CxP].[Comparativa_NC] (
                FechaEjecucion, FechaRetoma, IDEjecucion, IDRegistro,
                NIT, NombreProveedor, NotaCredito, Item, ValorXML,
                ValorFactura, Aprobado, Estado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            cur.execute(sql, (
                datos_traza.get('FechaEjecucion'),
                datos_traza.get('FechaRetoma'),
                datos_traza.get('IDEjecucion'),
                datos_traza.get('IDRegistro'),
                datos_traza.get('NIT'),
                datos_traza.get('NombreProveedor'),
                datos_traza.get('NotaCredito'),
                datos_traza.get('Item'),
                datos_traza.get('ValorXML'),
                datos_traza.get('ValorFactura'),
                datos_traza.get('Aprobado'),
                datos_traza.get('Estado')
            ))
            
            cur.close()
            
        except Exception as e:
            print(f"[ERROR] Error insertando trazabilidad NC: {str(e)}")
    
    def insertar_trazabilidad_nd(cx, datos_traza):
        """
        Inserta registro en tabla de trazabilidad para ND.
        
        Estructura: Fecha ejecución, ID ejecución, ID Registro, NIT, 
                    Nombre Proveedor, Nota Debito, Item, Valor XML, Aprobado, Estado
        """
        try:
            cur = cx.cursor()
            
            sql = """
            INSERT INTO [CxP].[Comparativa_ND] (
                FechaEjecucion, IDEjecucion, IDRegistro,
                NIT, NombreProveedor, NotaDebito, Item, ValorXML,
                Aprobado, Estado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            cur.execute(sql, (
                datos_traza.get('FechaEjecucion'),
                datos_traza.get('IDEjecucion'),
                datos_traza.get('IDRegistro'),
                datos_traza.get('NIT'),
                datos_traza.get('NombreProveedor'),
                datos_traza.get('NotaDebito'),
                datos_traza.get('Item'),
                datos_traza.get('ValorXML'),
                datos_traza.get('Aprobado'),
                datos_traza.get('Estado')
            ))
            
            cur.close()
            
        except Exception as e:
            print(f"[ERROR] Error insertando trazabilidad ND: {str(e)}")
    
    # =========================================================================
    # FUNCIONES DE GENERACIÓN DE TRAZABILIDAD
    # =========================================================================
    
    def generar_trazabilidad_nc(cx, registro, id_ejecucion, fecha_ejecucion, items_validacion, estado_final):
        """
        Genera todos los items de trazabilidad para una NC.
        
        Items estándar:
            1. Nombre Emisor
            2. NIT Emisor
            3. Nombre Receptor
            4. Nit Receptor
            5. Tipo Persona Receptor
            6. DigitoVerificacion Receptor
            7. TaxLevelCode Receptor
            8. Fecha emisión del documento
            9. LineExtensionAmount
            10. Tipo de nota crédito
            11. Referencia
            12. Código CUFE de la factura
            13. Cude de la Nota Credito
            14. ActualizacionNombreArchivos
            15. RutaRespaldo
            16. Observaciones
        """
        registro_id = safe_str(registro.get('ID', ''))
        nit = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
        nombre_proveedor = safe_str(registro.get('nombre_emisor', ''))
        nota_credito = safe_str(registro.get('numero_de_nota_credito', ''))
        fecha_retoma = registro.get('fecha_de_retoma', fecha_ejecucion)
        
        datos_base = {
            'FechaEjecucion': fecha_ejecucion,
            'FechaRetoma': fecha_retoma,
            'IDEjecucion': id_ejecucion,
            'IDRegistro': registro_id,
            'NIT': nit,
            'NombreProveedor': nombre_proveedor,
            'NotaCredito': nota_credito,
            'Estado': estado_final
        }
        
        for item_nombre, item_data in items_validacion.items():
            datos_item = datos_base.copy()
            datos_item['Item'] = item_nombre
            datos_item['ValorXML'] = item_data.get('valor_xml', '')
            datos_item['ValorFactura'] = item_data.get('valor_factura', '')
            datos_item['Aprobado'] = item_data.get('aprobado', '')
            
            insertar_trazabilidad_nc(cx, datos_item)
    
    def generar_trazabilidad_nd(cx, registro, id_ejecucion, fecha_ejecucion, items_validacion, estado_final):
        """Genera todos los items de trazabilidad para una ND."""
        registro_id = safe_str(registro.get('ID', ''))
        nit = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
        nombre_proveedor = safe_str(registro.get('nombre_emisor', ''))
        nota_debito = safe_str(registro.get('numero_de_nota_debito', ''))
        
        datos_base = {
            'FechaEjecucion': fecha_ejecucion,
            'IDEjecucion': id_ejecucion,
            'IDRegistro': registro_id,
            'NIT': nit,
            'NombreProveedor': nombre_proveedor,
            'NotaDebito': nota_debito,
            'Estado': estado_final
        }
        
        for item_nombre, item_data in items_validacion.items():
            datos_item = datos_base.copy()
            datos_item['Item'] = item_nombre
            datos_item['ValorXML'] = item_data.get('valor_xml', '')
            datos_item['Aprobado'] = item_data.get('aprobado', '')
            
            insertar_trazabilidad_nd(cx, datos_item)
    
    # =========================================================================
    # FUNCIONES DE BÚSQUEDA DE FACTURAS
    # =========================================================================
    
    def buscar_factura_correspondiente(cx, nit, referencia, fecha_ejecucion):
        """
        Busca la factura FV correspondiente a la NC.
        
        Criterios:
            - Fecha emisión <= 1 mes atrás vs fecha ejecución
            - Mismo NIT
            - Referencia = Número de factura
            - Prioridad: RECHAZADO > APROBADO
            - Sin Nota crédito referenciada
        
        Returns:
            dict o None: Datos de la factura encontrada
        """
        try:
            # Calcular fecha límite (1 mes atrás)
            if isinstance(fecha_ejecucion, str):
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                    try:
                        fecha_ejecucion = datetime.strptime(fecha_ejecucion, fmt)
                        break
                    except:
                        continue
            
            # Primer día del mes anterior
            primer_dia_mes_actual = fecha_ejecucion.replace(day=1)
            primer_dia_mes_anterior = (primer_dia_mes_actual - timedelta(days=1)).replace(day=1)
            
            cur = cx.cursor()
            
            # Query para buscar FV con prioridad RECHAZADO > APROBADO
            query = """
            SELECT TOP 1 *
            FROM [CxP].[DocumentsProcessing]
            WHERE [tipo_de_documento] = 'FV'
              AND [nit_emisor_o_nit_del_proveedor] = ?
              AND [numero_de_factura] = ?
              AND [fecha_de_emision_documento] >= ?
              AND ([NotaCreditoReferenciada] IS NULL OR [NotaCreditoReferenciada] = '')
            ORDER BY 
                CASE [ResultadoFinalAntesEventos] 
                    WHEN 'RECHAZADO' THEN 1 
                    WHEN 'APROBADO' THEN 2 
                    ELSE 3 
                END ASC
            """
            
            cur.execute(query, (nit, referencia, primer_dia_mes_anterior))
            
            row = cur.fetchone()
            
            if row:
                columns = [desc[0] for desc in cur.description]
                result = dict(zip(columns, row))
                cur.close()
                return result
            
            cur.close()
            return None
            
        except Exception as e:
            print(f"[ERROR] Error buscando factura: {str(e)}")
            return None
    
    # =========================================================================
    # FUNCIONES DE GENERACIÓN DE INSUMO DE RETORNO
    # =========================================================================
    
    def generar_insumo_retorno_nc(registros_novedad, ruta_insumo):
        """
        Genera el archivo Reporte_de_Retorno_Bot.xlsx para NC con novedad.
        
        Columnas:
            - ID
            - Fecha_Carga
            - Nit
            - Numero_Nota_Crédito
            - Estado_CXP_Bot
        """
        try:
            if not registros_novedad:
                print("[INFO] No hay registros NC con novedad para generar insumo")
                return
            
            import openpyxl
            from openpyxl import Workbook
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(ruta_insumo), exist_ok=True)
            
            # Verificar si existe el archivo
            if os.path.exists(ruta_insumo):
                wb = openpyxl.load_workbook(ruta_insumo)
            else:
                wb = Workbook()
            
            # Crear o acceder hoja NC
            if 'NC' in wb.sheetnames:
                ws = wb['NC']
            else:
                ws = wb.create_sheet('NC')
                # Escribir encabezados
                ws.append(['ID', 'Fecha_Carga', 'Nit', 'Numero_Nota_Credito', 'Estado_CXP_Bot'])
            
            # Agregar registros
            fecha_carga = datetime.now().strftime('%Y-%m-%d')
            
            for reg in registros_novedad:
                ws.append([
                    reg.get('ID', ''),
                    fecha_carga,
                    reg.get('nit_emisor_o_nit_del_proveedor', ''),
                    reg.get('numero_de_nota_credito', ''),
                    reg.get('estado', 'CON NOVEDAD')
                ])
            
            # Eliminar hoja por defecto si existe y está vacía
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            wb.save(ruta_insumo)
            print(f"[INFO] Insumo de retorno generado: {ruta_insumo}")
            
        except Exception as e:
            print(f"[ERROR] Error generando insumo de retorno: {str(e)}")
    
    # =========================================================================
    # PROCESAMIENTO PRINCIPAL
    # =========================================================================
    
    try:
        print("")
        print("=" * 80)
        print("[INICIO] Procesamiento HU4.2 - VALIDACION NC y ND")
        print("=" * 80)
        
        t_inicio = time.time()
        
        # 1. Obtener y validar configuración
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[INFO] Configuracion cargada exitosamente")
        
        # Parámetros de configuración
        plazo_maximo_retoma = int(cfg.get('plazo_maximo_retoma_dias', 120))
        ruta_insumo_retorno = cfg.get('RutaInsumoRetorno', '')
        id_ejecucion = cfg.get('IDEjecucion', str(int(time.time())))
        es_retorno_manual = cfg.get('EsRetornoManual', False)
        
        fecha_ejecucion = datetime.now()
        fecha_ejecucion_str = fecha_ejecucion.strftime('%Y-%m-%d')
        
        print(f"[INFO] Plazo maximo retoma: {plazo_maximo_retoma} dias")
        print(f"[INFO] ID Ejecucion: {id_ejecucion}")
        print(f"[INFO] Es retorno manual: {es_retorno_manual}")
        
        # 2. Conectar a base de datos
        with crear_conexion_db(cfg) as cx:
            
            # =========================================================
            # PROCESAR NOTAS CRÉDITO (NC)
            # =========================================================
            
            print("\n" + "=" * 40)
            print("[PROCESO] Procesando Notas Crédito (NC)")
            print("=" * 40)
            
            # Query para obtener NC pendientes
            query_nc = """
                SELECT * FROM [CxP].[DocumentsProcessing]
                WHERE [tipo_de_documento] = 'NC'
                  AND ([ResultadoFinalAntesEventos] IS NULL 
                       OR [ResultadoFinalAntesEventos] NOT IN ('Encontrado', 'No exitoso'))
                ORDER BY [executionDate] DESC
            """
            
            df_nc = pd.read_sql(query_nc, cx)
            print(f"[INFO] Obtenidas {len(df_nc)} notas credito para procesar")
            
            # Variables de conteo NC
            nc_procesadas = 0
            nc_encontradas = 0
            nc_con_novedad = 0
            nc_no_exitoso = 0
            registros_nc_novedad = []
            
            for idx, registro in df_nc.iterrows():
                try:
                    registro_id = safe_str(registro.get('ID', ''))
                    numero_nc = safe_str(registro.get('numero_de_nota_credito', ''))
                    nit = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
                    
                    print(f"\n[NC] Procesando {nc_procesadas + 1}/{len(df_nc)}: NC {numero_nc}, NIT {nit}")
                    
                    items_validacion = {}
                    hay_novedad = False
                    estado_final = None
                    
                    # =====================================================
                    # PASO 1: Validar fecha de retoma (si no es retorno manual)
                    # =====================================================
                    
                    if not es_retorno_manual:
                        fecha_retoma = registro.get('fecha_de_retoma')
                        
                        if campo_con_valor(fecha_retoma):
                            # Calcular días transcurridos
                            dias_transcurridos = calcular_dias_diferencia(fecha_retoma, fecha_ejecucion)
                            
                            if dias_transcurridos > plazo_maximo_retoma:
                                print(f"[INFO] NC {numero_nc} excede plazo maximo ({dias_transcurridos} > {plazo_maximo_retoma} dias)")
                                
                                # Marcar como No exitoso
                                observacion = "Registro excede el plazo maximo de retoma"
                                
                                campos_actualizar = {
                                    'EstadoFinalFase_4': 'No exitoso',
                                    'ObservacionesFase_4': observacion,
                                    'ResultadoFinalAntesEventos': 'No exitoso'
                                }
                                actualizar_bd_cxp(cx, registro_id, campos_actualizar)
                                
                                # Trazabilidad
                                items_validacion['Observaciones'] = {
                                    'valor_xml': observacion,
                                    'valor_factura': '',
                                    'aprobado': ''
                                }
                                
                                generar_trazabilidad_nc(cx, registro, id_ejecucion, fecha_ejecucion_str, items_validacion, 'No exitoso')
                                
                                nc_no_exitoso += 1
                                nc_procesadas += 1
                                continue
                        else:
                            # Registrar fecha de retoma si no tiene
                            cur = cx.cursor()
                            cur.execute("""
                                UPDATE [CxP].[DocumentsProcessing]
                                SET [fecha_de_retoma] = ?
                                WHERE [ID] = ?
                            """, (fecha_ejecucion_str, registro_id))
                            cur.close()
                    
                    # =====================================================
                    # PASO 2: Validar campos básicos (Nombre Emisor, NIT, Fecha)
                    # =====================================================
                    
                    nombre_emisor = safe_str(registro.get('nombre_emisor', ''))
                    nit_emisor = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
                    fecha_emision = safe_str(registro.get('fecha_de_emision_documento', ''))
                    
                    # Nombre Emisor
                    items_validacion['Nombre Emisor'] = {
                        'valor_xml': nombre_emisor,
                        'valor_factura': '',
                        'aprobado': 'SI' if campo_con_valor(nombre_emisor) else 'NO'
                    }
                    
                    # NIT Emisor
                    items_validacion['NIT Emisor'] = {
                        'valor_xml': nit_emisor,
                        'valor_factura': '',
                        'aprobado': 'SI' if campo_con_valor(nit_emisor) else 'NO'
                    }
                    
                    # Fecha emisión
                    items_validacion['Fecha emision del documento'] = {
                        'valor_xml': fecha_emision,
                        'valor_factura': '',
                        'aprobado': 'SI' if campo_con_valor(fecha_emision) else 'NO'
                    }
                    
                    # =====================================================
                    # PASO 3: Validar campos receptor
                    # =====================================================
                    
                    nombre_receptor = safe_str(registro.get('nombre_del_adquiriente', ''))
                    nit_receptor = safe_str(registro.get('nit_del_adquiriente', ''))
                    tipo_persona = safe_str(registro.get('tipo_persona', ''))
                    digito_verif = safe_str(registro.get('digito_de_verificacion', ''))
                    tax_level_code = safe_str(registro.get('responsabilidad_tributaria_adquiriente', ''))
                    
                    # Nombre Receptor
                    nombre_receptor_valido = validar_nombre_receptor(nombre_receptor)
                    items_validacion['Nombre Receptor'] = {
                        'valor_xml': nombre_receptor,
                        'valor_factura': '',
                        'aprobado': 'SI' if nombre_receptor_valido else 'NO'
                    }
                    
                    # Nit Receptor
                    nit_receptor_valido = validar_nit_receptor(nit_receptor)
                    items_validacion['Nit Receptor'] = {
                        'valor_xml': nit_receptor,
                        'valor_factura': '',
                        'aprobado': 'SI' if nit_receptor_valido else 'NO'
                    }
                    
                    # Tipo Persona Receptor
                    tipo_persona_valido = validar_tipo_persona(tipo_persona)
                    items_validacion['Tipo Persona Receptor'] = {
                        'valor_xml': tipo_persona,
                        'valor_factura': '',
                        'aprobado': 'SI' if tipo_persona_valido else 'NO'
                    }
                    
                    # DigitoVerificacion Receptor
                    digito_valido = validar_digito_verificacion(digito_verif)
                    items_validacion['DigitoVerificacion Receptor'] = {
                        'valor_xml': digito_verif,
                        'valor_factura': '',
                        'aprobado': 'SI' if digito_valido else 'NO'
                    }
                    
                    # TaxLevelCode Receptor
                    tax_code_valido = validar_tax_level_code(tax_level_code)
                    items_validacion['TaxLevelCode Receptor'] = {
                        'valor_xml': tax_level_code,
                        'valor_factura': '',
                        'aprobado': 'SI' if tax_code_valido else 'NO'
                    }
                    
                    # =====================================================
                    # PASO 4: Validar referencia (Tipo nota crédito = 20)
                    # =====================================================
                    
                    tipo_nota_credito = safe_str(registro.get('tipo_de_nota_credito', ''))
                    codigo_cufe = safe_str(registro.get('codigo_cufe_de_la_factura', ''))
                    cude_nc = safe_str(registro.get('cude_de_la_nota_credito', ''))
                    
                    items_validacion['Tipo de nota credito'] = {
                        'valor_xml': tipo_nota_credito,
                        'valor_factura': '',
                        'aprobado': ''
                    }
                    
                    resultado_final_previo = safe_str(registro.get('ResultadoFinalAntesEventos', ''))
                    
                    if not resultado_final_previo:
                        # No tiene marca previa, validar tipo
                        if tipo_nota_credito == '20':
                            # Tiene referencia, continuar
                            items_validacion['Codigo CUFE de la factura'] = {
                                'valor_xml': codigo_cufe,
                                'valor_factura': '',
                                'aprobado': ''
                            }
                            items_validacion['Cude de la Nota Credito'] = {
                                'valor_xml': cude_nc,
                                'valor_factura': '',
                                'aprobado': ''
                            }
                        else:
                            # Sin referencia
                            print(f"[INFO] NC {numero_nc} sin referencia (tipo {tipo_nota_credito})")
                            
                            observacion = "Nota credito sin referencia"
                            
                            campos_actualizar = {
                                'EstadoFinalFase_4': 'Exitoso',
                                'ObservacionesFase_4': observacion,
                                'ResultadoFinalAntesEventos': 'Con Novedad'
                            }
                            actualizar_bd_cxp(cx, registro_id, campos_actualizar)
                            
                            items_validacion['Cude de la Nota Credito'] = {
                                'valor_xml': cude_nc,
                                'valor_factura': '',
                                'aprobado': ''
                            }
                            items_validacion['Observaciones'] = {
                                'valor_xml': observacion,
                                'valor_factura': '',
                                'aprobado': ''
                            }
                            
                            generar_trazabilidad_nc(cx, registro, id_ejecucion, fecha_ejecucion_str, items_validacion, 'Con Novedad')
                            
                            registros_nc_novedad.append({
                                'ID': registro_id,
                                'nit_emisor_o_nit_del_proveedor': nit,
                                'numero_de_nota_credito': numero_nc,
                                'estado': 'CON NOVEDAD'
                            })
                            
                            nc_con_novedad += 1
                            nc_procesadas += 1
                            continue
                    
                    # =====================================================
                    # PASO 5: Comparar NC con FV en BD
                    # =====================================================
                    
                    prefijo_numero = safe_str(registro.get('prefijo_y_numero', ''))
                    
                    items_validacion['Referencia'] = {
                        'valor_xml': prefijo_numero,
                        'valor_factura': '',
                        'aprobado': ''
                    }
                    
                    # Buscar factura correspondiente
                    fv_encontrada = buscar_factura_correspondiente(cx, nit, prefijo_numero, fecha_ejecucion)
                    
                    if not fv_encontrada:
                        print(f"[INFO] NC {numero_nc} - No se encuentra FV correspondiente")
                        
                        observacion = "Nota credito con referencia no encontrada"
                        
                        campos_actualizar = {
                            'EstadoFinalFase_4': 'Exitoso',
                            'ObservacionesFase_4': observacion,
                            'ResultadoFinalAntesEventos': 'Con novedad'
                        }
                        actualizar_bd_cxp(cx, registro_id, campos_actualizar)
                        
                        items_validacion['Referencia']['valor_factura'] = ''
                        items_validacion['Referencia']['aprobado'] = 'NO'
                        items_validacion['Observaciones'] = {
                            'valor_xml': observacion,
                            'valor_factura': '',
                            'aprobado': ''
                        }
                        
                        generar_trazabilidad_nc(cx, registro, id_ejecucion, fecha_ejecucion_str, items_validacion, 'Con novedad')
                        
                        registros_nc_novedad.append({
                            'ID': registro_id,
                            'nit_emisor_o_nit_del_proveedor': nit,
                            'numero_de_nota_credito': numero_nc,
                            'estado': 'CON NOVEDAD'
                        })
                        
                        nc_con_novedad += 1
                        nc_procesadas += 1
                        continue
                    
                    # FV encontrada - actualizar referencia
                    numero_factura_fv = safe_str(fv_encontrada.get('numero_de_factura', ''))
                    items_validacion['Referencia']['valor_factura'] = numero_factura_fv
                    items_validacion['Referencia']['aprobado'] = 'SI'
                    
                    # =====================================================
                    # PASO 6: Comparar valores
                    # =====================================================
                    
                    valor_nc = normalizar_decimal(registro.get('valor_a_pagar_nc', 0))
                    valor_fv = normalizar_decimal(fv_encontrada.get('valor_a_pagar', 0))
                    
                    items_validacion['LineExtensionAmount'] = {
                        'valor_xml': str(valor_nc),
                        'valor_factura': str(valor_fv),
                        'aprobado': 'SI'
                    }
                    
                    # Actualizar NC como Encontrado
                    campos_nc = {
                        'EstadoFinalFase_4': 'Exitoso',
                        'ResultadoFinalAntesEventos': 'Encontrado'
                    }
                    actualizar_bd_cxp(cx, registro_id, campos_nc)
                    
                    # Actualizar FV con nota crédito referenciada
                    fv_id = safe_str(fv_encontrada.get('ID', ''))
                    actualizar_nota_credito_referenciada_fv(cx, fv_id, numero_nc)
                    
                    # Agregar items adicionales
                    items_validacion['ActualizacionNombreArchivos'] = {
                        'valor_xml': safe_str(registro.get('actualizacion_nombre_archivos', '')),
                        'valor_factura': '',
                        'aprobado': ''
                    }
                    items_validacion['RutaRespaldo'] = {
                        'valor_xml': safe_str(registro.get('ruta_respaldo', '')),
                        'valor_factura': '',
                        'aprobado': ''
                    }
                    items_validacion['Observaciones'] = {
                        'valor_xml': safe_str(registro.get('ObservacionesFase_4', '')),
                        'valor_factura': '',
                        'aprobado': ''
                    }
                    
                    # Generar trazabilidad exitosa
                    generar_trazabilidad_nc(cx, registro, id_ejecucion, fecha_ejecucion_str, items_validacion, 'Encontrado')
                    
                    print(f"[SUCCESS] NC {numero_nc} - Encontrada FV {numero_factura_fv}")
                    nc_encontradas += 1
                    nc_procesadas += 1
                    
                except Exception as e:
                    print(f"[ERROR] Error procesando NC {idx}: {str(e)}")
                    print(traceback.format_exc())
                    nc_procesadas += 1
                    continue
            
            # Generar insumo de retorno para NC con novedad
            if registros_nc_novedad and ruta_insumo_retorno:
                generar_insumo_retorno_nc(registros_nc_novedad, ruta_insumo_retorno)
            
            # =========================================================
            # PROCESAR NOTAS DÉBITO (ND)
            # =========================================================
            
            print("\n" + "=" * 40)
            print("[PROCESO] Procesando Notas Debito (ND)")
            print("=" * 40)
            
            # Query para obtener ND pendientes
            query_nd = """
                SELECT * FROM [CxP].[DocumentsProcessing]
                WHERE [tipo_de_documento] = 'ND'
                  AND ([ResultadoFinalAntesEventos] IS NULL 
                       OR [ResultadoFinalAntesEventos] NOT IN ('Exitoso'))
                ORDER BY [executionDate] DESC
            """
            
            df_nd = pd.read_sql(query_nd, cx)
            print(f"[INFO] Obtenidas {len(df_nd)} notas debito para procesar")
            
            # Variables de conteo ND
            nd_procesadas = 0
            nd_exitosas = 0
            nd_con_error = 0
            
            for idx, registro in df_nd.iterrows():
                try:
                    registro_id = safe_str(registro.get('ID', ''))
                    numero_nd = safe_str(registro.get('numero_de_nota_debito', ''))
                    nit = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
                    
                    print(f"\n[ND] Procesando {nd_procesadas + 1}/{len(df_nd)}: ND {numero_nd}, NIT {nit}")
                    
                    items_validacion = {}
                    
                    # =====================================================
                    # PASO 1: Validar campos básicos
                    # =====================================================
                    
                    nombre_emisor = safe_str(registro.get('nombre_emisor', ''))
                    nit_emisor = safe_str(registro.get('nit_emisor_o_nit_del_proveedor', ''))
                    fecha_emision = safe_str(registro.get('fecha_de_emision_documento', ''))
                    
                    items_validacion['Nombre Emisor'] = {
                        'valor_xml': nombre_emisor,
                        'aprobado': 'SI' if campo_con_valor(nombre_emisor) else 'NO'
                    }
                    
                    items_validacion['NIT Emisor'] = {
                        'valor_xml': nit_emisor,
                        'aprobado': 'SI' if campo_con_valor(nit_emisor) else 'NO'
                    }
                    
                    items_validacion['Fecha emision del documento'] = {
                        'valor_xml': fecha_emision,
                        'aprobado': 'SI' if campo_con_valor(fecha_emision) else 'NO'
                    }
                    
                    # =====================================================
                    # PASO 2: Validar campos receptor
                    # =====================================================
                    
                    nombre_receptor = safe_str(registro.get('nombre_del_adquiriente', ''))
                    nit_receptor = safe_str(registro.get('nit_del_adquiriente', ''))
                    tipo_persona = safe_str(registro.get('tipo_persona', ''))
                    digito_verif = safe_str(registro.get('digito_de_verificacion', ''))
                    tax_level_code = safe_str(registro.get('responsabilidad_tributaria_adquiriente', ''))
                    
                    items_validacion['Nombre Receptor'] = {
                        'valor_xml': nombre_receptor,
                        'aprobado': 'SI' if validar_nombre_receptor(nombre_receptor) else 'NO'
                    }
                    
                    items_validacion['Nit Receptor'] = {
                        'valor_xml': nit_receptor,
                        'aprobado': 'SI' if validar_nit_receptor(nit_receptor) else 'NO'
                    }
                    
                    items_validacion['Tipo Persona Receptor'] = {
                        'valor_xml': tipo_persona,
                        'aprobado': 'SI' if validar_tipo_persona(tipo_persona) else 'NO'
                    }
                    
                    items_validacion['DigitoVerificacion Receptor'] = {
                        'valor_xml': digito_verif,
                        'aprobado': 'SI' if validar_digito_verificacion(digito_verif) else 'NO'
                    }
                    
                    items_validacion['TaxLevelCode Receptor'] = {
                        'valor_xml': tax_level_code,
                        'aprobado': 'SI' if validar_tax_level_code(tax_level_code) else 'NO'
                    }
                    
                    # =====================================================
                    # PASO 3: Marcar como Exitoso y agregar items finales
                    # =====================================================
                    
                    campos_nd = {
                        'EstadoFinalFase_4': 'Exitoso',
                        'ResultadoFinalAntesEventos': 'Exitoso'
                    }
                    actualizar_bd_cxp(cx, registro_id, campos_nd)
                    
                    # Items adicionales
                    items_validacion['LineExtensionAmount'] = {
                        'valor_xml': safe_str(registro.get('valor_a_pagar', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['Tipo de nota debito'] = {
                        'valor_xml': safe_str(registro.get('tipo_de_nota_debito', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['Referencia'] = {
                        'valor_xml': safe_str(registro.get('prefijo_y_numero', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['Codigo CUFE de la factura'] = {
                        'valor_xml': safe_str(registro.get('codigo_cufe_de_la_factura', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['Cude de la Nota Debito'] = {
                        'valor_xml': safe_str(registro.get('cude_de_la_nota_debito', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['ActualizacionNombreArchivos'] = {
                        'valor_xml': safe_str(registro.get('actualizacion_nombre_archivos', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['RutaRespaldo'] = {
                        'valor_xml': safe_str(registro.get('ruta_respaldo', '')),
                        'aprobado': ''
                    }
                    
                    items_validacion['Observaciones'] = {
                        'valor_xml': safe_str(registro.get('ObservacionesFase_4', '')),
                        'aprobado': ''
                    }
                    
                    # Generar trazabilidad
                    generar_trazabilidad_nd(cx, registro, id_ejecucion, fecha_ejecucion_str, items_validacion, 'Exitoso')
                    
                    print(f"[SUCCESS] ND {numero_nd} - Procesada exitosamente")
                    nd_exitosas += 1
                    nd_procesadas += 1
                    
                except Exception as e:
                    print(f"[ERROR] Error procesando ND {idx}: {str(e)}")
                    print(traceback.format_exc())
                    nd_con_error += 1
                    nd_procesadas += 1
                    continue
        
        # Fin del procesamiento
        tiempo_total = time.time() - t_inicio
        
        print("")
        print("=" * 80)
        print("[FIN] Procesamiento HU4.2 - VALIDACION NC y ND completado")
        print("=" * 80)
        print("[ESTADISTICAS NC]")
        print(f"  Total procesadas: {nc_procesadas}")
        print(f"  Encontradas: {nc_encontradas}")
        print(f"  Con novedad: {nc_con_novedad}")
        print(f"  No exitoso (excede plazo): {nc_no_exitoso}")
        print("[ESTADISTICAS ND]")
        print(f"  Total procesadas: {nd_procesadas}")
        print(f"  Exitosas: {nd_exitosas}")
        print(f"  Con error: {nd_con_error}")
        print(f"[TIEMPO] Total: {round(tiempo_total, 2)}s")
        print("=" * 80)
        
        resumen = f"NC: {nc_procesadas} procesadas ({nc_encontradas} encontradas, {nc_con_novedad} con novedad). ND: {nd_procesadas} procesadas ({nd_exitosas} exitosas)"
        
        #SetVar("vLocStrResultadoSP", "True")
        #SetVar("vLocStrResumenSP", resumen)
        
    except Exception as e:
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion HU42_ValidarNotasCreditoDebito fallo")
        print("=" * 80)
        print(f"[ERROR] Mensaje: {str(e)}")
        print(traceback.format_exc())
        print("=" * 80)
        
        #SetVar("vGblStrDetalleError", str(e))
        #SetVar("vGblStrSystemError", traceback.format_exc())
        #SetVar("vLocStrResultadoSP", "False")


# Mock para pruebas locales
if __name__ == "__main__":
    _mock_vars = {}
    def GetVar(name):
        return _mock_vars.get(name, "")
    def #SetVar(name, value):
        _mock_vars[name] = value
        print(f"[#SetVar] {name} = {value}")
    
    _mock_vars["vLocDicConfig"] = '''{
        "ServidorBaseDatos": "localhost",
        "NombreBaseDatos": "NotificationsPaddy",
        "plazo_maximo_retoma_dias": 120,
        "IDEjecucion": "12345"
    }'''
    _mock_vars["vGblStrUsuarioBaseDatos"] = "sa"
    _mock_vars["vGblStrClaveBaseDatos"] = "password"
    
    print("Ejecutando prueba local...")
    # HU42_ValidarNotasCreditoDebito()
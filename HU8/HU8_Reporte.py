"""
================================================================================
SCRIPT: HU8_Reporte.py
================================================================================

Descripcion General:
--------------------
    Genera reportes consolidados y organiza archivos del proceso CxP.
    Crea estructura de carpetas en File Server, mueve archivos segun estado,
    y genera multiples reportes Excel (diarios, mensuales y anuales).

Autor: Diego Ivan Lopez Ochoa
Version: 1.0 - 12 Enero 2026
Plataforma: RocketBot RPA

================================================================================
REPORTES GENERADOS
================================================================================

    DIARIOS:
        - Reporte_de_ejecucion_CXP
        - Reporte_de_ejecucion_COMERCIALIZADOS

    MENSUALES (ejecutan un dia especifico del mes):
        - Reporte_de_ejecucion_GRANOS
        - Reporte_de_ejecucion_MAIZ
        - Reporte_KPIs_CXP
        - Consolidado_FV_CXP_ConNovedad
        - Consolidado_CXP_NoExitososRechazados
        - Consolidado_CXP_Pendientes
        - Consolidado_NC_ND_CXP

    ANUALES:
        - Consolidado_Global_CXP

================================================================================
DIAGRAMA DE FLUJO
================================================================================

    +-------------------------------------------------------------+
    |                        INICIO                               |
    |              HU8_GenerarReportesCxP()                       |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  1. Cargar configuracion desde vLocDicConfig                |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  2. Crear arbol de carpetas en File Server                  |
    |     - CONSOLIDADOS                                          |
    |     - INSUMO DE RETORNO                                     |
    |     - RESULTADOS BOT CXP                                    |
    |     - Carpetas por estado (14 categorias)                   |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  3. Conectar a BD y consultar DocumentsProcessing           |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  4. Para cada registro:                                     |
    |  +-------------------------------------------------------+  |
    |  |  a. Verificar archivos XML/PDF en ruta respaldo       |  |
    |  |  b. Determinar carpeta destino segun estado           |  |
    |  |  c. Mover/copiar archivos a carpeta correspondiente   |  |
    |  +-------------------------------------------------------+  |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  5. Generar reportes diarios                                |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  6. SI es dia de reportes mensuales:                        |
    |     -> Generar reportes mensuales y KPIs                    |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  7. SI es mes de reporte anual:                             |
    |     -> Generar Consolidado_Global_CXP                       |
    +-----------------------------+-------------------------------+
                                  |
                                  v
    +-------------------------------------------------------------+
    |  8. Retornar estadisticas a RocketBot                       |
    +-------------------------------------------------------------+

================================================================================
ESTRUCTURA DE CARPETAS CREADAS
================================================================================

    {RutaFileServer}/
    |
    +-- {Anio}/
        |
        +-- {Mes}/
        |   |
        |   +-- CONSOLIDADOS/
        |   +-- INSUMO DE RETORNO/
        |   +-- {Dia}/
        |       |
        |       +-- RESULTADOS BOT CXP/
        |       +-- EJECUCION {N} CXP/
        |           |
        |           +-- CXP/
        |               |
        |               +-- INSUMOS/
        |                   |
        |                   +-- EN ESPERA/
        |                   +-- CON NOVEDAD NO CONTADO/
        |                   +-- CON NOVEDAD NO CONTADO/EXCLUIDOS CONTABILIZACION/
        |                   +-- CON NOVEDAD CONTADO/
        |                   +-- CON NOVEDAD CONTADO/EXCLUIDOS CONTABILIZACION/
        |                   +-- APROBADOS NO CONTADO/
        |                   +-- APROBADOS CONTADO/
        |                   +-- APROBADOS SIN CONTABILIZACION/
        |                   +-- APROBADO CONTADO Y O EVENTO MANUAL/
        |                   +-- NO EXITOSOS/
        |                   +-- PENDIENTES/
        |                   +-- RECLASIFICADOS/
        |                   +-- RECHAZADOS/
        |                   +-- ND EXITOSOS/
        |                   +-- NC ENCONTRADOS/
        |
        +-- MATERIA PRIMA GRANOS/
        |   +-- {Anio}/
        |       +-- {Mes}/
        |           +-- INSUMO/
        |           +-- RESULTADO/
        |
        +-- MATERIA PRIMA MAIZ/
        |   +-- {Anio}/
        |       +-- {Mes}/
        |           +-- INSUMO/
        |           +-- RESULTADO/
        |
        +-- COMERCIALIZADOS/
            +-- {Anio}/
                +-- {Mes}/
                    +-- {Dia}/
                        +-- INSUMO/
                        |   +-- CON NOVEDAD/
                        |   +-- EN ESPERA/
                        +-- RESULTADO/

================================================================================
CARPETAS DE DESTINO POR ESTADO
================================================================================

    ResultadoFinalAntesEventos        ->  Carpeta Destino
    ----------------------------          ---------------
    EN ESPERA                         ->  EN ESPERA
    CON NOVEDAD                       ->  CON NOVEDAD NO CONTADO
    CON NOVEDAD CONTADO               ->  CON NOVEDAD CONTADO
    CON NOVEDAD EXCLUIDO CONTAB       ->  CON NOVEDAD NO CONTADO/EXCLUIDOS...
    APROBADO                          ->  APROBADOS NO CONTADO
    APROBADO CONTADO                  ->  APROBADOS CONTADO
    APROBADO SIN CONTABILIZACION      ->  APROBADOS SIN CONTABILIZACION
    APROBADO CONTADO Y/O EVENTO       ->  APROBADO CONTADO Y O EVENTO MANUAL
    NO EXITOSO                        ->  NO EXITOSOS
    PENDIENTE                         ->  PENDIENTES
    RECLASIFICADO                     ->  RECLASIFICADOS
    RECHAZADO                         ->  RECHAZADOS
    
    Notas Debito (ND):
        EXITOSO                       ->  ND EXITOSOS
        NO EXITOSO                    ->  NO EXITOSOS
        
    Notas Credito (NC):
        ENCONTRADO                    ->  NC ENCONTRADOS

================================================================================
VARIABLES DE ENTRADA (RocketBot)
================================================================================

    vLocDicConfig : str | dict
        Configuracion JSON con parametros:
        - ServidorBaseDatos: Servidor SQL Server
        - NombreBaseDatos: Nombre de la base de datos
        - UsuarioBaseDatos: Usuario SQL (opcional)
        - ClaveBaseDatos: Contrasena SQL (opcional)
        - RutaFileServer: Ruta base del File Server
        - NumeroEjecucion: Numero de ejecucion actual
        - DiaReporteMensual: Dia del mes para generar reportes mensuales
        - MesReporteAnual: Mes para generar reporte anual

================================================================================
VARIABLES DE SALIDA (RocketBot)
================================================================================

    vLocStrResultadoSP : str
        "True" si exitoso, "False" si error critico

    vLocStrResumenSP : str
        "HU8 completada. reportes generados"

    vGblStrDetalleError : str
        Traceback en caso de error critico

    vGblStrSystemError : str
        "ErrorHU4_4.1" en caso de error

================================================================================
FUNCIONES PRINCIPALES
================================================================================

    crear_arbol_carpetas(ruta_base, ult_numero):
        Crea toda la estructura de carpetas necesaria
        Retorna dict con rutas importantes
        
    determinar_carpeta_destino(resultado_final, tipo_documento):
        Determina carpeta de destino segun estado y tipo
        
    verificar_archivos_insumo(ruta_respaldo, nombre_archivos):
        Verifica existencia de archivos XML y PDF
        
    mover_archivos_a_destino(ruta_xml, carpeta_destino, ...):
        Mueve/copia archivos a carpeta correspondiente
        
    crear_excel_desde_df(df, ruta_archivo, nombre_hoja):
        Crea Excel formateado desde DataFrame
        
    crear_excel_multihoja(hojas_data, ruta_archivo):
        Crea Excel con multiples hojas

================================================================================
FUNCIONES DE GENERACION DE REPORTES
================================================================================

    generar_reporte_ejecucion_cxp():
        Reporte diario de ejecucion CxP
        
    generar_reporte_granos():
        Reporte mensual materia prima granos
        
    generar_reporte_maiz():
        Reporte mensual materia prima maiz
        
    generar_reporte_comercializados():
        Reporte diario comercializados (OC inicia con 50)
        
    generar_reporte_kpis():
        KPIs mensuales del proceso
        
    generar_consolidado_fv_novedad():
        Consolidado mensual FV con novedad
        
    generar_consolidado_no_exitosos():
        Consolidado mensual no exitosos y rechazados
        
    generar_consolidado_pendientes():
        Consolidado mensual pendientes
        
    generar_consolidado_nc_nd():
        Consolidado mensual notas credito y debito
        
    generar_reporte_anual_global():
        Consolidado anual de todo el proceso

================================================================================
FORMATO EXCEL
================================================================================

    Encabezados:
        - Fondo azul (#4472C4)
        - Texto blanco en negrita
        - Alineacion centrada
        - Bordes finos
        
    Columnas:
        - Ancho automatico (max 50 caracteres)
        
    Datos:
        - Valores nulos se muestran como vacio

================================================================================
EJEMPLOS DE USO
================================================================================

    # Configurar variables en RocketBot
    SetVar("vLocDicConfig", json.dumps({
        "ServidorBaseDatos": "servidor.ejemplo.com",
        "NombreBaseDatos": "NotificationsPaddy",
        "RutaFileServer": "\\\\fileserver\\CxP",
        "NumeroEjecucion": 1,
        "DiaReporteMensual": 1,
        "MesReporteAnual": 1
    }))
    
    # Ejecutar funcion
    HU8_GenerarReportesCxP()
    
    # Verificar resultado
    resultado = GetVar("vLocStrResultadoSP")

================================================================================
NOTAS TECNICAS
================================================================================

    - Carpetas se crean con os.makedirs(exist_ok=True)
    - Archivos se copian con shutil.copy2 (preserva metadata)
    - Comercializados: OC que inicia con '50' se copia adicionalmente
    - Meses en espanol sin tildes para nombres de carpetas
    - Reportes mensuales solo se generan en dia configurado
    - Reporte anual solo se genera en mes configurado

================================================================================
CONSTANTES
================================================================================

    MESES_ESPANOL:
        Diccionario de meses {1: '01. Enero', 2: '02. Febrero', ...}
        
    CARPETAS_INSUMOS:
        Lista de las 14 carpetas de clasificacion por estado

================================================================================
"""

def HU8_GenerarReportesCxP():
    """
    Función para generar reportes y organizar archivos del proceso CxP.
    
    VERSIÓN: 1.0 - 12 Enero 2026
    
    FUNCIONALIDADES:
        1. Crear árbol de carpetas en File Server
        2. Identificar y verificar archivos XML/PDF
        3. Mover archivos según estado del registro
        4. Generar reportes:
           - Reporte_de_ejecución_CXP (diario)
           - Reporte_de_ejecución_GRANOS (mensual)
           - Reporte_de_ejecución_MAÍZ (mensual)
           - Reporte_de_ejecución_COMERCIALIZADOS (diario)
           - Reporte_KPIs_CXP (mensual)
           - Consolidado_FV_CXP_ConNovedad (mensual)
           - Consolidado_CXP_NoExitososRechazados (mensual)
           - Consolidado_CXP_Pendientes (mensual)
           - Consolidado_Global_CXP (anual)
           - Consolidado_NC_ND_CXP (mensual)
    
    Returns:
        None: Actualiza variables globales en RocketBot
    """
    
    # =========================================================================
    # IMPORTS
    # =========================================================================
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime, timedelta
    from contextlib import contextmanager
    import time
    import warnings
    import re
    import os
    import shutil
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    warnings.filterwarnings('ignore')
    
    # =========================================================================
    # CONSTANTES
    # =========================================================================
    
    MESES_ESPANOL = {
        1: '01. Enero', 2: '02. Febrero', 3: '03. Marzo', 4: '04. Abril',
        5: '05. Mayo', 6: '06. Junio', 7: '07. Julio', 8: '08. Agosto',
        9: '09. Septiembre', 10: '10. Octubre', 11: '11. Noviembre', 12: '12. Diciembre'
    }
    
    CARPETAS_INSUMOS = [
        'EN ESPERA',
        'CON NOVEDAD NO CONTADO',
        'CON NOVEDAD NO CONTADO/EXCLUIDOS CONTABILIZACION',
        'CON NOVEDAD CONTADO',
        'CON NOVEDAD CONTADO/EXCLUIDOS CONTABILIZACION',
        'APROBADOS NO CONTADO',
        'APROBADOS CONTADO',
        'APROBADOS SIN CONTABILIZACION',
        'APROBADO CONTADO Y O EVENTO MANUAL',
        'NO EXITOSOS',
        'PENDIENTES',
        'RECLASIFICADOS',
        'RECHAZADOS',
        'ND EXITOSOS',
        'NC ENCONTRADOS'
    ]
    
    # =========================================================================
    # FUNCIONES AUXILIARES BÁSICAS
    # =========================================================================
    
    def safe_str(v):
        """Convierte un valor a string de manera segura."""
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        if isinstance(v, bytes):
            try:
                return v.decode('latin-1', errors='replace').strip()
            except:
                return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (np.isnan(v) or pd.isna(v)):
                return ""
            return str(v)
        try:
            return str(v).strip()
        except:
            return ""
    
    def parse_config(raw):
        """Parsea la configuración desde RocketBot."""
        if isinstance(raw, dict):
            if not raw:
                raise ValueError("Config vacia (dict)")
            return raw
        text = safe_str(raw)
        if not text:
            raise ValueError("vLocDicConfig vacio")
        try:
            config = json.loads(text)
            if not config:
                raise ValueError("Config vacia (JSON)")
            return config
        except json.JSONDecodeError:
            pass
        try:
            config = ast.literal_eval(text)
            if not config:
                raise ValueError("Config vacia (literal)")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError(f"Config invalida: {str(e)}")
    
    def campo_vacio(valor):
        """Verifica si un campo está vacío."""
        valor_str = safe_str(valor)
        return valor_str == "" or valor_str.lower() in ('null', 'none', 'nan')
    
    def campo_con_valor(valor):
        """Verifica si un campo tiene valor."""
        return not campo_vacio(valor)
    
    # =========================================================================
    # CONEXIÓN A BASE DE DATOS
    # =========================================================================
    
    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        """
        Context Manager para establecer una conexion segura y resiliente a SQL Server.

        Implementa una estrategia de reintentos (Retries) y prueba dos metodos
        de autenticacion secuencialmente, garantizando maxima compatibilidad
        con diferentes configuraciones de servidor.

        Args:
            cfg (dict): Diccionario con credenciales y parametros de conexion.
                Claves requeridas:
                    - ServidorBaseDatos (str): Hostname o IP del servidor SQL.
                    - NombreBaseDatos (str): Nombre de la base de datos.
                Claves opcionales:
                    - UsuarioBaseDatos (str): Usuario para autenticacion SQL.
                    - ClaveBaseDatos (str): Contrasena del usuario SQL.
            max_retries (int, optional): Numero maximo de intentos por cada
                metodo de autenticacion. Por defecto 3.

        Yields:
            pyodbc.Connection: Objeto de conexion activo con autocommit deshabilitado.
                - Se hace commit automatico al salir del contexto sin errores.
                - Se hace rollback automatico si ocurre una excepcion.

        Raises:
            ValueError: Si faltan parametros obligatorios (ServidorBaseDatos,
                NombreBaseDatos) en el diccionario de configuracion.
            pyodbc.Error: Si no se logra conectar tras agotar todos los intentos
                con ambos metodos de autenticacion.
            Exception: Cualquier error que ocurra durante operaciones de BD.

        Examples:
            Uso basico con autenticacion SQL::

                cfg = {
                    "ServidorBaseDatos": "sqlserver.empresa.com",
                    "NombreBaseDatos": "CxP_Produccion",
                    "UsuarioBaseDatos": "app_user",
                    "ClaveBaseDatos": "SecurePass123"
                }
                with crear_conexion_db(cfg) as conexion:
                    cursor = conexion.cursor()
                    cursor.execute("SELECT * FROM Tabla")
                    resultados = cursor.fetchall()
                # Commit automatico al salir del bloque 'with'

            Manejo de errores con rollback automatico::

                try:
                    with crear_conexion_db(cfg) as conexion:
                        cursor = conexion.cursor()
                        cursor.execute("INSERT INTO Tabla VALUES (...)")
                        raise ValueError("Error intencional")
                except ValueError:
                    # El rollback ya se ejecuto automaticamente
                    print("Transaccion revertida")

        Note:
            Estrategia de conexion:
            
            1. **Fase 1**: Intenta autenticacion SQL (UID/PWD) hasta max_retries veces.
            2. **Fase 2**: Si Fase 1 falla, intenta Trusted Connection (Windows Auth).
            3. Hay una pausa de 1 segundo entre reintentos fallidos.
            4. Usa ODBC Driver 17 for SQL Server.
            5. Timeout de conexion: 30 segundos.

        Warning:
            - La conexion debe usarse dentro de un bloque ``with``.
            - No reutilizar el objeto de conexion fuera del contexto.
            - El commit/rollback se maneja automaticamente.
        """
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing:
            raise ValueError(f"Parametros faltantes: {', '.join(missing)}")

        usuario = cfg['UsuarioBaseDatos']
        contrasena = cfg['ClaveBaseDatos']
        
        # Cadenas de conexion para los dos metodos soportados
        conn_str_auth = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={cfg['ServidorBaseDatos']};"
            f"DATABASE={cfg['NombreBaseDatos']};"
            f"UID={usuario};"
            f"PWD={contrasena};"
            "autocommit=False;"
        )
        
        conn_str_trusted = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={cfg['ServidorBaseDatos']};"
            f"DATABASE={cfg['NombreBaseDatos']};"
            "Trusted_Connection=yes;"
            "autocommit=False;"
        )

        cx = None
        conectado = False
        excepcion_final = None

        # Fase 1: Intentar Autenticacion SQL
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str_auth, timeout=30)
                cx.autocommit = False
                conectado = True
                break
            except pyodbc.Error as e:
                excepcion_final = e
                if attempt < max_retries - 1:
                    time.sleep(1)

        # Fase 2: Intentar Trusted Connection (solo si fallo la anterior)
        if not conectado:
            for attempt in range(max_retries):
                try:
                    cx = pyodbc.connect(conn_str_trusted, timeout=30)
                    cx.autocommit = False
                    conectado = True
                    break
                except pyodbc.Error as e:
                    excepcion_final = e
                    if attempt < max_retries - 1:
                        time.sleep(1)

        if not conectado:
            raise excepcion_final or Exception("No se pudo conectar a la base de datos con ningun metodo")
        
        try:
            yield cx
            if cx:
                cx.commit()  # Commit final si todo salio bien
        except Exception as e:
            if cx:
                cx.rollback()  # Rollback en caso de error dentro del bloque 'with'
                print(f"[ERROR] Rollback por error: {str(e)}")
            raise
        finally:
            if cx:
                try:
                    cx.close()
                except:
                    pass
                
    def actualizar_insumos_comparativa(registro, cx, nit, factura, nombre_item, valor_insumo):
        """
        Actualiza específicamente los ítems de insumos (InsumoPDF, InsumoXML) 
        en la tabla [dbo].[CxP.Comparativa].
        
        Args:
            registro (dict | pd.Series): Registro con ID_dp.
            cx (pyodbc.Connection): Conexión activa.
            nit (str): NIT del proveedor.
            factura (str): Número de factura.
            nombre_item (str): 'InsumoPDF' o 'InsumoXML'.
            valor_insumo (str): El valor a guardar (ej: 'SI', 'NO', o la ruta).
        """
        cur = cx.cursor()
        id_reg = registro.get('ID_dp', '')

        def safe_db_val(v):
            if v is None: return None
            s = str(v).strip()
            return None if s.lower() in ['none', 'null', ''] else s

        val_final = safe_db_val(valor_insumo)

        try:
            # Verificar si ya existe el registro para este ítem específico
            query_check = """
            SELECT COUNT(*) FROM [dbo].[CxP.Comparativa]
            WHERE NIT = ? AND Factura = ? AND Item = ? AND ID_registro = ?
            """
            cur.execute(query_check, (nit, factura, nombre_item, id_reg))
            existe = cur.fetchone()[0] > 0

            if existe:
                # Actualizar solo el campo Valor_XML
                update_query = """
                UPDATE [dbo].[CxP.Comparativa]
                SET Valor_XML = ?
                WHERE NIT = ? AND Factura = ? AND Item = ? AND ID_registro = ?
                """
                cur.execute(update_query, (val_final, nit, factura, nombre_item, id_reg))
            else:
                # Insertar registro nuevo con la información base del documento
                insert_query = """
                INSERT INTO [dbo].[CxP.Comparativa] (
                    Fecha_de_retoma_antes_de_contabilizacion, Tipo_de_Documento, 
                    Orden_de_Compra, Nombre_Proveedor, ID_registro, NIT, Factura, 
                    Item, Valor_XML
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                cur.execute(insert_query, (
                    registro.get('Fecha_de_retoma_antes_de_contabilizacion_dp',''),
                    registro.get('documenttype_dp',''),
                    registro.get('numero_de_liquidacion_u_orden_de_compra_dp',''),
                    registro.get('nombre_emisor_dp',''), 
                    id_reg, nit, factura, nombre_item, val_final
                ))
            
            cx.commit()
        except Exception as e:
            print(f"[ERROR] Error actualizando insumo {nombre_item}: {str(e)}")
            cx.rollback()
        finally:
            cur.close()
    
    # =========================================================================
    # FUNCIONES DE GESTIÓN DE CARPETAS
    # =========================================================================
    
    def verificar_acceso_ruta(ruta_base):
        """Verifica si se tiene acceso a la ruta del File Server."""
        try:
            if os.path.exists(ruta_base):
                # Intentar crear un archivo temporal para verificar permisos de escritura
                test_file = os.path.join(ruta_base, '.test_access')
                try:
                    with open(test_file, 'w') as f:
                        f.write('test')
                    os.remove(test_file)
                    return True
                except:
                    print(f"[WARNING] Acceso de solo lectura a {ruta_base}")
                    return True  # Al menos tiene acceso de lectura
            return False
        except Exception as e:
            print(f"[ERROR] Error verificando acceso: {str(e)}")
            return False
    
    def crear_arbol_carpetas(ruta_base, fecha_ejecucion, ult_numero):
        """
        Crea el árbol completo de carpetas según la estructura requerida.
        
        Estructura:
            /AÑO/##. MES/CONSOLIDADOS
            /AÑO/##. MES/INSUMO DE RETORNO
            /AÑO/##. MES/DÍA/RESULTADOS BOT CXP
            /AÑO/##. MES/DÍA/EJECUCION ## CXP/CXP/INSUMOS/[subcarpetas]
            /AÑO/MATERIA PRIMA GRANOS/AÑO/##. MES/INSUMO|RESULTADO
            /AÑO/MATERIA PRIMA MAÍZ/AÑO/##. MES/INSUMO|RESULTADO
            /AÑO/COMERCIALIZADOS/AÑO/##. MES/DÍA/INSUMO|RESULTADO
        """
        try:
            anio = fecha_ejecucion.year
            mes = fecha_ejecucion.month
            dia = fecha_ejecucion.day
            
            try:
                mes_nombre = MESES_ESPANOL.get(mes)
                dia_str = f'{dia:02d}'
                if not mes_nombre or not dia_str:
                    raise Exception('No fue posible extraer el nombre del mes o del dia para la creacion de las carpetas')
            except Exception as e:
                raise e
            
            ult_numero = f"EJECUCION {ult_numero} CXP"
            
            # Estructura principal
            rutas_crear = [
                # Consolidados y retorno
                os.path.join(ruta_base, str(anio), mes_nombre, 'CONSOLIDADOS'),
                os.path.join(ruta_base, str(anio), mes_nombre, 'INSUMO DE RETORNO'),
                
                # Resultados del día
                os.path.join(ruta_base, str(anio), mes_nombre, dia_str, 'RESULTADOS BOT CXP'),
                
                # Carpetas de insumos por estado
            ]
            
            # Agregar carpetas de insumos
            ruta_insumos = os.path.join(ruta_base, str(anio), mes_nombre, dia_str, ult_numero, 'CXP', 'INSUMOS')
            for carpeta in CARPETAS_INSUMOS:
                rutas_crear.append(os.path.join(ruta_insumos, carpeta))
            
            # Materia Prima Granos
            rutas_crear.extend([
                os.path.join(ruta_base, str(anio), 'MATERIA PRIMA GRANOS', str(anio), mes_nombre, 'INSUMO'),
                os.path.join(ruta_base, str(anio), 'MATERIA PRIMA GRANOS', str(anio), mes_nombre, 'RESULTADO'),
            ])
            
            # Materia Prima Maíz
            rutas_crear.extend([
                os.path.join(ruta_base, str(anio), 'MATERIA PRIMA MAIZ', str(anio), mes_nombre, 'INSUMO'),
                os.path.join(ruta_base, str(anio), 'MATERIA PRIMA MAIZ', str(anio), mes_nombre, 'RESULTADO'),
            ])
            
            # Comercializados
            rutas_crear.extend([
                os.path.join(ruta_base, str(anio), 'COMERCIALIZADOS', str(anio), mes_nombre, dia_str, 'INSUMO', 'CON NOVEDAD'),
                os.path.join(ruta_base, str(anio), 'COMERCIALIZADOS', str(anio), mes_nombre, dia_str, 'INSUMO', 'EN ESPERA'),
                os.path.join(ruta_base, str(anio), 'COMERCIALIZADOS', str(anio), mes_nombre, dia_str, 'RESULTADO'),
            ])
            
            # Crear todas las carpetas
            for ruta in rutas_crear:
                os.makedirs(ruta, exist_ok=True)
            
            print(f"[INFO] Arbol de carpetas creado/verificado exitosamente")
            
            # Retornar rutas importantes
            return {
                'consolidados': os.path.join(ruta_base, str(anio), mes_nombre, 'CONSOLIDADOS'),
                'insumo_retorno': os.path.join(ruta_base, str(anio), mes_nombre, 'INSUMO DE RETORNO'),
                'resultados_dia': os.path.join(ruta_base, str(anio), mes_nombre, dia_str, 'RESULTADOS BOT CXP'),
                'insumos_cxp': ruta_insumos,
                'granos_resultado': os.path.join(ruta_base, str(anio), 'MATERIA PRIMA GRANOS', str(anio), mes_nombre, 'RESULTADO'),
                'maiz_resultado': os.path.join(ruta_base, str(anio), 'MATERIA PRIMA MAIZ', str(anio), mes_nombre, 'RESULTADO'),
                'comercializados_resultado': os.path.join(ruta_base, str(anio), 'COMERCIALIZADOS', str(anio), mes_nombre, dia_str, 'RESULTADO'),
                'comercializados_insumo': os.path.join(ruta_base, str(anio), 'COMERCIALIZADOS', str(anio), mes_nombre, dia_str, 'INSUMO'),
                'global_anual': os.path.join(ruta_base, str(anio)),
            }
            
        except Exception as e:
            print(f"[ERROR] Error creando arbol de carpetas: {str(e)}")
            raise
    
    def determinar_carpeta_destino(resultado_final, tipo_documento):
        """
        Determina la carpeta de destino según el resultado final y tipo de documento.
        
        Args:
            resultado_final: Estado del registro (ResultadoFinalAntesEventos)
            tipo_documento: FV, NC o ND
        
        Returns:
            str: Nombre de la carpeta destino
        """
        resultado = safe_str(resultado_final).upper()
        tipo = safe_str(tipo_documento).upper()
        
        # ND siempre va a ND EXITOSOS si es exitoso
        if tipo == 'ND':
            if 'NO EXITOSO' in resultado:
                return 'NO EXITOSOS'
            elif 'EXITOSO' in resultado:
                return 'ND EXITOSOS'
            elif 'PENDIENTE' in resultado:
                return 'PENDIENTES'
            else:
                return ''
        
        # NC
        if tipo == 'NC':
            if 'ENCONTRADO' in resultado:
                return 'NC ENCONTRADOS'
        
        # FV y NC común
        if 'NO EXITOSO' in resultado:
            return 'NO EXITOSOS'
        
        if 'PENDIENTE' in resultado:
            return 'PENDIENTES'
        
        if 'EN ESPERA' in resultado:
            return 'EN ESPERA'
        
        if 'RECHAZADO' in resultado:
            return 'RECHAZADOS'
        
        if 'RECLASIFICADO' in resultado:
            return 'RECLASIFICADOS'
        
        # CON NOVEDAD
        if 'CON NOVEDAD' in resultado:
            if 'EXCLUIDOS CONTABILIZACION' in resultado or 'EXCLUIDO CONTABILIZACION' in resultado:
                if 'CONTADO' in resultado:
                    return 'CON NOVEDAD CONTADO/EXCLUIDOS CONTABILIZACION'
                else:
                    return 'CON NOVEDAD NO CONTADO/EXCLUIDOS CONTABILIZACION'
            elif 'CONTADO' in resultado:
                return 'CON NOVEDAD CONTADO'
            else:
                return 'CON NOVEDAD NO CONTADO'
        
        # APROBADO
        if 'APROBADO' in resultado:
            if 'CONTADO Y/O EVENTO MANUAL' in resultado or 'CONTADO Y O EVENTO MANUAL' in resultado:
                return 'APROBADO CONTADO Y O EVENTO MANUAL'
            elif 'SIN CONTABILIZACION' in resultado:
                return 'APROBADOS SIN CONTABILIZACION'
            elif 'CONTADO' in resultado:
                return 'APROBADOS CONTADO'
            else:
                return 'APROBADOS NO CONTADO'
        
        # Por defecto
        return 'PENDIENTES'
    
    # =========================================================================
    # FUNCIONES DE MANEJO DE ARCHIVOS
    # =========================================================================
    
    def verificar_archivos_insumo(ruta_respaldo, nombre_archivos):
        """
        Verifica la existencia de archivos XML y PDF a partir de un nombre que ya incluye extensión.
        
        Returns:
            tuple: (xml_encontrado, pdf_encontrado, ruta_xml, ruta_pdf)
        """
        xml_encontrado = False
        pdf_encontrado = False
        ruta_xml = None
        ruta_pdf = None
        
        try:
            if campo_vacio(ruta_respaldo) or campo_vacio(nombre_archivos):
                return xml_encontrado, pdf_encontrado, ruta_xml, ruta_pdf
            
            # 1. Limpiamos el nombre y separamos la extensión
            nombre_completo = safe_str(nombre_archivos)
            # os.path.splitext separa 'archivo.xml' en ('archivo', '.xml')
            nombre_base, ext = os.path.splitext(nombre_completo)
            
            # 2. Definimos las rutas finales
            # Forzamos las extensiones correctas partiendo del nombre base
            ruta_xml = os.path.join(ruta_respaldo, f"{nombre_base}.xml")
            ruta_pdf = os.path.join(ruta_respaldo, f"{nombre_base}.pdf")
            
            # 3. Verificamos XML
            if os.path.exists(ruta_xml):
                xml_encontrado = True
            else:
                ruta_xml = None # Limpiamos si no existe
                
            # 4. Verificamos PDF
            if os.path.exists(ruta_pdf):
                pdf_encontrado = True
            else:
                ruta_pdf = None # Limpiamos si no existe
                
        except Exception as e:
            print(f"[ERROR] Error verificando archivos: {str(e)}")
        
        return xml_encontrado, pdf_encontrado, ruta_xml, ruta_pdf
    
    def mover_archivos_a_destino(ruta_xml, carpeta_destino, numero_oc=None, ruta_comercializados=None):
        """
        Mueve los archivos XML y PDF a la carpeta de destino.
        También copia a comercializados si OC inicia con 50.
        
        Returns:
            str: Nueva ruta de respaldo
        """
        nueva_ruta = None
        
        try:
            os.makedirs(carpeta_destino, exist_ok=True)
            
            archivos_movidos = []
            
            if ruta_xml and os.path.exists(ruta_xml):
                nombre_xml = os.path.basename(ruta_xml)
                destino_xml = os.path.join(carpeta_destino, nombre_xml)
                shutil.copy2(ruta_xml, destino_xml)
                archivos_movidos.append(destino_xml)
                
                # Copiar a comercializados si aplica
                if numero_oc and safe_str(numero_oc).startswith('50') and ruta_comercializados:
                    os.makedirs(ruta_comercializados, exist_ok=True)
                    shutil.copy2(ruta_xml, os.path.join(ruta_comercializados, nombre_xml))
            
            if archivos_movidos:
                nueva_ruta = carpeta_destino
                
            return nueva_ruta
            
        except Exception as e:
            print(f"[ERROR] Error moviendo archivos: {str(e)}")
            return None
        
    
    # =========================================================================
    # FUNCIONES DE FORMATO EXCEL
    # =========================================================================
    
    def aplicar_formato_encabezado(ws, num_columnas):
        """Aplica formato de encabezado a la primera fila."""
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, num_columnas + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def ajustar_ancho_columnas(ws):
        """Ajusta automáticamente el ancho de las columnas al contenido."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column].width = adjusted_width
    
    def crear_excel_desde_df(df, ruta_archivo, nombre_hoja='Datos'):
        """Crea un archivo Excel formateado desde un DataFrame."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = nombre_hoja
            
            # Escribir encabezados
            for col_idx, column in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=column)
            
            # Escribir datos
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    if pd.isna(value):
                        ws.cell(row=row_idx, column=col_idx, value='')
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Aplicar formato
            aplicar_formato_encabezado(ws, len(df.columns))
            ajustar_ancho_columnas(ws)
            
            wb.save(ruta_archivo)
            print(f"[INFO] Archivo creado: {ruta_archivo}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Error creando Excel: {str(e)}")
            return False
    
    def crear_excel_multihoja(hojas_data, ruta_archivo):
        """
        Crea un archivo Excel con múltiples hojas.
        
        Args:
            hojas_data: dict {nombre_hoja: DataFrame}
            ruta_archivo: Ruta del archivo a crear
        """
        try:
            wb = Workbook()
            
            # Eliminar hoja por defecto
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            for nombre_hoja, df in hojas_data.items():
                ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel limita a 31 caracteres
                
                # Escribir encabezados
                for col_idx, column in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=column)
                
                # Escribir datos
                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        if pd.isna(value):
                            ws.cell(row=row_idx, column=col_idx, value='')
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Aplicar formato
                aplicar_formato_encabezado(ws, len(df.columns))
                ajustar_ancho_columnas(ws)
            
            wb.save(ruta_archivo)
            print(f"[INFO] Archivo multihoja creado: {ruta_archivo}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Error creando Excel multihoja: {str(e)}")
            return False
    
    # =========================================================================
    # FUNCIONES DE GENERACIÓN DE REPORTES
    # =========================================================================
    
    def generar_reporte_cxp(df_main, df_detalles, df_historico, rutabase):
        """
        Toma los 3 dataframes (simulando las 3 tablas SQL), hace los cruces (merges)
        y genera un Excel formateado.
        
        Args:
            df_main: DataFrame de DocumentsProcessing
            df_detalles: DataFrame de CxP.Comparativa
            df_historico: DataFrame de HistoricoOrdenesCompra
            rutabase: Ruta donde se guardará el archivo
        
        Returns:
            str: Ruta completa del archivo generado
        """
        
        # =========================================================================
        # 1. DEFINICIÓN DE COLUMNAS POR HOJA
        # =========================================================================
        
        # Hoja FACTURAS (20 columnas)
        cols_facturas = [
            'Fecha de ejecución',
            'Fecha primera revisión antes de Contab.',
            'ID ejecución',
            'ID Registro',
            'Tipo de documento',
            'Orden de Compra',
            'Clase de pedido',
            'NIT',
            'Nombre Proveedor',
            'Factura',
            'Item',
            'Valor XML',
            'Valor Orden de Compra',
            'Valor Orden de Compra Comercializados',
            'Aprobado',
            'Estado validación antes de eventos',
            'Fecha primera revisión para Contab.',
            'Estado contabilización',
            'Fecha primera revisión para Compensación',
            'Estado compensación'
        ]
        
        # Hoja NC (12 columnas)
        cols_nc = [
            'Fecha de ejecución',
            'Fecha primera revisión',
            'ID ejecución',
            'ID Registro',
            'NIT',
            'Nombre Proveedor',
            'Nota Credito',
            'Item',
            'Valor XML',
            'Valor Factura',
            'Aprobado',
            'Estado'
        ]
        
        # Hoja ND (11 columnas)
        cols_nd = [
            'Fecha de ejecución',
            'Fecha primera revisión',
            'ID ejecución',
            'ID Registro',
            'NIT',
            'Nombre Proveedor',
            'Nota Debito',
            'Item',
            'Valor XML',
            'Aprobado',
            'Estado'
        ]
        
        # =========================================================================
        # 2. PREPARACIÓN DE DATOS (LÓGICA SQL EN PANDAS)
        # =========================================================================
        
        print("📊 Preparando datos para el reporte...")
        
        # Copias para no modificar los originales
        df_main_copy = df_main.copy()
        df_detalles_copy = df_detalles.copy()
        df_historico_copy = df_historico.copy()
        
        # -----------------------------------------------------------------
        # Paso A: Limpieza de claves para asegurar que los cruces funcionen
        # -----------------------------------------------------------------
        df_main_copy['nit_join'] = df_main_copy['nit_emisor_o_nit_del_proveedor'].astype(str).str.strip()
        df_main_copy['factura_join'] = df_main_copy['numero_de_factura'].astype(str).str.strip()
        df_main_copy['doc_compra_join'] = df_main_copy['numero_de_liquidacion_u_orden_de_compra'].astype(str).str.strip()
        
        df_detalles_copy['nit_join'] = df_detalles_copy['NIT'].astype(str).str.strip()
        df_detalles_copy['factura_join'] = df_detalles_copy['Factura'].astype(str).str.strip()
        
        df_historico_copy['nit_join'] = df_historico_copy['NitCedula'].astype(str).str.strip()
        df_historico_copy['doc_compra_join'] = df_historico_copy['DocCompra'].astype(str).str.strip()
        
        # -----------------------------------------------------------------
        # Paso B: Lógica de la Tabla 3 (HistoricoOrdenesCompra)
        # Solo necesitamos el primer valor -> eliminamos duplicados
        # -----------------------------------------------------------------
        df_historico_unique = df_historico_copy.drop_duplicates(
            subset=['nit_join', 'doc_compra_join'], 
            keep='first'
        )
        
        # -----------------------------------------------------------------
        # Paso C: Cruce Principal (Main + Detalles)
        # Expande la fila principal por cada Item en la tabla de detalles
        # -----------------------------------------------------------------
        # Seleccionar solo columnas necesarias de detalles para evitar duplicados
        cols_detalles_necesarias = ['nit_join', 'factura_join', 'Item', 'Valor_XML', 
                                    'Valor_Orden_de_Compra', 'Valor_Orden_de_Compra_Comercializados', 'Aprobado']
        cols_detalles_existentes = [c for c in cols_detalles_necesarias if c in df_detalles_copy.columns]
        
        df_merged = pd.merge(
            df_main_copy,
            df_detalles_copy[cols_detalles_existentes],
            how='left',
            on=['nit_join', 'factura_join']
        )
        
        # -----------------------------------------------------------------
        # Paso D: Cruce con Histórico (para obtener ClaseDePedido)
        # -----------------------------------------------------------------
        df_final = pd.merge(
            df_merged,
            df_historico_unique[['nit_join', 'doc_compra_join', 'ClaseDePedido']],
            how='left',
            on=['nit_join', 'doc_compra_join']
        )
        
        # Eliminar columnas de join temporales
        cols_drop = ['nit_join', 'factura_join', 'doc_compra_join']
        df_final = df_final.drop(columns=[c for c in cols_drop if c in df_final.columns], errors='ignore')
        
        print(f"   Registros después del cruce: {len(df_final)}")
        
        # =========================================================================
        # 3. MAPEO Y RENOMBRADO DE COLUMNAS
        # =========================================================================
        
        # Mapeo base común
        base_mapping = {
            'executionDate': 'Fecha de ejecución',
            'Fecha_de_retoma_antes_de_contabilizacion': 'Fecha primera revisión antes de Contab.',
            'executionNum': 'ID ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'ClaseDePedido': 'Clase de pedido',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'Item': 'Item',
            'Valor_XML': 'Valor XML',
            'Valor_Orden_de_Compra': 'Valor Orden de Compra',
            'Valor_Orden_de_Compra_Comercializados': 'Valor Orden de Compra Comercializados',
            'Aprobado': 'Aprobado',
            'ResultadoFinalAntesEventos': 'Estado validación antes de eventos',
            'Fecha_retoma_contabilizacion': 'Fecha primera revisión para Contab.',
            'Estado_contabilizacion': 'Estado contabilización',
            'FechaRetomaCompensacion_Fase7': 'Fecha primera revisión para Compensación',
            'EstadoCompensacionFase_7': 'Estado compensación'
        }
        
        # Renombrar columnas existentes
        cols_a_renombrar = {k: v for k, v in base_mapping.items() if k in df_final.columns}
        df_final = df_final.rename(columns=cols_a_renombrar)
        
        # =========================================================================
        # 4. FUNCIÓN PARA PREPARAR CADA HOJA
        # =========================================================================
        
        def preparar_hoja(df_input, target_columns, doc_type, mapping_adicional=None):
            """
            Prepara un DataFrame para una hoja específica.
            """
            # Filtrar por tipo de documento
            if doc_type:
                df = df_input[df_input['Tipo de documento'] == doc_type].copy()
            else:
                df = df_input.copy()
            
            if df.empty:
                return pd.DataFrame(columns=target_columns)
            
            # Aplicar mapeo adicional si existe
            if mapping_adicional:
                df = df.rename(columns=mapping_adicional)
            
            # Crear columnas faltantes con valores vacíos
            for col in target_columns:
                if col not in df.columns:
                    df[col] = None
            
            # Formatear fechas
            fecha_cols = [c for c in target_columns if 'Fecha' in c]
            for col in fecha_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            # Retornar con las columnas en el orden exacto
            return df[target_columns]
        
        # =========================================================================
        # 5. PROCESAMIENTO DE CADA HOJA
        # =========================================================================
        
        print("   Procesando hojas...")
        
        # --- Hoja FACTURAS ---
        df_facturas = preparar_hoja(df_final, cols_facturas, 'FV')
        print(f"   - FACTURAS: {len(df_facturas)} registros")
        
        # --- Hoja NC ---
        # Mapeo especial para NC
        map_nc = {
            'Fecha primera revisión antes de Contab.': 'Fecha primera revisión',
            'Factura': 'Nota Credito',
            'Valor Orden de Compra': 'Valor Factura',
            'Estado validación antes de eventos': 'Estado'
        }
        df_nc = preparar_hoja(df_final, cols_nc, 'NC', map_nc)
        print(f"   - NC: {len(df_nc)} registros")
        
        # --- Hoja ND ---
        # Mapeo especial para ND
        map_nd = {
            'Fecha primera revisión antes de Contab.': 'Fecha primera revisión',
            'Factura': 'Nota Debito',
            'Estado validación antes de eventos': 'Estado'
        }
        df_nd = preparar_hoja(df_final, cols_nd, 'ND', map_nd)
        print(f"   - ND: {len(df_nd)} registros")
        
        # =========================================================================
        # 6. GENERACIÓN DEL ARCHIVO EXCEL
        # =========================================================================
        
        # Generar nombre del archivo con formato ddmmaaaa_HHmm
        ahora = datetime.now()
        str_fecha_hora = ahora.strftime('%d%m%Y_%H%M')
        nombre_archivo = f"Reporte_de_ejecución_CXP_{str_fecha_hora}.xlsx"
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        # Configuración de hojas
        sheets_config = {
            'FACTURAS': df_facturas,
            'NC': df_nc,
            'ND': df_nd
        }
        
        # =========================================================================
        # 7. ESCRITURA CON FORMATO
        # =========================================================================
        
        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # -----------------------------------------------------------------
            # Definir estilos
            # -----------------------------------------------------------------
            
            # Formato de encabezado (azul corporativo)
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#1F4E78',
                'font_color': '#FFFFFF',
                'border': 1
            })
            
            # Formato para fechas
            date_format = workbook.add_format({
                'num_format': 'yyyy-mm-dd',
                'border': 1,
                'align': 'center'
            })
            
            # Formato para moneda/valores
            money_format = workbook.add_format({
                'num_format': '#,##0.00',
                'border': 1,
                'align': 'right'
            })
            
            # Formato para texto general
            text_format = workbook.add_format({
                'border': 1,
                'text_wrap': True,
                'valign': 'vcenter'
            })
            
            # Formato para texto centrado
            text_center_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # -----------------------------------------------------------------
            # Procesar cada hoja
            # -----------------------------------------------------------------
            
            for sheet_name, df_data in sheets_config.items():
                if df_data is None or len(df_data) == 0:
                    print(f"   ⚠️ Hoja {sheet_name} vacía, omitiendo...")
                    continue
                
                # Escribir datos empezando en fila 0 para encabezados
                df_data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                
                worksheet = writer.sheets[sheet_name]
                (max_row, max_col) = df_data.shape
                
                # ---------------------------------------------------------
                # Crear Tabla con Filtros
                # ---------------------------------------------------------
                if max_row > 0:
                    table_name = f'Tabla_{sheet_name}'
                    column_settings = [{'header': col} for col in df_data.columns]
                    worksheet.add_table(0, 0, max_row, max_col - 1, {
                        'columns': column_settings,
                        'style': 'TableStyleMedium2',
                        'name': table_name
                    })
                
                # ---------------------------------------------------------
                # Aplicar formato de encabezados (azul corporativo)
                # ---------------------------------------------------------
                for col_num, value in enumerate(df_data.columns):
                    worksheet.write(0, col_num, value, header_format)
                
                # ---------------------------------------------------------
                # Ajustar ancho de columnas según contenido y tipo
                # ---------------------------------------------------------
                for i, col in enumerate(df_data.columns):
                    # Calcular ancho basado en el contenido
                    if not df_data.empty:
                        max_len_contenido = df_data[col].astype(str).map(len).max()
                    else:
                        max_len_contenido = 0
                    
                    col_width = max(max_len_contenido, len(col)) + 2
                    col_width = max(col_width, 10)   # Mínimo 10
                    col_width = min(col_width, 50)   # Máximo 50
                    
                    # Ajustes específicos por tipo de columna
                    if 'Fecha' in col:
                        worksheet.set_column(i, i, 14, date_format)
                    elif 'Valor' in col:
                        worksheet.set_column(i, i, 18, money_format)
                    elif col in ['NIT', 'ID ejecución', 'ID Registro']:
                        worksheet.set_column(i, i, 12, text_center_format)
                    elif col in ['Aprobado', 'Tipo de documento', 'Clase de pedido']:
                        worksheet.set_column(i, i, 12, text_center_format)
                    elif col == 'Nombre Proveedor':
                        worksheet.set_column(i, i, 35, text_format)
                    elif col == 'Orden de Compra':
                        worksheet.set_column(i, i, 40, text_format)
                    elif 'Estado' in col:
                        worksheet.set_column(i, i, 30, text_center_format)
                    else:
                        worksheet.set_column(i, i, col_width, text_format)
                
                # ---------------------------------------------------------
                # Ajustar altura de fila de encabezados
                # ---------------------------------------------------------
                worksheet.set_row(0, 40)
                
                # ---------------------------------------------------------
                # Congelar panel superior
                # ---------------------------------------------------------
                worksheet.freeze_panes(1, 0)
        
        print(f"\n✅ Reporte generado exitosamente: {ruta_completa}")
        print(f"   - FACTURAS: {len(df_facturas)} registros")
        print(f"   - NC: {len(df_nc)} registros")
        print(f"   - ND: {len(df_nd)} registros")
        
        return ruta_completa

    def generar_reporte_granos(df_main, df_detalles, rutabase):
        """
        Genera el reporte de GRANOS en un archivo Excel con UNA SOLA HOJA.
        Combina la tabla principal (DocumentsProcessing) con los detalles (Comparativa).
        """
        
        # ---------------------------------------------------------
        # 1. PREPARACIÓN Y CRUCE DE DATOS
        # ---------------------------------------------------------
        
        # A. Limpieza de claves para asegurar que el cruce (Join) funcione perfecto
        # Convertimos a string y quitamos espacios y decimales extraños (.0) si existen
        df_main['nit_join'] = df_main['nit_emisor_o_nit_del_proveedor'].apply(lambda x: str(x).strip().replace('.0', '') if pd.notnull(x) else '')
        df_main['factura_join'] = df_main['numero_de_factura'].astype(str).str.strip()
        
        df_detalles['nit_join'] = df_detalles['NIT'].apply(lambda x: str(x).strip().replace('.0', '') if pd.notnull(x) else '')
        df_detalles['factura_join'] = df_detalles['Factura'].astype(str).str.strip()

        # B. Cruce (Merge)
        # Hacemos un LEFT JOIN para mantener todos los registros de la tabla principal
        # y traer los detalles (Item, Valor XML, etc.) si existen.
        df_final = pd.merge(
            df_main,
            df_detalles, 
            how='left',
            left_on=['nit_join', 'factura_join'],
            right_on=['nit_join', 'factura_join']
        )

        # ---------------------------------------------------------
        # 2. DEFINICIÓN Y RENOMBRADO DE COLUMNAS
        # ---------------------------------------------------------
        # Mapeamos las columnas SQL a nombres bonitos para el Excel
        column_mapping = {
            # --- Tabla Principal ---
            'executionDate': 'Fecha de ejecución',
            'Fecha_de_retoma_antes_de_contabilizacion': 'Fecha de retoma',
            'executionNum': 'ID ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ResultadoFinalAntesEventos': 'Estado validación antes de eventos',
            
            # --- Tabla Comparativa (Detalles) ---
            'Item': 'Item',
            'Valor_XML': 'Valor XML',
            'Valor_Orden_de_Compra': 'Valor Orden de Compra',
            'Valor_Orden_de_Compra_Comercializados': 'Valor OC Comercializados',
            'Aprobado': 'Aprobado'
        }

        # Filtramos solo las columnas que existen y las renombramos
        cols_to_keep = [c for c in column_mapping.keys() if c in df_final.columns]
        df_final = df_final[cols_to_keep].rename(columns=column_mapping)

        # Ordenamos las columnas de forma lógica para la lectura
        desired_order = [
            'Fecha de ejecución', 'Fecha de retoma', 'ID ejecución', 'ID Registro', 
            'Tipo de documento', 'Orden de Compra', 'NIT', 'Nombre Proveedor', 'Factura', 
            'Item', 'Valor XML', 'Valor Orden de Compra', 'Valor OC Comercializados', 
            'Aprobado', 'Estado validación antes de eventos'
        ]
        # Seleccionamos solo las que logramos obtener
        final_cols = [c for c in desired_order if c in df_final.columns]
        df_final = df_final[final_cols]

        # ---------------------------------------------------------
        # 3. GENERACIÓN DEL EXCEL (DISEÑO)
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%d%m%Y_%H%M')

        # 3. Construir el nombre completo
        nombre_archivo = f"Reporte_de_ejecución_GRANOS_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        sheet_name = "Facturas" # Única hoja solicitada

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # --- Estilos Personalizados ---
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#1F4E78', # Azul oscuro elegante
                'font_color': '#FFFFFF',
                'border': 1
            })
            
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            money_format = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            # --- Escribir Datos ---
            # Escribimos el dataframe empezando en la fila 2 (row 1) para dejar espacio al encabezado de tabla
            df_final.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
            worksheet = writer.sheets[sheet_name]
            (max_row, max_col) = df_final.shape
            
            # --- Crear Tabla de Excel ---
            # Esto añade automáticamente los filtros y el diseño de bandas
            column_settings = [{'header': col} for col in df_final.columns]
            worksheet.add_table(0, 0, max_row, max_col - 1, {
                'columns': column_settings,
                'style': 'TableStyleMedium2', # Estilo azul medio
                'name': 'TablaGranos'
            })
            
            # Sobrescribimos los encabezados para aplicar nuestro color azul oscuro exacto
            for col_num, value in enumerate(df_final.columns):
                worksheet.write(0, col_num, value, header_format)

            # --- Ajuste Automático de Ancho de Columnas ---
            for i, col in enumerate(df_final.columns):
                # Calcular ancho basado en el contenido más largo o el título
                col_len = max(df_final[col].astype(str).map(len).max(), len(col)) + 3
                col_len = min(col_len, 60) # Poner un límite para que no sea gigante
                
                # Aplicar formato según el tipo de dato
                if 'Fecha' in col:
                    worksheet.set_column(i, i, 14, date_format)
                elif 'Valor' in col:
                    worksheet.set_column(i, i, 18, money_format)
                else:
                    worksheet.set_column(i, i, col_len, text_format)

        print(f"✅ Reporte generado exitosamente: {ruta_completa}")
    
    def generar_reporte_maiz(df_main, df_detalles, rutabase):
        """
        Genera el reporte de GRANOS en un archivo Excel con UNA SOLA HOJA.
        Combina la tabla principal (DocumentsProcessing) con los detalles (Comparativa).
        """
        
        # ---------------------------------------------------------
        # 1. PREPARACIÓN Y CRUCE DE DATOS
        # ---------------------------------------------------------
        
        # A. Limpieza de claves para asegurar que el cruce (Join) funcione perfecto
        # Convertimos a string y quitamos espacios y decimales extraños (.0) si existen
        df_main['nit_join'] = df_main['nit_emisor_o_nit_del_proveedor'].apply(lambda x: str(x).strip().replace('.0', '') if pd.notnull(x) else '')
        df_main['factura_join'] = df_main['numero_de_factura'].astype(str).str.strip()
        
        df_detalles['nit_join'] = df_detalles['NIT'].apply(lambda x: str(x).strip().replace('.0', '') if pd.notnull(x) else '')
        df_detalles['factura_join'] = df_detalles['Factura'].astype(str).str.strip()

        # B. Cruce (Merge)
        # Hacemos un LEFT JOIN para mantener todos los registros de la tabla principal
        # y traer los detalles (Item, Valor XML, etc.) si existen.
        df_final = pd.merge(
            df_main,
            df_detalles, 
            how='left',
            left_on=['nit_join', 'factura_join'],
            right_on=['nit_join', 'factura_join']
        )

        # ---------------------------------------------------------
        # 2. DEFINICIÓN Y RENOMBRADO DE COLUMNAS
        # ---------------------------------------------------------
        # Mapeamos las columnas SQL a nombres bonitos para el Excel
        column_mapping = {
            # --- Tabla Principal ---
            'executionDate': 'Fecha de ejecución',
            'Fecha_de_retoma_antes_de_contabilizacion': 'Fecha de retoma',
            'executionNum': 'ID ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ResultadoFinalAntesEventos': 'Estado validación antes de eventos',
            
            # --- Tabla Comparativa (Detalles) ---
            'Item': 'Item',
            'Valor_XML': 'Valor XML',
            'Valor_Orden_de_Compra': 'Valor Orden de Compra',
            'Valor_Orden_de_Compra_Comercializados': 'Valor OC Comercializados',
            'Aprobado': 'Aprobado'
        }

        # Filtramos solo las columnas que existen y las renombramos
        cols_to_keep = [c for c in column_mapping.keys() if c in df_final.columns]
        df_final = df_final[cols_to_keep].rename(columns=column_mapping)

        # Ordenamos las columnas de forma lógica para la lectura
        desired_order = [
            'Fecha de ejecución', 'Fecha de retoma', 'ID ejecución', 'ID Registro', 
            'Tipo de documento', 'Orden de Compra', 'NIT', 'Nombre Proveedor', 'Factura', 
            'Item', 'Valor XML', 'Valor Orden de Compra', 'Valor OC Comercializados', 
            'Aprobado', 'Estado validación antes de eventos'
        ]
        # Seleccionamos solo las que logramos obtener
        final_cols = [c for c in desired_order if c in df_final.columns]
        df_final = df_final[final_cols]

        # ---------------------------------------------------------
        # 3. GENERACIÓN DEL EXCEL (DISEÑO)
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%d%m%Y_%H%M')

        # 3. Construir el nombre completo
        nombre_archivo = f"Reporte_de_ejecución_MAIZ_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        sheet_name = "Facturas" # Única hoja solicitada

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # --- Estilos Personalizados ---
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#1F4E78', # Azul oscuro elegante
                'font_color': '#FFFFFF',
                'border': 1
            })
            
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            money_format = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            # --- Escribir Datos ---
            # Escribimos el dataframe empezando en la fila 2 (row 1) para dejar espacio al encabezado de tabla
            df_final.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
            worksheet = writer.sheets[sheet_name]
            (max_row, max_col) = df_final.shape
            
            # --- Crear Tabla de Excel ---
            # Esto añade automáticamente los filtros y el diseño de bandas
            column_settings = [{'header': col} for col in df_final.columns]
            worksheet.add_table(0, 0, max_row, max_col - 1, {
                'columns': column_settings,
                'style': 'TableStyleMedium2', # Estilo azul medio
                'name': 'TablaGranos'
            })
            
            # Sobrescribimos los encabezados para aplicar nuestro color azul oscuro exacto
            for col_num, value in enumerate(df_final.columns):
                worksheet.write(0, col_num, value, header_format)

            # --- Ajuste Automático de Ancho de Columnas ---
            for i, col in enumerate(df_final.columns):
                # Calcular ancho basado en el contenido más largo o el título
                col_len = max(df_final[col].astype(str).map(len).max(), len(col)) + 3
                col_len = min(col_len, 60) # Poner un límite para que no sea gigante
                
                # Aplicar formato según el tipo de dato
                if 'Fecha' in col:
                    worksheet.set_column(i, i, 14, date_format)
                elif 'Valor' in col:
                    worksheet.set_column(i, i, 18, money_format)
                else:
                    worksheet.set_column(i, i, col_len, text_format)

        print(f"✅ Reporte generado exitosamente: {ruta_completa}")
    
    def generar_reporte_comercializados(df_main, df_detalles, rutabase):
        """
        Toma los 3 dataframes (simulando las 3 tablas SQL), hace los cruces (merges)
        y genera un Excel formateado.
        """
        
        # ---------------------------------------------------------
        # 1. PREPARACIÓN DE DATOS (LÓGICA SQL EN PANDAS)
        # ---------------------------------------------------------
        
        # Paso A: Limpieza de claves para asegurar que los cruces funcionen
        # Convertimos a string y quitamos espacios para evitar errores de llave
        df_main['nit_join'] = df_main['nit_emisor_o_nit_del_proveedor'].astype(str).str.strip()
        df_main['factura_join'] = df_main['numero_de_factura'].astype(str).str.strip()
        df_main['doc_compra_join'] = df_main['numero_de_liquidacion_u_orden_de_compra'].astype(str).str.strip()

        df_detalles['nit_join'] = df_detalles['NIT'].astype(str).str.strip()
        df_detalles['factura_join'] = df_detalles['Factura'].astype(str).str.strip()


        # Paso B: Lógica de la Tabla 3 (HistoricoOrdenesCompra)
        # "Solo necesitamos extraer el primer valor obtenido" -> eliminamos duplicados manteniendo el primero
        df_historico_unique = df_historico.drop_duplicates(subset=['nit_join', 'doc_compra_join'], keep='first')

        # Paso C: Cruce Principal (Main + Detalles)
        # Esto expande la fila principal por cada Item encontrado en la tabla de detalles
        df_merged = pd.merge(
            df_main,
            df_detalles,
            how='left', # Usamos left para no perder la cabecera si no hay detalles, o 'inner' si es estricto
            left_on=['nit_join', 'factura_join'],
            right_on=['nit_join', 'factura_join']
        )

        # ---------------------------------------------------------
        # 2. RENOMBRADO Y SELECCIÓN DE COLUMNAS (MAPPING)
        # ---------------------------------------------------------
        # Mapeamos las columnas de SQL a los nombres bonitos del Excel final
        column_mapping = {
            'executionDate': 'Fecha de ejecución',
            'Fecha_de_retoma_antes_de_contabilizacion': 'Fecha 1ra Revisión',
            'executionNum': 'ID Ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo Documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'ClaseDePedido': 'Clase de Pedido', # Viene de la tabla 3
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'Item': 'Item', # Viene de la tabla 2
            'Valor_XML': 'Valor XML',
            'Valor_Orden_de_Compra': 'Valor OC',
            'Valor_Orden_de_Compra_Comercializados': 'Valor OC Comercializados',
            'Aprobado': 'Aprobado',
            'ResultadoFinalAntesEventos': 'Estado Validación',
        }

        # Seleccionamos solo las columnas que existen en el mapping y las renombramos
        cols_to_keep = [c for c in column_mapping.keys() if c in df_final.columns]
        df_final = df_final[cols_to_keep].rename(columns=column_mapping)

        # Reordenamos para que quede lógico (opcional)
        desired_order = list(column_mapping.values())
        # Filtramos desired_order para asegurarnos de que solo pedimos columnas que existen
        final_cols = [c for c in desired_order if c in df_final.columns]
        df_final = df_final[final_cols]

        # ---------------------------------------------------------
        # 3. GENERACIÓN DEL EXCEL "HERMOSO"
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%d%m%Y_%H%M')

        # 3. Construir el nombre completo
        nombre_archivo = f"Reporte_de_ejecución_COMERCIALIZADOS_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        # Diccionario para separar las hojas
        sheets_config = {
            'FACTURAS': 'FV'
        }

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # --- Estilos Personalizados ---
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#1F4E78', # Azul oscuro profesional
                'font_color': '#FFFFFF',
                'border': 1
            })
            
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            money_format = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            for sheet_name, doc_type in sheets_config.items():
                # Filtramos por tipo de documento
                df_sheet = df_final[df_final['Tipo Documento'] == doc_type].copy()
                
                if df_sheet.empty:
                    continue # Si no hay datos de ese tipo, saltamos

                # Escribimos los datos en Excel (sin el índice de pandas)
                df_sheet.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                
                # --- Formateo de la Hoja ---
                worksheet = writer.sheets[sheet_name]
                
                # Obtener dimensiones
                (max_row, max_col) = df_sheet.shape
                
                # Crear una tabla de Excel real (añade filtros y estilo automático)
                column_settings = [{'header': column} for column in df_sheet.columns]
                worksheet.add_table(0, 0, max_row, max_col - 1, {
                    'columns': column_settings,
                    'style': 'TableStyleMedium2', # Estilo azul/gris limpio
                    'name': f'Tabla_{sheet_name}'
                })

                # Sobrescribir los encabezados con nuestro formato personalizado (opcional, para forzar el azul oscuro)
                for col_num, value in enumerate(df_sheet.columns):
                    worksheet.write(0, col_num, value, header_format)

                # Ajuste de ancho de columnas y formatos de celdas
                for i, col in enumerate(df_sheet.columns):
                    # Ancho estimado basado en la longitud del encabezado + un poco más
                    column_len = max(df_sheet[col].astype(str).map(len).max(), len(col)) + 2
                    # Tope máximo de ancho
                    column_len = min(column_len, 50) 
                    
                    # Aplicar formatos según el nombre de la columna
                    if 'Fecha' in col:
                        worksheet.set_column(i, i, 12, date_format)
                    elif 'Valor' in col:
                        worksheet.set_column(i, i, 15, money_format)
                    else:
                        worksheet.set_column(i, i, column_len, text_format)

        print(f"✅ Reporte generado exitosamente: {ruta_completa}")

    def generar_consolidado_novedades(df_historico_novedades, df_docs_processing, df_historico_ordenes, rutabase):
        """
        Genera el reporte Consolidado CXP con 2 hojas:
        1. Total Mensual: Filtra df_historico_novedades por el mes anterior a la fecha actual.
        2. Vigentes: Cruza df_docs_processing con df_historico_ordenes.
        """
        
        # ---------------------------------------------------------
        # 1. LÓGICA HOJA 1: "TOTAL MENSUAL" (MES ANTERIOR)
        # ---------------------------------------------------------
        
        # Calcular el mes y año objetivo (mes anterior al actual)
        hoy = datetime.now()
        primer_dia_este_mes = hoy.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_este_mes - timedelta(days=1)
        
        target_month = ultimo_dia_mes_anterior.month
        target_year = ultimo_dia_mes_anterior.year
        
        print(f"📅 Generando reporte para el periodo: {target_year}-{target_month:02d}")

        # Asegurar formato fecha
        df_historico_novedades['Fecha_ejecucion'] = pd.to_datetime(df_historico_novedades['Fecha_ejecucion'])
        
        # Filtrar registros del mes anterior
        mask_mes_anterior = (
            (df_historico_novedades['Fecha_ejecucion'].dt.month == target_month) & 
            (df_historico_novedades['Fecha_ejecucion'].dt.year == target_year)
        )
        df_total_mensual = df_historico_novedades[mask_mes_anterior].copy()

        # Mapeo de columnas (SQL -> Excel)
        cols_map_1 = {
            'Fecha_ejecucion': 'Fecha de ejecución',
            'Fecha_de_retoma': 'Fecha de primera revisión antes de contab.',
            'ID_ejecucion': 'ID ejecución',
            'ID_registro': 'ID Registro',
            'Nit': 'NIT',
            'Nombre_Proveedor': 'Nombre Proveedor',
            'Orden_de_compra': 'Orden de Compra',
            'Factura': 'Factura',
            'Fec_Doc': 'Fec.Doc',
            'Fec_Reg': 'Fec.Reg',
            'Observaciones': 'Observaciones'
        }
        
        # Seleccionar y renombrar
        # Nota: Si falta alguna columna en el SQL original (como Tipo documento), no la incluimos para no romperlo
        cols_existentes_1 = [c for c in cols_map_1.keys() if c in df_total_mensual.columns]
        df_total_mensual = df_total_mensual[cols_existentes_1].rename(columns=cols_map_1)


        # ---------------------------------------------------------
        # 2. LÓGICA HOJA 2: "VIGENTES" (CRUCE DE TABLAS)
        # ---------------------------------------------------------
        
        # Limpieza de claves para el Join
        # Convertimos a string y quitamos .0 y espacios
        def limpiar_clave(val):
            return str(val).strip().replace('.0', '') if pd.notnull(val) else ''

        # Preparar claves Tabla 1 (DocumentsProcessing)
        df_docs_processing['nit_join'] = df_docs_processing['nit_emisor_o_nit_del_proveedor'].apply(limpiar_clave)
        df_docs_processing['doc_compra_join'] = df_docs_processing['numero_de_liquidacion_u_orden_de_compra'].apply(limpiar_clave)
        
        # Preparar claves Tabla 2 (HistoricoOrdenesCompra)
        # Aseguramos que no haya duplicados para el cruce (opcional, pero recomendado)
        # df_historico_ordenes = df_historico_ordenes.drop_duplicates(subset=['NitCedula', 'DocCompra']) 
        df_historico_ordenes['nit_join'] = df_historico_ordenes['NitCedula'].apply(limpiar_clave)
        df_historico_ordenes['doc_compra_join'] = df_historico_ordenes['DocCompra'].apply(limpiar_clave)

        # Realizar el cruce (Left Join para mantener todos los vigentes)
        df_vigentes_full = pd.merge(
            df_docs_processing,
            df_historico_ordenes[['nit_join', 'doc_compra_join', 'FecDoc', 'FecReg']], # Solo traemos lo necesario
            how='left',
            left_on=['nit_join', 'doc_compra_join'],
            right_on=['nit_join', 'doc_compra_join']
        )

        # Mapeo de columnas (SQL -> Excel)
        cols_map_2 = {
            'executionDate': 'Fecha de ejecución',
            'executionNum': 'ID ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'numero_de_factura': 'Factura',
            'FecDoc': 'Fec.Doc', # Viene del cruce
            'FecReg': 'Fec.Reg', # Viene del cruce
            'ObservacionesFase_4': 'Observaciones'
        }

        cols_existentes_2 = [c for c in cols_map_2.keys() if c in df_vigentes_full.columns]
        df_vigentes = df_vigentes_full[cols_existentes_2].rename(columns=cols_map_2)

        # ---------------------------------------------------------
        # 3. GENERACIÓN DEL EXCEL
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%Y%m')

        # 3. Construir el nombre completo
        nombre_archivo = f"Consolidado_FV_CXP_ConNovedad_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # Estilos
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#1F4E78', 'font_color': '#FFFFFF', 'border': 1
            })
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            # Configuración de Hojas
            hojas = {
                'Total Mensual': df_total_mensual,
                'Vigentes': df_vigentes
            }

            for sheet_name, df_sheet in hojas.items():
                if df_sheet.empty:
                    # Crear hoja vacía con encabezados si no hay datos
                    pd.DataFrame(columns=df_sheet.columns).to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                else:
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                
                worksheet = writer.sheets[sheet_name]
                (max_row, max_col) = df_sheet.shape
                
                # Tabla Excel
                if max_row > 0:
                    column_settings = [{'header': col} for col in df_sheet.columns]
                    worksheet.add_table(0, 0, max_row, max_col - 1, {
                        'columns': column_settings,
                        'style': 'TableStyleMedium2',
                        'name': f'Tabla_{sheet_name.replace(" ", "")}'
                    })
                else:
                    # Escribir solo encabezados manuales si está vacía
                    for col_num, value in enumerate(df_sheet.columns):
                        worksheet.write(0, col_num, value, header_format)

                # Escribir encabezados con formato (sobreescribe los de la tabla para asegurar color)
                for col_num, value in enumerate(df_sheet.columns):
                    worksheet.write(0, col_num, value, header_format)

                # Ajustar columnas
                for i, col in enumerate(df_sheet.columns):
                    col_len = max(df_sheet[col].astype(str).map(len).max(), len(col)) + 2 if not df_sheet.empty else len(col) + 2
                    col_len = min(col_len, 50)
                    
                    if 'Fec' in col or 'Fecha' in col:
                        worksheet.set_column(i, i, 14, date_format)
                    else:
                        worksheet.set_column(i, i, col_len, text_format)

        print(f"✅ Reporte generado: {ruta_completa}")
    
    def generar_consolidado_no_exitosos_rechazados(df_no_exitosos_sql, df_rechazados_sql, rutabase):
        """
        Genera el reporte Consolidado CXP No Exitosos y Rechazados.
        Filtra ambos dataframes por el MES ANTERIOR a la fecha actual.
        """
        
        # ---------------------------------------------------------
        # 1. CÁLCULO DEL MES ANTERIOR
        # ---------------------------------------------------------
        hoy = datetime.now()
        primer_dia_este_mes = hoy.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_este_mes - timedelta(days=1)
        
        target_month = ultimo_dia_mes_anterior.month
        target_year = ultimo_dia_mes_anterior.year
        
        print(f"📅 Generando reporte para el periodo: {target_year}-{target_month:02d}")
        
        # ---------------------------------------------------------
        # 2. PROCESAMIENTO DE DATAFRAMES
        # ---------------------------------------------------------
        
        # Mapeo de columnas SQL -> Excel
        # Nota: Omitimos 'Fecha_de_retoma...' ya que no aparece en tus plantillas CSV de este reporte
        col_map = {
            'executionDate': 'Fecha de ejecución',
            'executionNum': 'ID ejecución',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ObservacionesFase_4': 'Observaciones'
        }

        def procesar_df(df_input):
            # Convertir fecha
            df_input['executionDate'] = pd.to_datetime(df_input['executionDate'])
            
            # Filtrar mes anterior
            mask = (df_input['executionDate'].dt.month == target_month) & \
                (df_input['executionDate'].dt.year == target_year)
            df_filtered = df_input[mask].copy()
            
            # Renombrar columnas
            cols_existentes = [c for c in col_map.keys() if c in df_filtered.columns]
            df_final = df_filtered[cols_existentes].rename(columns=col_map)
            
            # Asegurar orden (opcional, pero se ve mejor)
            orden_deseado = [
                'Fecha de ejecución', 'ID ejecución', 'ID Registro', 'Tipo de documento',
                'Orden de Compra', 'NIT', 'Nombre Proveedor', 'Factura', 'Observaciones'
            ]
            cols_finales = [c for c in orden_deseado if c in df_final.columns]
            return df_final[cols_finales]

        # Procesamos ambas tablas
        df_sheet_no_exitosos = procesar_df(df_no_exitosos_sql)
        df_sheet_rechazados = procesar_df(df_rechazados_sql)

        # ---------------------------------------------------------
        # 3. GENERACIÓN DEL EXCEL
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%Y%m')

        # 3. Construir el nombre completo
        nombre_archivo = f"Consolidado_CXP_NoExitososRechazados_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        # Definir las hojas y sus datos correspondientes
        # Query 1 (NO EXITOSO) -> Hoja "No Exitosos Vigentes"
        # Query 2 (RECHAZADO)  -> Hoja "Rechazados Total MES Con Evento"
        sheets_config = {
            'No Exitosos Vigentes': df_sheet_no_exitosos,
            'Rechazados Total MES Con Evento': df_sheet_rechazados
        }

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # Estilos
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#1F4E78', 'font_color': '#FFFFFF', 'border': 1
            })
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            for sheet_name, df_data in sheets_config.items():
                # Escribir datos
                if df_data.empty:
                    pd.DataFrame(columns=df_data.columns).to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                else:
                    df_data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                
                worksheet = writer.sheets[sheet_name]
                (max_row, max_col) = df_data.shape
                
                # Tabla Excel
                if max_row > 0:
                    column_settings = [{'header': col} for col in df_data.columns]
                    worksheet.add_table(0, 0, max_row, max_col - 1, {
                        'columns': column_settings,
                        'style': 'TableStyleMedium2',
                        'name': f'T_{sheet_name.split()[0]}' # Nombre tabla simple
                    })
                else:
                    for col_num, value in enumerate(df_data.columns):
                        worksheet.write(0, col_num, value, header_format)

                # Formatear encabezados
                for col_num, value in enumerate(df_data.columns):
                    worksheet.write(0, col_num, value, header_format)

                # Ajustar columnas
                for i, col in enumerate(df_data.columns):
                    col_len = max(df_data[col].astype(str).map(len).max(), len(col)) + 2 if not df_data.empty else len(col) + 2
                    col_len = min(col_len, 50)
                    
                    if 'Fecha' in col:
                        worksheet.set_column(i, i, 14, date_format)
                    else:
                        worksheet.set_column(i, i, col_len, text_format)

        print(f"✅ Reporte generado: {ruta_completa}")
    
    def generar_consolidado_pendientes(df_eventos_sql, df_compensacion_sql, df_contabilizacion_sql, rutabase):
        """
        Genera el reporte Consolidado CXP Pendientes con 3 hojas:
        1. Pendiente Eventos Vigentes (18 columnas con eventos)
        2. Pendiente Contab Vigentes (10 columnas)
        3. Pendiente Compen Vigentes (9 columnas)

        Args:
            df_eventos_sql: DataFrame con datos de eventos pendientes
            df_compensacion_sql: DataFrame con datos de compensacion pendiente
            df_contabilizacion_sql: DataFrame con datos de contabilizacion pendiente
            rutabase: Ruta donde se guardara el archivo
        """

        # =========================================================================
        # 1. MAPEOS DE COLUMNAS POR TIPO DE HOJA
        # =========================================================================

        # Mapeo para HOJA 1: Pendiente Eventos Vigentes (18 columnas)
        mapping_eventos = {
            'executionDate': 'Fecha de ejecucion',
            'executionNum': 'ID ejecucion',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ResultadoFinalAntesEventos': 'Estado validacion antes de eventos',
            'FechaHora_Evento_030': 'Fecha - hora Evento Acuse de Recibo',
            'Estado_Evento_030': 'Estado Evento Acuse de Recibo',
            'FechaHora_Evento_032': 'Fecha - hora Evento Recibo del bien y/o prestacion del servicio',
            'Estado_Evento_032': 'Estado Evento Recibo del bien y/o prestacion del servicio',
            'FechaHora_Evento_033': 'Fecha - hora Evento Aceptacion Expresa',
            'Estado_Evento_033': 'Estado Evento Aceptacion Expresa',
            'FechaHora_Evento_031': 'Fecha - hora Evento Reclamo de la Factura Electronica de Venta',
            'Estado_Evento_031': 'Estado Evento Reclamo',
            'ObservacionesFase_4': 'Observaciones'
        }

        # Orden de columnas para HOJA 1: Eventos (18 columnas)
        orden_eventos = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'Orden de Compra',
            'NIT',
            'Nombre Proveedor',
            'Factura',
            'Estado validacion antes de eventos',
            'Fecha - hora Evento Acuse de Recibo',
            'Estado Evento Acuse de Recibo',
            'Fecha - hora Evento Recibo del bien y/o prestacion del servicio',
            'Estado Evento Recibo del bien y/o prestacion del servicio',
            'Fecha - hora Evento Aceptacion Expresa',
            'Estado Evento Aceptacion Expresa',
            'Fecha - hora Evento Reclamo de la Factura Electronica de Venta',
            'Estado Evento Reclamo',
            'Observaciones'
        ]

        # Mapeo para HOJA 2: Pendiente Contab Vigentes (10 columnas)
        mapping_contabilizacion = {
            'executionDate': 'Fecha de ejecucion',
            'executionNum': 'ID ejecucion',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ResultadoFinalAntesEventos': 'Estado validacion antes de eventos',
            'ObservacionesFase_4': 'Observaciones'
        }

        # Orden de columnas para HOJA 2: Contabilizacion (10 columnas)
        orden_contabilizacion = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'Orden de Compra',
            'NIT',
            'Nombre Proveedor',
            'Factura',
            'Estado validacion antes de eventos',
            'Observaciones'
        ]

        # Mapeo para HOJA 3: Pendiente Compen Vigentes (9 columnas)
        mapping_compensacion = {
            'executionDate': 'Fecha de ejecucion',
            'executionNum': 'ID ejecucion',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_factura': 'Factura',
            'ObservacionesFase_4': 'Observaciones'
        }

        # Orden de columnas para HOJA 3: Compensacion (9 columnas)
        orden_compensacion = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'Orden de Compra',
            'NIT',
            'Nombre Proveedor',
            'Factura',
            'Observaciones'
        ]

        # =========================================================================
        # 2. FUNCION AUXILIAR PARA PREPARAR CADA HOJA
        # =========================================================================

        def preparar_hoja(df, mapping, orden_columnas):
            """
            Prepara un DataFrame para una hoja especifica del Excel.

            Args:
                df: DataFrame original con datos del SQL
                mapping: Diccionario de mapeo columna_sql -> columna_excel
                orden_columnas: Lista con el orden deseado de columnas

            Returns:
                DataFrame preparado con columnas renombradas y ordenadas
            """
            if df is None or df.empty:
                return pd.DataFrame(columns=orden_columnas)

            df_out = df.copy()

            # Formatear fechas si existen
            fecha_cols = ['executionDate', 'Fecha_de_retoma_antes_de_contabilizacion']
            for col in fecha_cols:
                if col in df_out.columns:
                    df_out[col] = pd.to_datetime(df_out[col], errors='coerce')

            # Formatear columnas fecha-hora de eventos
            evento_fecha_cols = [
                'FechaHora_Evento_030',
                'FechaHora_Evento_031',
                'FechaHora_Evento_032',
                'FechaHora_Evento_033'
            ]
            for col in evento_fecha_cols:
                if col in df_out.columns:
                    df_out[col] = pd.to_datetime(df_out[col], errors='coerce')
                    df_out[col] = df_out[col].apply(
                        lambda x: x.strftime('%d/%m/%Y - %H:%M') if pd.notna(x) else ''
                    )

            # Filtrar columnas existentes
            cols_existentes = [c for c in mapping.keys() if c in df_out.columns]
            df_out = df_out[cols_existentes].copy()

            # Renombrar columnas
            df_out = df_out.rename(columns=mapping)

            # Agregar columnas faltantes
            for col in orden_columnas:
                if col not in df_out.columns:
                    df_out[col] = ''

            # Ordenar columnas
            df_out = df_out[orden_columnas]

            return df_out
        
        # =========================================================================
        # 3. PROCESAMIENTO DE CADA HOJA
        # =========================================================================

        # Hoja 1: Pendiente Eventos Vigentes (18 columnas)
        df_sheet1 = preparar_hoja(df_eventos_sql, mapping_eventos, orden_eventos)

        # Hoja 2: Pendiente Contab Vigentes (10 columnas)
        df_sheet2 = preparar_hoja(df_contabilizacion_sql, mapping_contabilizacion, orden_contabilizacion)

        # Hoja 3: Pendiente Compen Vigentes (9 columnas)
        df_sheet3 = preparar_hoja(df_compensacion_sql, mapping_compensacion, orden_compensacion)

        # =========================================================================
        # 4. GENERACION DEL ARCHIVO EXCEL
        # =========================================================================

        # Generar nombre del archivo con formato YYYYMM
        ahora = datetime.now()
        str_fecha_hora = ahora.strftime('%Y%m')
        nombre_archivo = f"Consolidado_CXP_Pendientes_{str_fecha_hora}.xlsx"
        ruta_completa = os.path.join(rutabase, nombre_archivo)

        # Configuracion de hojas
        sheets_config = {
            'Pendiente Eventos Vigentes': df_sheet1,
            'Pendiente Contab Vigentes': df_sheet2,
            'Pendiente Compen Vigentes': df_sheet3
        }

        # =========================================================================
        # 5. ESCRITURA CON FORMATO
        # =========================================================================

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book

            # -----------------------------------------------------------------
            # Definir estilos
            # -----------------------------------------------------------------

            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#1F4E78',
                'font_color': '#FFFFFF',
                'border': 1
            })

            date_format = workbook.add_format({
                'num_format': 'yyyy-mm-dd',
                'border': 1,
                'align': 'center'
            })

            text_format = workbook.add_format({
                'border': 1,
                'text_wrap': True,
                'valign': 'vcenter'
            })

            text_center_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            # -----------------------------------------------------------------
            # Procesar cada hoja
            # -----------------------------------------------------------------

            for sheet_name, df_data in sheets_config.items():
                df_data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

                worksheet = writer.sheets[sheet_name]
                (max_row, max_col) = df_data.shape

                # ---------------------------------------------------------
                # Crear tabla con filtros (si hay datos)
                # ---------------------------------------------------------
                if max_row > 0:
                    table_name = f'T_{sheet_name.replace(" ", "").replace(".", "")}'

                    column_settings = [{'header': col} for col in df_data.columns]
                    worksheet.add_table(0, 0, max_row, max_col - 1, {
                        'columns': column_settings,
                        'style': 'TableStyleMedium2',
                        'name': table_name
                    })

                # ---------------------------------------------------------
                # Aplicar formato de encabezados
                # ---------------------------------------------------------
                for col_num, value in enumerate(df_data.columns):
                    worksheet.write(0, col_num, value, header_format)

                # ---------------------------------------------------------
                # Ajustar ancho de columnas
                # ---------------------------------------------------------
                for i, col in enumerate(df_data.columns):
                    if not df_data.empty:
                        max_len_contenido = df_data[col].astype(str).map(len).max()
                    else:
                        max_len_contenido = 0

                    col_width = max(max_len_contenido, len(col)) + 2
                    col_width = max(col_width, 10)
                    col_width = min(col_width, 50)

                    if 'Fecha de ejecucion' in col:
                        worksheet.set_column(i, i, 14, date_format)
                    elif 'Fecha - hora' in col:
                        worksheet.set_column(i, i, 22, text_center_format)
                    elif 'Estado' in col:
                        worksheet.set_column(i, i, 15, text_center_format)
                    elif col == 'NIT':
                        worksheet.set_column(i, i, 12, text_center_format)
                    elif col == 'ID ejecucion' or col == 'ID Registro':
                        worksheet.set_column(i, i, 12, text_center_format)
                    elif col == 'Observaciones':
                        worksheet.set_column(i, i, 60, text_format)
                    elif col == 'Nombre Proveedor':
                        worksheet.set_column(i, i, 35, text_format)
                    else:
                        worksheet.set_column(i, i, col_width, text_format)

                # ---------------------------------------------------------
                # Ajustar altura de encabezados
                # ---------------------------------------------------------
                worksheet.set_row(0, 40)

                # ---------------------------------------------------------
                # Congelar encabezados
                # ---------------------------------------------------------
                worksheet.freeze_panes(1, 0)

        print(f"Reporte generado exitosamente: {ruta_completa}")
        print(f"- Pendiente Eventos Vigentes: {len(df_sheet1)} registros")
        print(f"- Pendiente Contab Vigentes: {len(df_sheet2)} registros")
        print(f"- Pendiente Compen Vigentes: {len(df_sheet3)} registros")

        return ruta_completa


    def generar_consolidado_nc_nd_actualizado(df_nc_encontrados_sql, df_nc_novedad_sql, df_nd_sql, rutabase):
        """
        Genera el reporte Consolidado NC ND CXP actualizado con los nuevos campos.
        1. NC Encontrados-NoExitosos MES (Filtra Mes Anterior)
        2. NC Con Novedad Vigentes (NO filtra mes, todo el historial)
        3. ND Total Mes (Filtra Mes Anterior)
        """
        
        # ---------------------------------------------------------
        # 1. CÁLCULO DEL MES ANTERIOR
        # ---------------------------------------------------------
        hoy = datetime.now()
        primer_dia_este_mes = hoy.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_este_mes - timedelta(days=1)
        
        target_month = ultimo_dia_mes_anterior.month
        target_year = ultimo_dia_mes_anterior.year
        
        # Nombre del archivo con formato MMyyyy (Mes Anterior)
        str_periodo = f"{target_month:02d}{target_year}"
        print(f"📅 Periodo objetivo (Mes Anterior): {target_month:02d}-{target_year}")

        # ---------------------------------------------------------
        # 2. DEFINICIÓN DE COLUMNAS TARGET (EXCEL)
        # ---------------------------------------------------------
        # Definimos las columnas que tendrá cada hoja final
        cols_nc_encontrados = [
            'Fecha de ejecución', 'ID ejecución', 'ID Registro', 'NIT', 'Nombre Proveedor', 
            'Nota Credito', 'Tipo de nota crédito', 'Referencia', 'LineExtensionAmount', 
            'Estado', 'Observaciones'
        ]
        
        cols_nc_novedad = [
            'Fecha de ejecución', 'ID ejecución', 'ID Registro', 'NIT', 'Nombre Proveedor', 
            'Nota Credito', 'Tipo de nota crédito', 'Referencia', 'LineExtensionAmount', 
            'Observaciones'
        ]
        
        # Agregamos LineExtensionAmount aquí también porque tu query trae 'valor_a_pagar'
        cols_nd_total = [
            'Fecha de ejecución', 'ID ejecución', 'ID Registro', 'NIT', 'Nombre Proveedor', 
            'Nota Debito', 'Tipo de nota débito', 'Referencia', 'LineExtensionAmount',
            'Observaciones'
        ]

        # ---------------------------------------------------------
        # 3. FUNCIÓN DE PROCESAMIENTO
        # ---------------------------------------------------------
        
        def procesar_hoja(df_input, columnas_destino, filtrar_mes=False, tipo='NC'):
            df = df_input.copy()
            
            # 1. Convertir fecha
            if 'executionDate' in df.columns:
                df['executionDate'] = pd.to_datetime(df['executionDate'])
                
            # 2. Filtro de Mes (Solo si aplica)
            if filtrar_mes:
                mask = (df['executionDate'].dt.month == target_month) & \
                    (df['executionDate'].dt.year == target_year)
                df = df[mask].copy()
                
            # 3. Mapeo de columnas SQL -> Excel
            # Creamos un diccionario dinámico según el tipo (NC o ND)
            col_doc_num = 'Nota Credito' if tipo == 'NC' else 'Nota Debito'
            col_doc_type = 'Tipo de nota crédito' if tipo == 'NC' else 'Tipo de nota débito'
            col_valor = 'valor_a_pagar' if tipo == 'NC' else 'valor_a_pagar'
            
            mapping = {
                'executionDate': 'Fecha de ejecución',
                'executionNum': 'ID ejecución',
                'ID': 'ID Registro',
                'nit_emisor_o_nit_del_proveedor': 'NIT',
                'nombre_emisor': 'Nombre Proveedor',
                'Numero_de_nota_credito': col_doc_num,         # Nuevo campo SQL
                'Tipo_de_nota_cred_deb': col_doc_type,         # Nuevo campo SQL
                'NotaCreditoReferenciada': 'Referencia',       # Nuevo campo SQL
                col_valor: 'LineExtensionAmount',              # Nuevo campo SQL
                'ResultadoFinalAntesEventos': 'Estado',
                'ObservacionesFase_4': 'Observaciones'
            }
            
            df = df.rename(columns=mapping)
            
            # 4. Asegurar columnas faltantes (rellenar con vacío)
            for col in columnas_destino:
                if col not in df.columns:
                    df[col] = None
                    
            # 5. Retornar ordenado
            return df[columnas_destino]

        # ---------------------------------------------------------
        # 4. PROCESAMIENTO DE CADA DATAFRAME
        # ---------------------------------------------------------
        
        # Hoja 1: NC Encontrados-NoExitosos MES (Filtra Mes Anterior)
        df_s1 = procesar_hoja(df_nc_encontrados_sql, cols_nc_encontrados, filtrar_mes=True, tipo='NC')
        
        # Hoja 2: NC Con Novedad Vigentes (NO Filtra Mes, Histórico completo)
        df_s2 = procesar_hoja(df_nc_novedad_sql, cols_nc_novedad, filtrar_mes=False, tipo='NC')
        
        # Hoja 3: ND Total Mes (Filtra Mes Anterior)
        df_s3 = procesar_hoja(df_nd_sql, cols_nd_total, filtrar_mes=True, tipo='ND')

        # ---------------------------------------------------------
        # 5. GENERACIÓN DEL EXCEL
        # ---------------------------------------------------------
        # 1. Obtener el momento actual
        ahora = datetime.now()

        # 2. Formatear la fecha como string
        # %d = día, %m = mes, %Y = año (4 dígitos), %H = hora (24h), %M = minuto
        str_fecha_hora = ahora.strftime('%Y%m')

        # 3. Construir el nombre completo
        nombre_archivo = f"Consolidado_NC_ND_CXP_{str_fecha_hora}.xlsx"
        
        ruta_completa = os.path.join(rutabase, nombre_archivo)
        
        sheets_config = {
            'NC Encontrados-NoExitosos MES': df_s1,
            'NC Con Novedad Vigentes': df_s2,
            'ND Total Mes': df_s3
        }

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            workbook = writer.book
            
            # Estilos
            header_format = workbook.add_format({
                'bold': True, 'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
                'fg_color': '#1F4E78', 'font_color': '#FFFFFF', 'border': 1
            })
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd', 'border': 1})
            money_format = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
            text_format = workbook.add_format({'border': 1})

            for sheet_name, df_data in sheets_config.items():
                # Escribir datos
                start_row = 1
                df_data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
                
                worksheet = writer.sheets[sheet_name]
                (max_row, max_col) = df_data.shape
                
                # Crear Tabla
                if max_row > 0:
                    column_settings = [{'header': col} for col in df_data.columns]
                    worksheet.add_table(0, 0, max_row, max_col - 1, {
                        'columns': column_settings,
                        'style': 'TableStyleMedium2',
                        'name': f'T_{sheet_name.split()[0]}_{sheet_name.split()[-1]}'
                    })
                else:
                    for col_num, value in enumerate(df_data.columns):
                        worksheet.write(0, col_num, value, header_format)

                # Formatos de columnas
                for col_num, value in enumerate(df_data.columns):
                    worksheet.write(0, col_num, value, header_format)

                for i, col in enumerate(df_data.columns):
                    # Ancho automático
                    len_contenido = df_data[col].astype(str).map(len).max() if not df_data.empty else 0
                    col_len = max(len_contenido, len(col)) + 2
                    col_len = min(col_len, 50)
                    
                    # Aplicar formatos
                    if 'Fecha' in col:
                        worksheet.set_column(i, i, 14, date_format)
                    elif 'Amount' in col or 'Valor' in col:
                        worksheet.set_column(i, i, 18, money_format)
                    else:
                        worksheet.set_column(i, i, col_len, text_format)

        print(f"✅ Reporte generado: {ruta_completa}")
    
    def generar_reporte_anual_global(df_facturas_sql, df_nc_sql, df_nd_sql, rutabase):
        """
        Genera el reporte Consolidado Global CXP filtrando por el ANIO ANTERIOR.
        """

        # =========================================================================
        # 1. CALCULO DEL ANIO ANTERIOR
        # =========================================================================
        hoy = datetime.now()
        anio_anterior = hoy.year - 1

        print(f"Generando reporte anual para el anio: {anio_anterior}")

        # =========================================================================
        # 2. DEFINICION DE COLUMNAS POR HOJA
        # =========================================================================

        cols_facturas = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'NIT',
            'Nombre Proveedor',
            'Orden de Compra',
            'Factura',
            'Nota Credito',
            'Tipo de nota credito',
            'Nota Debito',
            'Tipo de nota debito',
            'Estado validacion antes de eventos',
            'Fecha - hora Evento Acuse de Recibo',
            'Estado Evento Acuse de Recibo',
            'Fecha - hora Evento Recibo del bien y/o prestacion del servicio',
            'Estado Evento Recibo del bien y/o prestacion del servicio',
            'Fecha - hora Evento Aceptacion Expresa',
            'Estado Evento Aceptacion Expresa',
            'Fecha - hora Evento Reclamo de la Factura Electronica de Venta',
            'Estado Evento Reclamo',
            'Estado contabilizacion',
            'Estado compensacion',
            'Observaciones'
        ]

        cols_nc = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'NIT',
            'Nombre Proveedor',
            'Nota Credito',
            'Tipo de nota credito',
            'Factura',
            'Estado validacion antes de eventos',
            'Observaciones'
        ]

        cols_nd = [
            'Fecha de ejecucion',
            'ID ejecucion',
            'ID Registro',
            'Tipo de documento',
            'NIT',
            'Nombre Proveedor',
            'Factura',
            'Nota Debito',
            'Tipo de nota debito',
            'Estado validacion antes de eventos',
            'Observaciones'
        ]

        # =========================================================================
        # 3. MAPEOS DE COLUMNAS
        # =========================================================================

        base_mapping = {
            'executionDate': 'Fecha de ejecucion',
            'executionNum': 'ID ejecucion',
            'ID': 'ID Registro',
            'documenttype': 'Tipo de documento',
            'nit_emisor_o_nit_del_proveedor': 'NIT',
            'nombre_emisor': 'Nombre Proveedor',
            'numero_de_liquidacion_u_orden_de_compra': 'Orden de Compra',
            'numero_de_factura': 'Factura',
            'ResultadoFinalAntesEventos': 'Estado validacion antes de eventos',
            'ObservacionesFase_4': 'Observaciones',
            'Estado_contabilizacion': 'Estado contabilizacion',
            'EstadoCompensacionFase_7': 'Estado compensacion',
            'FechaHora_Evento_030': 'Fecha - hora Evento Acuse de Recibo',
            'Estado_Evento_030': 'Estado Evento Acuse de Recibo',
            'FechaHora_Evento_032': 'Fecha - hora Evento Recibo del bien y/o prestacion del servicio',
            'Estado_Evento_032': 'Estado Evento Recibo del bien y/o prestacion del servicio',
            'FechaHora_Evento_033': 'Fecha - hora Evento Aceptacion Expresa',
            'Estado_Evento_033': 'Estado Evento Aceptacion Expresa',
            'FechaHora_Evento_031': 'Fecha - hora Evento Reclamo de la Factura Electronica de Venta',
            'Estado_Evento_031': 'Estado Evento Reclamo'
        }

        # =========================================================================
        # 4. FUNCION DE PROCESAMIENTO
        # =========================================================================

        def procesar_hoja(df_input, target_columns, mapping_adicional=None):
            if df_input is None or df_input.empty:
                return pd.DataFrame(columns=target_columns)

            df = df_input.copy()

            if 'executionDate' in df.columns:
                df['executionDate'] = pd.to_datetime(df['executionDate'], errors='coerce')
                df = df[df['executionDate'].dt.year == anio_anterior].copy()

            evento_fecha_cols = [
                'FechaHora_Evento_030',
                'FechaHora_Evento_031',
                'FechaHora_Evento_032',
                'FechaHora_Evento_033'
            ]
            for col in evento_fecha_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    df[col] = df[col].apply(
                        lambda x: x.strftime('%d/%m/%Y - %H:%M') if pd.notna(x) else ''
                    )

            mapping_completo = base_mapping.copy()
            if mapping_adicional:
                mapping_completo.update(mapping_adicional)

            cols_a_renombrar = {k: v for k, v in mapping_completo.items() if k in df.columns}
            df = df.rename(columns=cols_a_renombrar)

            for col in target_columns:
                if col not in df.columns:
                    df[col] = None

            return df[target_columns]

        # =========================================================================
        # 5. PROCESAMIENTO DE HOJAS
        # =========================================================================

        df_s1 = procesar_hoja(df_facturas_sql, cols_facturas, {
            'Numero_de_nota_credito': 'Nota Credito',
            'Tipo_de_nota_cred_deb': 'Tipo de nota credito'
        })

        df_s2 = procesar_hoja(df_nc_sql, cols_nc, {
            'Numero_de_nota_credito': 'Nota Credito',
            'Tipo_de_nota_cred_deb': 'Tipo de nota credito'
        })

        df_s3 = procesar_hoja(df_nd_sql, cols_nd, {
            'Numero_de_nota_credito': 'Nota Debito',
            'Tipo_de_nota_cred_deb': 'Tipo de nota debito'
        })

        # =========================================================================
        # 6. GENERACION ARCHIVO EXCEL
        # =========================================================================

        nombre_archivo = f"Consolidado_Global_CXP_{datetime.now().strftime('%Y')}.xlsx"
        ruta_completa = os.path.join(rutabase, nombre_archivo)

        sheets_config = {
            'Total Anual Facturas': df_s1,
            'Total Anual Notas Credito': df_s2,
            'Total Anual Notas Debito': df_s3
        }

        with pd.ExcelWriter(ruta_completa, engine='xlsxwriter') as writer:
            for sheet_name, df_data in sheets_config.items():
                df_data.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"Reporte generado exitosamente: {ruta_completa}")
        return ruta_completa

    # =========================================================================
    # PROCESAMIENTO PRINCIPAL
    # =========================================================================
    
    try:
        print("")
        print("=" * 80)
        print("[INICIO] HU8 - Generación de Reportes CxP")
        print("=" * 80)
        
        t_inicio = time.time()
        
        # 1. Obtener y validar configuración
        cfg = parse_config(GetVar("vLocDicConfig"))
        print("[INFO] Configuracion cargada exitosamente")
        
        # Parámetros de configuración
        ruta_base = cfg.get('RutaFileServer', r'\\172.16.250.222\BOT_Validacion_FV_NC_ND_CXP')
        #numero_ejecucion = int(cfg.get('NumeroEjecucion', 1))
        
        fecha_ejecucion = datetime.now()
        
        # 4. Conectar a BD y procesar
        with crear_conexion_db(cfg) as cx:
        
            print(f"[INFO] Ruta base: {ruta_base}")
            #print(f"[INFO] Numero ejecucion: {numero_ejecucion}")
            print(f"[INFO] Fecha ejecucion: {fecha_ejecucion}")
            
            query_num_ejecucion = """
                SELECT MAX(CAST([executionNum] AS INT)) as max_val 
                FROM [CxP].[DocumentsProcessing]
            """
            
            # Ejecutas el query
            df_resultado = pd.read_sql(query_num_ejecucion, cx)

            # Extraes el valor escalar
            ultimo_num = df_resultado['max_val'].iloc[0]
            
            # 2. Verificar acceso a ruta
            print("\n[PASO 1] Verificando acceso a File Server...")
            if not verificar_acceso_ruta(ruta_base):
                raise Exception(f"No se tiene acceso a la ruta: {ruta_base}")
            print("[OK] Acceso verificado")
            
            # 3. Crear árbol de carpetas
            print("\n[PASO 2] Creando/verificando arbol de carpetas...")
            rutas = crear_arbol_carpetas(ruta_base, fecha_ejecucion, ultimo_num)
            print("[OK] Arbol de carpetas listo")
        
            # 5. Procesar archivos de cada registro
            print("\n[PASO 3] Procesando archivos de registros...")
            
            query_registros_insumos = """
                SELECT *
                FROM [CxP].[DocumentsProcessing]
            """
            
            df_registros_insumos = pd.read_sql(query_registros_insumos, cx)
            print(f"[INFO] {len(df_registros_insumos)} registros para procesar archivos")
            
            archivos_procesados = 0
            
            for idx, reg in df_registros_insumos.iterrows():
                try:
                    registro_id = reg['ID']
                    tipo_doc = safe_str(reg['documenttype'])
                    nombre_archivos = safe_str(reg['actualizacionNombreArchivos'])
                    ruta_respaldo = safe_str(reg['RutaArchivo'])
                    resultado_final = safe_str(reg['ResultadoFinalAntesEventos'])
                    numero_oc = safe_str(reg['numero_de_liquidacion_u_orden_de_compra'])
                    nit = safe_str(reg['nit_emisor_o_nit_del_proveedor'])
                    factura = safe_str(reg['numero_de_factura'])
                    estado_xml = str(reg.get('Insumo_XML', '')).strip().lower()
                    estado_pdf = str(reg.get('Insumo_PDF', '')).strip().lower()
                    insumo_ubicado = str(reg.get('Insumo_reubicado', '')).strip().lower()
                    
                    if estado_xml in ['', 'none', 'nan'] or estado_pdf in ['', 'none', 'nan']:
                        # Verificar archivos
                        xml_enc, pdf_enc, ruta_xml, ruta_pdf = verificar_archivos_insumo(ruta_respaldo, nombre_archivos)
                        
                        # Actualizar BD con estado de insumos
                        cur = cx.cursor()
                        cur.execute("""
                            UPDATE [CxP].[DocumentsProcessing]
                            SET [Insumo_XML] = ?,
                                [Insumo_PDF] = ?
                            WHERE [ID] = ?
                        """, (
                            'ENCONTRADO' if xml_enc else 'NO ENCONTRADO',
                            'ENCONTRADO' if pdf_enc else 'NO ENCONTRADO',
                            registro_id
                        ))
                        cx.commit()
                        cur.close()
                        
                        actualizar_insumos_comparativa(reg, cx, nit, factura, 'InsumoPDF', 'ENCONTRADO' if pdf_enc else 'NO ENCONTRADO')
                        actualizar_insumos_comparativa(reg, cx, nit, factura, 'InsumoPDF', 'ENCONTRADO' if xml_enc else 'NO ENCONTRADO')
                        
                        if xml_enc or pdf_enc:
                            carpeta_destino = determinar_carpeta_destino(resultado_final, tipo_doc)
                            ruta_destino_completa = os.path.join(rutas['insumos_cxp'], carpeta_destino)
                            
                            # Determinar si copiar a comercializados
                            ruta_comercializados = None
                            if numero_oc.startswith('50'):
                                ruta_comercializadosHU4 = cfg['HU4RutaInsumos']
                            
                            if pdf_enc and not xml_enc:
                                nueva_ruta = mover_archivos_a_destino(
                                    ruta_pdf, ruta_destino_completa,
                                    numero_oc, ruta_comercializadosHU4
                                )
                            else:
                                nueva_ruta = mover_archivos_a_destino(
                                    ruta_pdf, ruta_destino_completa,
                                    numero_oc, ruta_comercializadosHU4
                                )
                            
                            if nueva_ruta:
                                cur = cx.cursor()
                                cur.execute("""
                                    UPDATE [CxP].[DocumentsProcessing]
                                    SET [Ruta_respaldo] = ?
                                    WHERE [ID] = ?
                                """, (nueva_ruta, registro_id))
                                cx.commit()
                                cur.close()
                            
                            archivos_procesados += 1
                    
                except Exception as e:
                    print("")
                    print("=" * 80)
                    print("[ERROR CRITICO] La funcion HU8_GenerarReportesCxP fallo")
                    print("=" * 80)
                    print(f"[ERROR] Mensaje: {str(e)}")
                    print(traceback.format_exc())
                    print("=" * 80)
                    
                    SetVar("vGblStrDetalleError", str(traceback.format_exc()))
                    SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                    SetVar("vLocStrResultadoSP", "False")
                    raise(e)
            
            
            #REPORTES DIARIOS
            try:
                print(f"[OK] {archivos_procesados} archivos procesados")
                
                
                with crear_conexion_db(cfg) as cx:
                    #CONSULTAS
                    # 2. Cargar Primera Tabla
                    QUERY_MAIN = """
                                SELECT 
                                    [executionDate], 
                                    [Fecha_de_retoma_antes_de_contabilizacion], 
                                    [executionNum], 
                                    [ID], 
                                    [documenttype], 
                                    [numero_de_liquidacion_u_orden_de_compra], 
                                    [nit_emisor_o_nit_del_proveedor],
                                    [nombre_emisor], 
                                    [numero_de_factura], 
                                    [ResultadoFinalAntesEventos], 
                                    [Fecha_retoma_contabilizacion], 
                                    [Estado_contabilizacion], 
                                    [FechaRetomaCompensacion_Fase7],
                                    [EstadoCompensacionFase_7]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]
                                """
                    df_main = pd.read_sql(QUERY_MAIN, cx)
                    
                    # 3. Cargar las otras tablas
                    query2 = "SELECT * FROM [NotificationsPaddy].[dbo].[CxP.Comparativa]"
                    df_detalles = pd.read_sql(query2, cx)

                    query3 = "SELECT * FROM [NotificationsPaddy].[CxP].[HistoricoOrdenesCompra]"
                    df_historico = pd.read_sql(query3, cx)
                

                    # 4. Generar reporte
                    generar_reporte_cxp(df_main, df_detalles, df_historico, rutas['resultados_dia'])
                    
                    #REPORTE GRANOS
                    
                    query1 = """SELECT 
                            [executionDate]
                            ,[Fecha_de_retoma_antes_de_contabilizacion]
                            ,[executionNum]
                            ,[ID]
                            ,[documenttype]
                            ,[numero_de_liquidacion_u_orden_de_compra]
                            ,[nit_emisor_o_nit_del_proveedor]
                            ,[nombre_emisor]
                            ,[numero_de_factura]
                            ,[ResultadoFinalAntesEventos]
                            ,[Fecha_retoma_contabilizacion]
                            ,[Estado_contabilizacion]
                        FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]
                        WHERE [agrupacion] LIKE '%MAPG%'"""
                    
                    df_main = pd.read_sql(query1, cx)
                    
                    generar_reporte_granos(df_main, df_detalles, rutas['granos_resultado'])
                    
                    
                    
                    # REPORTE MAIZ
                    query1 = """SELECT 
                            [executionDate]
                            ,[Fecha_de_retoma_antes_de_contabilizacion]
                            ,[executionNum]
                            ,[ID]
                            ,[documenttype]
                            ,[numero_de_liquidacion_u_orden_de_compra]
                            ,[nit_emisor_o_nit_del_proveedor]
                            ,[nombre_emisor]
                            ,[numero_de_factura]
                            ,[ResultadoFinalAntesEventos]
                            ,[Fecha_retoma_contabilizacion]
                            ,[Estado_contabilizacion]
                        FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]
                        WHERE [agrupacion] LIKE '%MAPM%'"""
                    
                    df_main = pd.read_sql(query1, cx)
                    
                    generar_reporte_maiz(df_main, df_detalles, rutas['granos_resultado'])
            
                
            
                    #REPORTE COMERCIALIZADOS:
                    query1 = """SELECT 
                            [executionDate]
                            ,[Fecha_de_retoma_antes_de_contabilizacion]
                            ,[executionNum]
                            ,[ID]
                            ,[documenttype]
                            ,[numero_de_liquidacion_u_orden_de_compra]
                            ,[nit_emisor_o_nit_del_proveedor]
                            ,[nombre_emisor]
                            ,[numero_de_factura]
                            ,[ResultadoFinalAntesEventos]
                        FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]
                        WHERE [numero_de_liquidacion_u_orden_de_compra] LIKE '50%'"""
                        
                    df_main = pd.read_sql(query1, cx)
                    
                    generar_reporte_comercializados(df_main, df_detalles, rutas['comercializados_resultado'])
                    
                    # Obtener el día actual
                    hoy = datetime.now()

                    if hoy.day == int(safe_str(cfg['DiaReporteMensualAnual'])):
                        
                        print("✅ Hoy es el dia seleccionado para ejecutar los reportes mensuales")
                        
                        query1 = """SELECT [Fecha_ejecucion]
                                    ,[Fecha_de_retoma]
                                    ,[ID_ejecucion]
                                    ,[ID_registro]
                                    ,[Nit]
                                    ,[Nombre_Proveedor]
                                    ,[Orden_de_compra]
                                    ,[Factura]
                                    ,[Fec_Doc]
                                    ,[Fec_Reg]
                                    ,[Observaciones]
                                FROM [NotificationsPaddy].[CxP].[HistoricoNovedades] ORDER BY Factura"""
                                
                        df_historico_novedades = pd.read_sql(query1, cx)
                        
                        query2 = """SELECT 
                                    [executionDate]
                                    ,[Fecha_de_retoma_antes_de_contabilizacion]
                                    ,[executionNum]
                                    ,[ID]
                                    ,[documenttype]
                                    ,[nit_emisor_o_nit_del_proveedor]
                                    ,[nombre_emisor]
                                    ,[numero_de_liquidacion_u_orden_de_compra]
                                    ,[numero_de_factura]
                                    ,[ObservacionesFase_4]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE ResultadoFinalAntesEventos LIKE '%CON NOVEDAD%'
                                """
                        
                        df_docs_processing = pd.read_sql(query2, cx)
                        
                        generar_consolidado_novedades(df_historico_novedades, df_docs_processing, df_historico, rutas['consolidados'])
                        
                        query1 = """SELECT 
                                    [executionDate]
                                    ,[Fecha_de_retoma_antes_de_contabilizacion]
                                    ,[executionNum]
                                    ,[ID]
                                    ,[documenttype]
                                    ,[nit_emisor_o_nit_del_proveedor]
                                    ,[nombre_emisor]
                                    ,[numero_de_liquidacion_u_orden_de_compra]
                                    ,[numero_de_factura]
                                    ,[ObservacionesFase_4]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE ResultadoFinalAntesEventos LIKE '%RECHAZADO%'
                                """
                        df_rechazados_sql = pd.read_sql(query1, cx)
                        
                        query2 = """SELECT 
                                    [executionDate]
                                    ,[Fecha_de_retoma_antes_de_contabilizacion]
                                    ,[executionNum]
                                    ,[ID]
                                    ,[documenttype]
                                    ,[nit_emisor_o_nit_del_proveedor]
                                    ,[nombre_emisor]
                                    ,[numero_de_liquidacion_u_orden_de_compra]
                                    ,[numero_de_factura]
                                    ,[ObservacionesFase_4]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE ResultadoFinalAntesEventos LIKE '%NO EXITOSO%'
                                """
                        
                        df_no_exitosos_sql = pd.read_sql(query2, cx)
                        
                        generar_consolidado_no_exitosos_rechazados(df_no_exitosos_sql, df_rechazados_sql, rutas['consolidados'])
                        
                        QUERY_EVENTOS = """
                                        SELECT 
                                            [executionDate],
                                            [Fecha_de_retoma_antes_de_contabilizacion],
                                            [executionNum],
                                            [ID],
                                            [documenttype],
                                            [nit_emisor_o_nit_del_proveedor],
                                            [nombre_emisor],
                                            [numero_de_liquidacion_u_orden_de_compra],
                                            [numero_de_factura],
                                            [ResultadoFinalAntesEventos],
                                            [FechaHora_Evento_030],
                                            [Estado_Evento_030],
                                            [FechaHora_Evento_032],
                                            [Estado_Evento_032],
                                            [FechaHora_Evento_033],
                                            [Estado_Evento_033],
                                            [FechaHora_Evento_031],
                                            [Estado_Evento_031],
                                            [ObservacionesFase_4]
                                        FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] 
                                        WHERE ResultadoFinalAntesEventos LIKE '%PENDIENTE%'
                                        """
                        
                        df_eventos_sql = pd.read_sql(QUERY_EVENTOS, cx)
                        
                        QUERY_CONTABILIZACION = """
                                                SELECT 
                                                    [executionDate],
                                                    [Fecha_de_retoma_antes_de_contabilizacion],
                                                    [executionNum],
                                                    [ID],
                                                    [documenttype],
                                                    [nit_emisor_o_nit_del_proveedor],
                                                    [nombre_emisor],
                                                    [numero_de_liquidacion_u_orden_de_compra],
                                                    [numero_de_factura],
                                                    [ResultadoFinalAntesEventos],
                                                    [ObservacionesFase_4]
                                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] 
                                                WHERE EstadoCompensacionFase_7 LIKE '%CONTABILIZACION PENDIENTE%'
                                                """
                                
                        df_contabilizacion_sql = pd.read_sql(QUERY_CONTABILIZACION, cx)
                        
                        QUERY_COMPENSACION = """
                                            SELECT 
                                                [executionDate],
                                                [Fecha_de_retoma_antes_de_contabilizacion],
                                                [executionNum],
                                                [ID],
                                                [documenttype],
                                                [nit_emisor_o_nit_del_proveedor],
                                                [nombre_emisor],
                                                [numero_de_liquidacion_u_orden_de_compra],
                                                [numero_de_factura],
                                                [ObservacionesFase_4]
                                            FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] 
                                            WHERE EstadoCompensacionFase_7 LIKE '%COMPENSACION PENDIENTE%'
                                            """  
                        df_compensacion_sql = pd.read_sql(QUERY_COMPENSACION, cx)
                                
                        generar_consolidado_pendientes(df_eventos_sql, df_compensacion_sql, df_contabilizacion_sql, rutas['consolidados'])
                        
                        query1 = """SELECT 
                                [executionDate]
                                ,[executionNum]
                                ,[ID]
                                ,[nit_emisor_o_nit_del_proveedor]
                                ,[nombre_emisor]
                                ,[Numero_de_nota_credito]
                                ,[Tipo_de_nota_cred_deb]
                                ,[NotaCreditoReferenciada]
                                ,[valor_a_pagar]
                                ,[ResultadoFinalAntesEventos]
                                ,[ObservacionesFase_4]
                            FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]  WHERE documenttype = 'NC' 
                            AND (ResultadoFinalAntesEventos LIKE '%ENCONTRADOS%' OR ResultadoFinalAntesEventos LIKE '%NO EXITOSOS%') """
                        
                        df_nc_encontrados_sql = pd.read_sql(query1, cx)
                        
                        quer2 = """SELECT 
                                    [executionDate]
                                    ,[executionNum]
                                    ,[ID]
                                    ,[nit_emisor_o_nit_del_proveedor]
                                    ,[nombre_emisor]
                                    ,[Numero_de_nota_credito]
                                    ,[Tipo_de_nota_cred_deb]
                                    ,[NotaCreditoReferenciada]
                                    ,[valor_a_pagar]
                                    ,[ResultadoFinalAntesEventos]
                                    ,[ObservacionesFase_4]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE documenttype = 'NC'
                                AND ResultadoFinalAntesEventos LIKE '%CON NOVEDAD%' 
                                """
                        
                        df_nc_novedad_sql = pd.read_sql(query2, cx)
                        
                        query3 = """SELECT 
                                    [executionDate]
                                    ,[executionNum]
                                    ,[ID]
                                    ,[nit_emisor_o_nit_del_proveedor]
                                    ,[nombre_emisor]
                                    ,[Numero_de_nota_credito]
                                    ,[Tipo_de_nota_cred_deb]
                                    ,[NotaCreditoReferenciada]
                                    ,[valor_a_pagar]
                                    ,[ObservacionesFase_4]
                                FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE documenttype = 'ND'"""
                        
                        df_nd_sql = pd.read_sql(query3, cx)
                        
                        generar_consolidado_nc_nd_actualizado(df_nc_encontrados_sql, df_nc_novedad_sql, df_nd_sql, rutas['consolidados'])
                        
                        hoy = datetime.now()

                        if hoy.month == int(safe_str(cfg['MesReporteAnual'])):
                            
                            query1 = """SELECT 
                                        [executionDate],
                                        [executionNum],
                                        [ID],
                                        [documenttype],
                                        [nit_emisor_o_nit_del_proveedor],
                                        [nombre_emisor],
                                        [numero_de_liquidacion_u_orden_de_compra],
                                        [numero_de_factura],
                                        [Numero_de_nota_credito],
                                        [Tipo_de_nota_cred_deb],
                                        [ResultadoFinalAntesEventos],
                                        [FechaHora_Evento_030],   
                                        [Estado_Evento_030],         
                                        [FechaHora_Evento_032],      
                                        [Estado_Evento_032],          
                                        [FechaHora_Evento_033],      
                                        [Estado_Evento_033],          
                                        [FechaHora_Evento_031],      
                                        [Estado_Evento_031],          
                                        [Estado_contabilizacion],
                                        [EstadoCompensacionFase_7],
                                        [ObservacionesFase_4]
                                    FROM [NotificationsPaddy].[CxP].[DocumentsProcessing]
                                    WHERE documenttype = 'FV'"""
                            
                            df_facturas_sql = pd.read_sql(query1, cx)
                            
                            query2 = """SELECT 
                                        [executionDate]
                                        ,[executionNum]
                                        ,[ID]
                                        ,[documenttype]
                                        ,[nit_emisor_o_nit_del_proveedor]
                                        ,[nombre_emisor]
                                        ,[numero_de_factura]
                                        ,[Numero_de_nota_credito]
                                        ,[Tipo_de_nota_cred_deb]
                                        ,[ResultadoFinalAntesEventos]
                                        ,[ObservacionesFase_4]
                                    FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE documenttype = 'NC'"""
                            
                            df_nc_sql = pd.read_sql(query2, cx)
                            
                            query3 = """SELECT 
                                        [executionDate]
                                        ,[executionNum]
                                        ,[ID]
                                        ,[documenttype]
                                        ,[nit_emisor_o_nit_del_proveedor]
                                        ,[nombre_emisor]
                                        ,[numero_de_factura]
                                        ,[Numero_de_nota_credito]
                                        ,[Tipo_de_nota_cred_deb]
                                        ,[ResultadoFinalAntesEventos]
                                        ,[ObservacionesFase_4]
                                    FROM [NotificationsPaddy].[CxP].[DocumentsProcessing] WHERE documenttype = 'ND'"""
                                    
                            df_nd_sql = pd.read_sql(query3, cx)
                            
                            generar_reporte_anual_global(df_facturas_sql, df_nc_sql, df_nd_sql, rutas['global_anual'])
                        
                    else:
                        print(f"❌ Hoy es día {hoy.day}, no es el dia para generar los reportes mensuales")
                        
            except Exception as e:
                print("")
                print("=" * 80)
                print("[ERROR CRITICO] La funcion HU8_GenerarReportesCxP fallo")
                print("=" * 80)
                print(f"[ERROR] Mensaje: {str(e)}")
                print(traceback.format_exc())
                print("=" * 80)
                
                SetVar("vGblStrDetalleError", str(traceback.format_exc()))
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                raise(e)
                        
        
        # Fin del procesamiento
        tiempo_total = time.time() - t_inicio
        
        print("")
        print("=" * 80)
        print("[FIN] HU8 - Generación de Reportes CxP completado")
        print("=" * 80)
        print("[ESTADISTICAS]")
        print(f"  Archivos procesados: {archivos_procesados}")
        print(f"  Tiempo total: {round(tiempo_total, 2)}s")
        print("=" * 80)
        
        resumen = f"HU8 completada. reportes generados"
        
        SetVar("vLocStrResultadoSP", "True")
        SetVar("vLocStrResumenSP", resumen)
        
    except Exception as e:
        print("")
        print("=" * 80)
        print("[ERROR CRITICO] La funcion HU8_GenerarReportesCxP fallo")
        print("=" * 80)
        print(f"[ERROR] Mensaje: {str(e)}")
        print(traceback.format_exc())
        print("=" * 80)
        
        SetVar("vGblStrDetalleError", str(traceback.format_exc()))
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        raise(e)


# Mock para pruebas locales
if __name__ == "__main__":
    _mock_vars = {}
    def GetVar(name):
        return _mock_vars.get(name, "")
    def SetVar(name, value):
        _mock_vars[name] = value
        print(f"[SetVar] {name} = {value}")
    
    _mock_vars["vLocDicConfig"] = '''{
        "ServidorBaseDatos": "localhost\SQLEXPRESS",
        "NombreBaseDatos": "NotificationsPaddy",
        "RutaFileServer": "C:/Users/diego/Desktop/insumosprueba",
        "NumeroEjecucion": 1,
        "UsuarioBaseDatos":"aa",
        "ClaveBaseDatos":"aa",
    }'''
    _mock_vars["vGblStrUsuarioBaseDatos"] = "sa"
    _mock_vars["vGblStrClaveBaseDatos"] = "password"
    
    print("Ejecutando prueba local...")
    HU8_GenerarReportesCxP()
# IF NOT EXISTS (SELECT * FROM sys.objects WHERE object_id = OBJECT_ID(N'[CxP].[Comparativa_NC]') AND type in (N'U'))
# BEGIN
#     CREATE TABLE [CxP].[Comparativa_NC](
#         [Fecha_de_ejecucion] [datetime] NULL,
#         [Fecha_de_retoma] [datetime] NULL,
#         [ID_ejecucion] [varchar](50) NULL,
#         [ID_Registro] [varchar](50) NULL,
#         [NIT] [varchar](50) NULL,
#         [Nombre_Proveedor] [varchar](255) NULL,
#         [Nota_Credito] [varchar](100) NULL,
#         [Item] [varchar](100) NULL,
#         [Valor_XML] [nvarchar](4000) NULL, -- Requerimiento especifico: 4000 caracteres
#         [Valor_Factura] [varchar](255) NULL,
#         [Aprobado] [varchar](50) NULL,
#         [Estado] [varchar](255) NULL
#     )
# END


# IF NOT EXISTS (SELECT * FROM sys.objects WHERE object_id = OBJECT_ID(N'[CxP].[Comparativa_ND]') AND type in (N'U'))
# BEGIN
#     CREATE TABLE [CxP].[Comparativa_ND](
#         [Fecha_de_ejecucion] [datetime] NULL,
#         [ID_ejecucion] [varchar](50) NULL,
#         [ID_Registro] [varchar](50) NULL,
#         [NIT] [varchar](50) NULL,
#         [Nombre_Proveedor] [varchar](255) NULL,
#         [Nota_Debito] [varchar](100) NULL,
#         [Item] [varchar](100) NULL,
#         [Valor_XML] [nvarchar](4000) NULL, -- Requisito: 4000 caracteres
#         [Aprobado] [varchar](50) NULL,
#         [Estado] [varchar](255) NULL
#     )
# END

def HU42_ValidarNotasCreditoDebito():
    """
    Funcion principal para procesar las validaciones de Notas Credito (NC) y Notas Debito (ND).
    
    Esta funcion orquesta todo el flujo de negocio:
    1. Conecta a la base de datos SQL Server.
    2. Limpia y repuebla las tablas de comparativa [CxP].[Comparativa_NC] y [CxP].[Comparativa_ND].
    3. Itera sobre las NC pendientes, busca sus Facturas (FV) correspondientes y valida montos y referencias.
    4. Itera sobre las ND pendientes y valida datos tributarios.
    5. Genera reportes en Excel con las novedades encontradas.
    
    VERSION: 1.0
    """
    
    import json
    import ast
    import traceback
    import pyodbc
    import pandas as pd
    import numpy as np
    from datetime import datetime, timedelta
    from contextlib import contextmanager
    import time
    import warnings
    import re
    import os
    import unicodedata
    from dateutil.relativedelta import relativedelta
    import random
    from openpyxl import load_workbook, Workbook 

    # Ignoramos advertencias de pandas sobre SQLAlchemy para mantener la consola limpia
    warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy')

    # =========================================================================
    # FUNCIONES AUXILIARES BASICAS
    # =========================================================================

    def safe_str(v):
        """
        Convierte de manera segura cualquier entrada a un string limpio.

        Maneja bytes (decodificando latin-1), numeros (float/int) y Nones.

        Args:
            v (any): El valor a convertir.

        Returns:
            str: Cadena de texto limpia (strip) o cadena vacia si la entrada es nula.
        """
        if v is None: return ""
        if isinstance(v, str): return v.strip()
        if isinstance(v, bytes):
            try: return v.decode('latin-1', errors='replace').strip()
            except: return str(v).strip()
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (np.isnan(v) or pd.isna(v)): return ""
            return str(v)
        try: return str(v).strip()
        except: return ""
    
    def truncar_observacion(obs, max_len=3900):
        """
        Trunca un texto de observacion para evitar errores de desbordamiento en BD.

        Args:
            obs (str): El texto de la observacion.
            max_len (int): Longitud maxima permitida (Default: 3900).

        Returns:
            str: Texto truncado con '...' al final si excedia la longitud.
        """
        if not obs:
            return ""
        obs_str = safe_str(obs)
        if len(obs_str) > max_len:
            return obs_str[:max_len - 3] + "..."
        return obs_str
    
    def parse_config(raw):
        """
        Analiza la configuracion de entrada (JSON o Dict string).

        Args:
            raw (str|dict): La configuracion cruda proveniente de Rocketbot.

        Returns:
            dict: Diccionario de configuracion python valido.

        Raises:
            ValueError: Si la configuracion esta vacia o tiene formato invalido.
        """
        if isinstance(raw, dict):
            if not raw: raise ValueError("Config vacia (dict)")
            return raw
        text = safe_str(raw)
        if not text: raise ValueError("vLocDicConfig vacio")
        try:
            config = json.loads(text)
            if not config: raise ValueError("Config vacia (JSON)")
            return config
        except json.JSONDecodeError: pass
        try:
            config = ast.literal_eval(text)
            if not config: raise ValueError("Config vacia (literal)")
            return config
        except (ValueError, SyntaxError) as e:
            raise ValueError(f"Config invalida: {str(e)}")
    
    def normalizar_decimal(valor):
        """
        Normaliza un valor numerico o string monetario a float estandar.

        Maneja formatos como '1.200,50' o '1200.50'.

        Args:
            valor (str|float): Valor a normalizar.

        Returns:
            float: Valor numerico puro. Retorna 0.0 si falla la conversion.
        """
        if pd.isna(valor) or valor == '' or valor is None: return 0.0
        try: return float(str(valor).strip().replace(',', '.').replace(r'[^\d.\-]', ''))
        except: return 0.0

    def campo_vacio(valor):
        """
        Verifica si un campo se considera 'vacio' para la logica de negocio.

        Args:
            valor (any): Valor a verificar.

        Returns:
            bool: True si es vacio, None, 'null' o 'nan'.
        """
        return safe_str(valor) == "" or safe_str(valor).lower() in ('null', 'none', 'nan')
    
    def campo_con_valor(valor):
        """
        Inverso de campo_vacio. Verifica si el campo tiene informacion util.

        Args:
            valor (any): Valor a verificar.

        Returns:
            bool: True si tiene datos validos.
        """
        return not campo_vacio(valor)

    def quitar_tildes(texto):
        """
        Elimina acentos y diacriticos de un texto para estandarizacion.

        Args:
            texto (str): Texto con posibles tildes.

        Returns:
            str: Texto ASCII limpio (ej: 'Cancion' -> 'Cancion').
        """
        if not texto: return ""
        return ''.join([c for c in unicodedata.normalize('NFKD', texto) if not unicodedata.combining(c)])

    def calcular_dias_diferencia(fecha_inicio, fecha_fin):
        """
        Calcula los dias transcurridos entre dos fechas soportando multiples formatos.

        Args:
            fecha_inicio (str|datetime): Fecha inicial.
            fecha_fin (str|datetime): Fecha final.

        Returns:
            int: Cantidad de dias de diferencia. Retorna 0 si falla el parseo.
        """
        try:
            # Lista de formatos soportados (orden de prioridad)
            formatos = [
                '%Y-%m-%d %H:%M:%S',    # Formato con hora estandar
                '%Y-%m-%d',             # Formato fecha estandar
                '%d/%m/%Y',             # Formato latino
                '%d-%m-%Y'              # Variante con guiones
            ]

            def convertir_a_datetime(fecha):
                if isinstance(fecha, datetime):
                    return fecha
                
                if isinstance(fecha, str):
                    # Limpiamos milisegundos si existen para evitar errores de formato
                    if '.' in fecha:
                        fecha = fecha.split('.')[0]
                    
                    for fmt in formatos:
                        try:
                            return datetime.strptime(fecha, fmt)
                        except ValueError:
                            continue
                return None

            dt_inicio = convertir_a_datetime(fecha_inicio)
            dt_fin = convertir_a_datetime(fecha_fin)

            if dt_inicio and dt_fin:
                return (dt_fin - dt_inicio).days
                
            return 0

        except Exception as e:
            return 0

    # =========================================================================
    # FUNCIONES DE CONEXION A BASE DE DATOS
    # =========================================================================

    @contextmanager
    def crear_conexion_db(cfg, max_retries=3):
        """
        Context Manager para establecer conexion segura a SQL Server.

        Intenta conexion por autenticacion SQL primero, y luego por Windows Auth (Trusted).
        Maneja commits y rollbacks automaticos.

        Args:
            cfg (dict): Diccionario con credenciales (Servidor, BD, Usuario, Clave).
            max_retries (int): Intentos maximos por metodo de conexion.

        Yields:
            pyodbc.Connection: Objeto de conexion activo.
        
        Raises:
            Exception: Si no logra conectar tras todos los intentos.
        """
        required = ["ServidorBaseDatos", "NombreBaseDatos"]
        missing = [k for k in required if not cfg.get(k)]
        if missing: raise ValueError(f"Parametros faltantes: {', '.join(missing)}")

        usuario = cfg.get('UsuarioBaseDatos', '')
        contrasena = cfg.get('ClaveBaseDatos', '')
        
        # String para Auth SQL Estandar
        conn_str_auth = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={cfg['ServidorBaseDatos']};"
            f"DATABASE={cfg['NombreBaseDatos']};"
            f"UID={usuario};PWD={contrasena};autocommit=False;"
        )
        
        # String para Auth Windows (Trusted)
        conn_str_trusted = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            f"SERVER={cfg['ServidorBaseDatos']};"
            f"DATABASE={cfg['NombreBaseDatos']};"
            "Trusted_Connection=yes;autocommit=False;"
        )

        cx = None
        conectado = False
        excepcion_final = None

        print("[INFO] Iniciando protocolo de conexion a Base de Datos...")
        
        # Intento 1: SQL Auth
        for attempt in range(max_retries):
            try:
                cx = pyodbc.connect(conn_str_auth, timeout=30)
                cx.autocommit = False
                conectado = True
                print(f"[INFO] Conexion SQL (Auth) establecida exitosamente (intento {attempt + 1})")
                break
            except pyodbc.Error as e:
                excepcion_final = e
                if attempt < max_retries - 1: time.sleep(1)

        # Intento 2: Windows Auth (Fallback)
        if not conectado:
            print("[WARNING] Fallo Auth SQL. Intentando Trusted Connection...")
            for attempt in range(max_retries):
                try:
                    cx = pyodbc.connect(conn_str_trusted, timeout=30)
                    cx.autocommit = False
                    conectado = True
                    print(f"[INFO] Conexion SQL (Trusted) establecida exitosamente (intento {attempt + 1})")
                    break
                except pyodbc.Error as e:
                    excepcion_final = e
                    if attempt < max_retries - 1: time.sleep(1)

        if not conectado:
            print(f"[ERROR] No se pudo conectar a la BD tras varios intentos.")
            raise excepcion_final or Exception("No se pudo conectar a la base de datos")
        
        try:
            yield cx
            if cx: cx.commit()
        except Exception as e:
            if cx: cx.rollback()
            raise
        finally:
            if cx:
                try: cx.close()
                except: pass

    # =========================================================================
    # FUNCIONES DE ACTUALIZACION DE BASE DE DATOS
    # =========================================================================

    def actualizar_bd_cxp(cx, registro_id, campos_actualizar):
        """
        Actualiza campos especificos en la tabla principal [CxP].[DocumentsProcessing].

        Args:
            cx (pyodbc.Connection): Conexion activa.
            registro_id (str): ID del registro a actualizar.
            campos_actualizar (dict): Diccionario {columna: valor}.

        Raises:
            Exception: Si falla la ejecucion SQL.
        """
        try:
            sets = []
            parametros = []
            for campo, valor in campos_actualizar.items():
                if valor is not None:
                    # Logica especial para concatenar observaciones en lugar de sobrescribir
                    if campo == 'ObservacionesFase_4':
                        sets.append(f"[{campo}] = CASE WHEN [{campo}] IS NULL OR [{campo}] = '' THEN ? ELSE [{campo}] + ', ' + ? END")
                        parametros.extend([valor, valor])
                    else:
                        sets.append(f"[{campo}] = ?")
                        parametros.append(valor)

            if sets:
                parametros.append(registro_id)
                sql = f"UPDATE [CxP].[DocumentsProcessing] SET {', '.join(sets)} WHERE [ID] = ?"
                cur = cx.cursor()
                cur.execute(sql, parametros)
                cx.commit()
                cur.close()
        except Exception as e:
            print(f"[ERROR] Fallo al actualizar DocumentsProcessing ID {registro_id}: {str(e)}")
            raise

    def actualizar_items_comparativa_nc(registro_id, cx, nit, nombre_item,
                                    actualizar_valor_xml=True, valor_xml=None,
                                    actualizar_aprobado=True, valor_aprobado=None, 
                                    actualizar_valor_factura=True, valor_factura=None, 
                                    actualizar_estado=True, valor_estado=None, 
                                    # Datos de contexto para INSERT
                                    id_ejecucion=None, fecha_retoma=None, nombre_proveedor=None):
        """
        Actualiza o inserta items en la tabla [CxP].[Comparativa_NC].
        
        Utiliza logica de CTE para manejar multiples items identicos (duplicados por splits).
        
        Args:
            registro_id (str): ID del registro unico.
            cx (pyodbc.Connection): Conexion.
            nit (str): NIT del emisor.
            nombre_item (str): Nombre del campo validado (ej: 'NIT Emisor').
            actualizar_valor_xml (bool): Flag para actualizar el valor XML.
            valor_xml (str): Valor extraido del XML/NC.
            actualizar_aprobado (bool): Flag para actualizar estado aprobado.
            valor_aprobado (str): 'SI' o 'NO'.
            actualizar_valor_factura (bool): Flag para actualizar valor contrapartida.
            valor_factura (str): Valor de la Factura encontrada.
            actualizar_estado (bool): Flag para actualizar estado general.
            valor_estado (str): Estado del proceso (ej: 'EXITOSO').
            id_ejecucion (str): ID de ejecucion del bot (solo para inserts).
            fecha_retoma (datetime): Fecha de retoma (solo para inserts).
            nombre_proveedor (str): Nombre del proveedor (solo para inserts).

        Raises:
            Exception: Si falla la consulta SQL.
        """
        try:
            cur = cx.cursor()
            
            def safe_db_val(v):
                if v is None: return None
                s = str(v).strip()
                if not s or s.lower() == 'none' or s.lower() == 'null': return None
                return s

            # 1. Verificamos cuantos registros existen para decidir entre UPDATE o INSERT
            query_count = """
            SELECT COUNT(*)
            FROM [CxP].[Comparativa_NC]
            WHERE NIT = ? AND Item = ? AND ID_Registro = ?
            """
            cur.execute(query_count, (nit, nombre_item, registro_id))
            count_existentes = cur.fetchone()[0]

            # 2. Preparamos las listas por si vienen valores concatenados con pipe '|'
            lista_factura = str(valor_factura).split('|') if valor_factura else []
            lista_xml = str(valor_xml).split('|') if valor_xml else []
            lista_aprob = str(valor_aprobado).split('|') if valor_aprobado else []
            lista_estado = str(valor_estado).split('|') if valor_estado else []

            maximo_conteo = max(len(lista_factura), len(lista_xml), len(lista_aprob), len(lista_estado))
            maximo_conteo = 1 if maximo_conteo == 0 else maximo_conteo

            # 3. Iteramos para procesar cada sub-valor
            for i in range(maximo_conteo):
                item_factura = lista_factura[i] if i < len(lista_factura) else None
                item_xml = lista_xml[i] if i < len(lista_xml) else None
                item_aprob = lista_aprob[i] if i < len(lista_aprob) else None
                item_estado = lista_estado[i] if i < len(lista_estado) else None

                val_factura = safe_db_val(item_factura)
                val_xml = safe_db_val(item_xml)
                val_aprob = safe_db_val(item_aprob)
                val_estado = safe_db_val(item_estado)

                if i < count_existentes:
                    # LOGICA UPDATE
                    set_clauses = []
                    params = []

                    if actualizar_valor_factura:
                        set_clauses.append("Valor_Factura = ?")
                        params.append(val_factura)
                    if actualizar_valor_xml:
                        set_clauses.append("Valor_XML = ?")
                        params.append(val_xml)
                    if actualizar_aprobado:
                        set_clauses.append("Aprobado = ?")
                        params.append(val_aprob)
                    if actualizar_estado:
                        set_clauses.append("Estado = ?")
                        params.append(val_estado)

                    if not set_clauses: continue

                    # Usamos CTE con ; al inicio para compatibilidad SQL Server
                    update_query = f"""
                    ;WITH CTE AS (
                        SELECT Valor_Factura, Valor_XML, Aprobado, Estado,
                            ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) as rn
                        FROM [CxP].[Comparativa_NC]
                        WHERE NIT = ?  AND Item = ? AND ID_Registro = ?
                    )
                    UPDATE CTE
                    SET {", ".join(set_clauses)}
                    WHERE rn = ?
                    """
                    final_params = params + [nit, nombre_item, registro_id, i + 1]
                    cur.execute(update_query, final_params)

                else:
                    # LOGICA INSERT
                    insert_query = """
                    INSERT INTO [CxP].[Comparativa_NC] (
                        Fecha_de_ejecucion, Fecha_de_retoma, ID_ejecucion, ID_Registro, 
                        NIT, Nombre_Proveedor, Item, 
                        Valor_Factura, Valor_XML, Aprobado, Estado
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                    cur.execute(insert_query, (
                        datetime.now(), fecha_retoma, id_ejecucion, registro_id,
                        nit, nombre_proveedor, nombre_item,
                        val_factura, val_xml, val_aprob, val_estado or 'PENDIENTE'
                    ))
            
            cx.commit()

        except Exception as e:
            print(f"[ERROR] Fallo en actualizar_items_comparativa_nc para ID {registro_id}: {e}")
            cx.rollback()
            raise e
   
    def actualizar_items_comparativa_nd(registro_id, cx, nit, nombre_item,
                                        actualizar_valor_xml=True, valor_xml=None,
                                        actualizar_aprobado=True, valor_aprobado=None, 
                                        actualizar_estado=True, valor_estado=None,
                                        # Datos de contexto para INSERT
                                        id_ejecucion=None, nombre_proveedor=None):
        """
        Actualiza o inserta items en la tabla [CxP].[Comparativa_ND].
        
        Args:
            registro_id (str): ID del registro unico.
            cx (pyodbc.Connection): Conexion activa.
            nit (str): NIT del emisor.
            nombre_item (str): Nombre del item validado.
            actualizar_valor_xml (bool): Flag valor XML.
            valor_xml (str): Valor del XML/ND.
            actualizar_aprobado (bool): Flag aprobado.
            valor_aprobado (str): 'SI' o 'NO'.
            actualizar_estado (bool): Flag estado.
            valor_estado (str): Estado del proceso.
            id_ejecucion (str): ID de ejecucion del bot.
            nombre_proveedor (str): Nombre del proveedor.

        Raises:
            Exception: Si falla la transaccion SQL.
        """
        try:
            cur = cx.cursor()
            
            def safe_db_val(v):
                if v is None: return None
                s = str(v).strip()
                if not s or s.lower() == 'none' or s.lower() == 'null': return None
                return s

            query_count = """
            SELECT COUNT(*)
            FROM [CxP].[Comparativa_ND]
            WHERE NIT = ? AND Item = ? AND ID_Registro = ?
            """
            cur.execute(query_count, (nit, nombre_item, registro_id))
            count_existentes = cur.fetchone()[0]

            lista_xml = str(valor_xml).split('|') if valor_xml else []
            lista_aprob = str(valor_aprobado).split('|') if valor_aprobado else []
            lista_estado = str(valor_estado).split('|') if valor_estado else []

            maximo_conteo = max(len(lista_xml), len(lista_aprob), len(lista_estado))
            maximo_conteo = 1 if maximo_conteo == 0 else maximo_conteo

            for i in range(maximo_conteo):
                item_xml = lista_xml[i] if i < len(lista_xml) else None
                item_aprob = lista_aprob[i] if i < len(lista_aprob) else None
                item_estado = lista_estado[i] if i < len(lista_estado) else None

                val_xml = safe_db_val(item_xml)
                val_aprob = safe_db_val(item_aprob)
                val_estado = safe_db_val(item_estado)

                if i < count_existentes:
                    set_clauses = []
                    params = []

                    if actualizar_valor_xml:
                        set_clauses.append("Valor_XML = ?")
                        params.append(val_xml)
                    if actualizar_aprobado:
                        set_clauses.append("Aprobado = ?")
                        params.append(val_aprob)
                    if actualizar_estado:
                        set_clauses.append("Estado = ?")
                        params.append(val_estado)

                    if not set_clauses: continue

                    update_query = f"""
                    ;WITH CTE AS (
                        SELECT Valor_XML, Aprobado, Estado,
                            ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) as rn
                        FROM [CxP].[Comparativa_ND]
                        WHERE NIT = ? AND Item = ? AND ID_Registro = ?
                    )
                    UPDATE CTE
                    SET {", ".join(set_clauses)}
                    WHERE rn = ?
                    """
                    final_params = params + [nit, nombre_item, registro_id, i + 1]
                    cur.execute(update_query, final_params)

                else:
                    insert_query = """
                    INSERT INTO [CxP].[Comparativa_ND] (
                        Fecha_de_ejecucion, ID_ejecucion, ID_Registro, 
                        NIT, Nombre_Proveedor, Item, 
                        Valor_XML, Aprobado, Estado
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                    cur.execute(insert_query, (
                        datetime.now(), id_ejecucion, registro_id,
                        nit, nombre_proveedor, nombre_item,
                        val_xml, val_aprob, val_estado or 'PENDIENTE'
                    ))
            
            cx.commit()

        except Exception as e:
            print(f"[ERROR] Fallo en actualizar_items_comparativa_nd para ID {registro_id}: {e}")
            cx.rollback()
            raise e
    
    def actualizar_estado_comparativa_nc(cx, nit, registro_id, estado):
        """
        Actualiza la columna 'Estado' masivamente para todos los items de una NC especifica.

        Args:
            cx (pyodbc.Connection): Conexion activa.
            nit (str): NIT del emisor.
            registro_id (str): ID unico del registro.
            estado (str): Nuevo estado (ej: 'ENCONTRADO', 'NO EXITOSO').

        Raises:
            Exception: Si falla la actualizacion.
        """
        try:
            cur = cx.cursor()
            
            update_sql = """
            UPDATE [CxP].[Comparativa_NC]
            SET Estado = ?
            WHERE NIT = ? AND [ID_Registro] = ?
            """
            cur.execute(update_sql, (estado, nit, registro_id))
            cx.commit()
            cur.close()
            
        except Exception as e:
            print(f"[ERROR] Error actualizando estado global en Comparativa_NC: {e}")
            cx.rollback()
            raise e
    
    def limpiar_tablas_comparativas(cx):
        """
        Ejecuta TRUNCATE TABLE sobre las tablas comparativas para iniciar la ejecucion limpia.
        
        Args:
            cx (pyodbc.Connection): Conexion activa.

        Raises:
            Exception: Si no se pueden truncar las tablas.
        """
        try:
            cursor = cx.cursor()
            
            print("[INFO] Limpiando tabla [CxP].[Comparativa_NC]...")
            cursor.execute("TRUNCATE TABLE [CxP].[Comparativa_NC]")
            
            print("[INFO] Limpiando tabla [CxP].[Comparativa_ND]...")
            cursor.execute("TRUNCATE TABLE [CxP].[Comparativa_ND]")
            
            cx.commit()
            print("[SUCCESS] Tablas comparativas limpiadas exitosamente.")
            
        except Exception as e:
            print(f"[ERROR] Fallo critico al limpiar tablas: {e}")
            cx.rollback()
            raise e
    
    def poblar_inicial_comparativa_nc(cx, df):
        """
        Realiza una insercion masiva (Bulk Insert) inicial en Comparativa_NC.
        
        Genera 16 registros por cada fila del DataFrame de entrada (uno por cada item a validar).
        
        Args:
            cx (pyodbc.Connection): Conexion activa.
            df (pandas.DataFrame): DataFrame con las NC pendientes.

        Raises:
            Exception: Si falla la insercion masiva.
        """
        try:
            # Lista items estandarizada (SIN TILDES)
            items_a_validar = [
                "Nombre Emisor",
                "NIT Emisor",
                "Nombre Receptor",
                "Nit Receptor",
                "Tipo Persona Receptor",
                "DigitoVerificacion Receptor",
                "TaxLevelCode Receptor",
                "Fecha emision del documento",
                "LineExtensionAmount",
                "Tipo de nota credito",
                "Referencia",
                "Codigo CUFE de la factura",
                "Cude de la Nota Credito", 
                "ActualizacionNombreArchivos",
                "RutaRespaldo",
                "Observaciones"
            ]

            fecha_actual = datetime.now()
            datos_para_insertar = []

            # Iteramos el DF usando itertuples por eficiencia
            for row in df.itertuples(index=False):
                
                id_registro = getattr(row, 'ID', None) 
                id_ejecucion = getattr(row, 'executionNum', None) 
                nit = getattr(row, 'nit_emisor_o_nit_del_proveedor', None)
                nombre_proveedor = getattr(row, 'nombre_emisor', None) 
                nota_credito = getattr(row, 'Numero_de_nota_credito', None)
                fecha_retoma = getattr(row, 'Fecha_retoma_contabilizacion', None)
                estado = getattr(row, 'ResultadoFinalAntesEventos', None)

                # Multiplicamos la fila por cada item a validar
                for item in items_a_validar:
                    tupla_sql = (
                        fecha_actual,       # Fecha_de_ejecucion
                        fecha_retoma,       # Fecha_de_retoma
                        str(id_ejecucion),  # ID_ejecucion
                        str(id_registro),   # ID_Registro
                        str(nit),           # NIT
                        str(nombre_proveedor), # Nombre_Proveedor
                        str(nota_credito),  # Nota_Credito
                        item,               # Item (El que cambia)
                        '',                 # Valor_XML (Vacio al inicio)
                        '',                 # Valor_Factura (Vacio al inicio)
                        '',                 # Aprobado (Vacio al inicio)
                        estado              # Estado
                    )
                    datos_para_insertar.append(tupla_sql)

            if datos_para_insertar:
                cursor = cx.cursor()
                
                sql_insert = """
                INSERT INTO [CxP].[Comparativa_NC] (
                    [Fecha_de_ejecucion], [Fecha_de_retoma], [ID_ejecucion], [ID_Registro], 
                    [NIT], [Nombre_Proveedor], [Nota_Credito], [Item], 
                    [Valor_XML], [Valor_Factura], [Aprobado], [Estado]
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                
                print(f"[INFO] Insertando {len(datos_para_insertar)} registros masivos en Comparativa_NC...")
                cursor.executemany(sql_insert, datos_para_insertar)
                cx.commit()
                print("[SUCCESS] Poblado inicial NC completado con exito.")
                
            else:
                print("[WARNING] El DataFrame de NC estaba vacio, no se insertaron registros.")

        except Exception as e:
            print(f"[ERROR] Fallo al poblar tabla inicial NC: {e}")
            cx.rollback()
            raise e

    def poblar_inicial_comparativa_nd(cx, df):
        """
        Realiza una insercion masiva (Bulk Insert) inicial en Comparativa_ND.
        
        Args:
            cx (pyodbc.Connection): Conexion activa.
            df (pandas.DataFrame): DataFrame con las ND pendientes.

        Raises:
            Exception: Si falla la insercion.
        """
        try:
            # Lista items estandarizada ND (SIN TILDES)
            items_a_validar = [
                "Nombre Emisor",
                "NIT Emisor",
                "Nombre Receptor",
                "Nit Receptor",
                "Tipo Persona Receptor",
                "DigitoVerificacion Receptor",
                "TaxLevelCode Receptor",
                "Fecha emision del documento",
                "LineExtensionAmount",
                "Tipo de nota debito",
                "Referencia",
                "Codigo CUFE de la factura",
                "Cude de la Nota Debito",
                "ActualizacionNombreArchivos",
                "RutaRespaldo",
                "Observaciones"
            ]

            fecha_actual = datetime.now()
            datos_para_insertar = []

            for row in df.itertuples(index=False):
                
                executionNum = getattr(row, 'executionNum', None)
                id_registro = getattr(row, 'ID', None)
                nit = getattr(row, 'nit_emisor_o_nit_del_proveedor', None)
                nombre_proveedor = getattr(row, 'nombre_emisor', None) 
                
                # Ajuste critico: Busca el numero de nota debito, con fallback a capitalizacion alternativa
                nota_debito = getattr(row, 'numero_de_nota_debito', None)
                if not nota_debito: nota_debito = getattr(row, 'Numero_de_nota_debito', None)

                for item in items_a_validar:
                    tupla_sql = (
                        fecha_actual,       # Fecha_de_ejecucion
                        str(executionNum),  # ID_ejecucion
                        str(id_registro),   # ID_Registro
                        str(nit),           # NIT
                        str(nombre_proveedor), # Nombre_Proveedor
                        str(nota_debito),   # Nota_Debito
                        item,               # Item
                        '',                 # Valor_XML (Vacio inicial)
                        '',                 # Aprobado (Vacio inicial)
                        'PENDIENTE'         # Estado
                    )
                    datos_para_insertar.append(tupla_sql)

            if datos_para_insertar:
                cursor = cx.cursor()
                
                sql_insert = """
                INSERT INTO [CxP].[Comparativa_ND] (
                    [Fecha_de_ejecucion], [ID_ejecucion], [ID_Registro], 
                    [NIT], [Nombre_Proveedor], [Nota_Debito], [Item], 
                    [Valor_XML], [Aprobado], [Estado]
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                
                print(f"[INFO] Insertando {len(datos_para_insertar)} registros masivos en Comparativa_ND...")
                cursor.executemany(sql_insert, datos_para_insertar)
                cx.commit()
                print("[SUCCESS] Poblado inicial de ND completado con exito.")
                
            else:
                print("[WARNING] El DataFrame de ND estaba vacio, no se insertaron registros.")

        except Exception as e:
            print(f"[ERROR] Fallo al poblar tabla inicial ND: {e}")
            cx.rollback()
            raise e
    
    
    # =========================================================================
    # VALIDACIONES DE NEGOCIO HU4.2
    # =========================================================================

    def validar_nombre_receptor(nombre):
        """Valida que el receptor sea DIANA o DICORP."""
        if campo_vacio(nombre): return False
        nombre = re.sub(r'[,.\s]', '', quitar_tildes(safe_str(nombre).upper()))
        return nombre in ['DIANACORPORACIONSAS', 'DICORPSAS']

    def validar_nit_receptor(nit):
        """Valida que el NIT receptor sea el estandar 860031606."""
        return re.sub(r'\D', '', safe_str(nit)) == '860031606'

    def validar_tipo_persona(tipo):
        """Valida codigo tipo persona '31'."""
        return safe_str(tipo) == '31'

    def validar_digito_verificacion(digito):
        """Valida digito verificacion '6'."""
        return safe_str(digito) == '6'

    def validar_tax_level_code(tax_code):
        """Valida que el codigo de responsabilidad contenga alguno de los valores permitidos."""
        if campo_vacio(tax_code): return False
        tax = safe_str(tax_code).upper()
        return any(v in tax for v in ['O-13', 'O-15', 'O-23', 'O-47', 'R-99-PN'])
   
    def generar_reporte_retorno_nc(cx, ruta_base, nombre_reporte):
        """
        Genera o actualiza un archivo Excel con el reporte de registros 'CON NOVEDAD'.

        Busca en la BD las NC marcadas y las agrega a un archivo Excel en la ruta 
        correspondiente al mes actual. Crea carpetas y archivos si no existen.

        Args:
            cx (pyodbc.Connection): Conexion activa.
            ruta_base (str): Ruta raiz del servidor donde se guardan reportes.
            nombre_reporte (str): Prefijo del nombre del archivo.

        Raises:
            Exception: Si falla la lectura de BD o escritura de Excel.
        """
        try:
            fecha_actual = datetime.now()
            anio = fecha_actual.year
            fecha_solo = fecha_actual.date()
            
            # Meses sin tildes para nombres de carpetas
            meses = {
                1: "01. Enero", 2: "02. Febrero", 3: "03. Marzo", 4: "04. Abril",
                5: "05. Mayo", 6: "06. Junio", 7: "07. Julio", 8: "08. Agosto",
                9: "09. Septiembre", 10: "10. Octubre", 11: "11. Noviembre", 12: "12. Diciembre"
            }
            nombre_mes = meses[fecha_actual.month]
            
            ruta_carpeta = os.path.join(ruta_base, str(anio), nombre_mes, "INSUMO DE RETORNO")
            
            # Creacion recursiva de directorios
            if not os.path.exists(ruta_carpeta):
                print(f"[INFO] La carpeta no existe. Creando ruta: {ruta_carpeta}")
                os.makedirs(ruta_carpeta)

            ruta_completa = os.path.join(ruta_carpeta, f"{nombre_reporte}_{fecha_solo}.xlsx")

            print(f"[INFO] Gestionando reporte en: {ruta_completa}")

            query_novedades = """
            SELECT 
                executionNum as ID,
                nit_emisor_o_nit_del_proveedor as Nit,
                numero_de_factura as Numero_Nota_Credito,
                ResultadoFinalAntesEventos as Estado_CXP_Bot
            FROM [CxP].[DocumentsProcessing]
            WHERE 
                documenttype = 'NC' 
                AND ResultadoFinalAntesEventos = 'CON NOVEDAD'
            """
            
            df_novedades = pd.read_sql(query_novedades, cx)

            if not df_novedades.empty:
                print(f"[INFO] Se encontraron {len(df_novedades)} registros CON NOVEDAD para reportar.")
                
                wb = None
                ws = None
                encabezados = ["ID", "Fecha_Carga", "Nit", "Numero_Nota_Credito", "Estado_CXP_Bot"]

                # Logica para crear o abrir archivo
                if not os.path.exists(ruta_completa):
                    print("[INFO] El archivo no existe. Se creara uno nuevo.")
                    wb = Workbook()
                    ws = wb.active
                    ws.title = 'NC'
                    ws.append(encabezados) 
                else:
                    print("[INFO] El archivo existe. Se agregaran datos.")
                    wb = load_workbook(ruta_completa)
                    
                    if 'NC' not in wb.sheetnames:
                        ws = wb.create_sheet('NC')
                        ws.append(encabezados)
                    else:
                        ws = wb['NC']

                fecha_carga = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
                
                # Escribimos filas
                for index, row in df_novedades.iterrows():
                    fila_excel = [
                        row['ID'],
                        fecha_carga,
                        row['Nit'],
                        row['Numero_Nota_Credito'],
                        row['Estado_CXP_Bot']
                    ]
                    ws.append(fila_excel)

                wb.save(ruta_completa)
                wb.close()
                print("[SUCCESS] Reporte guardado exitosamente.")
                
            else:
                print("[INFO] No hay novedades para reportar en este ciclo.")

        except Exception as e:
            print(f"[ERROR] Error generando reporte de retorno: {e}")
            raise e
        
        
    # =========================================================================
    # LOGICA PRINCIPAL DE ORQUESTACION
    # =========================================================================

    try:
        print("")
        print("=" * 80)
        print("[INFO] INICIO HU4.2 - Validacion NC/ND")
        print("=" * 80)

        # 1. Parsing de Configuracion
        cfg = parse_config(GetVar("vLocDicConfig"))

        plazo_max = int(cfg.get('PlazoMaximoRetoma', 120))
        now = datetime.now()
        
        # 2. Conexion a Base de Datos (Context Manager)
        with crear_conexion_db(cfg) as cx:

            # -----------------------------------------------------------------
            # FASE 1: PROCESAMIENTO DE NOTAS CREDITO (NC)
            # -----------------------------------------------------------------
            # Buscamos NC pendientes que no esten ya finalizadas
            df_nc = pd.read_sql("SELECT * FROM [CxP].[DocumentsProcessing] WHERE [tipo_de_documento]='NC' AND ([ResultadoFinalAntesEventos] IS NULL OR [ResultadoFinalAntesEventos] NOT IN ('ENCONTRADO', 'NO EXITOSO'))", cx)
            print(f"[INFO] Procesando {len(df_nc)} Notas Credito (NC)...")
            
            # Limpieza previa de tablas
            try:
                limpiar_tablas_comparativas(cx)
            except Exception as e:
                error_msg = traceback.format_exc() 
                SetVar("vGblStrDetalleError", error_msg)
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                print(f"[ERROR] Critico al limpiar tablas: {error_msg}")
                raise e
                
            # Poblado inicial (Snapshot)
            try:
                poblar_inicial_comparativa_nc(cx, df_nc)
            except Exception as e:
                error_msg = traceback.format_exc() 
                SetVar("vGblStrDetalleError", error_msg)
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                print(f"[ERROR] Critico al poblar NC: {error_msg}")
                raise e
            
            # Carga de Facturas (FV) para cruce
            try:
                fecha_ejecucion = datetime.now()

                # Ventana de tiempo: Mes actual + Mes anterior completo
                fecha_inicio_filtro = (fecha_ejecucion.replace(day=1) - relativedelta(months=1)).strftime('%Y-%m-%d')
                fecha_fin_filtro = fecha_ejecucion.strftime('%Y-%m-%d')

                print(f"[INFO] Buscando documentos emitidos entre: {fecha_inicio_filtro} y {fecha_fin_filtro}")

                query_fv = f"""
                            SELECT *
                            FROM [CxP].[DocumentsProcessing]
                            WHERE 
                                documenttype = 'FV'
                                AND fecha_de_emision_documento >= '{fecha_inicio_filtro}'
                                AND fecha_de_emision_documento <= '{fecha_fin_filtro}'
                                AND (NotaCreditoReferenciada IS NULL OR NotaCreditoReferenciada = '')
                        """
                
                df_fv = pd.read_sql(query_fv, cx)
                
                # Filtramos FV candidatas (Exitosas o Rechazadas)
                condicion = df_fv['ResultadoFinalAntesEventos'].str.contains('EXITOSO|RECHAZADO', case=False, na=False, regex=True)
                df_fv_final = df_fv[condicion].copy()
                
            except Exception as e:
                error_msg = traceback.format_exc() 
                SetVar("vGblStrDetalleError", error_msg)
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                print(f"[ERROR] Fallo al cargar Facturas candidatas: {error_msg}")
                raise e
                
            cnt_nc = 0
            list_nov = []
            
            # Iteracion sobre cada Nota Credito
            for idx, r in df_nc.iterrows():
                retorno_manual = True if 'EXITOSO' in str(r.get('EstadoFase_3').upper()) else False
                reg_id = safe_str(r.get('ID', ''))
                nit = safe_str(r.get('nit_emisor_o_nit_del_proveedor', ''))
                num_nc = safe_str(r.get('numero_de_nota_credito', ''))
                oc = safe_str(r.get('numero_de_liquidacion_u_orden_de_compra', ''))
                num_factura = r.get('numero_de_factura') 
                obs_anterior = safe_str(r.get('ObservacionesFase_4')) 
                
                fecha_retoma = r.get('Fecha_retoma_contabilizacion')
                tipo_doc = r.get('tipo_de_documento')
                nombre_prov = r.get('nombre_emisor')

                print(f"[INFO] Analizando NC: {num_nc} (ID: {reg_id})")

                # Regla 1: Validacion de plazos de Retoma
                if not retorno_manual:
                    f_ret = r.get('Fecha_retoma_contabilizacion')
                    if campo_con_valor(f_ret):
                        dias = calcular_dias_diferencia(f_ret, now)
                        if dias > plazo_max:
                            obs = f"Registro excede el plazo maximo de retoma, {obs_anterior}"
                            # Marcamos como No Exitoso en BD y Comparativa
                            actualizar_bd_cxp(cx, reg_id, {
                                'EstadoFinalFase_4': 'VALIDACION DATOS DE FACTURACION: No exitoso',
                                'ObservacionesFase_4': truncar_observacion(obs),
                                'ResultadoFinalAntesEventos': 'NO EXITOSO'
                            })
                            
                            actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Observaciones',
                                                            valor_xml=truncar_observacion(obs),
                                                            valor_aprobado=None,
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                            
                            actualizar_estado_comparativa_nc(cx, nit, reg_id, 'NO EXITOSO')
                            cnt_nc += 1
                            continue
                    else:
                        # Asignamos fecha retoma inicial si no existe
                        cx.cursor().execute("UPDATE [CxP].[DocumentsProcessing] SET [Fecha_de_retoma_antes_de_contabilizacion]=? WHERE [ID]=?", (now.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3], reg_id))

                # Regla 2: Validaciones Tributarias (Emisor/Receptor)
                validaciones = {
                    'Nombre Emisor': {'val': r.get('nombre_emisor'), 'check': campo_con_valor},
                    'NIT Emisor': {'val': r.get('nit_emisor_o_nit_del_proveedor'), 'check': campo_con_valor},
                    'Fecha emision del documento': {'val': r.get('fecha_de_emision_documento'), 'check': campo_con_valor},
                    'Nombre Receptor': {'val': r.get('nombre_del_adquiriente'), 'check': validar_nombre_receptor},
                    'Nit Receptor': {'val': r.get('nit_del_adquiriente'), 'check': validar_nit_receptor},
                    'Tipo Persona Receptor': {'val': r.get('tipo_persona'), 'check': validar_tipo_persona},
                    'DigitoVerificacion Receptor': {'val': r.get('digito_de_verificacion'), 'check': validar_digito_verificacion},
                    'TaxLevelCode Receptor': {'val': r.get('responsabilidad_tributaria_adquiriente'), 'check': validar_tax_level_code}
                }
                
                # Actualizamos tabla comparativa con resultados
                for k, v in validaciones.items():
                    actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item=k,
                                                            valor_xml=safe_str(v['val']),
                                                            valor_aprobado='SI' if v['check'](v['val']) else 'NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                
                tiene_estado_final = True if safe_str(r.get('ResultadoFinalAntesEventos')) else False
                
                # Regla 3: Validaciones de Tipo Nota Credito y Referencia
                if not tiene_estado_final:
                    tipo_nc = safe_str(r.get('tipo_de_nota_credito', ''))
                    
                    # Si es Tipo 20 (NC sin referencia o devolucion total/parcial especifica)
                    if tipo_nc == '20':
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Tipo de nota credito',
                                                            valor_xml=r.get('Tipo_de_nota_cred_deb'),
                                                            valor_aprobado='SI' if campo_con_valor(r.get('Tipo_de_nota_cred_deb')) else 'NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                        
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Codigo CUFE de la factura',
                                                            valor_xml=r.get('cufeuuid', ''),
                                                            valor_aprobado='SI' if campo_con_valor(r.get('cufeuuid')) else 'NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                        
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Cude de la Nota Credito',
                                                            valor_xml=r.get('cufe_fe', ''),
                                                            valor_aprobado='SI' if campo_con_valor(r.get('cufe_fe')) else 'NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                    else:
                        # Si no es tipo 20 y no tiene referencia, es Novedad
                        obs = f"Nota credito sin referencia, {obs_anterior}"
                        actualizar_bd_cxp(cx, reg_id, {'EstadoFinalFase_4': 'VALIDACION DATOS DE FACTURACION: Exitoso', 'ObservacionesFase_4': truncar_observacion(obs), 'ResultadoFinalAntesEventos': 'CON NOVEDAD'})
                        
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Tipo de nota credito',
                                                            valor_xml=r.get('Tipo_de_nota_cred_deb'),
                                                            valor_aprobado='NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                        
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Observaciones',
                                                            valor_xml=truncar_observacion(obs),
                                                            valor_aprobado='NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)
                        
                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                            nombre_item='Cude de la Nota Credito',
                                                            valor_xml=r.get('cufe_fe', ''),
                                                            valor_aprobado='NO',
                                                            valor_factura=None,
                                                            nombre_proveedor=None,
                                                            fecha_retoma=None)

                        actualizar_estado_comparativa_nc(cx, nit, reg_id, 'CON NOVEDAD')
                        cnt_nc += 1
                        continue
                        
                # Regla 4: Cruce contra Facturas (Match de Referencia y Monto)
                condicion = (df_fv_final['numero_de_factura'] == r.get('PrefijoYNumero')) & (df_fv_final['nit_emisor_o_nit_del_proveedor'] == nit)
                resultado = df_fv_final[condicion]
                
                if not resultado.empty:
                    print(f"[SUCCESS] Encontrada FV para NC {r.get('numero_de_factura')}")
                    
                    # Marcamos referencia encontrada
                    actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                    nombre_item='Referencia',
                                                    valor_xml=r.get('PrefijoYNumero'),
                                                    valor_aprobado='SI',
                                                    valor_factura=r.get('numero_de_factura'),
                                                    nombre_proveedor=None,
                                                    fecha_retoma=None)
                    
                    fv_encontrada = resultado.iloc[0]
                    
                    # Normalizacion y comparacion de montos (con tolerancia 0.01)
                    v_nc = normalizar_decimal(r.get('valor_a_pagar_nc'))
                    v_fv = normalizar_decimal(fv_encontrada['valor_a_pagar'])
                    
                    es_coincidencia = abs(v_nc - v_fv) < 0.01
                    
                    if es_coincidencia:
                        obs = f"Nota credito con referencia no encontrada, {obs_anterior}" # Nota: Texto heredado de logica original
                        actualizar_bd_cxp(cx, fv_encontrada['ID'], {'EstadoFinalFase_4': 'VALIDACION DATOS DE FACTURACION: Exitoso','ResultadoFinalAntesEventos': 'ENCONTRADO','NotaCreditoReferenciada':f"{r.get('Numero_de_nota_credito')}"})

                        actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                    nombre_item='LineExtensionAmount',
                                                    valor_xml=r.get('valor_a_pagar_nc'),
                                                    valor_aprobado='SI',
                                                    valor_factura=r.get('valor_a_pagar'),
                                                    nombre_proveedor=None,
                                                    fecha_retoma=None)
                        
                        actualizar_estado_comparativa_nc(cx, nit, reg_id, 'ENCONTRADO')
                    
                else:
                    # No se encontro factura cruce
                    obs = f"Nota credito con referencia no encontrada, {obs_anterior}"
                    actualizar_bd_cxp(cx, reg_id, {'EstadoFinalFase_4': 'VALIDACION DATOS DE FACTURACION: Exitoso', 'ObservacionesFase_4': truncar_observacion(obs), 'ResultadoFinalAntesEventos': 'CON NOVEDAD'})
                    
                    actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                    nombre_item='Referencia',
                                                    valor_xml=r.get('PrefijoYNumero'),
                                                    valor_aprobado='NO',
                                                    valor_factura=r.get('numero_de_factura'),
                                                    nombre_proveedor=None,
                                                    fecha_retoma=None)
                    
                    actualizar_items_comparativa_nc(reg_id, cx, nit,
                                                    nombre_item='Observaciones',
                                                    valor_xml=obs,
                                                    valor_aprobado=None,
                                                    valor_factura=None,
                                                    nombre_proveedor=None,
                                                    fecha_retoma=None)
                    
                    actualizar_estado_comparativa_nc(cx, nit, reg_id, 'CON NOVEDAD')
                    cnt_nc += 1
                    continue
                
            # Generacion de Reporte Excel al final de procesar NCs
            try:
                generar_reporte_retorno_nc(cx, cfg['RutaBaseReporteNC'], cfg['NombreReporteNC'])
            except Exception as e:
                error_msg = traceback.format_exc() 
                SetVar("vGblStrDetalleError", error_msg)
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                print(f"[ERROR] Critico generando reporte: {error_msg}")
            
            # -----------------------------------------------------------------
            # FASE 2: PROCESAMIENTO DE NOTAS DEBITO (ND)
            # -----------------------------------------------------------------
            df_nd = pd.read_sql("SELECT * FROM [CxP].[DocumentsProcessing] WHERE [tipo_de_documento]='ND' AND ([ResultadoFinalAntesEventos] IS NULL OR [ResultadoFinalAntesEventos] NOT IN ('Exitoso')) ORDER BY [executionDate] DESC", cx)
            print(f"[INFO] Procesando {len(df_nd)} Notas Debito (ND)...")
            
            try:
                poblar_inicial_comparativa_nd(cx, df_nd)
            except Exception as e:
                error_msg = traceback.format_exc() 
                SetVar("vGblStrDetalleError", error_msg)
                SetVar("vGblStrSystemError", "ErrorHU4_4.1")
                SetVar("vLocStrResultadoSP", "False")
                print(f"[ERROR] Critico al poblar ND: {error_msg}")
                raise e
            
            cnt_nd = 0
            for idx, r in df_nd.iterrows():
                try:
                    reg_id = safe_str(r.get('ID', ''))
                    nit = safe_str(r.get('nit_emisor_o_nit_del_proveedor', ''))
                    num_nd = safe_str(r.get('numero_de_nota_debito', ''))
                    
                    nombre_prov = r.get('nombre_emisor')
                    
                    print(f"[INFO] Analizando ND: {num_nd} (ID: {reg_id})")

                    validaciones = {
                        'Nombre Emisor': {'val': r.get('nombre_emisor'), 'check': campo_con_valor},
                        'NIT Emisor': {'val': r.get('nit_emisor_o_nit_del_proveedor'), 'check': campo_con_valor},
                        'Fecha emision del documento': {'val': r.get('fecha_de_emision_documento'), 'check': campo_con_valor},
                        'Nombre Receptor': {'val': r.get('nombre_del_adquiriente'), 'check': validar_nombre_receptor},
                        'Nit Receptor': {'val': r.get('nit_del_adquiriente'), 'check': validar_nit_receptor},
                        'Tipo Persona Receptor': {'val': r.get('tipo_persona'), 'check': validar_tipo_persona},
                        'DigitoVerificacion Receptor': {'val': r.get('digito_de_verificacion'), 'check': validar_digito_verificacion},
                        'TaxLevelCode Receptor': {'val': r.get('responsabilidad_tributaria_adquiriente'), 'check': validar_tax_level_code}
                    }

                    # Validaciones Tributarias
                    for k, v in validaciones.items():
                        actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item=k,
                                                            valor_xml=safe_str(v['val']),
                                                            valor_aprobado='SI' if v['check'](v['val']) else 'NO',
                                                            nombre_proveedor=None)

                    # Actualizacion Exitosa en BD Principal
                    actualizar_bd_cxp(cx, reg_id, {'EstadoFinalFase_4': 'VALIDACION DATOS DE FACTURACION: Exitoso', 'ResultadoFinalAntesEventos': 'EXITOSO'})
                    
                    # Actualizacion de campos informativos finales en comparativa
                    actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item='LineExtensionAmount',
                                                            valor_xml=safe_str(r.get('valor_a_pagar')),
                                                            valor_aprobado=None,
                                                            nombre_proveedor=None)
                    
                    actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item='Tipo de nota debito',
                                                            valor_xml=safe_str(r.get('Tipo_de_nota_cred_deb')),
                                                            valor_aprobado=None,
                                                            nombre_proveedor=None)
                    
                    actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item='Referencia',
                                                            valor_xml=safe_str(r.get('PrefijoYNumero')),
                                                            valor_aprobado=None,
                                                            nombre_proveedor=None)
                    
                    actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item='Codigo CUFE de la factura',
                                                            valor_xml=safe_str(r.get('cufeuuid')),
                                                            valor_aprobado=None,
                                                            nombre_proveedor=None)
                    
                    actualizar_items_comparativa_nd(reg_id, cx, nit,
                                                            nombre_item='Cude de la Nota Debito',
                                                            valor_xml=safe_str(r.get('cufe_fe')),
                                                            valor_aprobado=None,
                                                            nombre_proveedor=None)
                    cnt_nd += 1
                except Exception as e:
                    print(f"[ERROR] Procesando ND {r.get('ID')}: {e}")
                    print(traceback.format_exc())
                    cnt_nd += 1

            print(f"[SUCCESS] Procesamiento completado. NC procesadas: {cnt_nc}, ND procesadas: {cnt_nd}")

            SetVar("vLocStrResultadoSP", "True")
            SetVar("vLocStrResumenSP", f"Procesamiento Finalizado. NC: {cnt_nc}, ND: {cnt_nd}")

    except Exception as e:
        print("")
        print("=" * 80)
        print("[ERROR] FALLO CRITICO EN HU4.2")
        print(f"Mensaje: {str(e)}")
        print(traceback.format_exc())
        print("=" * 80)

        SetVar("vGblStrDetalleError", traceback.format_exc())
        SetVar("vGblStrSystemError", "ErrorHU4_4.1")
        SetVar("vLocStrResultadoSP", "False")
        raise e